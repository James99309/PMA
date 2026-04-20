"""add pma_xlsx_style skill to cli_skills

Revision ID: xlsx_skill_system_20260420
Revises: reply_type_20260420
Create Date: 2026-04-20 23:00:00.000000

"""
from alembic import op
import sqlalchemy as sa

revision = 'xlsx_skill_system_20260420'
down_revision = 'reply_type_20260420'
branch_labels = None
depends_on = None

XLSX_SKILL_BODY = r'''
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ALT_ROW_BG = "EBF3FB"
BORDER_CLR = "B8CCE4"

COMPANY_NAMES = {
    "SP8D": "和源通信（上海）股份有限公司",
    "OVS":  "Evertac Solutions",
}

def _thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def init_workbook(db_type="SP8D", sheet_title="Sheet1"):
    """初始化 PMA 样式 Workbook，返回 (wb, ws)"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws

def write_title(ws, title, row=1, col_span=8):
    """写标题行（深蓝色大字，跨列居中），返回下一可用行号"""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(name="Microsoft YaHei", bold=True, size=14, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=col_span
        )
    return row + 1

def write_header_row(ws, headers, row=2):
    """写表头行（深蓝底白字），返回下一可用行号"""
    fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill      = fill
        c.font      = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin_border()
    ws.row_dimensions[row].height = 22
    return row + 1

def write_data_row(ws, row, values, alt=False):
    """写数据行，alt=True 时隔行浅蓝背景，返回下一可用行号"""
    fill = (
        PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
        if alt else None
    )
    font = Font(name="Microsoft YaHei", size=10)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border    = _thin_border()
        if fill:
            c.fill = fill
    return row + 1

def write_subheader_row(ws, label, row, col_span=8):
    """写分组小标题行（中蓝色背景），返回下一可用行号"""
    fill = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill      = fill
    cell.font      = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    cell.alignment = Alignment(vertical="center")
    cell.border    = _thin_border()
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=col_span
        )
    return row + 1

def auto_column_width(ws, min_w=8, max_w=40):
    """根据内容自动调整列宽（中文字符按2倍计算）"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                    max_len = max(max_len, length)
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def freeze_header(ws, freeze_row=2):
    """冻结表头（freeze_row 行以上）"""
    ws.freeze_panes = f"A{freeze_row + 1}"
'''

XLSX_SKILL_DESCRIPTION = """[PMA Excel 样式库]
生成 Excel 时用 export_to_excel 工具，python_code 参数使用以下函数：
  from pma_xlsx_style import (init_workbook, write_title, write_header_row, write_data_row,
      write_subheader_row, auto_column_width, freeze_header, COMPANY_NAMES)
  wb, ws = init_workbook(db_type=DB_TYPE, sheet_title="Sheet名")   # DB_TYPE 从 [运行环境] 获取
  row = write_title(ws, "标题", row=1, col_span=列数)               # 返回下一行号
  row = write_header_row(ws, ["列1","列2",...], row=row)            # 深蓝底白字表头
  for i, item in enumerate(data):
      row = write_data_row(ws, row, [v1, v2,...], alt=(i%2==1))    # 隔行浅蓝背景
  row = write_subheader_row(ws, "分组标题", row=row, col_span=列数)  # 中蓝分组行（可选）
  auto_column_width(ws)   # 自动列宽（中文按2倍）
  freeze_header(ws)       # 冻结表头
  wb.save("__OUTPUT__")   # 必须以此结尾
多 Sheet：ws2 = wb.create_sheet("明细") 后对 ws2 调用同样函数。"""


def upgrade():
    conn = op.get_bind()

    existing = conn.execute(
        sa.text("SELECT id FROM cli_skills WHERE name='pma_xlsx_style'")
    ).scalar()
    if not existing:
        conn.execute(sa.text("""
            INSERT INTO cli_skills
                (name, title, description, parameters, queries, output_format,
                 skill_type, skill_body, scope, is_active, created_at, updated_at)
            VALUES
                ('pma_xlsx_style',
                 'PMA Excel 样式库',
                 :desc,
                 '[]'::jsonb,
                 '[]'::jsonb,
                 NULL,
                 'xlsx',
                 :body,
                 'global',
                 true,
                 CURRENT_TIMESTAMP,
                 CURRENT_TIMESTAMP)
        """), {'body': XLSX_SKILL_BODY, 'desc': XLSX_SKILL_DESCRIPTION})

        skill_id = conn.execute(
            sa.text("SELECT id FROM cli_skills WHERE name='pma_xlsx_style'")
        ).scalar()
        conn.execute(sa.text("""
            INSERT INTO cli_skill_versions (skill_id, version, skill_body, change_note, created_at)
            VALUES (:sid, 1, :body, '初始版本 v1.0 — openpyxl PMA 品牌样式库', CURRENT_TIMESTAMP)
        """), {'sid': skill_id, 'body': XLSX_SKILL_BODY})


def downgrade():
    conn = op.get_bind()
    skill_id = conn.execute(
        sa.text("SELECT id FROM cli_skills WHERE name='pma_xlsx_style'")
    ).scalar()
    if skill_id:
        conn.execute(sa.text("DELETE FROM cli_skill_versions WHERE skill_id = :sid"), {'sid': skill_id})
        conn.execute(sa.text("DELETE FROM cli_skills WHERE id = :sid"), {'sid': skill_id})
