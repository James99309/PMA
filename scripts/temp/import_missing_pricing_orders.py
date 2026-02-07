#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel导入缺失的批价单/结算单数据到NAS数据库
生成SQL INSERT语句文件，通过SSH在NAS上执行

数据源: /Users/nijie/Downloads/线上缺少项目明细.xlsx
目标: pma_synology 数据库 (NAS)
"""
import sys
import os
from datetime import datetime

# 路径修正
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

PROJECT_ROOT = get_project_root()

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

# ============================================================
# ID 映射表（全部来自计划中确认的数据）
# ============================================================

# 用户映射：销售归属 → user_id
USER_MAP = {
    '范敬': 16,
    '杨俊杰': 14,
    '李华伟': 15,
    '周裔锦': 17,
    '郭小会': 13,
}

# 结算方映射：Excel名称 → company_id
DISTRIBUTOR_MAP = {
    '上海淳泊': 282,
    '上海瑞康': 496,
}

# 提货方映射：Excel名称 → company_id
DEALER_MAP = {
    '敦力（南京）科技': 21,
    '上海福玛': 288,
    '上海瀚网': 497,
    '浙江航博智能': 319,
    '成都天皓': 340,
    '广州宇洪智能技术': 112,
    '重庆大鹏鸟科技': 507,
    '合肥兴和通讯': 508,
    '南通顺通': 632,
    '中铁二局集团有限公司': 246,
    '北京联航讯达': 11,
}

# 项目类型映射
FLOW_TYPE_MAP = {
    '渠道跟进': 'channel_follow',
    '销售重点': 'sales_focus',
}

# 项目 → (project_id, quotation_id) 完整映射
PROJECT_MAP = {
    '荡口古镇太师府酒店建设项目智能化工程': (154, 353),
    '上海名人苑': (17, 505),
    '上海嘉定集成电路研发中心增补': (136, 536),
    '上海市松江区巨人科技园B楼项目弱电智能化投标项目': (129, 349),
    '上海静安太保家园': (130, 350),
    '东莞市博物馆项目': (679, 735),
    '中欧轨道智能交通国际研创基地启动区项目智城大厦': (573, 655),
    '九星城项目': (187, 546),
    '华虹张江工厂配套用房': (61, 514),
    '台泥公亮大楼酒店': (416, 719),
    '合肥机场东区国际货站扩建工程民航专业工程': (99, 427),
    '太保家园成都国际颐养社区项目': (71, 425),
    '学院路科技园东升园（H地块）项目': (529, 602),
    '张江人工智能岛': (135, 535),
    '张江创新药基地B03C-02B03K-03': (102, 530),
    '御桥12C-18地块': (18, 506),
    '招商282 &01-10': (784, 808),
    '招商银行深圳总部大厦': (132, 571),
    '招商银行深圳总部大厦（增补）': (27, 560),
    '无锡奥林匹克体育产业中心二期项目': (127, 348),
    '昆山高新区前进路南侧、江浦路西侧商住用房新建项目': (666, 729),
    '杭州博览会议中心二期': (5, 453),
    '松江海螺水泥': (517, 595),
    '松江站服务中心': (278, 555),
    '海宁经开智创园': (36, 455),
    '海宁经开智创园增补': (605, 679),
    '海灏生物创新港': (26, 559),
    '深圳市星河智善科技有限公司集采-东湾复建': (599, 667),
    '美团科技中心': (518, 596),
    '舜宇12英寸透明衬底晶圆AR眼镜微纳光学产品': (39, 509),
    '重庆芯联微电子': (415, 713),
    '静安假日酒店': (6, 454),
    '静安工人文化宫': (163, 463),
    '马来西亚万国数据中心': (156, 537),
    '龙舟计划二期--增补': (9, 337),
}

# 需更新为 signed 的项目（当前为 quoted 或 awarded）
PROJECTS_TO_SIGN = [679, 573, 529, 135, 102, 784, 666, 517, 415, 163, 99]


def escape_sql(val):
    """转义SQL字符串中的单引号"""
    if val is None:
        return 'NULL'
    s = str(val).replace("'", "''")
    return f"'{s}'"


def read_excel():
    """读取Excel，按项目分组返回产品明细"""
    excel_path = '/Users/nijie/Downloads/线上缺少项目明细.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    projects = {}  # {project_name: {'rows': [...], 'meta': {...}}}

    for row_idx in range(2, ws.max_row + 1):
        project_name = ws.cell(row_idx, 3).value  # C: 项目名称
        if not project_name:
            continue
        project_name = str(project_name).strip()

        sales_person = str(ws.cell(row_idx, 4).value or '').strip()  # D: 销售归属
        project_type = str(ws.cell(row_idx, 5).value or '').strip()  # E: 项目类型
        distributor = str(ws.cell(row_idx, 7).value or '').strip()   # G: 结算方
        dealer = str(ws.cell(row_idx, 8).value or '').strip()        # H: 提货方

        settlement_discount = ws.cell(row_idx, 9).value   # I: 结算折扣
        pickup_discount = ws.cell(row_idx, 10).value       # J: 提货折扣
        mn = ws.cell(row_idx, 11).value                    # K: MN号
        product_name = ws.cell(row_idx, 12).value          # L: 产品名称
        product_model = ws.cell(row_idx, 13).value         # M: 产品型号
        product_spec = ws.cell(row_idx, 14).value          # N: 产品规格
        brand = ws.cell(row_idx, 15).value                 # O: 品牌
        unit = ws.cell(row_idx, 16).value                  # P: 单位
        quantity = ws.cell(row_idx, 18).value               # R: 数量
        retail_price = ws.cell(row_idx, 19).value           # S: 零售单价
        settlement_unit_price = ws.cell(row_idx, 21).value  # U: 结算单价
        settlement_total = ws.cell(row_idx, 22).value       # V: 结算批价额
        pickup_unit_price = ws.cell(row_idx, 23).value      # W: 提货单价
        pickup_total = ws.cell(row_idx, 24).value            # X: 提货批价额

        if project_name not in projects:
            projects[project_name] = {
                'meta': {
                    'sales_person': sales_person,
                    'project_type': project_type,
                    'distributor': distributor,
                    'dealer': dealer,
                },
                'rows': []
            }

        projects[project_name]['rows'].append({
            'mn': str(mn).strip() if mn else None,
            'product_name': str(product_name).strip() if product_name else '',
            'product_model': str(product_model).strip() if product_model else None,
            'product_spec': str(product_spec).strip() if product_spec else None,
            'brand': str(brand).strip() if brand else None,
            'unit': str(unit).strip() if unit else None,
            'quantity': int(quantity) if quantity else 0,
            'retail_price': float(retail_price) if retail_price else 0.0,
            'pickup_discount': float(pickup_discount) if pickup_discount else 1.0,
            'pickup_unit_price': float(pickup_unit_price) if pickup_unit_price else 0.0,
            'pickup_total': float(pickup_total) if pickup_total else 0.0,
            'settlement_discount': float(settlement_discount) if settlement_discount else 1.0,
            'settlement_unit_price': float(settlement_unit_price) if settlement_unit_price else 0.0,
            'settlement_total': float(settlement_total) if settlement_total else 0.0,
        })

    wb.close()
    return projects


def generate_sql(projects):
    """根据Excel数据生成完整的SQL INSERT语句"""
    sql_lines = []
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sql_lines.append('-- ============================================================')
    sql_lines.append('-- 批价单/结算单批量导入 SQL')
    sql_lines.append(f'-- 生成时间: {now_str}')
    sql_lines.append(f'-- 项目数: {len(projects)}, 总明细数: {sum(len(p["rows"]) for p in projects.values())}')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')
    sql_lines.append('BEGIN;')
    sql_lines.append('')

    po_seq = 1  # PO202602-001 ~ 035
    total_details = 0
    errors = []

    # 按项目名称排序处理
    for project_name in sorted(projects.keys()):
        proj_data = projects[project_name]
        meta = proj_data['meta']
        rows = proj_data['rows']

        # 查找映射
        if project_name not in PROJECT_MAP:
            errors.append(f"项目未找到映射: {project_name}")
            continue

        project_id, quotation_id = PROJECT_MAP[project_name]
        user_id = USER_MAP.get(meta['sales_person'])
        if not user_id:
            errors.append(f"销售人员未找到: {meta['sales_person']} (项目: {project_name})")
            continue

        flow_type = FLOW_TYPE_MAP.get(meta['project_type'])
        if not flow_type:
            errors.append(f"项目类型未找到: {meta['project_type']} (项目: {project_name})")
            continue

        dealer_id = DEALER_MAP.get(meta['dealer']) if meta['dealer'] else None
        distributor_id = DISTRIBUTOR_MAP.get(meta['distributor']) if meta['distributor'] else None

        po_number = f'PO202602-{po_seq:03d}'
        so_number = f'SO202602-{po_seq:03d}'

        # 计算汇总金额
        pricing_total = sum(r['pickup_total'] for r in rows)
        settlement_total = sum(r['settlement_total'] for r in rows)
        market_total = sum(r['retail_price'] * r['quantity'] for r in rows)

        pricing_discount_rate = pricing_total / market_total if market_total > 0 else 1.0
        settlement_discount_rate = settlement_total / market_total if market_total > 0 else 1.0

        sql_lines.append(f'-- ========== 项目 #{po_seq}: {project_name} (project_id={project_id}) ==========')
        sql_lines.append(f'-- 明细数: {len(rows)}, 提货总额: {pricing_total:.2f}, 结算总额: {settlement_total:.2f}')
        sql_lines.append('')

        # 1. INSERT pricing_orders
        sql_lines.append(f"INSERT INTO pricing_orders (order_number, project_id, quotation_id, dealer_id, distributor_id, "
                         f"is_direct_contract, is_factory_pickup, approval_flow_type, status, current_approval_step, "
                         f"pricing_total_amount, pricing_total_discount_rate, settlement_total_amount, settlement_total_discount_rate, "
                         f"approved_by, approved_at, created_by, created_at, updated_at, currency) VALUES (")
        sql_lines.append(f"  {escape_sql(po_number)}, {project_id}, {quotation_id}, "
                         f"{'NULL' if dealer_id is None else dealer_id}, "
                         f"{'NULL' if distributor_id is None else distributor_id}, "
                         f"false, false, {escape_sql(flow_type)}, 'approved', 1, "
                         f"{pricing_total:.2f}, {pricing_discount_rate:.6f}, "
                         f"{settlement_total:.2f}, {settlement_discount_rate:.6f}, "
                         f"5, NOW(), {user_id}, NOW(), NOW(), 'CNY'")
        sql_lines.append(');')
        sql_lines.append('')

        # 2. INSERT pricing_order_details (批量)
        sql_lines.append(f"-- 批价单明细 ({len(rows)} 条)")
        sql_lines.append(f"INSERT INTO pricing_order_details (pricing_order_id, product_name, product_model, product_desc, "
                         f"brand, unit, product_mn, market_price, unit_price, quantity, discount_rate, total_price, "
                         f"source_type, currency) VALUES")

        detail_values = []
        for r in rows:
            val = (f"  (currval('pricing_orders_id_seq'), {escape_sql(r['product_name'])}, "
                   f"{escape_sql(r['product_model'])}, {escape_sql(r['product_spec'])}, "
                   f"{escape_sql(r['brand'])}, {escape_sql(r['unit'])}, {escape_sql(r['mn'])}, "
                   f"{r['retail_price']:.2f}, {r['pickup_unit_price']:.2f}, {r['quantity']}, "
                   f"{r['pickup_discount']:.6f}, {r['pickup_total']:.2f}, 'manual', 'CNY')")
            detail_values.append(val)

        sql_lines.append(',\n'.join(detail_values) + ';')
        sql_lines.append('')

        # 3. INSERT settlement_orders
        sql_lines.append(f"INSERT INTO settlement_orders (order_number, pricing_order_id, project_id, quotation_id, "
                         f"distributor_id, dealer_id, is_direct_contract, is_factory_pickup, "
                         f"total_amount, total_discount_rate, status, settlement_status, "
                         f"approved_by, approved_at, created_by, created_at, updated_at, currency) VALUES (")
        sql_lines.append(f"  {escape_sql(so_number)}, currval('pricing_orders_id_seq'), {project_id}, {quotation_id}, "
                         f"{'NULL' if distributor_id is None else distributor_id}, "
                         f"{'NULL' if dealer_id is None else dealer_id}, "
                         f"false, false, "
                         f"{settlement_total:.2f}, {settlement_discount_rate:.6f}, "
                         f"'approved', 'pending', "
                         f"5, NOW(), {user_id}, NOW(), NOW(), 'CNY'")
        sql_lines.append(');')
        sql_lines.append('')

        # 4. INSERT settlement_order_details
        # 需要引用 pricing_order_details 的 id，使用子查询
        sql_lines.append(f"-- 结算单明细 ({len(rows)} 条)")
        sql_lines.append(f"-- 使用子查询关联批价单明细ID")
        sql_lines.append(f"INSERT INTO settlement_order_details (pricing_order_id, settlement_order_id, "
                         f"product_name, product_model, product_desc, brand, unit, product_mn, "
                         f"market_price, unit_price, quantity, discount_rate, total_price, "
                         f"pricing_detail_id, settlement_status, currency) ")

        # 用 WITH 子句获取刚插入的 pricing_order_details IDs
        sql_lines.append(f"SELECT currval('pricing_orders_id_seq'), currval('settlement_orders_id_seq'), "
                         f"pod.product_name, pod.product_model, pod.product_desc, pod.brand, pod.unit, pod.product_mn, "
                         f"pod.market_price, ")
        # 结算单价和折扣率可能不同于批价单，需要用 CASE 或预计算值
        # 简化方案：逐条插入结算明细
        sql_lines.pop()  # 去掉上面的SELECT行
        sql_lines.pop()  # 去掉INSERT行

        # 改为逐条INSERT结算明细
        sql_lines.append(f"-- 结算单明细 ({len(rows)} 条)")
        sql_lines.append(f"INSERT INTO settlement_order_details (pricing_order_id, settlement_order_id, "
                         f"product_name, product_model, product_desc, brand, unit, product_mn, "
                         f"market_price, unit_price, quantity, discount_rate, total_price, "
                         f"pricing_detail_id, settlement_status, currency) VALUES")

        sod_values = []
        for i, r in enumerate(rows):
            # pricing_detail_id 需要引用对应的 pricing_order_details ID
            # 使用子查询：根据刚插入的 pricing_order 的明细，按 id 排序取第 i+1 条
            pricing_detail_subq = (f"(SELECT id FROM pricing_order_details WHERE pricing_order_id = currval('pricing_orders_id_seq') "
                                   f"ORDER BY id OFFSET {i} LIMIT 1)")

            val = (f"  (currval('pricing_orders_id_seq'), currval('settlement_orders_id_seq'), "
                   f"{escape_sql(r['product_name'])}, {escape_sql(r['product_model'])}, "
                   f"{escape_sql(r['product_spec'])}, {escape_sql(r['brand'])}, "
                   f"{escape_sql(r['unit'])}, {escape_sql(r['mn'])}, "
                   f"{r['retail_price']:.2f}, {r['settlement_unit_price']:.2f}, {r['quantity']}, "
                   f"{r['settlement_discount']:.6f}, {r['settlement_total']:.2f}, "
                   f"{pricing_detail_subq}, 'draft', 'CNY')")
            sod_values.append(val)

        sql_lines.append(',\n'.join(sod_values) + ';')
        sql_lines.append('')

        # 5. INSERT approval record
        sql_lines.append(f"INSERT INTO pricing_order_approval_records (pricing_order_id, step_order, step_name, "
                         f"approver_role, approver_id, action, comment, approved_at, is_fast_approval) VALUES (")
        sql_lines.append(f"  currval('pricing_orders_id_seq'), 1, '审批', 'approver', 5, 'approve', "
                         f"'线下已审批，系统批量导入补录 (2025年)', NOW(), false")
        sql_lines.append(');')
        sql_lines.append('')

        total_details += len(rows)
        po_seq += 1

    # 6. 更新项目阶段
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- 更新项目阶段为 signed 并锁定')
    sql_lines.append('-- ============================================================')
    project_ids_str = ', '.join(str(pid) for pid in PROJECTS_TO_SIGN)
    sql_lines.append(f"UPDATE projects SET")
    sql_lines.append(f"  current_stage = 'signed',")
    sql_lines.append(f"  is_locked = true,")
    sql_lines.append(f"  locked_reason = '项目已签约，自动锁定',")
    sql_lines.append(f"  locked_by = 5,")
    sql_lines.append(f"  locked_at = NOW()")
    sql_lines.append(f"WHERE id IN ({project_ids_str})")
    sql_lines.append(f"AND current_stage IN ('quoted', 'awarded');")
    sql_lines.append('')

    # 7. 更新序列值（确保下次自增不冲突）
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- 更新序列值')
    sql_lines.append('-- ============================================================')
    sql_lines.append("SELECT setval('pricing_orders_id_seq', (SELECT MAX(id) FROM pricing_orders));")
    sql_lines.append("SELECT setval('pricing_order_details_id_seq', (SELECT MAX(id) FROM pricing_order_details));")
    sql_lines.append("SELECT setval('settlement_orders_id_seq', (SELECT MAX(id) FROM settlement_orders));")
    sql_lines.append("SELECT setval('settlement_order_details_id_seq', (SELECT MAX(id) FROM settlement_order_details));")
    sql_lines.append("SELECT setval('pricing_order_approval_records_id_seq', (SELECT MAX(id) FROM pricing_order_approval_records));")
    sql_lines.append('')

    sql_lines.append('COMMIT;')
    sql_lines.append('')

    # 验证查询
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- 验证查询')
    sql_lines.append('-- ============================================================')
    sql_lines.append("SELECT '批价单数量' as metric, COUNT(*) as value FROM pricing_orders WHERE order_number LIKE 'PO202602%';")
    sql_lines.append("SELECT '批价单明细数' as metric, COUNT(*) as value FROM pricing_order_details pd JOIN pricing_orders po ON pd.pricing_order_id = po.id WHERE po.order_number LIKE 'PO202602%';")
    sql_lines.append("SELECT '结算单数量' as metric, COUNT(*) as value FROM settlement_orders WHERE order_number LIKE 'SO202602%';")
    sql_lines.append("SELECT '结算单明细数' as metric, COUNT(*) as value FROM settlement_order_details sd JOIN settlement_orders so ON sd.settlement_order_id = so.id WHERE so.order_number LIKE 'SO202602%';")
    sql_lines.append("SELECT '审批记录数' as metric, COUNT(*) as value FROM pricing_order_approval_records WHERE comment LIKE '%批量导入补录%';")
    sql_lines.append("SELECT id, project_name, current_stage, is_locked FROM projects WHERE id IN ({});".format(project_ids_str))

    if errors:
        print("\n⚠️ 映射错误:")
        for e in errors:
            print(f"  - {e}")
        return None

    return '\n'.join(sql_lines)


def main():
    print("读取 Excel 数据...")
    projects = read_excel()
    print(f"读取完成: {len(projects)} 个项目, {sum(len(p['rows']) for p in projects.values())} 条明细")

    # 验证所有映射
    print("\n验证映射...")
    ok = True
    for pname in projects:
        if pname not in PROJECT_MAP:
            print(f"  ❌ 项目未映射: {pname}")
            ok = False
        meta = projects[pname]['meta']
        if meta['sales_person'] not in USER_MAP:
            print(f"  ❌ 销售未映射: {meta['sales_person']}")
            ok = False
        if meta['dealer'] and meta['dealer'] not in DEALER_MAP:
            print(f"  ❌ 提货方未映射: {meta['dealer']}")
            ok = False
        if meta['distributor'] and meta['distributor'] not in DISTRIBUTOR_MAP:
            print(f"  ❌ 结算方未映射: {meta['distributor']}")
            ok = False

    if not ok:
        print("\n存在未映射的数据，请检查后重试")
        sys.exit(1)

    print("  ✅ 所有映射验证通过")

    print("\n生成 SQL...")
    sql = generate_sql(projects)
    if not sql:
        print("SQL 生成失败")
        sys.exit(1)

    output_path = os.path.join(PROJECT_ROOT, 'scripts', 'temp', 'import_pricing_data.sql')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f"✅ SQL 已生成: {output_path}")
    print(f"   文件大小: {os.path.getsize(output_path) / 1024:.1f} KB")

    # 统计
    po_count = sql.count("INSERT INTO pricing_orders (")
    pod_count = sql.count("INSERT INTO pricing_order_details (")
    so_count = sql.count("INSERT INTO settlement_orders (")
    sod_count = sql.count("INSERT INTO settlement_order_details (")
    approval_count = sql.count("INSERT INTO pricing_order_approval_records (")
    print(f"\n📊 SQL 统计:")
    print(f"   批价单 INSERT: {po_count}")
    print(f"   批价单明细 INSERT: {pod_count}")
    print(f"   结算单 INSERT: {so_count}")
    print(f"   结算单明细 INSERT: {sod_count}")
    print(f"   审批记录 INSERT: {approval_count}")
    print(f"   项目阶段 UPDATE: {len(PROJECTS_TO_SIGN)} 个项目")


if __name__ == '__main__':
    main()
