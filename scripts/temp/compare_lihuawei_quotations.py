#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
李华伟批价单数据差异分析脚本
对比绩效采集表和PMA导出数据,找出金额差异原因
"""
import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 路径修正
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

PROJECT_ROOT = get_project_root()
sys.path.insert(0, PROJECT_ROOT)

# 文件路径
PERFORMANCE_FILE = os.path.expanduser('~/Downloads/2025年绩效数据采集表（12.31）.xlsx')
PMA_DATA_FILE = os.path.expanduser('~/Downloads/李华伟_批价单数据清单_2025年度_已清理.xlsx')
OUTPUT_FILE = os.path.join(PROJECT_ROOT, 'data/temp/李华伟_批价单差异分析_20260116.xlsx')

def read_performance_data():
    """从绩效表读取李华伟批价单数据"""
    print("📊 读取绩效采集表...")

    try:
        df = pd.read_excel(PERFORMANCE_FILE, sheet_name='订单业绩12.31', header=None)

        # 查找李华伟所在的行（第9行，索引为9）
        lihuawei_row = None
        for idx, row in df.iterrows():
            if len(row) > 1 and row.iloc[1] == '李华伟':
                lihuawei_row = idx
                print(f"✅ 找到李华伟数据行: {lihuawei_row}")

                # 提取数据 - 根据结构：
                # 列3 (索引3): 2025年度目标
                # 列5 (索引5): Q1完成金额
                # 列8 (索引8): Q2完成金额
                # 列11 (索引11): Q3完成金额
                # 列14 (索引14): Q4完成金额
                # 列17 (索引16): 全年完成金额

                performance_data = {
                    '2025年度目标': float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0,
                    'Q1完成': float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
                    'Q2完成': float(row.iloc[8]) if pd.notna(row.iloc[8]) else 0,
                    'Q3完成': float(row.iloc[11]) if pd.notna(row.iloc[11]) else 0,
                    'Q4完成': float(row.iloc[14]) if pd.notna(row.iloc[14]) else 0,
                    '全年完成': float(row.iloc[16]) if pd.notna(row.iloc[16]) else 0,
                }

                print(f"   - 2025年度目标: ¥{performance_data['2025年度目标']:,.2f}")
                print(f"   - Q1完成: ¥{performance_data['Q1完成']:,.2f}")
                print(f"   - Q2完成: ¥{performance_data['Q2完成']:,.2f}")
                print(f"   - Q3完成: ¥{performance_data['Q3完成']:,.2f}")
                print(f"   - Q4完成: ¥{performance_data['Q4完成']:,.2f}")
                print(f"   - 全年完成: ¥{performance_data['全年完成']:,.2f}")

                return performance_data

        if lihuawei_row is None:
            print("❌ 未找到李华伟数据")
            return None

    except Exception as e:
        print(f"❌ 读取绩效表错误: {e}")
        import traceback
        traceback.print_exc()
        return None

def read_pma_data():
    """从PMA导出数据读取35个项目的批价单明细"""
    print("📋 读取PMA导出数据...")

    try:
        # 跳过前2行（标题和空行）,使用第3行作为header
        df = pd.read_excel(PMA_DATA_FILE, sheet_name='去重合并结果', skiprows=2)

        print(f"✅ 读取到 {len(df)} 条项目记录")
        print(f"   列名: {df.columns.tolist()}")

        # 提取关键列：序号, 项目名称, 状态, 线上金额, 线下金额, 最终用金额
        projects = []
        total_amount = 0

        for idx, row in df.iterrows():
            # 跳过NaN行
            if pd.isna(row.iloc[0]):
                continue

            project_data = {
                '序号': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
                '项目名': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                '状态': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                '线上金额': float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0,
                '线下金额': float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0,
                '最终用金额': float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0,
            }

            projects.append(project_data)
            total_amount += project_data['最终用金额']

        print(f"   项目总数: {len(projects)}")
        print(f"   总金额: ¥{total_amount:,.2f}")

        return projects, df

    except Exception as e:
        print(f"❌ 读取PMA数据错误: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def analyze_differences(performance_data, pma_projects):
    """分析数据差异"""
    print("🔍 分析数据差异...")

    analysis = {
        'performance': performance_data,
        'pma_total': 0,
        'pma_projects': pma_projects,
        'projects_count': len(pma_projects)
    }

    # 计算绩效表总额（使用全年完成额）
    analysis['performance_total'] = performance_data.get('全年完成', 0)

    # 计算PMA总额
    for project in pma_projects:
        if isinstance(project.get('最终用金额'), (int, float)):
            analysis['pma_total'] += float(project['最终用金额'])

    # 计算差异
    analysis['difference'] = analysis['pma_total'] - analysis['performance_total']
    analysis['difference_rate'] = (
        (analysis['difference'] / analysis['performance_total'] * 100)
        if analysis['performance_total'] > 0 else 0
    )

    return analysis

def generate_report(analysis):
    """生成Excel差异分析报告"""
    print("📝 生成差异分析报告...")

    wb = Workbook()

    # 删除默认sheet
    wb.remove(wb.active)

    # Sheet1: 金额对比
    ws1 = wb.create_sheet('金额对比')
    generate_summary_sheet(ws1, analysis)

    # Sheet2: PMA项目清单
    ws2 = wb.create_sheet('PMA项目清单')
    generate_projects_sheet(ws2, analysis['pma_projects'])

    # Sheet3: 差异分析
    ws3 = wb.create_sheet('差异分析')
    generate_analysis_sheet(ws3, analysis)

    # 保存文件
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"✅ 报告已生成: {OUTPUT_FILE}")

def generate_summary_sheet(ws, analysis):
    """生成金额对比sheet"""
    # 标题
    ws['A1'] = '李华伟批价单金额对比'
    ws['A1'].font = Font(bold=True, size=14)

    # 数据
    headers = ['数据源', '总金额', '项目/月份数']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col).value = header
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # 绩效表数据
    ws['A4'] = '绩效采集表'
    ws['B4'] = analysis['performance_total']
    ws['C4'] = '年度'

    # PMA数据
    ws['A5'] = 'PMA导出数据'
    ws['B5'] = analysis['pma_total']
    ws['C5'] = analysis['projects_count']

    # 差异
    ws['A6'] = '差异'
    ws['B6'] = analysis['difference']
    ws['C6'] = f"{analysis['difference_rate']:.1f}%"
    ws['B6'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 格式化金额列
    for row in [4, 5, 6]:
        ws[f'B{row}'].number_format = '#,##0.00'

def generate_projects_sheet(ws, projects):
    """生成PMA项目清单sheet"""
    # 标题
    ws['A1'] = f'PMA导出的{len(projects)}个项目明细'
    ws['A1'].font = Font(bold=True, size=12)

    # 表头
    headers = ['序号', '项目名称', '状态', '线上金额', '线下金额', '最终用金额']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col).value = header
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

    # 数据
    total = 0
    for idx, project in enumerate(projects, 1):
        ws[f'A{idx+3}'] = project.get('序号', idx)
        ws[f'B{idx+3}'] = project.get('项目名', '')
        ws[f'C{idx+3}'] = project.get('状态', '')

        online = float(project.get('线上金额', 0))
        offline = float(project.get('线下金额', 0))
        final = float(project.get('最终用金额', 0))

        ws[f'D{idx+3}'] = online
        ws[f'D{idx+3}'].number_format = '#,##0.00'
        ws[f'E{idx+3}'] = offline
        ws[f'E{idx+3}'].number_format = '#,##0.00'
        ws[f'F{idx+3}'] = final
        ws[f'F{idx+3}'].number_format = '#,##0.00'

        total += final

    # 合计行
    last_row = len(projects) + 4
    ws[f'B{last_row}'] = '合计'
    ws[f'B{last_row}'].font = Font(bold=True)
    ws[f'F{last_row}'] = total
    ws[f'F{last_row}'].number_format = '#,##0.00'
    ws[f'F{last_row}'].font = Font(bold=True)
    ws[f'F{last_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def generate_analysis_sheet(ws, analysis):
    """生成差异分析sheet"""
    # 标题
    ws['A1'] = '差异原因分析'
    ws['A1'].font = Font(bold=True, size=12)

    # 基本统计
    ws['A3'] = '统计数据'
    ws['A3'].font = Font(bold=True)

    rows = [
        ('绩效表统计金额', f"¥{analysis['performance_total']:,.2f}"),
        ('PMA实际金额', f"¥{analysis['pma_total']:,.2f}"),
        ('差异总额', f"¥{analysis['difference']:,.2f}"),
        ('差异率', f"{analysis['difference_rate']:.2f}%"),
    ]

    for idx, (label, value) in enumerate(rows, 1):
        ws[f'A{idx+3}'] = label
        ws[f'B{idx+3}'] = value

    # 可能原因
    ws['A9'] = '可能的差异原因'
    ws['A9'].font = Font(bold=True)

    reasons = [
        '1. 绩效表统计时间与PMA导出时间不一致',
        '2. 绩效表可能只包含特定类型的批价单',
        '3. PMA中可能包含试验、样品或特殊订单',
        '4. 数据录入或统计方式存在差异',
        '5. 部分项目可能被划分到其他销售人员名下',
    ]

    for idx, reason in enumerate(reasons, 1):
        ws[f'A{idx+9}'] = reason

    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20

def main():
    """主程序"""
    print("\n" + "="*60)
    print("李华伟批价单数据差异分析")
    print("="*60)

    # 验证文件存在
    if not os.path.exists(PERFORMANCE_FILE):
        print(f"❌ 绩效表不存在: {PERFORMANCE_FILE}")
        return

    if not os.path.exists(PMA_DATA_FILE):
        print(f"❌ PMA数据文件不存在: {PMA_DATA_FILE}")
        return

    # 读取数据
    performance_data = read_performance_data()
    pma_projects, pma_df = read_pma_data()

    if not performance_data or not pma_projects:
        print("❌ 数据读取失败")
        return

    # 分析差异
    analysis = analyze_differences(performance_data, pma_projects)

    # 打印摘要
    print("\n📊 分析结果摘要:")
    print(f"  绩效表总额: ¥{analysis['performance_total']:,.2f}")
    print(f"  PMA总额: ¥{analysis['pma_total']:,.2f}")
    print(f"  差异: ¥{analysis['difference']:,.2f} ({analysis['difference_rate']:.2f}%)")
    print(f"  项目数: {len(pma_projects)}")

    # 生成报告
    generate_report(analysis)

    print("\n✅ 分析完成！")

if __name__ == '__main__':
    main()
