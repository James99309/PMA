#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比较两个Excel文件中李华伟的数据
找出问题所在
"""
import openpyxl
import sys

# 打开两个文件
file1 = '/Users/nijie/Downloads/2025年绩效数据采集表（12.31）.xlsx'
file2 = '/Users/nijie/Downloads/李华伟_批价单数据清单_2025年度_已清理.xlsx'

wb1 = openpyxl.load_workbook(file1, data_only=True)
wb2 = openpyxl.load_workbook(file2, data_only=True)

print("=" * 150)
print("对比分析：2025年绩效数据采集表 vs 李华伟_批价单数据清单")
print("=" * 150)
print()

# ========== 分析文件1 ==========
print("【文件1】2025年绩效数据采集表（12.31）.xlsx")
print("-" * 150)
print("Sheet列表:", wb1.sheetnames)
print()

# 查找李华伟的数据
for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    print(f"\n📄 分析Sheet: {sheet_name}")
    print("-" * 150)

    # 打印前20行来理解结构
    for row_idx in range(1, min(25, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(str(cell.value)[:20] if cell.value else '')

        print(f"行{row_idx}: {' | '.join(row_data)}")

    # 查找李华伟的数据
    print(f"\n查找李华伟的数据:")
    lhw_rows = []
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value and '李华伟' in str(cell_value):
                lhw_rows.append(row_idx)
                break

    if lhw_rows:
        print(f"找到李华伟在第 {lhw_rows} 行")
        for row_idx in lhw_rows:
            print(f"\n第{row_idx}行的完整数据:")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  {col_letter}{row_idx}: {cell.value}")

# ========== 分析文件2 ==========
print("\n\n" + "=" * 150)
print("【文件2】李华伟_批价单数据清单_2025年度_已清理.xlsx")
print("-" * 150)
print("Sheet列表:", wb2.sheetnames)
print()

for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"\n📄 Sheet: {sheet_name}")

    # 统计数据
    data_rows = []
    for row_idx in range(4, ws.max_row):  # 从第4行开始（跳过标题）
        first_col = ws.cell(row=row_idx, column=1).value
        if first_col and str(first_col).isdigit():
            data_rows.append(row_idx)
        elif first_col and '合计' in str(first_col):
            break

    if data_rows:
        print(f"  数据行数: {len(data_rows)}")

        # 显示数据摘要
        total_amount = 0
        print(f"\n  样本数据（前5条）:")
        print(f"  {'序号':<5} {'项目名称':<40} {'金额':<15}")
        print(f"  {'-' * 65}")

        for idx, row_idx in enumerate(data_rows[:5]):
            seq = ws.cell(row=row_idx, column=1).value
            proj = ws.cell(row=row_idx, column=2).value
            amount = ws.cell(row=row_idx, column=6).value  # 最终用金额在第6列

            print(f"  {seq:<5} {str(proj)[:38]:<40} ¥{amount:>13,.0f}" if amount else f"  {seq:<5} {str(proj)[:38]:<40}")

        # 计算总金额
        for row_idx in data_rows:
            amount = ws.cell(row=row_idx, column=6).value
            if isinstance(amount, (int, float)):
                total_amount += amount

        print(f"\n  统计：{len(data_rows)}条项目，金额总计: ¥{total_amount:,.0f}")

# ========== 关键问题分析 ==========
print("\n\n" + "=" * 150)
print("🔍 关键问题分析")
print("=" * 150)

# 重点检查第一个文件的数据结构
ws_perf = wb1[wb1.sheetnames[0]]

print("\n1. 检查绩效数据采集表中的李华伟数据:")
print("-" * 150)

lhw_perf_data = None
for row_idx in range(1, ws_perf.max_row + 1):
    for col_idx in range(1, ws_perf.max_column + 1):
        if ws_perf.cell(row=row_idx, column=col_idx).value and '李华伟' in str(ws_perf.cell(row=row_idx, column=col_idx).value):
            print(f"找到李华伟在第{row_idx}行")
            print("\n该行的所有数据:")

            for c_idx in range(1, min(20, ws_perf.max_column + 1)):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                value = ws_perf.cell(row=row_idx, column=c_idx).value
                print(f"  {col_letter}: {value}")

            lhw_perf_data = row_idx
            break

print("\n2. 对比两个文件的关键差异:")
print("-" * 150)

# 获取清理后文件的总金额
ws_clean = wb2['去重合并结果']
clean_total = 0
clean_count = 0

for row_idx in range(4, ws_clean.max_row):
    amount = ws_clean.cell(row=row_idx, column=6).value
    if isinstance(amount, (int, float)) and amount > 0:
        clean_total += amount
        clean_count += 1

print(f"清理后文件的数据:")
print(f"  项目数: {clean_count}个")
print(f"  总金额: ¥{clean_total:,.0f}")

print(f"\n绩效采集表中的数据:")
print(f"  (需要在表格中找到具体的项目数和金额)")

print("\n3. 可能存在的问题:")
print("-" * 150)
print("  ❓ 是否还有其他隐藏的项目？")
print("  ❓ 项目名称是否完全匹配？")
print("  ❓ 金额的折扣方式是否相同？")
print("  ❓ 是否还有其他分类方式的项目？")

wb1.close()
wb2.close()
