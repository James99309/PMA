#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出 April（骆永融）2026年3月工作日志 Excel
将"其他"类按内容重新归入市场科目，已分类的保持不变
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# 原始数据（从 CN NAS work_items 表导出）
# ============================================================

# 已分类的条目 (work_type, date, title, hours)
classified_items = [
    # 行政事务
    ("行政事务", "3/13", "新办公室展厅规划推进", 3.5),
    ("行政事务", "3/17", "新办公室展厅内容规划", 2.5),
    ("行政事务", "3/17", "新办公室装修进度沟通", 0.3),
    ("行政事务", "3/18", "新办公室展厅方案推进", 3.0),
    ("行政事务", "3/19", "展厅设计落地沟通推进", 3.0),
    ("行政事务", "3/20", "新办公室展厅事宜", 4.7),
    ("行政事务", "3/23", "新办公室展厅建设", 2.5),
    ("行政事务", "3/24", "新办公室展厅设计与施工对接", 1.0),
    ("行政事务", "3/26", "新办公室展厅画面设计沟通", 1.0),
    ("行政事务", "3/27", "新办公室展厅平面布局优化沟通", 1.0),
    ("行政事务", "3/30", "新办公室展厅方案优化与供应商对接", 2.0),
    ("行政事务", "3/31", "新办公室展厅方案优化与供应商对接", 1.5),
    # 售前支持
    ("售前支持", "3/17", "赋能销售的物料优化", 1.5),
    ("售前支持", "3/17", "成功案例成果物的制作对应与跟进", 0.8),
    ("售前支持", "3/23", "销售转化支持", 0.5),
    ("售前支持", "3/23", "成功案例制作推进", 2.0),
    ("售前支持", "3/24", "产品泰国认证配合", 2.5),
    ("售前支持", "3/24", "销售赋能与市场化内容优化", 2.5),
    ("售前支持", "3/25", "泰国产品认证配合", 2.5),
    ("售前支持", "3/25", "客户展厅物料对接（船舶行业）", 0.5),
    ("售前支持", "3/25", "中策平台汇报方案完善与提交", 1.5),
    ("售前支持", "3/26", "泰国产品认证配合", 2.0),
    ("售前支持", "3/26", "销售支持资料优化", 1.5),
    ("售前支持", "3/27", "市场工作重点梳理与方向规划", 3.5),
    ("售前支持", "3/30", "中策平台使用机制搭建与数据追踪规划", 2.0),
    ("售前支持", "3/31", "石油石化行业市场策略探索（AI辅助）", 1.8),
    # 品牌活动
    ("品牌活动", "3/13", "海外数据中心展会方案推进", 2.5),
    ("品牌活动", "3/16", "数据中心展会沟通会", 1.0),
    ("品牌活动", "3/17", "数据中心展会资源沟通与确认", 1.5),
    ("品牌活动", "3/19", "展会资源跟进与多方协同推进", 1.5),
    ("品牌活动", "3/23", "DCA-HK数据中心展会合同审核与流程推进", 0.7),
    ("品牌活动", "3/25", "海外展会沟通推进", 0.5),
    ("品牌活动", "3/26", "马来西亚9月展会信息整理与汇报", 1.0),
    ("品牌活动", "3/27", "海外数据中心行业拓展", 3.0),
    ("品牌活动", "3/30", "石油石化行业市场拓展", 1.0),
    ("品牌活动", "3/31", "香港数据中心展会合同推进", 1.0),
    # 商务洽谈
    ("商务洽谈", "3/13", "大数据平台API对接与能力评估", 0.9),
    ("商务洽谈", "3/16", "展会主办方信息跟进", 4.5),
    ("商务洽谈", "3/19", "中策平台API服务报价跟进", 1.0),
    ("商务洽谈", "3/20", "DCA数据中心展会合同审核与风险确认", 0.8),
    ("商务洽谈", "3/20", "中策平台合作条款沟通", 0.7),
    ("商务洽谈", "3/24", "中策数据平台评估与商务谈判推进", 1.3),
    ("商务洽谈", "3/26", "中策平台费用沟通", 1.0),
    # 社媒运营
    ("社媒运营", "3/26", "数据中心的素材切片准备", 1.0),
    ("社媒运营", "3/30", "官网内容建设", 2.0),
    ("社媒运营", "3/31", "官网内容建设与结构优化推进", 2.5),
    # 会议
    ("会议", "3/16", "内容营销物料优化沟通", 0.5),
    ("会议", "3/16", "市场部接下来的内容制作方向沟通", 1.0),
    ("会议", "3/18", "中策API功能的沟通会", 1.5),
    ("会议", "3/19", "新办公室展厅方案沟通及重点工作对齐", 1.5),
    # 物料设计
    ("物料设计", "3/16", "成功案例内容制作推进", 1.5),
    ("物料设计", "3/18", "项目案例物料推进", 1.0),
    ("物料设计", "3/20", "内容发布与案例内容制作", 1.5),
    ("物料设计", "3/30", "项目流程跟进", 0.3),
    # 需求分析
    ("需求分析", "3/18", "AI工具测试与自动化信息抓取测试", 2.0),
    # 客户维护
    ("客户维护", "3/23", "分销商证书外框选择方案", 1.5),
    # 服务响应
    ("服务响应", "3/16", "其他跟进事项", 1.0),
    # 视频制作
    ("视频制作", "3/13", "视频内容优化推进", 0.3),
    ("视频制作", "3/31", "数据中心AI视频优化沟通", 0.3),
    # 资产管理
    ("资产管理", "3/17", "项目素材资料收集与归档", 0.5),
    # 渠道活动
    ("渠道活动", "3/13", "石油石化行业活动跟进", 0.5),
]

# ============================================================
# "其他"类条目 → 按内容重新分类到市场科目
# ============================================================

reclassified_other = [
    # 品牌活动/展会
    ("品牌活动", "3/3", "石油石化行业活动推进", 2.0),
    ("品牌活动", "3/3", "行业展会数据分析与筛选决策", 2.0),
    ("品牌活动", "3/4", "海外数据中心活动营销", 4.0),
    ("品牌活动", "3/5", "海外数据中心活动营销推进", 2.0),
    ("品牌活动", "3/5", "石油石化行业活动推进沟通", 1.0),
    ("品牌活动", "3/6", "海外数据中心活动营销推进", 1.0),
    ("品牌活动", "3/9", "行业活动情况跟进", 2.0),
    ("品牌活动", "3/10", "数据中心行业展会评估与决策沟通", 2.0),
    ("品牌活动", "3/10", "展会主办方沟通与资源确认", 1.5),
    ("品牌活动", "3/11", "海外数据中心展会方案更新", 2.8),
    ("品牌活动", "3/12", "海外数据中心展会沟通与方案推进", 2.5),
    ("品牌活动", "3/12", "石油石化行业活动客户邀约情况", 1.0),

    # 物料设计/内容营销
    ("物料设计", "3/3", "石油石化行业线上营销内容", 0.8),
    ("物料设计", "3/4", "成功案例的营销内容推进", 1.0),
    ("物料设计", "3/5", "内容营销与物料制作推进", 1.0),
    ("物料设计", "3/6", "内容营销与物料制作推进", 4.5),
    ("物料设计", "3/9", "营销物料制作推进", 3.5),
    ("物料设计", "3/11", "产品技术资料优化", 1.0),
    ("物料设计", "3/12", "销售支持物料", 1.0),

    # 商务洽谈/平台合作
    ("商务洽谈", "3/4", "2026年大数据平台选择决策汇报", 2.5),
    ("商务洽谈", "3/5", "大数据平台筛选与续约决策推进", 3.0),
    ("商务洽谈", "3/10", "中策平台对接", 1.5),
    ("商务洽谈", "3/11", "中策大数据平台对接沟通", 1.0),
    ("商务洽谈", "3/12", "AI平台接口对接沟通", 1.0),

    # 视频制作
    ("视频制作", "3/10", "视频物料优化推进", 1.0),
    ("视频制作", "3/11", "产品视频优化推进", 1.0),

    # 社媒运营
    ("社媒运营", "3/2", "PMA数字化系统产品化营销", 4.0),
    ("社媒运营", "3/6", "PMA系统宣传推进", 2.0),

    # 行政事务
    ("行政事务", "3/10", "合同及流程事务处理", 0.5),
    ("行政事务", "3/11", "新办公室展厅规划沟通", 1.3),
    ("行政事务", "3/12", "新办公室展厅规划推进", 1.5),
    ("行政事务", "3/25", "新办公室展厅建设", 3.0),
    ("行政事务", "3/12", "网站运维续费沟通", 0.3),

    # 需求分析
    ("需求分析", "3/2", "数字化平台选型", 1.0),
    ("需求分析", "3/9", "PMA系统强化利用", 1.0),

    # 会议
    ("会议", "3/10", "管理层沟通与工作对齐", 1.0),

    # 售前支持
    ("售前支持", "3/3", "3月重点项目资料更新梳理", 2.0),
    ("售前支持", "3/2", "区域市场战略", 3.0),

    # 其他（无法归类）
    ("其他", "3/5", "其他", 0.7),
    ("其他", "3/9", "其他工作推进", 1.0),
]

# ============================================================
# 合并全部数据
# ============================================================

all_items = classified_items + reclassified_other

# 排序：按类型 → 日期
type_order = {
    "品牌活动": 1, "物料设计": 2, "售前支持": 3, "商务洽谈": 4,
    "行政事务": 5, "社媒运营": 6, "视频制作": 7, "会议": 8,
    "需求分析": 9, "客户维护": 10, "服务响应": 11, "资产管理": 12,
    "渠道活动": 13, "其他": 14
}

def sort_key(item):
    cat, date_str, title, hours = item
    month, day = date_str.split("/")
    return (type_order.get(cat, 99), int(day))

all_items.sort(key=sort_key)

# ============================================================
# 生成 Excel
# ============================================================

wb = Workbook()

# ---- Sheet 1: 分类明细 ----
ws = wb.active
ws.title = "分类明细"

# 样式定义
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
cat_font = Font(bold=True, size=11, color="1E40AF")
cat_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
reclassified_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # 黄底=从其他重分类
normal_font = Font(size=10)
number_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# 标题行
ws.append(["April（骆永融）2026年3月工作日志 — 分类明细"])
ws.merge_cells("A1:E1")
ws["A1"].font = Font(bold=True, size=14)
ws.append([])
ws.append(["黄色底纹 = 原\"其他\"类重新归入的条目"])
ws["A3"].font = Font(size=9, italic=True, color="92400E")
ws.append([])

# 表头
headers = ["工作类型", "日期", "工作内容", "工时(h)", "来源"]
ws.append(headers)
header_row = ws.max_row
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# 构建"原其他"集合用于标记
other_titles = set()
for cat, date_str, title, hours in reclassified_other:
    other_titles.add((date_str, title))

# 按类型分组写入
current_cat = None
cat_start_rows = {}

for cat, date_str, title, hours in all_items:
    if cat != current_cat:
        current_cat = cat
        cat_start_rows[cat] = ws.max_row + 1

    is_reclassified = (date_str, title) in other_titles
    source = "重分类" if is_reclassified else "原分类"

    ws.append([cat, date_str, title, hours, source])
    row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if is_reclassified:
            cell.fill = reclassified_fill
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4).number_format = '0.0'
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

# 列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 48
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10

# ---- Sheet 2: 汇总统计 ----
ws2 = wb.create_sheet("汇总统计")

# 按类型汇总
from collections import defaultdict
cat_summary = defaultdict(lambda: {"count": 0, "hours": 0, "reclassified_count": 0, "reclassified_hours": 0})

for cat, date_str, title, hours in all_items:
    cat_summary[cat]["count"] += 1
    cat_summary[cat]["hours"] += hours
    if (date_str, title) in other_titles:
        cat_summary[cat]["reclassified_count"] += 1
        cat_summary[cat]["reclassified_hours"] += hours

# 排序
sorted_cats = sorted(cat_summary.items(), key=lambda x: -x[1]["hours"])
total_hours = sum(v["hours"] for v in cat_summary.values())

ws2.append(["April（骆永融）2026年3月工作日志 — 汇总统计"])
ws2.merge_cells("A1:G1")
ws2["A1"].font = Font(bold=True, size=14)
ws2.append([])

sum_headers = ["工作类型", "条目数", "总工时(h)", "占比", "其中重分类条目", "重分类工时(h)", "原分类工时(h)"]
ws2.append(sum_headers)
header_row2 = ws2.max_row
for col_idx, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=header_row2, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for cat, data in sorted_cats:
    pct = data["hours"] / total_hours * 100
    original_hours = data["hours"] - data["reclassified_hours"]
    ws2.append([
        cat, data["count"], round(data["hours"], 1),
        f"{pct:.1f}%",
        data["reclassified_count"], round(data["reclassified_hours"], 1),
        round(original_hours, 1)
    ])
    row = ws2.max_row
    for col in range(1, 8):
        cell = ws2.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if col >= 2:
            cell.alignment = Alignment(horizontal="center")

# 合计行
ws2.append(["合计", sum(d["count"] for _, d in sorted_cats),
            round(total_hours, 1), "100%",
            sum(d["reclassified_count"] for _, d in sorted_cats),
            round(sum(d["reclassified_hours"] for _, d in sorted_cats), 1),
            round(sum(d["hours"] - d["reclassified_hours"] for _, d in sorted_cats), 1)])
row = ws2.max_row
for col in range(1, 8):
    cell = ws2.cell(row=row, column=col)
    cell.font = Font(bold=True, size=10)
    cell.border = thin_border
    cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    if col >= 2:
        cell.alignment = Alignment(horizontal="center")

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 14

# ---- Sheet 3: 每日明细（按日期） ----
ws3 = wb.create_sheet("每日明细")

ws3.append(["April（骆永融）2026年3月工作日志 — 按日期排列"])
ws3.merge_cells("A1:D1")
ws3["A1"].font = Font(bold=True, size=14)
ws3.append([])

day_headers = ["日期", "工作类型", "工作内容", "工时(h)"]
ws3.append(day_headers)
header_row3 = ws3.max_row
for col_idx, h in enumerate(day_headers, 1):
    cell = ws3.cell(row=header_row3, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# 按日期排序
day_sorted = sorted(all_items, key=lambda x: (int(x[1].split("/")[1]), type_order.get(x[0], 99)))
current_date = None
for cat, date_str, title, hours in day_sorted:
    is_reclassified = (date_str, title) in other_titles
    ws3.append([date_str, cat, title, hours])
    row = ws3.max_row
    for col in range(1, 5):
        cell = ws3.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if is_reclassified:
            cell.fill = reclassified_fill
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    ws3.cell(row=row, column=4).number_format = '0.0'

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 48
ws3.column_dimensions['D'].width = 10

# 保存
output_path = os.path.expanduser("~/Desktop/April_骆永融_2026年3月工作日志.xlsx")
wb.save(output_path)
print(f"✅ 已保存到: {output_path}")
