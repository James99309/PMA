#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
线上线下批价单数据对比分析

数据源1: 云端SP8D系统（线上）- 31个批价单
数据源2: 线下Excel文件 - 336行产品明细

输出: 对比报告Excel
"""
import sys
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 路径修正
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

project_root = get_project_root()
sys.path.insert(0, project_root)


# ============================================================
# 配置
# ============================================================

# 线下Excel文件路径
OFFLINE_EXCEL_PATH = '/Users/nijie/Downloads/线下批价单合计.xlsx'

# 销售人员名称映射（线下中文 -> 线上用户名）
SALES_NAME_MAPPING = {
    '杨俊杰': 'yangjj',
    '李华伟': 'lihuawei',
    '郭小会': 'gxh',
    '范敬': 'fanjing',
}

# 反向映射
SALES_USERNAME_TO_NAME = {v: k for k, v in SALES_NAME_MAPPING.items()}

# 目标销售人员用户名列表
TARGET_SALES_USERNAMES = ['lihuawei', 'gxh', 'fanjing', 'yangjj']


# ============================================================
# 读取线下Excel数据
# ============================================================

def read_offline_excel():
    """读取线下批价单Excel文件"""
    print("\n" + "="*100)
    print("步骤1: 读取线下Excel数据")
    print("="*100)

    if not os.path.exists(OFFLINE_EXCEL_PATH):
        print(f"❌ 文件不存在: {OFFLINE_EXCEL_PATH}")
        return None

    print(f"✓ 文件存在: {OFFLINE_EXCEL_PATH}")

    # 读取Excel，获取Sheet信息
    xl = pd.ExcelFile(OFFLINE_EXCEL_PATH)
    print(f"✓ Sheet列表: {xl.sheet_names}")

    # 读取目标Sheet（线上缺失的批价单明细）
    target_sheet = xl.sheet_names[0]
    print(f"✓ 读取Sheet: {target_sheet}")

    # 读取数据，header=1 跳过第一行空行，第二行是真正的表头
    df = pd.read_excel(OFFLINE_EXCEL_PATH, sheet_name=target_sheet, header=1)

    print(f"✓ 数据行数: {len(df)}")
    print(f"✓ 数据列: {list(df.columns)}")

    # 显示前几行数据
    print(f"\n前5行数据预览:")
    print(df.head())

    return df


def aggregate_offline_data(df):
    """按项目聚合线下数据"""
    print("\n" + "="*100)
    print("步骤2: 按项目聚合线下数据")
    print("="*100)

    # 检查必要的列
    required_cols = ['项目名称', '销售归属']
    for col in required_cols:
        if col not in df.columns:
            print(f"❌ 缺少必要列: {col}")
            print(f"   可用列: {list(df.columns)}")
            return None

    # 清理数据
    df = df.copy()
    df['项目名称'] = df['项目名称'].astype(str).str.strip()
    df['销售归属'] = df['销售归属'].astype(str).str.strip()

    # 过滤目标销售人员的数据
    target_sales_names = list(SALES_NAME_MAPPING.keys())
    df_filtered = df[df['销售归属'].isin(target_sales_names)].copy()
    print(f"✓ 过滤后数据行数（四个销售人员）: {len(df_filtered)}")

    # 识别金额列 - 线下Excel使用"批价单"列
    amount_col = None
    for col in ['批价单', '批价单金额', '批价金额', '金额']:
        if col in df.columns:
            amount_col = col
            break

    retail_col = None
    for col in ['零售价', '零售价格', '市场价']:
        if col in df.columns:
            retail_col = col
            break

    discount_col = None
    for col in ['提货折扣', '折扣', '折扣率']:
        if col in df.columns:
            discount_col = col
            break

    print(f"✓ 金额列: {amount_col}")
    print(f"✓ 零售价列: {retail_col}")
    print(f"✓ 折扣列: {discount_col}")

    # 转换金额为数值（处理可能的公式或文本）
    if amount_col:
        df_filtered[amount_col] = pd.to_numeric(df_filtered[amount_col], errors='coerce').fillna(0)
    if retail_col:
        df_filtered[retail_col] = pd.to_numeric(df_filtered[retail_col], errors='coerce').fillna(0)
    if discount_col:
        df_filtered[discount_col] = pd.to_numeric(df_filtered[discount_col], errors='coerce').fillna(0)

    # 如果有数量列
    if '数量' in df_filtered.columns:
        df_filtered['数量'] = pd.to_numeric(df_filtered['数量'], errors='coerce').fillna(0)

    # 按 项目名称 + 销售归属 分组聚合
    agg_config = {'行数': ('项目名称', 'count')}
    if amount_col:
        agg_config['批价单金额'] = (amount_col, 'sum')
    if retail_col:
        agg_config['零售价总和'] = (retail_col, 'sum')
    if discount_col:
        agg_config['平均折扣'] = (discount_col, 'mean')
    if '数量' in df_filtered.columns:
        agg_config['数量'] = ('数量', 'sum')

    # 执行聚合
    df_agg = df_filtered.groupby(['项目名称', '销售归属']).agg(**agg_config).reset_index()

    # 添加用户名列
    df_agg['用户名'] = df_agg['销售归属'].map(SALES_NAME_MAPPING)

    print(f"\n✓ 聚合后项目数: {len(df_agg)}")
    print(f"\n按销售人员统计:")
    for name in target_sales_names:
        count = len(df_agg[df_agg['销售归属'] == name])
        if '批价单金额' in df_agg.columns:
            total = df_agg[df_agg['销售归属'] == name]['批价单金额'].sum()
            print(f"   {name}: {count}个项目, ¥{total:,.0f}")
        else:
            print(f"   {name}: {count}个项目")

    return df_agg


# ============================================================
# 读取线上SP8D数据
# ============================================================

def read_online_sp8d_data():
    """读取云端SP8D系统的批价单数据"""
    print("\n" + "="*100)
    print("步骤3: 读取云端SP8D数据")
    print("="*100)

    # 加载云端SP8D配置
    env_file = os.path.join(project_root, '.env.sp8d.local')
    if os.path.exists(env_file):
        load_dotenv(env_file, override=True)
        print(f"✅ 已加载SP8D配置: {env_file}")
    else:
        print(f"❌ 找不到SP8D配置文件: {env_file}")
        return None

    from app import create_app, db
    from app.models.user import User
    from app.models.pricing_order import PricingOrder
    from app.models.project import Project

    app = create_app()
    with app.app_context():
        all_results = {}

        for username in TARGET_SALES_USERNAMES:
            # 找到用户
            user = User.query.filter(User.username == username).first()
            if not user:
                print(f"❌ 未找到用户: {username}")
                all_results[username] = {
                    'found': False,
                    'pricing_orders': []
                }
                continue

            print(f"✓ 找到用户: {user.real_name or user.username} (ID: {user.id})")

            # 查询该用户作为项目销售经理的所有批价单
            pricing_orders = PricingOrder.query.join(
                Project, PricingOrder.project_id == Project.id
            ).filter(
                Project.vendor_sales_manager_id == user.id
            ).all()

            # 过滤条件应用
            filtered_pos = []

            for po in pricing_orders:
                # 检查创建时间是否在2025年
                if po.created_at is None:
                    continue

                if po.created_at.tzinfo is None:
                    created_time = po.created_at
                else:
                    created_time = po.created_at.astimezone(ZoneInfo('Asia/Shanghai'))

                if created_time.year != 2025:
                    continue

                # 检查审批状态
                if po.status != 'approved':
                    continue

                # 获取关联的项目
                if not po.project:
                    continue

                # 排除测试项目
                if '测试' in po.project.project_name:
                    continue

                # 计算零售价总和
                retail_total = 0.0
                for detail in po.pricing_details:
                    if detail.market_price and detail.quantity:
                        retail_total += (detail.market_price * detail.quantity)

                # 获取折扣
                pricing_discount_rate = po.pricing_total_discount_rate or 1.0

                po_dict = {
                    'id': po.id,
                    'po_number': po.order_number,
                    'project_name': po.project.project_name,
                    'retail_price_total': retail_total,
                    'pricing_discount_rate': pricing_discount_rate * 100,
                    'pricing_amount': po.pricing_total_amount or 0.0,
                    'status': po.status,
                    'created_at': po.created_at,
                    'username': username,
                    'sales_name': SALES_USERNAME_TO_NAME.get(username, username)
                }

                filtered_pos.append(po_dict)

            print(f"   符合条件的批价单: {len(filtered_pos)} 个")

            all_results[username] = {
                'found': True,
                'user_id': user.id,
                'real_name': user.real_name or user.username,
                'pricing_orders': filtered_pos
            }

        # 汇总线上数据为DataFrame
        online_records = []
        for username, result in all_results.items():
            if result['found']:
                for po in result['pricing_orders']:
                    online_records.append({
                        '用户名': username,
                        '销售归属': SALES_USERNAME_TO_NAME.get(username, username),
                        '项目名称': po['project_name'],
                        '批价单号': po['po_number'],
                        '零售价总和': po['retail_price_total'],
                        '批价折扣率': po['pricing_discount_rate'],
                        '批价单金额': po['pricing_amount'],
                        '创建时间': po['created_at'],
                        '状态': po['status']
                    })

        df_online = pd.DataFrame(online_records)
        print(f"\n✓ 线上数据总计: {len(df_online)} 个批价单")

        return df_online, all_results


# ============================================================
# 对比分析
# ============================================================

def compare_data(df_online, df_offline_agg):
    """对比线上线下数据"""
    print("\n" + "="*100)
    print("步骤4: 对比分析")
    print("="*100)

    # 按项目名称建立映射
    online_projects = set(df_online['项目名称'].unique()) if len(df_online) > 0 else set()
    offline_projects = set(df_offline_agg['项目名称'].unique()) if len(df_offline_agg) > 0 else set()

    print(f"\n线上项目数: {len(online_projects)}")
    print(f"线下项目数: {len(offline_projects)}")

    # 交集、差集
    common_projects = online_projects & offline_projects
    online_only = online_projects - offline_projects
    offline_only = offline_projects - online_projects

    print(f"\n共同项目数（精确匹配）: {len(common_projects)}")
    print(f"仅线上项目数: {len(online_only)}")
    print(f"仅线下项目数: {len(offline_only)}")

    # 标记数据来源
    df_online['数据来源'] = '线上'
    df_online['匹配状态'] = df_online['项目名称'].apply(
        lambda x: '两边都有' if x in common_projects else '仅线上'
    )

    df_offline_agg['数据来源'] = '线下'
    df_offline_agg['匹配状态'] = df_offline_agg['项目名称'].apply(
        lambda x: '两边都有' if x in common_projects else '仅线下'
    )

    return {
        'common_projects': common_projects,
        'online_only': online_only,
        'offline_only': offline_only,
        'df_online': df_online,
        'df_offline': df_offline_agg
    }


def generate_summary(df_online, df_offline_agg):
    """生成汇总数据"""
    print("\n" + "="*100)
    print("步骤5: 生成汇总数据")
    print("="*100)

    summary_data = []

    for username in TARGET_SALES_USERNAMES:
        sales_name = SALES_USERNAME_TO_NAME.get(username, username)

        # 线上数据
        online_data = df_online[df_online['用户名'] == username] if len(df_online) > 0 else pd.DataFrame()
        online_count = len(online_data)
        online_amount = online_data['批价单金额'].sum() if len(online_data) > 0 else 0

        # 线下数据
        offline_data = df_offline_agg[df_offline_agg['销售归属'] == sales_name] if len(df_offline_agg) > 0 else pd.DataFrame()
        offline_count = len(offline_data)

        # 线下金额（统一使用"批价单金额"列名，已在聚合时标准化）
        offline_amount = offline_data['批价单金额'].sum() if (len(offline_data) > 0 and '批价单金额' in df_offline_agg.columns) else 0

        # 合计
        total_count = online_count + offline_count
        total_amount = online_amount + offline_amount

        # 差异
        diff_count = offline_count
        diff_amount = offline_amount
        diff_ratio = (diff_amount / online_amount * 100) if online_amount > 0 else 0

        summary_data.append({
            '销售人员': sales_name,
            '用户名': username,
            '线上批价单数': online_count,
            '线上金额': online_amount,
            '线下批价单数': offline_count,
            '线下金额': offline_amount,
            '合计批价单数': total_count,
            '合计金额': total_amount,
            '差异批价单数': diff_count,
            '差异金额': diff_amount,
            '差异占比': f"{diff_ratio:.1f}%"
        })

        print(f"\n{sales_name} ({username}):")
        print(f"   线上: {online_count}个批价单, ¥{online_amount:,.0f}")
        print(f"   线下: {offline_count}个项目(批价单), ¥{offline_amount:,.0f}")
        print(f"   合计: {total_count}个, ¥{total_amount:,.0f}")
        print(f"   差异: {diff_count}个, ¥{diff_amount:,.0f} ({diff_ratio:.1f}%)")

    df_summary = pd.DataFrame(summary_data)

    # 添加总计行
    total_row = {
        '销售人员': '总计',
        '用户名': '-',
        '线上批价单数': df_summary['线上批价单数'].sum(),
        '线上金额': df_summary['线上金额'].sum(),
        '线下批价单数': df_summary['线下批价单数'].sum(),
        '线下金额': df_summary['线下金额'].sum(),
        '合计批价单数': df_summary['合计批价单数'].sum(),
        '合计金额': df_summary['合计金额'].sum(),
        '差异批价单数': df_summary['差异批价单数'].sum(),
        '差异金额': df_summary['差异金额'].sum(),
        '差异占比': f"{(df_summary['差异金额'].sum() / df_summary['线上金额'].sum() * 100) if df_summary['线上金额'].sum() > 0 else 0:.1f}%"
    }
    df_summary = pd.concat([df_summary, pd.DataFrame([total_row])], ignore_index=True)

    print(f"\n总计:")
    print(f"   线上: {total_row['线上批价单数']}个批价单, ¥{total_row['线上金额']:,.0f}")
    print(f"   线下: {total_row['线下批价单数']}个项目(批价单), ¥{total_row['线下金额']:,.0f}")
    print(f"   合计: {total_row['合计批价单数']}个, ¥{total_row['合计金额']:,.0f}")

    return df_summary


# ============================================================
# 生成Excel报告
# ============================================================

def generate_excel_report(df_summary, df_online, df_offline_agg, compare_result):
    """生成对比报告Excel"""
    print("\n" + "="*100)
    print("步骤6: 生成Excel报告")
    print("="*100)

    # 输出路径
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f'/Users/nijie/Documents/线上线下批价单对比报告_{timestamp}.xlsx'

    # 创建工作簿
    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # ========== Sheet 1: 汇总 ==========
    ws1 = wb.active
    ws1.title = '汇总'

    # 写入标题
    ws1['A1'] = '线上线下批价单对比汇总'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:K1')

    ws1['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A3'] = ''

    # 写入汇总数据
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 4:  # 标题行
                cell.font = header_font_white
                cell.fill = header_fill
            cell.alignment = center_align

    # 调整列宽
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 14
    ws1.column_dimensions['J'].width = 15
    ws1.column_dimensions['K'].width = 12

    # ========== Sheet 2: 线上已有 ==========
    ws2 = wb.create_sheet('线上已有')

    if len(df_online) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(df_online, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
    else:
        ws2['A1'] = '无数据'

    # ========== Sheet 3: 线下缺失 ==========
    ws3 = wb.create_sheet('线下缺失（待补录）')

    if len(df_offline_agg) > 0:
        # 只显示仅线下的数据
        df_offline_only = df_offline_agg[df_offline_agg['匹配状态'] == '仅线下']

        if len(df_offline_only) > 0:
            for r_idx, row in enumerate(dataframe_to_rows(df_offline_only, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    if r_idx == 1:
                        cell.font = header_font_white
                        cell.fill = header_fill
        else:
            ws3['A1'] = '无需要补录的数据'
    else:
        ws3['A1'] = '无数据'

    # ========== Sheet 4: 完整视图 ==========
    ws4 = wb.create_sheet('完整视图')

    # 合并线上线下数据
    combined_data = []

    # 线上数据
    if len(df_online) > 0:
        for _, row in df_online.iterrows():
            combined_data.append({
                '数据来源': '线上',
                '销售人员': row.get('销售归属', ''),
                '用户名': row.get('用户名', ''),
                '项目名称': row.get('项目名称', ''),
                '批价单号': row.get('批价单号', ''),
                '批价单金额': row.get('批价单金额', 0),
                '匹配状态': row.get('匹配状态', '')
            })

    # 线下数据（仅线下）
    if len(df_offline_agg) > 0:
        df_offline_only = df_offline_agg[df_offline_agg['匹配状态'] == '仅线下']
        for _, row in df_offline_only.iterrows():
            # 查找金额列
            amount = 0
            for col in ['批价单金额', '批价金额', '金额']:
                if col in row.index:
                    amount = row[col]
                    break

            combined_data.append({
                '数据来源': '线下',
                '销售人员': row.get('销售归属', ''),
                '用户名': row.get('用户名', ''),
                '项目名称': row.get('项目名称', ''),
                '批价单号': '-',
                '批价单金额': amount,
                '匹配状态': row.get('匹配状态', '')
            })

    if combined_data:
        df_combined = pd.DataFrame(combined_data)
        for r_idx, row in enumerate(dataframe_to_rows(df_combined, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
    else:
        ws4['A1'] = '无数据'

    # ========== Sheet 5: 项目对比明细 ==========
    ws5 = wb.create_sheet('项目对比明细')

    # 按项目聚合对比
    project_compare = []

    # 所有项目
    all_projects = set()
    if len(df_online) > 0:
        all_projects.update(df_online['项目名称'].unique())
    if len(df_offline_agg) > 0:
        all_projects.update(df_offline_agg['项目名称'].unique())

    for project_name in sorted(all_projects):
        # 线上
        online_match = df_online[df_online['项目名称'] == project_name] if len(df_online) > 0 else pd.DataFrame()
        online_amount = online_match['批价单金额'].sum() if len(online_match) > 0 else 0
        online_sales = online_match['销售归属'].iloc[0] if len(online_match) > 0 else ''

        # 线下
        offline_match = df_offline_agg[df_offline_agg['项目名称'] == project_name] if len(df_offline_agg) > 0 else pd.DataFrame()
        offline_amount = 0
        offline_sales = ''
        if len(offline_match) > 0:
            for col in ['批价单金额', '批价金额', '金额']:
                if col in offline_match.columns:
                    offline_amount = offline_match[col].sum()
                    break
            offline_sales = offline_match['销售归属'].iloc[0]

        # 状态
        if len(online_match) > 0 and len(offline_match) > 0:
            status = '两边都有'
        elif len(online_match) > 0:
            status = '仅线上'
        else:
            status = '仅线下'

        project_compare.append({
            '项目名称': project_name,
            '销售人员': online_sales or offline_sales,
            '状态': status,
            '线上金额': online_amount,
            '线下金额': offline_amount,
            '差异': offline_amount - online_amount if status == '两边都有' else (offline_amount if status == '仅线下' else -online_amount)
        })

    if project_compare:
        df_project_compare = pd.DataFrame(project_compare)
        for r_idx, row in enumerate(dataframe_to_rows(df_project_compare, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill

    # 保存
    wb.save(output_path)
    print(f"\n✅ 报告已生成: {output_path}")

    return output_path


# ============================================================
# 主函数
# ============================================================

def main():
    print("="*100)
    print("线上线下批价单数据对比分析")
    print("="*100)
    print(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"线下文件: {OFFLINE_EXCEL_PATH}")
    print("="*100)

    # 1. 读取线下Excel
    df_offline = read_offline_excel()
    if df_offline is None:
        print("\n❌ 无法读取线下Excel文件，程序终止")
        return

    # 2. 聚合线下数据
    df_offline_agg = aggregate_offline_data(df_offline)
    if df_offline_agg is None:
        print("\n❌ 无法聚合线下数据，程序终止")
        return

    # 3. 读取线上数据
    result = read_online_sp8d_data()
    if result is None:
        print("\n❌ 无法读取线上数据，程序终止")
        return

    df_online, all_results = result

    # 4. 对比分析
    compare_result = compare_data(df_online, df_offline_agg)

    # 5. 生成汇总
    df_summary = generate_summary(df_online, df_offline_agg)

    # 6. 生成Excel报告
    output_path = generate_excel_report(df_summary, df_online, df_offline_agg, compare_result)

    print("\n" + "="*100)
    print("✅ 对比分析完成")
    print("="*100)
    print(f"报告路径: {output_path}")


if __name__ == '__main__':
    main()
