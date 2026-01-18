#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出批价员的详细批价单信息
包含：项目名称、批价单号、金额、人员名称
"""
import sys
import os

# 路径修正
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

sys.path.insert(0, get_project_root())

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("❌ 需要安装openpyxl")
    sys.exit(1)

def query_online_quotations():
    """查询线上sp8d数据库的批价单详情"""
    print(f"\n📊 查询sp8d数据库批价单详情...")

    try:
        from sqlalchemy import create_engine, text
        from dotenv import load_dotenv

        env_path = get_project_root() + '/.env.sp8d.local'
        load_dotenv(env_path, override=True)

        db_url = os.getenv('DATABASE_URL')
        if not db_url:
            print("   ❌ 未设置 DATABASE_URL")
            return None

        engine = create_engine(db_url, echo=False)

        with engine.connect() as conn:
            result = conn.execute(text('SELECT 1'))

        print("   ✅ 数据库连接成功")

        user_mappings = {
            'gxh': '郭小会',
            'yangjj': '杨俊杰',
            'lihuawei': '李华伟',
            'fanj': '范敬'
        }

        all_quotations = []

        for username, real_name in user_mappings.items():
            try:
                user_sql = text("""
                    SELECT id FROM users
                    WHERE username = :username
                """)

                with engine.connect() as conn:
                    user_result = conn.execute(user_sql, {'username': username})
                    user_row = user_result.fetchone()

                    if not user_row:
                        print(f"   {username}: 用户不存在")
                        continue

                    user_id = user_row[0]

                    # 查询批价单详情
                    quotation_sql = text("""
                        SELECT
                            q.quotation_number,
                            p.project_name,
                            COALESCE(q.implant_total_amount, q.amount, 0) as amount,
                            q.approval_status,
                            q.created_at,
                            :real_name as person_name
                        FROM quotations q
                        LEFT JOIN projects p ON q.project_id = p.id
                        WHERE q.owner_id = :user_id
                        AND (q.approval_status = 'quoted_approved' OR q.approval_status = 'approved')
                        ORDER BY q.created_at DESC
                    """)

                    with engine.connect() as conn2:
                        q_result = conn2.execute(
                            quotation_sql,
                            {'user_id': user_id, 'real_name': real_name}
                        )
                        quotations = q_result.fetchall()

                        print(f"   {username} ({real_name}): {len(quotations)} 个批价单")

                        for q in quotations:
                            q_number, project_name, amount, status, created_at, person = q
                            all_quotations.append({
                                '人员名称': person,
                                '批价单号': q_number or '',
                                '项目名称': project_name or '(项目已删除)',
                                '批价金额': float(amount) if amount else 0,
                                '审批状态': status or '',
                                '创建时间': str(created_at) if created_at else '',
                                '数据来源': '线上系统'
                            })

            except Exception as e:
                print(f"   {username}: 查询失败 - {e}")

        return all_quotations
    except Exception as e:
        print(f"❌ 数据库连接失败: {e}")
        return None

def extract_offline_quotations():
    """从Excel提取线下批价单信息"""
    print(f"\n🔍 从Excel提取线下批价单...")

    file_path = '/Users/nijie/Downloads/线上缺失的批价单内容（2025年）.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['线上缺失的批价单明细']

        rows = list(ws.iter_rows(values_only=True))
        headers = rows[1]

        # 查找列索引
        sales_col_idx = None
        project_col_idx = None
        amount_col_idx = None
        mn_col_idx = None

        for i, h in enumerate(headers):
            if h:
                h_str = str(h)
                if '销售' in h_str:
                    sales_col_idx = i
                if '项目名称' in h_str:
                    project_col_idx = i
                if '结算批价额' in h_str and '绩效' in h_str:
                    amount_col_idx = i
                if 'MN' in h_str:
                    mn_col_idx = i

        if amount_col_idx is None:
            for i, h in enumerate(headers):
                if h and '金额' in str(h) and '绩效' in str(h):
                    amount_col_idx = i
                    break

        print(f"   列索引 - 销售:{sales_col_idx}, 项目:{project_col_idx}, 金额:{amount_col_idx}, MN:{mn_col_idx}")

        all_records = []

        # 处理数据行
        for row_idx, row in enumerate(rows[2:], start=3):
            if not row:
                continue

            try:
                sales_person = str(row[sales_col_idx]).strip() if sales_col_idx is not None and sales_col_idx < len(row) else ''
                project_name = str(row[project_col_idx]).strip() if project_col_idx is not None and project_col_idx < len(row) else ''
                amount = row[amount_col_idx] if amount_col_idx is not None and amount_col_idx < len(row) else 0
                mn_number = str(row[mn_col_idx]).strip() if mn_col_idx is not None and mn_col_idx < len(row) else ''

                # 匹配员工
                matched_user = None
                if sales_person == '杨俊杰':
                    matched_user = '杨俊杰'
                elif sales_person == '郭小会':
                    matched_user = '郭小会'
                elif sales_person == '李华伟':
                    matched_user = '李华伟'
                elif sales_person == '范敬':
                    matched_user = '范敬'

                if matched_user:
                    try:
                        amount = float(amount) if amount else 0
                        if amount > 0 and project_name:
                            all_records.append({
                                '人员名称': matched_user,
                                '批价单号': mn_number,
                                '项目名称': project_name,
                                '批价金额': amount,
                                '审批状态': '已批价',
                                '创建时间': '',
                                '数据来源': '线下Excel'
                            })
                    except (ValueError, TypeError):
                        pass
            except Exception:
                pass

        print(f"   ✅ 提取完成: {len(all_records)} 条记录")

        return all_records
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return []

def create_export_file(online_data, offline_data):
    """创建导出Excel文件"""
    print(f"\n💾 生成导出文件...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        # 创建新的workbook
        wb = openpyxl.Workbook()

        # 1. 线上数据
        ws_online = wb.active
        ws_online.title = '线上数据'

        online_data_sorted = sorted(online_data, key=lambda x: x['人员名称'])

        headers = ['人员名称', '批价单号', '项目名称', '批价金额', '审批状态', '创建时间']
        ws_online.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws_online[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充数据
        for record in online_data_sorted:
            ws_online.append([
                record['人员名称'],
                record['批价单号'],
                record['项目名称'],
                record['批价金额'],
                record['审批状态'],
                record['创建时间']
            ])

        # 调整列宽
        ws_online.column_dimensions['A'].width = 12
        ws_online.column_dimensions['B'].width = 20
        ws_online.column_dimensions['C'].width = 40
        ws_online.column_dimensions['D'].width = 15
        ws_online.column_dimensions['E'].width = 12
        ws_online.column_dimensions['F'].width = 20

        # 2. 线下数据
        ws_offline = wb.create_sheet('线下数据')

        offline_data_sorted = sorted(offline_data, key=lambda x: x['人员名称'])

        ws_offline.append(headers)

        for cell in ws_offline[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for record in offline_data_sorted:
            ws_offline.append([
                record['人员名称'],
                record['批价单号'],
                record['项目名称'],
                record['批价金额'],
                record['审批状态'],
                record['创建时间']
            ])

        ws_offline.column_dimensions['A'].width = 12
        ws_offline.column_dimensions['B'].width = 20
        ws_offline.column_dimensions['C'].width = 40
        ws_offline.column_dimensions['D'].width = 15
        ws_offline.column_dimensions['E'].width = 12
        ws_offline.column_dimensions['F'].width = 20

        # 3. 汇总数据
        ws_summary = wb.create_sheet('汇总统计', 0)

        summary_headers = ['人员名称', '线上笔数', '线上金额', '线下笔数', '线下金额', '合计金额']
        ws_summary.append(summary_headers)

        for cell in ws_summary[1]:
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        users = ['郭小会', '杨俊杰', '李华伟', '范敬']

        for user in users:
            online_records = [r for r in online_data if r['人员名称'] == user]
            offline_records = [r for r in offline_data if r['人员名称'] == user]

            online_count = len(online_records)
            online_amt = sum(r['批价金额'] for r in online_records)

            offline_count = len(offline_records)
            offline_amt = sum(r['批价金额'] for r in offline_records)

            total_amt = online_amt + offline_amt

            ws_summary.append([
                user,
                online_count,
                online_amt,
                offline_count,
                offline_amt,
                total_amt
            ])

        ws_summary.column_dimensions['A'].width = 12
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_summary.column_dimensions[col].width = 15

        # 保存文件
        result_file = '/Users/nijie/Documents/PMA/data/批价员详细信息导出_核对表.xlsx'
        os.makedirs(os.path.dirname(result_file), exist_ok=True)

        wb.save(result_file)

        print(f"\n✅ 导出文件已生成: {result_file}")
        print(f"\n📋 文件包含三个Sheet:")
        print(f"   1. 汇总统计 - 各员工的统计数据")
        print(f"   2. 线上数据 - sp8d系统中已批准的批价单 ({len(online_data)} 条)")
        print(f"   3. 线下数据 - Excel中的批价单 ({len(offline_data)} 条)")

        return result_file
    except Exception as e:
        print(f"❌ 生成导出文件失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    print("=" * 70)
    print("批价员详细信息导出工具")
    print("=" * 70)

    # 查询线上数据
    online_data = query_online_quotations()
    if online_data is None:
        print("❌ 无法查询线上数据")
        return

    # 提取线下数据
    offline_data = extract_offline_quotations()

    # 创建导出文件
    result_file = create_export_file(online_data, offline_data)

    if result_file:
        print(f"\n" + "=" * 70)
        print(f"✅ 导出完成!")
        print(f"=" * 70)
        print(f"文件位置: {result_file}")
        print(f"\n您可以下载该文件进行线下核对。")

if __name__ == '__main__':
    main()
