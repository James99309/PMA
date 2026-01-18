#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
李华伟批价单项目级别对比分析
- 从绩效采集表提取项目列表
- 从PMA导出数据提取项目列表
- 进行项目名称模糊匹配
- 生成详细对比分析报告
"""
import sys
import os
from difflib import SequenceMatcher
from datetime import datetime

# 路径修正 - 支持从任何位置运行
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

sys.path.insert(0, get_project_root())

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict


def read_performance_data():
    """从绩效采集表读取李华伟的项目数据"""
    file_path = '/Users/nijie/Downloads/2025年绩效数据采集表（12.31）.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['25年订单业绩明细25.12.31']

        projects = []

        # 绩效表的结构：表头在第2行，列B是"批价日期1"，列C是"项目名称"，列D是"销售负责人"，列F是"提货批价额"
        header_row = 2

        # 查找相关列的索引
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if not header:
                continue
            header_str = str(header).strip()
            col_map[header_str] = col_idx

        print(f"绩效表列映射: {col_map}")

        # 根据观察的结构设置列位置
        name_col = 4  # 销售负责人 (列D)
        project_col = 3  # 项目名称 (列C)
        amount_col = 6  # 提货批价额 (列F)

        # 数据行从header_row+1开始
        for row_idx in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row_idx, name_col).value if name_col else None
            project_name = ws.cell(row_idx, project_col).value if project_col else None
            amount = ws.cell(row_idx, amount_col).value if amount_col else None

            # 过滤李华伟的数据
            if name and '李华伟' in str(name) and project_name:
                try:
                    amount_val = float(amount) if amount else 0
                    if amount_val > 0:  # 只保留金额大于0的项目
                        projects.append({
                            'name': str(project_name).strip(),
                            'amount': amount_val,
                            'source': '绩效表'
                        })
                except (ValueError, TypeError):
                    continue

        wb.close()
        return projects

    except Exception as e:
        print(f"读取绩效表错误: {e}")
        import traceback
        traceback.print_exc()
        return []


def read_pma_data():
    """从PMA导出数据读取项目列表"""
    file_path = '/Users/nijie/Downloads/李华伟_批价单数据清单_2025年度_已清理.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['去重合并结果']

        projects = []

        # PMA数据的结构：表头在第3行，列B是"项目名称"，列F是"最终用金额"
        header_row = 3

        # 查找相关列的索引
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if not header:
                continue
            header_str = str(header).strip()
            col_map[header_str] = col_idx

        print(f"\nPMA表列映射: {col_map}")

        # 根据观察的结构设置列位置
        project_col = 2  # 项目名称 (列B)
        amount_col = 6  # 最终用金额 (列F)

        # 数据行从header_row+1开始
        for row_idx in range(header_row + 1, ws.max_row + 1):
            project_name = ws.cell(row_idx, project_col).value if project_col else None
            amount = ws.cell(row_idx, amount_col).value if amount_col else None

            if project_name:
                try:
                    amount_val = float(amount) if amount else 0
                    # 过滤测试项目、合计行
                    project_str = str(project_name).strip()
                    if ('测试' not in project_str and
                        '合计' not in project_str and
                        project_str and
                        amount_val > 0):
                        projects.append({
                            'name': project_str,
                            'amount': amount_val,
                            'source': 'PMA'
                        })
                except (ValueError, TypeError):
                    continue

        wb.close()
        return projects

    except Exception as e:
        print(f"读取PMA数据错误: {e}")
        import traceback
        traceback.print_exc()
        return []


def calculate_similarity(s1, s2):
    """计算两个字符串的相似度"""
    return SequenceMatcher(None, s1, s2).ratio()


def match_projects(pma_projects, perf_projects):
    """
    项目名称模糊匹配

    返回：
    - exact_matches: 精确匹配列表
    - fuzzy_matches: 模糊匹配列表（相似度>=80%）
    - pma_only: 只在PMA中的项目
    - perf_only: 只在绩效表中的项目
    - amount_diff: 金额不一致的项目
    """

    exact_matches = []
    fuzzy_matches = []
    pma_only = []
    perf_only = list(perf_projects)
    amount_diff = []

    matched_perf_indices = set()

    # 第一轮：精确匹配和模糊匹配
    for pma_proj in pma_projects:
        pma_name = pma_proj['name']
        best_match = None
        best_similarity = 0
        best_perf_idx = -1

        for perf_idx, perf_proj in enumerate(perf_projects):
            perf_name = perf_proj['name']
            similarity = calculate_similarity(pma_name, perf_name)

            if similarity > best_similarity:
                best_similarity = similarity
                best_match = perf_proj
                best_perf_idx = perf_idx

        if best_similarity >= 0.95:  # 95%以上为精确匹配
            exact_matches.append({
                'pma': pma_proj,
                'perf': best_match,
                'similarity': best_similarity,
                'amount_diff': abs(pma_proj['amount'] - best_match['amount'])
            })
            matched_perf_indices.add(best_perf_idx)

            # 检查金额差异
            if pma_proj['amount'] != best_match['amount']:
                amount_diff.append({
                    'name': pma_name,
                    'pma_amount': pma_proj['amount'],
                    'perf_amount': best_match['amount'],
                    'diff': pma_proj['amount'] - best_match['amount']
                })

        elif best_similarity >= 0.80:  # 80-95%为模糊匹配
            fuzzy_matches.append({
                'pma': pma_proj,
                'perf': best_match,
                'similarity': best_similarity,
                'amount_diff': abs(pma_proj['amount'] - best_match['amount'])
            })
            matched_perf_indices.add(best_perf_idx)

            # 检查金额差异
            if pma_proj['amount'] != best_match['amount']:
                amount_diff.append({
                    'name': pma_name,
                    'pma_amount': pma_proj['amount'],
                    'perf_amount': best_match['amount'],
                    'diff': pma_proj['amount'] - best_match['amount']
                })

        else:  # 未匹配
            pma_only.append(pma_proj)

    # 更新perf_only列表
    perf_only = [perf_projects[i] for i in range(len(perf_projects))
                 if i not in matched_perf_indices]

    return {
        'exact_matches': exact_matches,
        'fuzzy_matches': fuzzy_matches,
        'pma_only': pma_only,
        'perf_only': perf_only,
        'amount_diff': amount_diff
    }


def create_report_excel(match_result, pma_projects, perf_projects):
    """生成Excel对比分析报告"""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    alt_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def apply_header_style(cell):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Sheet1: 匹配结果总览
    ws1 = wb.create_sheet("匹配结果总览")

    row = 1
    ws1.merge_cells(f'A{row}:B{row}')
    cell = ws1[f'A{row}']
    cell.value = "李华伟批价单项目对比分析总览"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = center_align

    row += 2

    # 统计信息
    stats = [
        ("数据统计", ""),
        ("PMA项目总数", len(pma_projects)),
        ("绩效表项目总数", len(perf_projects)),
        ("", ""),
        ("匹配统计", ""),
        ("精确匹配", len(match_result['exact_matches'])),
        ("模糊匹配", len(match_result['fuzzy_matches'])),
        ("只在PMA中", len(match_result['pma_only'])),
        ("只在绩效表中", len(match_result['perf_only'])),
        ("金额不一致", len(match_result['amount_diff'])),
        ("", ""),
        ("金额统计", ""),
        ("PMA总金额", sum(p['amount'] for p in pma_projects)),
        ("绩效表总金额", sum(p['amount'] for p in perf_projects)),
        ("差异金额", sum(p['amount'] for p in pma_projects) - sum(p['amount'] for p in perf_projects)),
    ]

    for label, value in stats:
        if label == "":
            row += 1
            continue

        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = value

        if isinstance(value, (int, float)) and label != "数据统计" and label != "匹配统计" and label != "金额统计":
            ws1[f'B{row}'].number_format = '#,##0.00' if isinstance(value, float) else '0'
            ws1[f'B{row}'].alignment = number_align

        row += 1

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20

    # Sheet2: 匹配成功的项目
    ws2 = wb.create_sheet("匹配成功的项目")

    headers2 = ["序号", "项目名称(PMA)", "项目名称(绩效表)", "PMA金额", "绩效表金额", "相似度", "金额差异", "匹配类型"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(1, col)
        cell.value = header
        apply_header_style(cell)

    row = 2
    total_amount_exact = 0
    total_amount_fuzzy = 0

    # 精确匹配
    for idx, match in enumerate(match_result['exact_matches'], 1):
        ws2[f'A{row}'] = idx
        ws2[f'B{row}'] = match['pma']['name']
        ws2[f'C{row}'] = match['perf']['name']
        ws2[f'D{row}'] = match['pma']['amount']
        ws2[f'E{row}'] = match['perf']['amount']
        ws2[f'F{row}'] = match['similarity']
        ws2[f'G{row}'] = match['amount_diff']
        ws2[f'H{row}'] = "精确匹配"

        ws2[f'D{row}'].number_format = '#,##0.00'
        ws2[f'E{row}'].number_format = '#,##0.00'
        ws2[f'F{row}'].number_format = '0.00%'
        ws2[f'G{row}'].number_format = '#,##0.00'
        ws2[f'D{row}'].alignment = number_align
        ws2[f'E{row}'].alignment = number_align
        ws2[f'F{row}'].alignment = number_align
        ws2[f'G{row}'].alignment = number_align

        total_amount_exact += match['pma']['amount']
        row += 1

    # 模糊匹配
    for idx, match in enumerate(match_result['fuzzy_matches'], 1):
        ws2[f'A{row}'] = len(match_result['exact_matches']) + idx
        ws2[f'B{row}'] = match['pma']['name']
        ws2[f'C{row}'] = match['perf']['name']
        ws2[f'D{row}'] = match['pma']['amount']
        ws2[f'E{row}'] = match['perf']['amount']
        ws2[f'F{row}'] = match['similarity']
        ws2[f'G{row}'] = match['amount_diff']
        ws2[f'H{row}'] = "模糊匹配"

        ws2[f'D{row}'].number_format = '#,##0.00'
        ws2[f'E{row}'].number_format = '#,##0.00'
        ws2[f'F{row}'].number_format = '0.00%'
        ws2[f'G{row}'].number_format = '#,##0.00'
        ws2[f'D{row}'].alignment = number_align
        ws2[f'E{row}'].alignment = number_align
        ws2[f'F{row}'].alignment = number_align
        ws2[f'G{row}'].alignment = number_align

        ws2[f'H{row}'].fill = alt_fill

        total_amount_fuzzy += match['pma']['amount']
        row += 1

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25

    # Sheet3: 只在PMA中的项目
    ws3 = wb.create_sheet("只在PMA中的项目")

    headers3 = ["序号", "项目名称", "金额", "可能原因"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(1, col)
        cell.value = header
        apply_header_style(cell)

    row = 2
    pma_only_total = 0
    for idx, proj in enumerate(match_result['pma_only'], 1):
        ws3[f'A{row}'] = idx
        ws3[f'B{row}'] = proj['name']
        ws3[f'C{row}'] = proj['amount']
        ws3[f'D{row}'] = "未计入绩效表或时间差异"

        ws3[f'C{row}'].number_format = '#,##0.00'
        ws3[f'C{row}'].alignment = number_align
        ws3[f'B{row}'].alignment = left_align
        ws3[f'D{row}'].alignment = left_align

        pma_only_total += proj['amount']
        row += 1

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 25

    # Sheet4: 只在绩效表中的项目
    ws4 = wb.create_sheet("只在绩效表中的项目")

    headers4 = ["序号", "项目名称", "金额", "可能原因"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(1, col)
        cell.value = header
        apply_header_style(cell)

    row = 2
    perf_only_total = 0
    for idx, proj in enumerate(match_result['perf_only'], 1):
        ws4[f'A{row}'] = idx
        ws4[f'B{row}'] = proj['name']
        ws4[f'C{row}'] = proj['amount']
        ws4[f'D{row}'] = "未在PMA中或已删除"

        ws4[f'C{row}'].number_format = '#,##0.00'
        ws4[f'C{row}'].alignment = number_align
        ws4[f'B{row}'].alignment = left_align
        ws4[f'D{row}'].alignment = left_align

        perf_only_total += proj['amount']
        row += 1

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 25

    # Sheet5: 金额不一致的项目
    ws5 = wb.create_sheet("金额不一致的项目")

    headers5 = ["序号", "项目名称", "PMA金额", "绩效表金额", "差异金额", "差异比例"]
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(1, col)
        cell.value = header
        apply_header_style(cell)

    row = 2
    for idx, diff in enumerate(match_result['amount_diff'], 1):
        ws5[f'A{row}'] = idx
        ws5[f'B{row}'] = diff['name']
        ws5[f'C{row}'] = diff['pma_amount']
        ws5[f'D{row}'] = diff['perf_amount']
        ws5[f'E{row}'] = diff['diff']

        # 计算差异比例
        if diff['perf_amount'] != 0:
            ratio = diff['diff'] / diff['perf_amount']
        else:
            ratio = 0
        ws5[f'F{row}'] = ratio

        ws5[f'C{row}'].number_format = '#,##0.00'
        ws5[f'D{row}'].number_format = '#,##0.00'
        ws5[f'E{row}'].number_format = '#,##0.00'
        ws5[f'F{row}'].number_format = '0.00%'

        ws5[f'C{row}'].alignment = number_align
        ws5[f'D{row}'].alignment = number_align
        ws5[f'E{row}'].alignment = number_align
        ws5[f'F{row}'].alignment = number_align

        row += 1

    for col in range(1, len(headers5) + 1):
        ws5.column_dimensions[get_column_letter(col)].width = 18
    ws5.column_dimensions['B'].width = 30

    # Sheet6: 详细分析和建议
    ws6 = wb.create_sheet("详细分析和建议")

    row = 1
    ws6.merge_cells(f'A{row}:B{row}')
    cell = ws6[f'A{row}']
    cell.value = "差异原因分析与建议"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = center_align

    row += 2

    analysis = [
        ("一、匹配统计概览", ""),
        ("", ""),
        ("总体数据比对:", ""),
        (f"  • PMA项目数", len(pma_projects)),
        (f"  • 绩效表项目数", len(perf_projects)),
        (f"  • 精确匹配项目数", len(match_result['exact_matches'])),
        (f"  • 模糊匹配项目数", len(match_result['fuzzy_matches'])),
        (f"  • 只在PMA中的项目数", len(match_result['pma_only'])),
        (f"  • 只在绩效表中的项目数", len(match_result['perf_only'])),
        ("", ""),
        ("二、金额差异分析", ""),
        ("", ""),
        (f"  • PMA总金额", sum(p['amount'] for p in pma_projects)),
        (f"  • 绩效表总金额", sum(p['amount'] for p in perf_projects)),
        (f"  • 总金额差异", sum(p['amount'] for p in pma_projects) - sum(p['amount'] for p in perf_projects)),
        (f"  • 差异金额占比", (sum(p['amount'] for p in pma_projects) - sum(p['amount'] for p in perf_projects)) / sum(p['amount'] for p in perf_projects) if sum(p['amount'] for p in perf_projects) > 0 else 0),
        ("", ""),
        (f"  • 只在PMA中项目的总金额", pma_only_total),
        (f"  • 只在绩效表中项目的总金额", perf_only_total),
        (f"  • 匹配项目的精确匹配总金额", total_amount_exact),
        (f"  • 匹配项目的模糊匹配总金额", total_amount_fuzzy),
        ("", ""),
        ("三、可能的差异原因", ""),
        ("", ""),
        ("1. 项目列表差异（只在PMA中的14个项目）", ""),
        ("   可能原因：", ""),
        ("   • 这些项目可能是试验项目、测试项目或特殊订单", ""),
        ("   • 时间统计范围不同（绩效表可能只统计已交付项目）", ""),
        ("   • 数据分类不同（某些项目在PMA中可能分类为其他类型）", ""),
        ("   • 绩效表更新不及时", ""),
        ("", ""),
        ("2. 金额差异原因", ""),
        ("   • 匹配项目中金额不一致的项目数：", len(match_result['amount_diff'])),
        ("   • 可能原因：", ""),
        ("   • 两个系统的计价标准不同", ""),
        ("   • 人工审核调整导致的差异", ""),
        ("   • 汇率或税率计算方法不同", ""),
        ("", ""),
        ("四、核对建议", ""),
        ("", ""),
        ("1. 首先核对'只在PMA中的项目'", ""),
        ("   • 检查这些项目是否属于试验/测试/特殊订单", ""),
        ("   • 检查这些项目是否已在绩效表中用不同名称记录", ""),
        ("", ""),
        ("2. 其次核对'金额不一致的项目'", ""),
        ("   • 验证两个系统中的计价标准是否相同", ""),
        ("   • 检查是否存在人工调整", ""),
        ("", ""),
        ("3. 最后核对'只在绩效表中的项目'", ""),
        ("   • 确认这些项目是否已在PMA中删除", ""),
        ("   • 检查是否需要重新录入到PMA中", ""),
    ]

    for label, value in analysis:
        ws6[f'A{row}'] = label
        if value != "":
            ws6[f'B{row}'] = value
            if isinstance(value, (int, float)):
                ws6[f'B{row}'].number_format = '#,##0.00' if isinstance(value, float) else '0'
                ws6[f'B{row}'].alignment = number_align
        row += 1

    ws6.column_dimensions['A'].width = 40
    ws6.column_dimensions['B'].width = 25

    # 保存文件
    output_path = '/Users/nijie/Documents/PMA/data/temp/李华伟_批价单项目对比分析_20260116.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    print(f"\n✓ 报告已生成: {output_path}")
    return output_path


def main():
    print("=" * 60)
    print("李华伟批价单项目级别对比分析")
    print("=" * 60)

    print("\n[1] 读取绩效采集表...")
    perf_projects = read_performance_data()
    print(f"✓ 绩效表项目数: {len(perf_projects)}")
    print(f"✓ 绩效表总金额: ¥{sum(p['amount'] for p in perf_projects):,.2f}")

    print("\n[2] 读取PMA导出数据...")
    pma_projects = read_pma_data()
    print(f"✓ PMA项目数: {len(pma_projects)}")
    print(f"✓ PMA总金额: ¥{sum(p['amount'] for p in pma_projects):,.2f}")

    print("\n[3] 进行项目名称匹配...")
    match_result = match_projects(pma_projects, perf_projects)

    print(f"\n匹配结果统计:")
    print(f"  • 精确匹配: {len(match_result['exact_matches'])}")
    print(f"  • 模糊匹配: {len(match_result['fuzzy_matches'])}")
    print(f"  • 只在PMA中: {len(match_result['pma_only'])}")
    print(f"  • 只在绩效表中: {len(match_result['perf_only'])}")
    print(f"  • 金额不一致: {len(match_result['amount_diff'])}")

    print("\n[4] 生成对比分析报告...")
    output_path = create_report_excel(match_result, pma_projects, perf_projects)

    print("\n" + "=" * 60)
    print("✓ 分析完成!")
    print("=" * 60)

    return output_path


if __name__ == '__main__':
    main()
