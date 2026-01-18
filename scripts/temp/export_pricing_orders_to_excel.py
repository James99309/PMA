#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将云端SP8D系统的批价单数据导出为Excel文件
"""
import sys
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 路径修正
def get_project_root():
    current = os.path.dirname(os.path.abspath(__file__))
    while current != '/':
        if os.path.exists(os.path.join(current, 'app')) and \
           os.path.exists(os.path.join(current, 'run.py')):
            return current
        current = os.path.dirname(current)
    raise RuntimeError("无法找到项目根目录")

project_root = get_project_root()
sys.path.insert(0, project_root)

# 加载云端SP8D配置
env_file = os.path.join(project_root, '.env.sp8d.local')
if os.path.exists(env_file):
    load_dotenv(env_file, override=True)

from app import create_app, db
from app.models.user import User
from app.models.pricing_order import PricingOrder
from app.models.project import Project

def query_sales_manager_pricing_orders():
    """查询四个销售人员的批价单及详细对比数据"""
    app = create_app()
    with app.app_context():
        target_sales_people = ['lihuawei', 'gxh', 'fanjing', 'yangjj']
        all_results = {}

        for username in target_sales_people:
            user = User.query.filter(User.username == username).first()
            if not user:
                all_results[username] = {
                    'found': False,
                    'user_id': None,
                    'pricing_orders': []
                }
                continue

            pricing_orders = PricingOrder.query.join(
                Project, PricingOrder.project_id == Project.id
            ).filter(
                Project.vendor_sales_manager_id == user.id
            ).all()

            filtered_pos = []

            for po in pricing_orders:
                if po.created_at is None:
                    continue

                if po.created_at.tzinfo is None:
                    created_time = po.created_at
                else:
                    created_time = po.created_at.astimezone(ZoneInfo('Asia/Shanghai'))

                if created_time.year != 2025:
                    continue

                if po.status != 'approved':
                    continue

                if not po.project:
                    continue

                if '测试' in po.project.project_name:
                    continue

                retail_total = 0.0
                for detail in po.pricing_details:
                    if detail.market_price and detail.quantity:
                        retail_total += (detail.market_price * detail.quantity)

                pricing_discount_rate = po.pricing_total_discount_rate or 1.0
                settlement_discount_rate = po.settlement_total_discount_rate or 1.0

                po_dict = {
                    'id': po.id,
                    'po_number': po.order_number,
                    'project_name': po.project.project_name,
                    'retail_price_total': retail_total,
                    'pricing_discount_rate': pricing_discount_rate,
                    'pricing_amount': po.pricing_total_amount or 0.0,
                    'settlement_discount_rate': settlement_discount_rate,
                    'settlement_amount': po.settlement_total_amount or 0.0,
                    'created_at': po.created_at,
                }

                filtered_pos.append(po_dict)

            all_results[username] = {
                'found': True,
                'user_id': user.id,
                'real_name': user.real_name or user.username,
                'pricing_orders': filtered_pos
            }

        return all_results


def create_excel_report(all_results, output_path):
    """创建Excel报告"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 样式定义
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    summary_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # Sheet1: 汇总表
    ws1 = wb.create_sheet('汇总表', 0)
    ws1['A1'] = '云端SP8D系统 - 四个销售人员批价单数据汇总'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:H1')

    headers = ['销售人员', '批价单数', '零售价总和', '批价折扣率', '批价单金额', '结算折扣率', '结算单金额', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    row = 4
    grand_total_retail = 0
    grand_total_pricing = 0
    grand_total_settlement = 0
    total_po_count = 0

    for username in ['lihuawei', 'gxh', 'fanjing', 'yangjj']:
        result = all_results[username]
        if not result['found']:
            continue

        pos = result['pricing_orders']
        if not pos:
            continue

        po_count = len(pos)
        total_retail = sum(p['retail_price_total'] for p in pos)
        total_pricing = sum(p['pricing_amount'] for p in pos)
        total_settlement = sum(p['settlement_amount'] for p in pos)

        avg_pricing_discount = sum(p['pricing_discount_rate'] * 100 for p in pos) / len(pos) if pos else 0
        avg_settlement_discount = sum(p['settlement_discount_rate'] * 100 for p in pos) / len(pos) if pos else 0

        grand_total_retail += total_retail
        grand_total_pricing += total_pricing
        grand_total_settlement += total_settlement
        total_po_count += po_count

        ws1.cell(row=row, column=1).value = result['real_name']
        ws1.cell(row=row, column=2).value = po_count
        ws1.cell(row=row, column=3).value = total_retail
        ws1.cell(row=row, column=4).value = avg_pricing_discount / 100
        ws1.cell(row=row, column=5).value = total_pricing
        ws1.cell(row=row, column=6).value = avg_settlement_discount / 100
        ws1.cell(row=row, column=7).value = total_settlement
        ws1.cell(row=row, column=8).value = ''

        for col in range(1, 9):
            cell = ws1.cell(row=row, column=col)
            cell.border = border
            if col in [3, 5, 7]:
                cell.number_format = '#,##0.00'
            elif col in [4, 6]:
                cell.number_format = '0.0%'

        row += 1

    # 合计行
    ws1.cell(row=row, column=1).value = '合计'
    ws1.cell(row=row, column=2).value = total_po_count
    ws1.cell(row=row, column=3).value = grand_total_retail
    ws1.cell(row=row, column=4).value = (sum(
        sum(p['pricing_discount_rate'] * 100 for p in all_results[u]['pricing_orders'])
        for u in ['lihuawei', 'gxh', 'fanjing', 'yangjj']
        if all_results[u]['found']
    ) / total_po_count / 100) if total_po_count else 0
    ws1.cell(row=row, column=5).value = grand_total_pricing
    ws1.cell(row=row, column=6).value = (sum(
        sum(p['settlement_discount_rate'] * 100 for p in all_results[u]['pricing_orders'])
        for u in ['lihuawei', 'gxh', 'fanjing', 'yangjj']
        if all_results[u]['found']
    ) / total_po_count / 100) if total_po_count else 0
    ws1.cell(row=row, column=7).value = grand_total_settlement

    for col in range(1, 9):
        cell = ws1.cell(row=row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font
        cell.border = border
        if col in [3, 5, 7]:
            cell.number_format = '#,##0.00'
        elif col in [4, 6]:
            cell.number_format = '0.0%'

    # 列宽
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 20

    # Sheet2-5: 详细表
    sheet_order = ['lihuawei', 'gxh', 'fanjing', 'yangjj']
    for sheet_idx, username in enumerate(sheet_order, 1):
        result = all_results[username]
        if not result['found']:
            ws = wb.create_sheet(f'{result["real_name"] or username}', sheet_idx)
            ws['A1'] = '用户不存在'
            continue

        ws = wb.create_sheet(result['real_name'], sheet_idx)
        pos = result['pricing_orders']

        ws['A1'] = f"{result['real_name']} - 批价单详细列表"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:H1')

        detail_headers = ['序号', '批价单号', '项目名称', '零售价总和', '批价折扣率', '批价单金额', '结算折扣率', '结算单金额']
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        detail_row = 4
        for idx, po in enumerate(pos, 1):
            ws.cell(row=detail_row, column=1).value = idx
            ws.cell(row=detail_row, column=2).value = po['po_number']
            ws.cell(row=detail_row, column=3).value = po['project_name']
            ws.cell(row=detail_row, column=4).value = po['retail_price_total']
            ws.cell(row=detail_row, column=5).value = po['pricing_discount_rate']
            ws.cell(row=detail_row, column=6).value = po['pricing_amount']
            ws.cell(row=detail_row, column=7).value = po['settlement_discount_rate']
            ws.cell(row=detail_row, column=8).value = po['settlement_amount']

            for col in range(1, 9):
                cell = ws.cell(row=detail_row, column=col)
                cell.border = border
                if col in [4, 6, 8]:
                    cell.number_format = '#,##0.00'
                elif col in [5, 7]:
                    cell.number_format = '0.0%'
                if col == 1:
                    cell.alignment = center_align

            detail_row += 1

        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15

    wb.save(output_path)
    print(f"✅ Excel报告已生成: {output_path}")
    return output_path


def main():
    # 查询数据
    print("正在查询云端SP8D系统数据...")
    all_results = query_sales_manager_pricing_orders()

    # 生成Excel报告
    output_dir = os.path.join(project_root, 'data', 'temp')
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'四个销售人员批价单对比_SP8D_{timestamp}.xlsx')

    create_excel_report(all_results, output_path)
    print(f"\n📊 报告已保存到: {output_path}")


if __name__ == '__main__':
    main()
