import msoffcrypto
import io
from openpyxl import load_workbook

def read_excel_file(file_path, password):
    try:
        print(f"\n正在读取文件: {file_path}")
        print("="*50)
        
        # 创建一个临时的字节流来存储解密后的文件
        decrypted = io.BytesIO()
        
        # 打开加密的文件
        with open(file_path, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            # 使用密码解密
            office_file.load_key(password=password)
            # 将解密后的内容写入字节流
            office_file.decrypt(decrypted)
        
        # 将字节流的位置重置到开始
        decrypted.seek(0)
        
        # 使用openpyxl读取解密后的内容
        wb = load_workbook(filename=decrypted, data_only=True)
        
        # 获取所有工作表名称
        sheet_names = wb.sheetnames
        print("\n工作表列表：")
        for i, name in enumerate(sheet_names, 1):
            print(f"{i}. {name}")
            
            # 读取工作表内容
            ws = wb[name]
            print(f"\n工作表 {name} 的内容:")
            
            # 获取列名（第一行）
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    headers.append(f"Unnamed_{len(headers)}")
            print("\n列名:")
            print(headers)
            
            # 获取前10行数据
            print("\n前10行数据:")
            for idx, row in enumerate(list(ws.rows)[:10], 1):
                row_data = [cell.value for cell in row]
                print(f"第{idx}行: {row_data}")
            
            print("\n" + "-"*50)
        
        print(f"\n总共有 {len(sheet_names)} 个工作表")
        
    except Exception as e:
        print(f"读取文件时出错: {str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()
        if 'decrypted' in locals():
            decrypted.close()

if __name__ == "__main__":
    file_path = "evertac-pricelist.xlsm"
    password = "1505562299AaBb"  # Excel文件的密码
    read_excel_file(file_path, password) 