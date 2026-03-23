from flask import json

# Flask 2.3+ JSON兼容层
try:
    from flask.json import jsonify, loads, dumps
except (ImportError, AttributeError):
    from flask import current_app
    
    def jsonify(*args, **kwargs):
        return current_app.json.response(*args, **kwargs)
    
    def dumps(*args, **kwargs):
        return current_app.json.dumps(*args, **kwargs)
    
    def loads(*args, **kwargs):
        return current_app.json.loads(*args, **kwargs)

# 分开导入其他组件
from flask import Blueprint, request, render_template, flash, redirect, url_for, session
from flask_babel import gettext as _
import logging

# 设置日志记录器
logger = logging.getLogger(__name__)
try:
    from flask.json.provider import JSONProvider  
except ImportError:
    # 兼容低版本的Flask
    JSONProvider = None
    
from app.models.product import Product
from app.models.product_code import ProductSubcategory, ProductCategory
from app.extensions import db
from app.utils.product_helpers import get_products_by_name
from app.utils.dictionary_helpers import get_currency_type_options
from config import Config
import logging
from decimal import Decimal, InvalidOperation
from flask_login import login_required, current_user
from datetime import datetime
from sqlalchemy import func, and_, or_, text, case
from flask import url_for
from app.decorators import permission_required, admin_required
import os
import io
import uuid
from PIL import Image
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from werkzeug.utils import secure_filename
from flask import current_app
from app.utils.supabase_client import get_supabase_client
from app.services.product_file_service import get_product_file_service
from app.utils.query_filters import (
    extract_filter_params, apply_filters_to_query, extract_sort_params,
    extract_pagination_params
)
from sqlalchemy.orm import joinedload

logger = logging.getLogger(__name__)
# 创建蓝图 - 标准产品库
# 蓝图名 'product' 用于 url_for('product.xxx')
bp = Blueprint('product', __name__)


# ============================================================================
# MN编号重复检查函数 (从已废弃的product_management.py迁移 2025-12-26)
# ============================================================================

def check_mn_code_duplicate(mn_code, exclude_dev_product_id=None, exclude_product_id=None, exclude_config_id=None):
    """
    检查MN编号全局唯一性（跨研发产品库、标准产品库和规格配置版本）

    Args:
        mn_code: 要检查的MN编号
        exclude_dev_product_id: 排除的研发产品ID（研发库编辑时使用）
        exclude_product_id: 排除的产品ID（产品库编辑时使用）
        exclude_config_id: 排除的配置版本ID（规格配置编辑时使用）

    Returns:
        dict: {
            'is_duplicate': bool,
            'dev_products': [{...}],  # 研发库重复产品列表
            'standard_products': [{...}],  # 产品库重复产品列表
            'config_versions': [{...}],  # 规格配置版本重复列表
            'total_duplicates': int
        }
    """
    if not mn_code:
        return {'is_duplicate': False, 'dev_products': [], 'standard_products': [], 'config_versions': []}

    try:
        from app.models.dev_product import DevProduct

        duplicate_dev_products = []
        duplicate_standard_products = []

        # 检查研发产品库（已废弃，但保留历史数据检查）
        dev_query = DevProduct.query.filter(DevProduct.mn_code == mn_code)
        if exclude_dev_product_id:
            dev_query = dev_query.filter(DevProduct.id != exclude_dev_product_id)

        dev_duplicates = dev_query.all()

        for product in dev_duplicates:
            duplicate_dev_products.append({
                'id': product.id,
                'model': product.model,
                'name': product.name,
                'status': product.status,
                'category': product.category.name if product.category else '未知',
                'subcategory': product.subcategory.name if product.subcategory else '未知',
                'created_at': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '未知',
                'creator': product.creator.username if product.creator else '未知',
                'mn_code': product.mn_code,
                'source': '研发产品库'
            })

        # 检查标准产品库
        prod_query = Product.query.filter(Product.product_mn == mn_code)
        if exclude_product_id:
            prod_query = prod_query.filter(Product.id != exclude_product_id)

        standard_duplicates = prod_query.all()

        for product in standard_duplicates:
            category_name = product.category_obj.name if product.category_obj else (product.category or '未知')

            duplicate_standard_products.append({
                'id': product.id,
                'model': product.model,
                'name': product.name,
                'status': product.status,
                'category': category_name,
                'type': product.type,
                'created_at': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '未知',
                'owner': product.owner.username if product.owner else '未知',
                'mn_code': product.product_mn,
                'source': '标准产品库'
            })

        # 检查规格配置版本
        from app.models.spec_template import ProductConfiguration
        duplicate_config_versions = []

        config_query = ProductConfiguration.query.filter(
            ProductConfiguration.mn_code == mn_code,
            ProductConfiguration.deleted_at.is_(None)
        )
        if exclude_config_id:
            config_query = config_query.filter(ProductConfiguration.id != exclude_config_id)

        config_duplicates = config_query.all()

        for config in config_duplicates:
            duplicate_config_versions.append({
                'id': config.id,
                'config_code': config.config_code,
                'template_name': config.template.name if config.template else '未知',
                'region': config.region_name or config.region or '未知',
                'status': config.status,
                'mn_code': config.mn_code,
                'created_at': config.created_at.strftime('%Y-%m-%d %H:%M:%S') if config.created_at else '未知',
                'creator': config.creator.username if config.creator else '未知',
                'source': '规格配置版本'
            })

        is_duplicate = len(duplicate_dev_products) > 0 or len(duplicate_standard_products) > 0 or len(duplicate_config_versions) > 0

        if is_duplicate:
            logger.warning(f"检测到MN编号 {mn_code} 重复，研发产品: {len(duplicate_dev_products)}个, 标准产品: {len(duplicate_standard_products)}个, 规格配置: {len(duplicate_config_versions)}个")

        return {
            'is_duplicate': is_duplicate,
            'dev_products': duplicate_dev_products,
            'standard_products': duplicate_standard_products,
            'config_versions': duplicate_config_versions,
            'total_duplicates': len(duplicate_dev_products) + len(duplicate_standard_products) + len(duplicate_config_versions)
        }

    except Exception as e:
        logger.error(f"检查MN编号重复时出错: {str(e)}")
        return {'is_duplicate': False, 'dev_products': [], 'standard_products': [], 'config_versions': [], 'error': str(e)}


# ============================================================================
# 分类代码生成辅助函数 (用于跨系统导入时自动创建分类)
# ============================================================================

def generate_next_category_code_letter():
    """生成下一个可用的分类代码字母

    Returns:
        str: 下一个可用的大写字母 (A-Z)

    Raises:
        ValueError: 如果所有字母代码已用完
    """
    import string
    used = {c.code_letter for c in ProductCategory.query.all() if c.code_letter}
    for letter in string.ascii_uppercase:
        if letter not in used:
            return letter
    raise ValueError("所有分类字母代码已用完")


def generate_next_subcategory_code_letter(category_id):
    """生成下一个可用的子分类代码字符

    Args:
        category_id: 父分类ID

    Returns:
        str: 下一个可用的字符 (A-Z, 然后是 0-9)

    Raises:
        ValueError: 如果所有字符代码已用完
    """
    import string
    used = {s.code_letter for s in ProductSubcategory.query.filter_by(category_id=category_id).all() if s.code_letter}
    # 先尝试字母
    for letter in string.ascii_uppercase:
        if letter not in used:
            return letter
    # 再尝试数字
    for digit in string.digits:
        if digit not in used:
            return digit
    raise ValueError("所有子分类字符代码已用完")


# ============================================================
# 产品库筛选配置（与报价单/客户/项目管理保持一致的通用模式）
# ============================================================
# 注意：search 字段需要跨表搜索（subcategory名称），在查询中手动处理
PRODUCT_FILTER_CONFIG = {
    'product_type': {'type': 'exact', 'field': 'type'},
    'brand': {'type': 'exact', 'field': 'brand'},
    'status': {'type': 'exact', 'field': 'status'},
    # category 需要跨表筛选，在查询中手动处理
    # search 需要跨表查询，在查询中手动处理
}

# 文件验证和处理功能已移至 ProductFileService

# 本地PDF处理函数已移除，使用 ProductFileService


def _get_product_form_data():
    """获取产品表单所需的通用数据（分类、地区、品牌列表）

    Returns:
        dict: 包含 categories, regions, brands 的字典
    """
    from app.models.product_code import ProductCategory, ProductCodeField, ProductCodeFieldOption

    # 获取所有产品分类
    categories = ProductCategory.get_ordered_list()

    # 获取所有产品地区（从ProductCodeField获取，统一数据源）
    region_fields = ProductCodeField.query.filter_by(field_type='origin_location')\
                                          .order_by(ProductCodeField.position).all()

    regions = []
    for field in region_fields:
        # 获取编码（处理"?"情况）
        code = field.code or '0'
        if code == "?":
            # 从字段选项中获取第一个选项的编码
            option = ProductCodeFieldOption.query.filter_by(field_id=field.id).first()
            code = option.code if option else "0"

        regions.append({
            'id': field.id,
            'name': field.name,
            'code': code
        })

    # 获取现有品牌列表（去重、非空、排序）
    existing_brands = db.session.query(Product.brand)\
        .filter(Product.brand.isnot(None), Product.brand != '')\
        .distinct()\
        .order_by(Product.brand)\
        .all()
    brands = [b[0] for b in existing_brands]

    return {
        'categories': categories,
        'regions': regions,
        'brands': brands
    }


@bp.route('/products', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 产品库只检查product权限
def product_list():
    """产品列表页面 - 使用通用工具和Tailwind模板"""
    # ============================================================
    # 1. 使用通用工具提取参数
    # ============================================================
    filters = extract_filter_params(request.args, PRODUCT_FILTER_CONFIG)
    offset, limit = extract_pagination_params(request.args, default_limit=30, max_limit=100)

    # 提取变量（search 和 category 从 request.args 获取，因为需要跨表搜索）
    search = request.args.get('search', '').strip()
    product_type = filters.get('product_type', '')
    category = request.args.get('category', '').strip()
    brand = filters.get('brand', '')
    # 默认筛选"生产中"状态，用户手动选"全部状态"时 status 参数为空字符串
    status = filters.get('status', '') or ('active' if 'status' not in request.args else '')
    if status and status != filters.get('status', ''):
        filters['status'] = status
    category_manager = request.args.get('category_manager', '').strip()

    # 获取排序参数（默认按分类体系排序）
    valid_sort_fields = ['type', 'model', 'product_mn', 'brand', 'status', 'retail_price', 'created_at']
    sort_field, sort_order = extract_sort_params(
        request.args, default_sort='', default_order='asc',
        allowed_fields=valid_sort_fields
    )

    # ============================================================
    # 2. 构建查询（预加载关联数据）
    # ============================================================
    query = Product.query.options(
        joinedload(Product.category_obj),
        joinedload(Product.subcategory_obj),
        joinedload(Product.region_obj)
    )

    # 产品停产状态过滤：只有产品经理、解决方案经理和管理员可以查看停产产品
    if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
        query = query.filter(Product.status == 'active')

    # ============================================================
    # 3. 应用筛选（使用通用工具 + 手动处理特殊情况）
    # ============================================================
    # 应用通用筛选
    query = apply_filters_to_query(query, Product, filters, PRODUCT_FILTER_CONFIG)

    # 手动处理 search（跨表搜索）
    if search:
        search_term = f'%{search}%'
        query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
        query = query.filter(
            or_(
                ProductSubcategory.name.ilike(search_term),
                Product.product_mn.ilike(search_term),
                Product.model.ilike(search_term)
            )
        )

    # 手动处理 category（跨表筛选）
    if category:
        if not search:  # 如果没有 search，需要先 join
            query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
        query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
        query = query.filter(ProductCategory.name == category)

    # 按分类负责人筛选
    if category_manager:
        manager_id = int(category_manager)
        managed_cat_ids = [c.id for c in ProductCategory.query.filter_by(manager_id=manager_id).all()]
        if managed_cat_ids:
            query = query.filter(Product.category_id.in_(managed_cat_ids))
        else:
            query = query.filter(db.literal(False))

    # ============================================================
    # 4. 统计（在分页前）
    # ============================================================
    # 统计查询（单次条件聚合，不分页）
    stats_result = query.with_entities(
        func.count(Product.id).label('total'),
        func.count(case((Product.status == 'active', Product.id))).label('active'),
        func.count(case((Product.status == 'discontinued', Product.id))).label('discontinued'),
        func.count(case((Product.status == 'upcoming', Product.id))).label('upcoming'),
        func.coalesce(func.sum(Product.retail_price), 0).label('total_value'),
    ).first()

    total_count = stats_result.total or 0
    active_count = stats_result.active or 0
    discontinued_count = stats_result.discontinued or 0
    upcoming_count = stats_result.upcoming or 0
    total_value_result = stats_result.total_value or 0

    from app.utils.dictionary_helpers import prepare_stats_card_amount
    amount_data = prepare_stats_card_amount(float(total_value_result))

    # ============================================================
    # 5. 排序和分页
    # ============================================================
    if sort_field and hasattr(Product, sort_field):
        order_attr = getattr(Product, sort_field)
        query = query.order_by(order_attr.desc() if sort_order == 'desc' else order_attr.asc())
    else:
        # 默认排序：按分类体系排序（与导出功能一致）
        # 确保已经 join 了分类表
        if not search and not category:
            query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
            query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
        elif search and not category:
            query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
        query = query.order_by(
            ProductCategory.display_order.asc(),
            ProductCategory.id.asc(),
            ProductSubcategory.display_order.asc(),
            ProductSubcategory.id.asc(),
            Product.product_name.asc(),
            Product.id.asc()
        )

    # 分页
    products = query.offset(offset).limit(limit).all()
    has_more = (offset + limit) < total_count

    # ============================================================
    # 6. 检查是否为AJAX请求
    # ============================================================
    if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        rows_html = render_template('product/tw_list_rows.html', products=products)
        return jsonify({
            'html': rows_html,
            'has_more': has_more,
            'offset': offset,
            'total_count': total_count,
            'statistics': {
                'total': total_count,
                'active': active_count,
                'discontinued': discontinued_count,
                'upcoming': upcoming_count
            }
        })

    # ============================================================
    # 7. 获取筛选选项数据
    # ============================================================
    product_types = db.session.query(Product.type).distinct().filter(
        Product.type.isnot(None), Product.type != ''
    ).all()
    product_types = [{'value': t[0], 'label': t[0]} for t in product_types if t[0]]

    categories = ProductCategory.get_ordered_list()
    categories = [{'value': cat.name, 'label': (cat.name_en or cat.name) if Config.IS_OVS else cat.name} for cat in categories]

    brands = db.session.query(Product.brand).distinct().filter(
        Product.brand.isnot(None), Product.brand != ''
    ).all()
    brands = [{'value': b[0], 'label': b[0]} for b in brands if b[0]]

    status_options = [
        {'value': 'active', 'label': _('生产中')},
        {'value': 'discontinued', 'label': _('已停产')},
        {'value': 'upcoming', 'label': _('待上市')}
    ]

    # 获取有管理分类的负责人列表
    from app.models.user import User
    manager_ids = db.session.query(ProductCategory.manager_id).filter(
        ProductCategory.manager_id.isnot(None)
    ).distinct().all()
    manager_ids = [m[0] for m in manager_ids]
    managers = User.query.filter(User.id.in_(manager_ids)).all() if manager_ids else []
    manager_options = [
        {'value': str(m.id), 'label': m.real_name or m.username}
        for m in sorted(managers, key=lambda u: u.real_name or u.username)
    ]

    # ============================================================
    # 8. 构建配置
    # ============================================================
    filter_config = {
        'action_url': url_for('product.product_list'),
        'form_id': 'filterForm',
        'search_field': {
            'name': 'search',
            'label': _('搜索'),
            'placeholder': _('产品名称、MN号或型号'),
            'value': search,
        },
        'filter_fields': [
            {
                'name': 'product_type',
                'label': _('产品类型'),
                'all_option_text': _('全部类型'),
                'current_value': product_type,
                'options': product_types
            },
            {
                'name': 'category',
                'label': _('产品类别'),
                'all_option_text': _('全部类别'),
                'current_value': category,
                'options': categories
            },
            {
                'name': 'brand',
                'label': _('品牌'),
                'all_option_text': _('全部品牌'),
                'current_value': brand,
                'options': brands
            },
            {
                'name': 'status',
                'label': _('状态'),
                'all_option_text': _('全部状态'),
                'current_value': status,
                'options': status_options
            },
            {
                'name': 'category_manager',
                'label': _('负责人'),
                'all_option_text': _('全部负责人'),
                'current_value': category_manager,
                'options': manager_options
            }
        ],
    }

    list_config = {
        'module_name': 'product',
        'title': _('产品库管理'),
        'stats': {
            'cards': [
                {
                    'id': 'total',
                    'title': _('总产品数'),
                    'icon': 'fas fa-cube',
                    'value': total_count,
                    'amount': amount_data['value'],
                    'unit': _('个'),
                    'amount_unit': amount_data['unit'],
                    'color': 'primary',
                },
                {
                    'id': 'active',
                    'title': _('生产中'),
                    'icon': 'fas fa-play-circle',
                    'value': active_count,
                    'unit': _('个'),
                    'color': 'success',
                },
                {
                    'id': 'discontinued',
                    'title': _('已停产'),
                    'icon': 'fas fa-stop-circle',
                    'value': discontinued_count,
                    'unit': _('个'),
                    'color': 'danger',
                },
                {
                    'id': 'upcoming',
                    'title': _('待上市'),
                    'icon': 'fas fa-clock',
                    'value': upcoming_count,
                    'unit': _('个'),
                    'color': 'warning',
                }
            ]
        },
        'table': {
            # 列宽由 tw_data_table 组件自动计算：min=表头宽度, max=720px (45个中文字符)
            'columns': [
                {'key': 'product_name', 'field': 'product_name', 'label': _('产品名称')},
                {'key': 'model', 'field': 'model', 'label': _('型号')},
                {'key': 'specification', 'field': 'specification', 'label': _('规格')},
                {'key': 'brand', 'field': 'brand', 'label': _('品牌')},
                {'key': 'unit', 'field': 'unit', 'label': _('单位')},
                {'key': 'retail_price', 'field': 'retail_price', 'label': _('价格'), 'align': 'right'},
                {'key': 'product_mn', 'field': 'product_mn', 'label': _('MN号')},
                {'key': 'created_at', 'field': 'created_at', 'label': _('创建时间')}
            ]
        }
    }

    # ============================================================
    # 9. 获取模态框所需数据（分类带code_letter，区域带code）
    # ============================================================
    all_categories = ProductCategory.get_ordered_list()
    modal_categories = [
        {'id': cat.id, 'name': cat.name, 'name_en': cat.name_en or cat.name, 'code_letter': cat.code_letter or ''}
        for cat in all_categories
    ]

    # 获取销售区域（从ProductCodeField）
    from app.models.product_code import ProductCodeField, ProductCodeFieldOption
    region_fields = ProductCodeField.query.filter_by(field_type='origin_location')\
                                          .order_by(ProductCodeField.position).all()
    modal_regions = []
    for field in region_fields:
        code = field.code or '0'
        if code == "?":
            option = ProductCodeFieldOption.query.filter_by(field_id=field.id).first()
            code = option.code if option else "0"
        modal_regions.append({
            'id': field.id,
            'name': field.name,
            'name_en': field.name_en or field.name,
            'code': code
        })

    # 获取货币选项（使用字典定义）
    modal_currencies = get_currency_type_options()

    return render_template('product/tw_list.html',
                          list_config=list_config,
                          filter_config=filter_config,
                          products=products,
                          total_count=total_count,
                          offset=offset,
                          limit=limit,
                          has_more=has_more,
                          sort_field=sort_field,
                          sort_order=sort_order,
                          modal_categories=modal_categories,
                          modal_regions=modal_regions,
                          modal_currencies=modal_currencies)

@bp.route('/products/create', methods=['GET'])
@login_required
@permission_required('product', 'create')  # 添加产品创建权限装饰器
def create_product_page():
    """创建产品页面（标准产品）- 使用统一模板"""
    form_data = _get_product_form_data()
    form_data['product_type'] = 'standard'  # 标记为标准产品
    return render_template('product/create.html', **form_data)

@bp.route('/create', methods=['POST'])
@login_required
@permission_required('product', 'create')
def create():
    """处理产品创建表单提交（新版分类体系+规格管理）"""
    try:
        logger.debug('正在创建新产品（新版）...')

        # 获取表单数据
        product_type = request.form.get('type') or None
        product_status = request.form.get('status', 'active')
        category_id = request.form.get('category_id')
        subcategory_id = request.form.get('subcategory_id')
        region_id = request.form.get('region_id') or None
        # 兼容模态框的 model 字段和传统 API 的 product_model 字段
        product_model = request.form.get('model') or request.form.get('product_model')
        # 新产品：从前端获取 spec_mn 作为初始 product_mn
        spec_mn_from_form = request.form.get('spec_mn')
        product_mn = request.form.get('product_mn') or spec_mn_from_form
        # 兼容模态框的 name 字段和传统 API 的 product_name 字段
        product_name = request.form.get('name') or request.form.get('product_name') or None
        brand = request.form.get('brand') or None
        unit = request.form.get('unit')
        retail_price = request.form.get('retail_price')
        currency = request.form.get('currency', Config.DEFAULT_CURRENCY)
        description = request.form.get('description')
        is_vendor_product = request.form.get('is_vendor_product') == 'on'

        # 验证必填字段
        if not all([category_id, subcategory_id, region_id, product_model, product_mn]):
            return jsonify({'success': False, 'message': '请填写所有必填字段（包括销售区域）'}), 400

        # 验证MN号全局唯一性（跨研发库和产品库）
        duplicate_info = check_mn_code_duplicate(product_mn)
        if duplicate_info['is_duplicate']:
            # 构建详细错误信息
            error_parts = []
            if duplicate_info['dev_products']:
                dev_models = ', '.join([p['model'] for p in duplicate_info['dev_products']])
                error_parts.append(f"研发产品库: {dev_models}")
            if duplicate_info['standard_products']:
                std_models = ', '.join([p['model'] for p in duplicate_info['standard_products']])
                error_parts.append(f"产品库: {std_models}")

            error_msg = f"MN编号 {product_mn} 已被以下产品使用：" + "；".join(error_parts)
            return jsonify({'success': False, 'message': error_msg}), 400

        # 处理零售价格
        if retail_price:
            try:
                retail_price = Decimal(retail_price)
            except (InvalidOperation, ValueError):
                return jsonify({'success': False, 'message': '零售价格格式不正确'}), 400
        else:
            retail_price = Decimal('0.00')

        # 创建新产品
        new_product = Product(
            type=product_type,
            status=product_status,
            category_id=int(category_id),
            subcategory_id=int(subcategory_id),
            region_id=int(region_id) if region_id else None,
            model=product_model,
            product_mn=product_mn,
            product_name=product_name,  # 独立的产品名称
            brand=brand,
            unit=unit,
            retail_price=retail_price,
            currency=currency,
            specification=description,  # 使用旧字段存储描述
            source_type='manual',  # 手动创建
            owner_id=current_user.id,
            is_vendor_product=is_vendor_product
        )

        # 保存产品以获取ID
        db.session.add(new_product)
        db.session.flush()

        # 保存规格数据
        spec_names = request.form.getlist('spec_name[]')
        spec_values = request.form.getlist('spec_value[]')
        spec_codes = request.form.getlist('spec_option_codes[]')
        spec_field_ids = request.form.getlist('spec_field_ids[]')

        if spec_names:
            spec_data_list = []
            for i in range(len(spec_names)):
                if spec_names[i].strip():  # 只处理非空规格
                    spec_data_list.append({
                        'field_name': spec_names[i],
                        'field_value': spec_values[i] if i < len(spec_values) else '',
                        'field_code': spec_codes[i] if i < len(spec_codes) and spec_codes[i] != '0' else None,
                        'action': 'create'
                    })

            if spec_data_list:
                # 使用统一的 SpecService 保存规格
                from app.services.spec_service import SpecService
                result = SpecService.save_specs(SpecService.TYPE_PRODUCT, new_product.id, spec_data_list)
                if not result['success']:
                    db.session.rollback()
                    return jsonify({'success': False, 'message': f'保存规格数据失败: {result.get("message", "未知错误")}'}), 400

                logger.debug(f'保存了规格数据，spec_mn: {result.get("spec_mn")}')

                # SpecService 已经自动处理了 spec_mn 生成，检查并同步 product_mn
                if result.get('spec_mn'):
                    # 新产品：product_mn 也设置为规格MN（如果还没有设置）
                    if not new_product.product_mn:
                        new_product.product_mn = result['spec_mn']
                    logger.debug(f'生成规格MN: {result["spec_mn"]}, product_mn: {new_product.product_mn}')

                # 生成编码定义快照
                from app.utils.product_helpers import generate_product_snapshot
                snapshot = generate_product_snapshot(product=new_product, source="manual_create")
                if snapshot:
                    new_product.code_definition_snapshot = snapshot

        # 提交事务
        db.session.commit()

        logger.info(f'产品创建成功: ID={new_product.id}, MN={new_product.product_mn}, spec_mn={new_product.spec_mn}, 型号={new_product.model}')

        return jsonify({
            'success': True,
            'message': '产品创建成功',
            'redirect': url_for('product.product_list')
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f'创建产品时出错: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'message': f'创建产品失败: {str(e)}'}), 500

@bp.route('/products/<int:id>/edit', methods=['GET'])
@login_required
@permission_required('product', 'edit')  # 添加产品编辑权限装饰器
def edit_product_page(id):
    """编辑产品页面"""
    from sqlalchemy.orm import joinedload
    from app.models.product_code import ProductSubcategory
    from app.models.product_spec import ProductSpec

    # 使用joinedload预加载关联数据
    product = db.session.query(Product).options(
        joinedload(Product.category_obj),
        joinedload(Product.subcategory_obj),
        joinedload(Product.region_obj)
    ).filter_by(id=id).first_or_404()

    # 使用统一函数获取通用数据（分类、地区、品牌列表）
    form_data = _get_product_form_data()

    # 获取当前分类下的子分类列表
    subcategories = []
    if product.category_id:
        subcategories = ProductSubcategory.query.filter_by(
            category_id=product.category_id
        ).order_by(ProductSubcategory.display_order).all()

    # 获取产品规格数据
    specs_db = ProductSpec.query.filter_by(product_id=id).order_by(ProductSpec.display_order).all()

    # 预先获取该子分类下所有字段，用于查找field_id和use_in_code
    # field_info 结构: {name: {'id': field_id, 'use_in_code': bool}}
    from app.models.product_code import ProductCodeField
    field_info = {}
    if product.subcategory_id:
        # 子分类级字段
        subcat_fields = ProductCodeField.query.filter_by(
            subcategory_id=product.subcategory_id,
            field_type='spec'
        ).all()
        for field in subcat_fields:
            field_info[field.name] = {'id': field.id, 'use_in_code': field.use_in_code}

        # 分类级继承字段
        if product.category_id:
            cat_fields = ProductCodeField.query.filter(
                ProductCodeField.category_id == product.category_id,
                ProductCodeField.subcategory_id.is_(None),
                ProductCodeField.field_type == 'spec'
            ).all()
            for field in cat_fields:
                field_info[field.name] = {'id': field.id, 'use_in_code': field.use_in_code}

    # 将ProductSpec对象转换为字典列表（方便前端使用）
    specs = []
    for spec in specs_db:
        # 获取字段信息（用于判断是否为编码规格）
        spec_field_info = field_info.get(spec.field_name, {})
        is_coded = spec_field_info.get('use_in_code', False)

        spec_dict = {
            'field_name': spec.field_name,
            'field_value': spec.field_value,
            'field_code': spec.field_code,
            'field_id': spec_field_info.get('id'),  # 添加 field_id
            'display_order': spec.display_order,  # 用于分类内排序
            'is_saved': True,  # 标记为已保存的规格，前端将以只读方式显示
            'is_coded': is_coded,  # 是否为编码规格（基于 use_in_code 判断）
            'include_in_description': spec.include_in_description if spec.include_in_description is not None else False
        }
        specs.append(spec_dict)

    # 获取锁定状态
    is_mn_locked = product.is_mn_locked or False

    # 管理员绕过锁定限制
    is_admin = current_user.role == 'admin'

    # 检查是否被报价单引用（按MN编号检查，MN是唯一标识）
    from app.models.quotation import QuotationDetail
    is_referenced = QuotationDetail.query.filter(
        QuotationDetail.product_mn == product.product_mn
    ).count() > 0

    # 检查关键字段是否有值（用于智能锁定控制）
    has_category = product.category_id is not None
    has_subcategory = product.subcategory_id is not None
    has_region = product.region_id is not None
    has_model = bool(product.model and product.model.strip())
    has_specs = len(specs) > 0

    return render_template('product/create.html',
                         product=product,
                         subcategories=subcategories,
                         specs=specs,
                         product_type='standard',  # 标识为标准产品
                         is_mn_locked=is_mn_locked,
                         is_admin=is_admin,
                         is_referenced=is_referenced,
                         has_category=has_category,
                         has_subcategory=has_subcategory,
                         has_region=has_region,
                         has_model=has_model,
                         has_specs=has_specs,
                         **form_data)

# API路由
@bp.route('/products/ajax', methods=['GET'])
@login_required
@permission_required('product', 'view')
def product_list_ajax():
    """产品列表AJAX端点 - 支持无限滚动和Tailwind模板"""
    try:
        from app.models.product_code import ProductCategory, ProductSubcategory

        # ============================================================
        # 1. 提取筛选参数
        # ============================================================
        search = request.args.get('search', '').strip()
        product_type = request.args.get('product_type', '').strip()
        category = request.args.get('category', '').strip()
        brand = request.args.get('brand', '').strip()
        # 默认筛选"生产中"状态，用户手动选"全部状态"时 status 参数为空字符串
        status = request.args.get('status', '').strip() or ('active' if 'status' not in request.args else '')
        category_manager = request.args.get('category_manager', '').strip()
        # ============================================================
        # 2. 提取排序和分页参数（默认按分类体系排序）
        # ============================================================
        sort_field = request.args.get('sort', '')
        sort_order = request.args.get('order', 'asc')
        offset = request.args.get('offset', 0, type=int)
        limit = request.args.get('limit', 50, type=int)

        # ============================================================
        # 3. 构建基础查询（带关联预加载）
        # ============================================================
        query = Product.query.options(
            joinedload(Product.category_obj),
            joinedload(Product.subcategory_obj)
        ).filter(Product.is_deleted == False)

        # 产品停产状态过滤：只有产品经理、解决方案经理和管理员可以查看停产产品
        if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            query = query.filter(Product.status == 'active')

        # ============================================================
        # 4. 应用筛选条件
        # ============================================================
        # 应用搜索条件（搜索产品名称、MN、型号）
        if search:
            search_term = f'%{search}%'
            query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
            query = query.filter(
                or_(
                    ProductSubcategory.name.ilike(search_term),
                    Product.product_mn.ilike(search_term),
                    Product.model.ilike(search_term)
                )
            )

        if product_type:
            query = query.filter(Product.type == product_type)
        if category:
            if not search:
                query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
            query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
            query = query.filter(ProductCategory.name == category)
        if brand:
            query = query.filter(Product.brand == brand)
        if status:
            query = query.filter(Product.status == status)

        # 按分类负责人筛选
        if category_manager:
            manager_id = int(category_manager)
            managed_cat_ids = [c.id for c in ProductCategory.query.filter_by(manager_id=manager_id).all()]
            if managed_cat_ids:
                query = query.filter(Product.category_id.in_(managed_cat_ids))
            else:
                query = query.filter(db.literal(False))

        # ============================================================
        # 5. 应用排序
        # ============================================================
        if sort_field and hasattr(Product, sort_field):
            field = getattr(Product, sort_field)
            query = query.order_by(field.desc() if sort_order == 'desc' else field.asc())
        else:
            # 默认排序：按分类体系排序（与导出功能一致）
            # 确保已经 join 了分类表
            if not search and not category:
                query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
                query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
            elif search and not category:
                query = query.outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)
            query = query.order_by(
                ProductCategory.display_order.asc(),
                ProductCategory.id.asc(),
                ProductSubcategory.display_order.asc(),
                ProductSubcategory.id.asc(),
                Product.product_name.asc(),
                Product.id.asc()
            )

        # ============================================================
        # 6. 获取总数和分页数据
        # ============================================================
        total_count = query.count()
        products = query.offset(offset).limit(limit).all()
        has_more = (offset + limit) < total_count

        # ============================================================
        # 7. 计算统计数据
        # ============================================================
        active_count = Product.query.filter(Product.status == 'active').count()
        discontinued_count = Product.query.filter(Product.status == 'discontinued').count()
        upcoming_count = Product.query.filter(Product.status == 'upcoming').count()

        # ============================================================
        # 8. 渲染产品HTML（使用Tailwind模板）
        # ============================================================
        products_html = render_template('product/tw_list_rows.html', products=products)

        return jsonify({
            'html': products_html,
            'has_more': has_more,
            'offset': offset,
            'total_count': total_count,
            'statistics': {
                'total': total_count,
                'active': active_count,
                'discontinued': discontinued_count,
                'upcoming': upcoming_count
            }
        })

    except Exception as e:
        logger.error(f'获取产品列表AJAX失败: {str(e)}')
        import traceback
        logger.error(f'详细错误: {traceback.format_exc()}')
        return jsonify({
            'success': False,
            'message': f'获取产品列表失败: {str(e)}'
        }), 500

@bp.route('/api/products', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_products():
    """获取产品列表API"""
    try:
        logger.debug('正在获取产品列表...')
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search_term = request.args.get('search', '')
        sort_by = request.args.get('sort_by', 'id')  # 默认按ID排序
        sort_order = request.args.get('sort_order', 'asc')  # 默认升序
        filter_field = request.args.get('filter_field', '')
        filter_value = request.args.get('filter_value', '')
        
        # 构建查询
        query = Product.query
        
        # 去除数据所有权过滤
        # 如果用户通过了permission_required('product', 'view')装饰器，就应该能查看所有产品
        
        # 产品停产状态过滤：只有产品经理、解决方案经理和管理员可以查看停产产品
        if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            # 其他角色只能看到生产中的产品（status = 'active'）
            query = query.filter(Product.status == 'active')
        
        # 应用搜索条件
        if search_term:
            search_term = f'%{search_term}%'
            # 先join ProductSubcategory以支持新字段搜索
            query = query.outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)
            query = query.filter(
                or_(
                    ProductSubcategory.name.ilike(search_term),
                    Product.product_mn.ilike(search_term),
                    Product.model.ilike(search_term),
                    Product.type.ilike(search_term)
                )
            )
        
        # 应用字段筛选
        if filter_field and filter_value:
            if filter_field == 'status':
                # 状态字段的特殊处理
                query = query.filter(Product.status == filter_value)
            else:
                # 动态构建筛选条件
                filter_attr = getattr(Product, filter_field, None)
                if filter_attr:
                    query = query.filter(filter_attr.ilike(f'%{filter_value}%'))
        
        # 应用排序
        if sort_by and hasattr(Product, sort_by):
            sort_attr = getattr(Product, sort_by)
            if sort_order == 'desc':
                query = query.order_by(sort_attr.desc())
            else:
                query = query.order_by(sort_attr.asc())
        else:
            # 默认按ID升序排序
            query = query.order_by(Product.id.asc())
        
        # 执行分页查询
        pagination = query.paginate(page=page, per_page=per_page)
        
        # 转换为JSON格式
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        result = {
            'items': [],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        }
        
        # 创建用户ID到用户真实姓名的映射
        from app.models.user import User
        user_map = {}
        # 收集所有产品的所有者ID
        owner_ids = [p.owner_id for p in pagination.items if p.owner_id]
        # 如果有所有者ID则查询对应的用户信息
        if owner_ids:
            users = User.query.filter(User.id.in_(set(owner_ids))).all()
            for user in users:
                # 优先使用真实姓名，如果没有则使用用户名
                user_map[user.id] = user.real_name if user.real_name else user.username
        
        for p in pagination.items:
            try:
                product_dict = {
                    'id': p.id,
                    'type': p.type,
                    'category': p.category_name,  # 使用智能属性
                    'product_mn': p.product_mn,
                    'product_name': p.name,  # 使用智能属性
                    'product_name_en': p.name_en,  # 英文产品名称
                    'model': p.model,
                    'specification': p.specification,
                    'brand': p.brand,
                    'unit': p.unit,
                    'retail_price': decimal_to_float(p.retail_price) if p.retail_price else 0,
                    'currency': p.currency if hasattr(p, 'currency') else Config.DEFAULT_CURRENCY,  # 添加货币字段
                    'status': p.status,
                    'is_vendor_product': p.is_vendor_product if hasattr(p, 'is_vendor_product') else False,  # 添加厂商产品标记
                    'points': p.points,
                    'points_tier': p.points_tier,
                    'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None,
                    'updated_at': p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else None,
                    'owner_id': p.owner_id,  # 添加所有者ID
                    'owner_name': user_map.get(p.owner_id, '未指定') if p.owner_id else '未指定',  # 添加所有者名称
                    'image_path': p.image_path  # 添加图片路径
                }
                result['items'].append(product_dict)
            except Exception as e:
                logger.error(f'处理产品时出错: {p.id} - {str(e)}')
                continue
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f'获取产品列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/categories', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product_categories():
    """获取去重后的产品类别列表"""
    try:
        logger.debug('正在获取产品类别列表...')
        # 直接从ProductCategory表查询,按display_order排序
        from app.models.product_code import ProductCategory
        categories = ProductCategory.get_ordered_list()

        # 提取类别名称
        category_list = [cat.name for cat in categories]

        logger.debug(f'找到 {len(category_list)} 个类别')
        return jsonify(category_list)
        
    except Exception as e:
        logger.error(f'获取产品类别列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品类别列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/by-category', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_products_by_category():
    """获取指定类别的产品列表"""
    try:
        category = request.args.get('category', '')
        logger.debug(f'正在获取类别 "{category}" 的产品列表...')
        
        if not category:
            return jsonify([])
        
        # 查询指定类别的产品
        products = Product.query.filter_by(
            category=category,
            status='active'
        ).all()
        
        logger.debug(f'找到 {len(products)} 个产品')
        
        # 小数类型转换为浮点数
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        # 构建完整产品信息
        result = []
        for p in products:
            try:
                product_dict = {
                    'id': p.id,
                    'product_name': p.name,  # 使用智能属性
                    'model': p.model,
                    'specification': p.specification,
                    'brand': p.brand,
                    'unit': p.unit,
                    'retail_price': decimal_to_float(p.retail_price) if p.retail_price else 0,
                    'currency': p.currency if hasattr(p, 'currency') else Config.DEFAULT_CURRENCY,  # 添加货币字段
                    'product_mn': p.product_mn
                }
                result.append(product_dict)
                logger.debug(f'成功处理产品: {p.product_name}, MN: {p.product_mn}')
            except Exception as e:
                logger.error(f'处理产品时出错: {p.id} - {str(e)}')
                continue
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f'获取类别产品列表时出错: {str(e)}')
        return jsonify({
            'error': '获取类别产品列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/by-name', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_products_by_name():
    """按名称获取产品列表"""
    try:
        product_name = request.args.get('product_name', '')
        logger.debug(f'正在获取产品 "{product_name}" 的型号列表...')
        
        if not product_name:
            return jsonify([])

        # 查询指定产品名称的产品型号（使用公共辅助函数）
        products = get_products_by_name(product_name)
        
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        result = []
        for p in products:
            try:
                product_dict = {
                    'id': p.id,
                    'model': p.model,
                    'product_mn': p.product_mn,
                    'specification': p.specification,
                    'brand': p.brand,
                    'unit': p.unit,
                    'retail_price': decimal_to_float(p.retail_price) if p.retail_price else 0,
                    'currency': p.currency if hasattr(p, 'currency') else Config.DEFAULT_CURRENCY  # 添加货币字段
                }
                result.append(product_dict)
            except Exception as e:
                logger.error(f'处理产品时出错: {p.id} - {str(e)}')
                continue
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f'获取产品型号列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品型号列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/models', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product_models():
    """获取指定类别和产品名称的型号列表"""
    try:
        category = request.args.get('category', '')
        product_name = request.args.get('product_name', '')
        logger.debug(f'正在获取产品型号，类别: "{category}", 产品名称: "{product_name}"')
        
        if not category or not product_name:
            return jsonify([])
        
        # 查询指定类别和产品名称的产品（包括停产产品）
        products = Product.query.filter_by(
            category=category,
            product_name=product_name
        ).order_by(Product.id).all()
        
        logger.debug(f'找到 {len(products)} 个产品')
        
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        result = []
        for p in products:
            try:
                product_dict = {
                    'id': p.id,
                    'product_name': p.name,  # 使用智能属性
                    'category': p.category_name,  # 使用智能属性
                    'model': p.model,
                    'product_model': p.model,  # 兼容性字段
                    'product_mn': p.product_mn,
                    'mn': p.product_mn,  # 兼容性字段
                    'specification': p.specification,
                    'product_spec': p.specification,  # 兼容性字段
                    'product_desc': p.specification,  # 兼容性字段
                    'spec': p.specification,  # 兼容性字段
                    'brand': p.brand,
                    'unit': p.unit,
                    'retail_price': decimal_to_float(p.retail_price) if p.retail_price else 0,
                    'market_price': decimal_to_float(p.retail_price) if p.retail_price else 0,  # 兼容性字段
                    'currency': p.currency if hasattr(p, 'currency') else Config.DEFAULT_CURRENCY,
                    'status': p.status,
                    'product_status': p.status  # 兼容性字段
                }
                result.append(product_dict)
                logger.debug(f'成功处理产品: {p.product_name}, 型号: {p.model}, MN: {p.product_mn}')
            except Exception as e:
                logger.error(f'处理产品时出错: {p.id} - {str(e)}')
                continue
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f'获取产品型号列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品型号列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/<int:product_id>/upload-image', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def upload_product_image(product_id):
    """上传产品图片（支持分类共享）

    参数:
        update_category: true=覆盖分类文件, false=仅用于此产品, 不传=检查是否需要确认
    """
    file_service = get_product_file_service()
    image_file = request.files.get('image')

    if not image_file or not image_file.filename:
        return jsonify({'success': False, 'error': '请选择要上传的图片文件'}), 400

    # 获取 update_category 参数
    update_category_str = request.form.get('update_category')
    if update_category_str is None:
        update_category = None  # 需要检查是否需要确认
    else:
        update_category = update_category_str.lower() == 'true'

    result = file_service.upload_file_with_category(product_id, image_file, 'image', update_category)
    status_code = 200 if result.get('success') or result.get('need_confirm') else 400
    return jsonify(result), status_code

@bp.route('/api/products/<int:product_id>/upload-pdf', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def upload_product_pdf(product_id):
    """上传产品PDF文档（支持分类共享）

    参数:
        update_category: true=覆盖分类文件, false=仅用于此产品, 不传=检查是否需要确认
    """
    file_service = get_product_file_service()
    pdf_file = request.files.get('pdf')

    if not pdf_file or not pdf_file.filename:
        return jsonify({'success': False, 'error': '请选择要上传的PDF文件'}), 400

    # 获取 update_category 参数
    update_category_str = request.form.get('update_category')
    if update_category_str is None:
        update_category = None  # 需要检查是否需要确认
    else:
        update_category = update_category_str.lower() == 'true'

    result = file_service.upload_file_with_category(product_id, pdf_file, 'pdf', update_category)
    status_code = 200 if result.get('success') or result.get('need_confirm') else 400
    return jsonify(result), status_code


@bp.route('/api/products/<int:product_id>/clear-image', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def clear_product_image(product_id):
    """清除产品图片（仅清除产品自身文件，不影响分类共享文件）"""
    file_service = get_product_file_service()
    result = file_service.clear_file(product_id, 'image')
    status_code = 200 if result.get('success') else 400
    return jsonify(result), status_code


@bp.route('/api/products/<int:product_id>/clear-pdf', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def clear_product_pdf(product_id):
    """清除产品PDF（仅清除产品自身文件，不影响分类共享文件）"""
    file_service = get_product_file_service()
    result = file_service.clear_file(product_id, 'pdf')
    status_code = 200 if result.get('success') else 400
    return jsonify(result), status_code


@bp.route('/api/products/<int:product_id>/upload-svg', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def upload_product_svg(product_id):
    """上传产品SVG图标（直接存入数据库TEXT字段）"""
    product = Product.query.get_or_404(product_id)
    svg_file = request.files.get('svg')
    if not svg_file or not svg_file.filename:
        return jsonify({'success': False, 'error': '请选择SVG文件'}), 400
    if not svg_file.filename.lower().endswith('.svg'):
        return jsonify({'success': False, 'error': '仅支持SVG格式'}), 400
    svg_content = svg_file.read().decode('utf-8')
    if '<svg' not in svg_content.lower():
        return jsonify({'success': False, 'error': '无效的SVG文件'}), 400
    product.icon_svg = svg_content
    db.session.commit()
    return jsonify({'success': True, 'message': 'SVG图标已上传'})


@bp.route('/api/products/<int:product_id>/delete-svg', methods=['DELETE'])
@login_required
@permission_required('product', 'edit')
def delete_product_svg(product_id):
    """删除产品SVG图标"""
    product = Product.query.get_or_404(product_id)
    product.icon_svg = None
    db.session.commit()
    return jsonify({'success': True, 'message': 'SVG图标已删除'})


@bp.route('/api/products/<int:product_id>/category-file-status', methods=['GET'])
@login_required
@permission_required('product', 'view')
def get_category_file_status(product_id):
    """获取产品子分类的文件状态（用于判断是否需要确认覆盖）

    注意：image_path 和 pdf_path 属性在 ProductSubcategory 上，不在 ProductCategory 上
    """
    from app.models.product_code import ProductSubcategory

    product = Product.query.get_or_404(product_id)
    if not product.subcategory_id:
        return jsonify({
            'success': True,
            'has_category': False,
            'category_image': None,
            'category_pdf': None
        })

    subcategory = ProductSubcategory.query.get(product.subcategory_id)
    if not subcategory:
        return jsonify({
            'success': True,
            'has_category': False,
            'category_image': None,
            'category_pdf': None
        })

    return jsonify({
        'success': True,
        'has_category': True,
        'category_name': subcategory.name,
        'category_image': subcategory.image_path,
        'category_pdf': subcategory.pdf_path
    })

@bp.route('/api/products/<int:product_id>/upload-files', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def upload_product_files(product_id):
    """批量上传产品文件（保持向后兼容）"""
    file_service = get_product_file_service()
    image_file = request.files.get('image')
    pdf_file = request.files.get('pdf')
    
    if not image_file and not pdf_file:
        return jsonify({'success': False, 'error': '请至少选择一个文件上传'}), 400
    
    result = {'success': True, 'image_url': None, 'pdf_url': None, 'errors': []}
    
    # 上传图片
    if image_file and image_file.filename:
        image_result = file_service.upload_image(product_id, image_file)
        if image_result['success']:
            result['image_url'] = image_result['image_url']
        else:
            result['errors'].append(image_result['error'])
    
    # 上传PDF
    if pdf_file and pdf_file.filename:
        pdf_result = file_service.upload_pdf(product_id, pdf_file)
        if pdf_result['success']:
            result['pdf_url'] = pdf_result['pdf_url']
        else:
            result['errors'].append(pdf_result['error'])
    
    result['success'] = len(result['errors']) == 0
    status_code = 200 if result['success'] else 400
    return jsonify(result), status_code

@bp.route('/products/<int:product_id>/upload', methods=['GET'])
@login_required
@permission_required('product', 'edit')
def upload_files_page(product_id):
    """显示产品文件上传页面"""
    # 检查产品是否存在
    product = Product.query.get_or_404(product_id)
    
    return render_template('product/upload_files.html', 
                         product_id=product_id,
                         product=product)

@bp.route('/api/products/check-mn', methods=['GET'])
def check_product_mn():
    """检查产品MN号是否重复"""
    try:
        product_mn = request.args.get('product_mn', '')
        exclude_id = request.args.get('exclude_id', 0, type=int)
        
        if not product_mn:
            return jsonify({'valid': False, 'message': 'MN号不能为空'})
        
        # 检查MN号是否已存在
        query = Product.query.filter(Product.product_mn == product_mn)
        
        # 如果是编辑现有产品，排除自身
        if exclude_id > 0:
            query = query.filter(Product.id != exclude_id)
        
        existing_product = query.first()
        
        if existing_product:
            return jsonify({
                'valid': False, 
                'message': f'MN号已被产品 {existing_product.product_name} ({existing_product.model}) 使用'
            })
        
        return jsonify({'valid': True, 'message': 'MN号可用'})
        
    except Exception as e:
        logger.error(f'检查MN号时出错: {str(e)}')
        return jsonify({
            'error': '检查MN号失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/dashboard-data', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_dashboard_data():
    """获取仪表盘数据"""
    try:
        # 基础查询，根据用户角色筛选可见产品
        base_query = Product.query
        
        # 如果不是管理员、产品经理或解决方案经理，只显示生产中的产品
        if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            base_query = base_query.filter(Product.status == 'active')
        
        # 按分类统计产品数量
        from app.models.product_code import ProductCategory

        category_stats = db.session.query(
            ProductCategory.name,
            func.count(Product.id)
        ).join(
            Product, Product.category_id == ProductCategory.id
        )

        # 应用产品可见性筛选到类别统计
        if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            category_stats = category_stats.filter(Product.status == 'active')

        # 完成分组查询，按ProductCategory.id分组保持业务顺序
        category_stats = category_stats.group_by(
            ProductCategory.id,
            ProductCategory.name
        ).order_by(ProductCategory.id).all()

        category_data = [{'category': cat, 'count': count} for cat, count in category_stats]
        
        # 统计各状态产品数量
        # 对于管理员、产品经理和解决方案经理，显示所有产品的状态
        if current_user.role in ['admin', 'product_manager', 'solution_manager']:
            active_count = Product.query.filter(Product.status == 'active').count()
            discontinued_count = Product.query.filter(Product.status == 'discontinued').count()
            upcoming_count = Product.query.filter(Product.status == 'upcoming').count()
        else:
            # 对于其他用户，只统计生产中的产品，其他状态产品显示为0
            active_count = Product.query.filter(Product.status == 'active').count()
            discontinued_count = 0
            upcoming_count = 0
        
        status_stats = {
            'active': active_count,
            'discontinued': discontinued_count,
            'upcoming': upcoming_count
        }
        
        # 统计项目产品和渠道产品数量
        # 构建类型统计查询
        type_stats_query = db.session.query(
            Product.type, 
            func.count(Product.id)
        ).filter(
            Product.type.isnot(None)
        )
        
        # 应用产品可见性筛选
        if current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            # 对于其他用户，只统计生产中的产品
            type_stats_query = type_stats_query.filter(Product.status == 'active')
        
        # 完成分组查询
        type_stats = type_stats_query.group_by(
            Product.type
        ).all()
        
        # 确保type_stats包含所有产品类型，即使数量为0
        product_types = ["项目产品", "渠道产品", "第三方产品"]
        type_data = []
        
        # 创建一个字典，存储已有的统计数据
        existing_types = {t: count for t, count in type_stats}
        
        # 为每种产品类型创建一个条目
        for product_type in product_types:
            count = existing_types.get(product_type, 0)
            type_data.append({'type': product_type, 'count': count})
        
        # 总产品数量也应基于用户权限
        if current_user.role in ['admin', 'product_manager', 'solution_manager']:
            total_products = Product.query.count()
        else:
            total_products = Product.query.filter(Product.status == 'active').count()
        
        # 汇总所有数据
        dashboard_data = {
            'category_stats': category_data,
            'status_stats': status_stats,
            'type_stats': type_data,
            'total_products': total_products
        }
        
        return jsonify(dashboard_data)
        
    except Exception as e:
        logger.error(f'获取仪表盘数据时出错: {str(e)}')
        return jsonify({
            'error': '获取仪表盘数据失败',
            'message': str(e)
        }), 500

# 产品规格查询API
@bp.route('/api/product/<int:product_id>/specs', methods=['GET'])
@login_required
@permission_required('product', 'view')
def get_product_specs(product_id):
    """
    获取产品的所有规格数据

    Args:
        product_id: 产品ID
        grouped: 可选参数，设为1时返回按分类分组的数据

    Returns:
        grouped=0 (默认):
        JSON: {
            'specs': [
                {
                    'id': int,
                    'field_name': str,
                    'field_value': str,
                    'field_code': str,
                    'position': int  # 规格在MN编码中的位置
                }
            ]
        }

        grouped=1:
        JSON: {
            'spec_categories': [
                {
                    'id': int,
                    'name': str,
                    'name_en': str,
                    'specs': [...]
                }
            ]
        }
    """
    try:
        from app.models.product_spec import ProductSpec
        from app.models.product_code import ProductCodeField
        from app.models.product import Product
        from app.services.spec_service import SpecService

        # 验证产品存在性
        product = Product.query.get_or_404(product_id)

        # 检查是否需要分组返回
        grouped = request.args.get('grouped', '0') == '1'

        # 获取产品规格（按 display_order 排序，确保与子分类规格顺序一致）
        specs = ProductSpec.query.filter_by(product_id=product_id).order_by(ProductSpec.display_order).all()

        # 转换为字典并添加position信息
        spec_list = []
        for spec in specs:
            spec_dict = spec.to_dict()

            # 获取规格字段的position（用于MN编码排序）
            if product.subcategory_id and spec.field_name:
                field = ProductCodeField.query.filter_by(
                    subcategory_id=product.subcategory_id,
                    name=spec.field_name
                ).first()
                if field:
                    spec_dict['position'] = field.position
                else:
                    spec_dict['position'] = 999  # 未找到对应字段，使用默认值
            else:
                spec_dict['position'] = 999

            spec_list.append(spec_dict)

        logger.debug(f"为产品 {product_id} 找到 {len(spec_list)} 个规格")

        # 根据参数决定返回格式
        if grouped:
            spec_categories = SpecService.group_specs_by_category(spec_list)
            return jsonify({'spec_categories': spec_categories})
        else:
            return jsonify({'specs': spec_list})

    except Exception as e:
        logger.error(f"获取产品规格失败: {str(e)}")
        return jsonify({'error': str(e)}), 500


# 产品规格MN预览API（用于确认对话框显示）
@bp.route('/api/products/<int:product_id>/specs/preview', methods=['POST'])
@login_required
@permission_required('product', 'view')
def preview_product_specs_mn(product_id):
    """
    预览规格保存后的MN变化（冲突检测同时检查产品库和研发库）

    Args:
        product_id: 产品ID

    Request Body:
        {
            'specs': [
                {
                    'field_name': str,
                    'field_value': str,
                    'field_code': str or null
                }
            ]
        }

    Returns:
        JSON: {
            'success': bool,
            'current_product_mn': str,
            'current_spec_mn': str,
            'new_spec_mn': str,
            'is_mn_locked': bool,
            'missing_codes': [str],
            'conflict_product': dict or null
        }
    """
    try:
        data = request.get_json()
        specs = data.get('specs', [])

        from app.services.spec_service import SpecService
        result = SpecService.preview_specs_mn(SpecService.TYPE_PRODUCT, product_id, specs)

        return jsonify(result)

    except Exception as e:
        logger.error(f"预览规格MN失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


# 产品规格保存API（用于详情页内联编辑）
# 使用统一的 SpecService 处理规格保存
@bp.route('/api/products/<int:product_id>/specs', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def save_product_specs_api(product_id):
    """
    保存产品规格数据（支持批量更新、添加、删除）
    使用统一的 SpecService 处理

    Args:
        product_id: 产品ID

    Request Body:
        {
            'specs': [{'id', 'field_name', 'field_value', 'field_code', 'include_in_description'}],
            'deleted': [int, ...]
        }

    Returns:
        JSON: {'success': bool, 'message': str, 'specs': [...], 'spec_mn': str, ...}
    """
    try:
        from app.models.product import Product
        from app.services.spec_service import SpecService

        # 验证产品存在性
        product = Product.query.get_or_404(product_id)

        data = request.get_json()
        specs = data.get('specs', [])
        deleted_ids = data.get('deleted', [])

        # 使用统一的 SpecService 保存规格
        result = SpecService.save_specs(SpecService.TYPE_PRODUCT, product_id, specs, deleted_ids)

        if result['success']:
            # 补充 display_mn 字段（SpecService 不返回此字段）
            product = Product.query.get(product_id)  # 重新查询获取最新数据
            result['display_mn'] = product.display_mn if hasattr(product, 'display_mn') else None

            # 更新编码定义快照
            from app.utils.product_helpers import generate_product_snapshot
            snapshot = generate_product_snapshot(product=product, source="manual_update")
            if snapshot:
                product.code_definition_snapshot = snapshot
                db.session.commit()

            result['has_valid_snapshot'] = snapshot is not None
            return jsonify(result)
        else:
            return jsonify(result), 500

    except Exception as e:
        db.session.rollback()
        logger.error(f"保存产品规格失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@bp.route('/api/products/create', methods=['POST'])
@login_required
@permission_required('product', 'create')  # 添加产品创建权限装饰器
def create_product():
    """创建新产品"""
    try:
        logger.debug('正在创建新产品...')
        
        # 获取表单数据
        product_data = {
            'type': request.form.get('type'),
            'category': request.form.get('category'),
            'product_mn': request.form.get('product_mn'),
            'product_name': request.form.get('product_name'),
            'model': request.form.get('model'),
            'specification': request.form.get('specification'),
            'brand': request.form.get('brand'),
            'unit': request.form.get('unit'),
            'unit_en': request.form.get('unit_en', ''),
            'status': request.form.get('status', 'active'),  # 默认为生产中
            'retail_price': request.form.get('retail_price'),
            'currency': request.form.get('currency', Config.DEFAULT_CURRENCY)  # 默认货币
        }
        
        # 验证必填字段
        required_fields = ['product_name', 'model', 'product_mn']
        for field in required_fields:
            if not product_data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'{field} 字段为必填项'
                }), 400
        
        # 验证MN号唯一性
        existing_product = Product.query.filter_by(product_mn=product_data['product_mn']).first()
        if existing_product:
            return jsonify({
                'success': False,
                'message': f'MN号 {product_data["product_mn"]} 已存在'
            }), 400
        
        # 处理零售价格
        if product_data['retail_price']:
            try:
                product_data['retail_price'] = Decimal(product_data['retail_price'])
            except InvalidOperation:
                return jsonify({
                    'success': False,
                    'message': '零售价格格式不正确'
                }), 400
        else:
            product_data['retail_price'] = Decimal('0.00')
        
        # 使用前端传递的状态值，默认为在售
        product_data['status'] = request.form.get('status', 'active')
        logger.debug(f'创建产品，状态: {product_data["status"]}, 名称: {product_data["product_name"]}')
        
        # 获取厂商产品标记
        is_vendor_product = request.form.get('is_vendor_product') == 'on'

        # 获取配置来源信息（从配置引入产品时设置）
        source_configuration_id = request.form.get('source_configuration_id')
        logger.info(f"[DEBUG] 收到 source_configuration_id: '{source_configuration_id}' (type: {type(source_configuration_id)})")
        logger.info(f"[DEBUG] 所有表单数据: {dict(request.form)}")
        if source_configuration_id:
            try:
                source_configuration_id = int(source_configuration_id)
                logger.info(f"[DEBUG] 转换后 source_configuration_id: {source_configuration_id}")
            except (ValueError, TypeError):
                logger.warning(f"[DEBUG] source_configuration_id 转换失败")
                source_configuration_id = None

        # 获取配置来源类型
        source_type = request.form.get('source_type', 'manual')

        # 验证 source_type 值
        valid_source_types = ['manual', 'from_config', 'from_sp8d', 'from_dev', 'from_spec']
        if source_type not in valid_source_types:
            source_type = 'manual'

        # 对于 from_sp8d 类型，需要按名称匹配或创建本地分类
        if source_type == 'from_sp8d':
            logger.info(f"[DEBUG] from_sp8d 类型，清除 source_configuration_id（远程ID不可用于本地FK）")
            source_configuration_id = None

            # 从表单获取分类名称（用于按名称匹配）
            category_name = request.form.get('category_name', '').strip()
            category_name_en = request.form.get('category_name_en', '').strip()
            subcategory_name = request.form.get('subcategory_name', '').strip()
            subcategory_name_en = request.form.get('subcategory_name_en', '').strip()

            logger.info(f"[DEBUG] SP8D导入分类名称: category={category_name}/{category_name_en}, subcategory={subcategory_name}/{subcategory_name_en}")

            # 匹配或创建一级分类
            local_cat = None
            if category_name_en:
                # 优先按英文名匹配（忽略大小写）
                local_cat = ProductCategory.query.filter(
                    func.lower(ProductCategory.name_en) == func.lower(category_name_en)
                ).first()
            if not local_cat and category_name:
                # 回退：按中文名匹配
                local_cat = ProductCategory.query.filter_by(name=category_name).first()
            if not local_cat and (category_name or category_name_en):
                # 自动创建新分类
                try:
                    new_code_letter = generate_next_category_code_letter()
                    local_cat = ProductCategory(
                        name=category_name or category_name_en,
                        name_en=category_name_en or None,
                        code_letter=new_code_letter
                    )
                    db.session.add(local_cat)
                    db.session.flush()
                    logger.info(f"[DEBUG] 自动创建分类: {local_cat.name} (code={new_code_letter})")
                except ValueError as e:
                    logger.error(f"创建分类失败: {str(e)}")

            category_id = local_cat.id if local_cat else None

            # 匹配或创建二级分类
            local_sub = None
            if subcategory_name_en and category_id:
                # 优先按英文名匹配（在同一父分类下，忽略大小写）
                local_sub = ProductSubcategory.query.filter(
                    ProductSubcategory.category_id == category_id,
                    func.lower(ProductSubcategory.name_en) == func.lower(subcategory_name_en)
                ).first()
            if not local_sub and subcategory_name and category_id:
                # 回退：按中文名匹配
                local_sub = ProductSubcategory.query.filter(
                    ProductSubcategory.category_id == category_id,
                    ProductSubcategory.name == subcategory_name
                ).first()
            if not local_sub and category_id and (subcategory_name or subcategory_name_en):
                # 自动创建新子分类
                try:
                    new_sub_code = generate_next_subcategory_code_letter(category_id)
                    local_sub = ProductSubcategory(
                        category_id=category_id,
                        name=subcategory_name or subcategory_name_en,
                        name_en=subcategory_name_en or None,
                        code_letter=new_sub_code
                    )
                    db.session.add(local_sub)
                    db.session.flush()
                    logger.info(f"[DEBUG] 自动创建子分类: {local_sub.name} (code={new_sub_code})")
                except ValueError as e:
                    logger.error(f"创建子分类失败: {str(e)}")

            subcategory_id = local_sub.id if local_sub else None
            logger.info(f"[DEBUG] SP8D导入分类匹配结果: category_id={category_id}, subcategory_id={subcategory_id}")

            # === 根据 region 编码查找本地 region_id ===
            region_id = None  # 初始化，后续可能被匹配覆盖
            region_code = request.form.get('region_code', '').strip()
            if region_code:
                from app.models.product_code import ProductCodeField, ProductCodeFieldOption
                # 先尝试通过 ProductCodeField.code 匹配
                region_field = ProductCodeField.query.filter(
                    ProductCodeField.field_type == 'origin_location',
                    ProductCodeField.code == region_code
                ).first()
                if region_field:
                    region_id = region_field.id
                    logger.info(f"[DEBUG] 通过 region 编码 '{region_code}' 匹配到本地区域 ID: {region_id}")
                else:
                    # 如果 field.code 是 '?' 则通过 option.code 查找
                    region_option = ProductCodeFieldOption.query.join(ProductCodeField).filter(
                        ProductCodeField.field_type == 'origin_location',
                        ProductCodeFieldOption.code == region_code
                    ).first()
                    if region_option:
                        region_id = region_option.field_id
                        logger.info(f"[DEBUG] 通过 option 编码 '{region_code}' 匹配到本地区域 ID: {region_id}")

        else:
            # 非SP8D导入：直接使用前端传递的ID
            # 获取分类体系字段
            category_id = request.form.get('category_id')
            subcategory_id = request.form.get('subcategory_id')

            # 转换为整数
            try:
                category_id = int(category_id) if category_id else None
            except (ValueError, TypeError):
                category_id = None
            try:
                subcategory_id = int(subcategory_id) if subcategory_id else None
            except (ValueError, TypeError):
                subcategory_id = None

        # 获取区域字段
        # 注意：SP8D导入时 region_id 已在上面通过 region_code 匹配设置
        if source_type != 'from_sp8d':
            # 非SP8D导入：使用表单传递的 region_id
            region_id = request.form.get('region_id')
            try:
                region_id = int(region_id) if region_id else None
            except (ValueError, TypeError):
                region_id = None

        # 创建新产品
        new_product = Product(
            type=product_data['type'],
            category=product_data['category'],
            product_mn=product_data['product_mn'],
            product_name=product_data['product_name'],
            model=product_data['model'],
            specification=product_data['specification'],
            brand=product_data['brand'],
            unit=product_data['unit'],
            unit_en=product_data.get('unit_en', ''),
            retail_price=product_data['retail_price'],
            currency=product_data['currency'],
            status=product_data['status'],
            is_vendor_product=is_vendor_product,
            owner_id=current_user.id,
            # 分类体系字段
            category_id=category_id,
            subcategory_id=subcategory_id,
            region_id=region_id,
            # 配置来源信息
            source_configuration_id=source_configuration_id,
            source_type=source_type,
            # 从配置引入时锁定MN编码（本地或远程配置都锁定）
            is_mn_locked=True if source_type in ['from_config', 'from_sp8d', 'from_spec'] else False
        )
        
        # 处理产品图片上传到Supabase
        has_image = False
        if 'product_image' in request.files:
            product_image = request.files['product_image']
            if product_image.filename:  # 确保有文件被上传
                logger.debug(f'处理产品图片上传: {product_image.filename}')
                # 使用智能存储系统（NAS 优先，Supabase 备份）
                try:
                    # 先保存产品以获取ID
                    db.session.add(new_product)
                    db.session.flush()  # 获取ID但不提交

                    # 使用智能存储系统上传图片
                    from app.utils.smart_storage_manager import get_smart_product_storage
                    smart_storage = get_smart_product_storage()
                    image_url = smart_storage.upload_product_file(new_product.id, product_image, 'image', 'product')

                    if image_url:
                        new_product.image_path = image_url
                        has_image = True
                        logger.debug(f'产品图片上传成功: {image_url}')
                    else:
                        logger.warning('产品图片上传失败，将创建没有图片的产品')
                except Exception as e:
                    logger.error(f'产品图片上传异常: {str(e)}')
                    logger.warning('图片处理失败，将创建没有图片的产品')
        else:
            logger.debug('没有上传产品图片')
        
        # 处理PDF文件上传（智能选择本地或云端）
        has_pdf = False
        if 'product_pdf' in request.files:
            product_pdf = request.files['product_pdf']
            if product_pdf.filename:  # 确保有文件被上传
                logger.debug(f'处理产品PDF上传: {product_pdf.filename}')
                # 使用智能存储系统（NAS 优先，Supabase 备份）
                try:
                    # 如果产品还没有添加到会话，先添加并flush
                    if new_product not in db.session:
                        db.session.add(new_product)
                        db.session.flush()  # 获取ID但不提交

                    # 使用智能存储系统上传PDF
                    from app.utils.smart_storage_manager import get_smart_product_storage
                    smart_storage = get_smart_product_storage()
                    pdf_url = smart_storage.upload_product_file(new_product.id, product_pdf, 'pdf', 'product')

                    if pdf_url:
                        new_product.pdf_path = pdf_url
                        has_pdf = True
                        logger.debug(f'产品PDF上传成功: {pdf_url}')
                    else:
                        logger.warning('产品PDF上传失败，将创建没有PDF的产品')
                except Exception as e:
                    logger.error(f'产品PDF上传异常: {str(e)}')
                    logger.warning('PDF文件处理失败，将创建没有PDF的产品')
        else:
            logger.debug('没有上传产品PDF文件')
        
        # 保存新产品（如果还没有添加到会话）
        if new_product not in db.session:
            db.session.add(new_product)
        db.session.commit()
        
        logger.info(f'产品创建成功: ID={new_product.id}, MN={new_product.product_mn}, 名称={new_product.product_name}, 有图片={has_image}, 有PDF={has_pdf}')
        logger.info(f'[DEBUG] 保存后的配置信息: source_configuration_id={new_product.source_configuration_id}, source_type={new_product.source_type}, category_id={new_product.category_id}, subcategory_id={new_product.subcategory_id}, region_id={new_product.region_id}')

        # 构建响应
        response_data = {
            'success': True,
            'message': '产品创建成功',
            'product': {
                'id': new_product.id,
                'product_name': new_product.product_name,
                'product_mn': new_product.product_mn,
                'has_image': has_image,
                'has_pdf': has_pdf
            }
        }

        # 从配置引入时，重定向到产品详情页并自动打开规格引入模态框
        if source_type in ('from_config', 'from_sp8d'):
            redirect_url = url_for('product.view_product_detail', id=new_product.id) + '?import_specs=1'
            if source_type == 'from_sp8d':
                sp8d_config_id = request.form.get('sp8d_configuration_id')
                if sp8d_config_id:
                    redirect_url += f'&sp8d_config_id={sp8d_config_id}'
            response_data['redirect'] = redirect_url

        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'创建产品时出错: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'message': f'创建产品失败: {str(e)}'
        }), 500

@bp.route('/products/<int:id>/update', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def update_product(id):
    """更新产品信息（表单提交）"""
    try:
        logger.debug(f'正在更新产品 ID={id}（新版分类体系）...')

        # 查找产品
        product = Product.query.get_or_404(id)

        # 检查权限
        if product.owner_id != current_user.id and current_user.role not in ['admin', 'product_manager']:
            return jsonify({'success': False, 'message': '您没有权限编辑此产品'}), 403

        # 获取表单数据（新版字段）
        product_type = request.form.get('type') or None
        product_status = request.form.get('status', 'active')
        # 兼容模态框的 model 字段和传统 API 的 product_model 字段
        product_model = request.form.get('model') or request.form.get('product_model')
        # 兼容模态框的 name 字段和传统 API 的 product_name 字段
        product_name = request.form.get('name') or request.form.get('product_name') or None
        brand = request.form.get('brand') or None
        unit = request.form.get('unit')
        retail_price = request.form.get('retail_price')
        currency = request.form.get('currency', Config.DEFAULT_CURRENCY)
        description = request.form.get('description')

        # 分类字段从表单获取
        category_id = request.form.get('category_id')
        subcategory_id = request.form.get('subcategory_id')
        region_id = request.form.get('region_id') or None

        # 前端已通过disabled属性锁定关键字段，disabled字段不会被提交
        # 因此后端不需要重复检查MN锁定状态，直接信任前端提交的数据即可

        # 验证必填字段
        if not product_model:
            return jsonify({'success': False, 'message': '请填写产品型号'}), 400
        if not region_id and not product.region_id:
            return jsonify({'success': False, 'message': '销售区域为必填项'}), 400

        # 更新产品基本信息
        product.type = product_type
        product.status = product_status
        product.model = product_model
        product.product_name = product_name  # 独立的产品名称
        product.unit = unit
        product.currency = currency
        # 只在表单中包含这些字段时才更新（避免编辑模态框清空这些值）
        if 'brand' in request.form:
            product.brand = brand
        if 'description' in request.form:
            product.specification = description

        # 更新厂商产品标记（不影响MN编码，允许修改）
        is_vendor_product = request.form.get('is_vendor_product') == 'on'
        if product.is_vendor_product != is_vendor_product:
            product.is_vendor_product = is_vendor_product
            logger.debug(f'厂商产品标记从 {product.is_vendor_product} 更新为 {is_vendor_product}')

        # 更新分类字段（允许补充空值，管理员可以修改已有值）
        # 记录旧的分类ID，用于检测分类是否改变
        old_category_id = product.category_id
        old_subcategory_id = product.subcategory_id
        old_region_id = product.region_id

        if category_id and (not product.category_id or current_user.role == 'admin'):
            product.category_id = int(category_id)
        if subcategory_id and (not product.subcategory_id or current_user.role == 'admin'):
            product.subcategory_id = int(subcategory_id)
        if region_id and (not product.region_id or current_user.role == 'admin'):
            product.region_id = int(region_id)

        # 检测分类是否改变（类别或系列改变都算分类改变）
        category_changed = (old_category_id != product.category_id) or (old_subcategory_id != product.subcategory_id)
        # 检测区域是否改变
        region_changed = (old_region_id != product.region_id) and product.region_id is not None

        # 获取确认标记
        clear_specs_confirmed = request.form.get('clear_specs_confirmed') == 'true'
        region_change_confirmed = request.form.get('region_change_confirmed') == 'true'

        # 如果分类改变，检查前端是否确认清空规格
        if category_changed and not clear_specs_confirmed:
            # 检查是否有现有规格数据
            from app.models.product_spec import ProductSpec
            existing_specs = ProductSpec.query.filter_by(product_id=product.id).count()
            if existing_specs > 0:
                # 有规格数据但未确认，返回需要确认的响应
                return jsonify({
                    'success': False,
                    'require_confirmation': True,
                    'confirmation_type': 'category_change',
                    'message': _('改变产品分类将删除现有的规格和指标数据，是否继续？')
                }), 200

        # 如果区域改变，检查前端是否确认更新编码
        if region_changed and not region_change_confirmed and not category_changed:
            # 只有区域改变时才单独确认（如果分类也改变，分类确认已包含编码变化）
            if product.spec_mn:
                # 有编码数据，需要确认
                return jsonify({
                    'success': False,
                    'require_confirmation': True,
                    'confirmation_type': 'region_change',
                    'message': _('改变销售区域将更新产品编码，是否继续？')
                }), 200

        # 处理零售价格
        if retail_price:
            try:
                product.retail_price = Decimal(retail_price)
            except (InvalidOperation, ValueError):
                return jsonify({'success': False, 'message': '零售价格格式不正确'}), 400
        else:
            product.retail_price = Decimal('0.00')

        # 积分系数 (仅admin可设置)
        if current_user.role == 'admin':
            coeff_str = request.form.get('points_coefficient_override', '').strip()
            if coeff_str:
                try:
                    new_val = float(coeff_str)
                    if product.points_coefficient_override is None or float(product.points_coefficient_override) != new_val:
                        product.points_coefficient_override = new_val
                        product.points_coefficient_override_at = datetime.now()
                except (ValueError, TypeError):
                    pass
            else:
                product.points_coefficient_override = None
                product.points_coefficient_override_at = None

        # 如果分类改变且已确认，清空规格数据并更新编码
        if category_changed and clear_specs_confirmed:
            from app.models.product_spec import ProductSpec
            deleted_count = ProductSpec.query.filter_by(product_id=product.id).delete()
            db.session.flush()
            logger.info(f'分类改变，清空规格数据: 产品ID={product.id}, 删除了{deleted_count}条规格')

            # 清空规格相关的编码
            product.spec_mn = None
            product.specification = None
            product.code_definition_snapshot = None

            # 如果MN未锁定，也清空product_mn（等待新规格生成新编码）
            if not product.is_mn_locked:
                product.product_mn = None
                logger.info(f'分类改变，清空product_mn: 产品ID={product.id}')

        # 更新规格数据
        # 前端已对编码规格字段禁用（disabled），非编码规格始终可编辑
        spec_names = request.form.getlist('spec_name[]')
        spec_values = request.form.getlist('spec_value[]')
        spec_codes = request.form.getlist('spec_option_codes[]')
        include_in_descriptions = request.form.getlist('include_in_description_indexed[]')

        # 只有当前端提交了规格数据时才更新规格
        # 如果前端没有规格表单（如详情编辑页面），则保留现有规格不做修改
        if spec_names:
            # 先删除所有现有规格（简单粗暴但可靠的更新方式）
            from app.models.product_spec import ProductSpec
            ProductSpec.query.filter_by(product_id=product.id).delete()
            db.session.flush()
            spec_data_list = []
            for i in range(len(spec_names)):
                if spec_names[i].strip():
                    # 获取是否纳入描述的状态，值为 '1' 表示勾选
                    include_in_desc = (i < len(include_in_descriptions) and
                                      include_in_descriptions[i] == '1')
                    spec_data_list.append({
                        'field_name': spec_names[i],
                        'field_value': spec_values[i] if i < len(spec_values) else '',
                        'field_code': spec_codes[i] if i < len(spec_codes) and spec_codes[i] != '0' else None,
                        'include_in_description': include_in_desc,
                        'action': 'create'
                    })

            if spec_data_list:
                # 使用统一的 SpecService 保存规格
                from app.services.spec_service import SpecService
                old_spec_mn = product.spec_mn  # 保存旧的 spec_mn 用于同步判断
                result = SpecService.save_specs(SpecService.TYPE_PRODUCT, product.id, spec_data_list)
                if not result['success']:
                    db.session.rollback()
                    return jsonify({'success': False, 'message': f'保存规格数据失败: {result.get("message", "未知错误")}'}), 400

                # 更新编码定义快照（保存规格成功后）
                from app.utils.product_helpers import generate_product_snapshot
                snapshot = generate_product_snapshot(
                    product=product,
                    source="manual_update"
                )
                if snapshot:
                    product.code_definition_snapshot = snapshot
                    logger.info(f'更新编码快照成功: 产品ID={product.id}')
                else:
                    # 快照更新失败不阻止产品编辑，只记录警告
                    logger.warning(f'快照更新跳过: 产品ID={product.id}（规格数据可能不完整）')

                # SpecService 已经更新了 spec_mn，处理 product_mn 同步
                spec_mn = result.get('spec_mn')
                if spec_mn:
                    logger.debug(f'更新规格MN: {spec_mn}')

                    # 只有在 MN 未锁定时才考虑同步 product_mn
                    if not product.is_mn_locked:
                        # 如果 product_mn 与旧 spec_mn 一致，说明是自动同步的，继续同步
                        if product.product_mn == old_spec_mn:
                            product.product_mn = spec_mn
                            logger.debug(f'同步更新 product_mn: {spec_mn}')
                        else:
                            logger.debug(f'product_mn ({product.product_mn}) 与 old_spec_mn ({old_spec_mn}) 不一致，保留历史值')
                    else:
                        logger.debug(f'产品 ID={product.id} MN已锁定，跳过 product_mn 同步')
        else:
            # 没有提交规格数据（详情编辑页面），但如果区域改变了，需要更新编码
            if region_changed:
                from app.utils.product_helpers import generate_spec_mn
                old_spec_mn = product.spec_mn
                spec_mn = generate_spec_mn(product)
                if spec_mn:
                    product.spec_mn = spec_mn
                    logger.info(f'区域改变，更新规格MN: {old_spec_mn} -> {spec_mn}')

                    # 如果 MN 未锁定，同步更新 product_mn
                    if not product.is_mn_locked:
                        if product.product_mn == old_spec_mn:
                            product.product_mn = spec_mn
                            logger.info(f'区域改变，同步更新 product_mn: {spec_mn}')
                    else:
                        logger.debug(f'产品 ID={product.id} MN已锁定，跳过 product_mn 同步')

        # 提交事务
        db.session.commit()

        logger.info(f'产品更新成功: ID={product.id}, MN={product.product_mn}, 型号={product.model}')

        # 返回JSON响应（前端使用AJAX提交）
        return jsonify({
            'success': True,
            'message': '产品更新成功',
            'redirect': url_for('product.view_product_detail', id=product.id)
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f'更新产品时出错: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'message': f'更新产品失败: {str(e)}'
        }), 500


@bp.route('/products/<int:id>/unlock-mn', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def unlock_product_mn(id):
    """管理员解锁产品MN编码"""
    # 只有管理员可以解锁
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以解锁MN编码'}), 403

    product = Product.query.get_or_404(id)

    if not product.is_mn_locked:
        return jsonify({'success': False, 'message': '该产品的MN编码未锁定'}), 400

    product.is_mn_locked = False
    db.session.commit()

    logger.info(f'管理员 {current_user.username} 解锁了产品 ID={id} 的MN编码')
    return jsonify({'success': True, 'message': 'MN编码已解锁'})


@bp.route('/products/<int:id>/lock-mn', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def lock_product_mn(id):
    """管理员锁定产品MN编码"""
    # 只有管理员可以锁定
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以锁定MN编码'}), 403

    product = Product.query.get_or_404(id)

    if product.is_mn_locked:
        return jsonify({'success': False, 'message': '该产品的MN编码已锁定'}), 400

    product.is_mn_locked = True
    db.session.commit()

    logger.info(f'管理员 {current_user.username} 锁定了产品 ID={id} 的MN编码')
    return jsonify({'success': True, 'message': 'MN编码已锁定'})


@bp.route('/api/products/<int:id>/update', methods=['POST'])
@login_required
@permission_required('product', 'edit')  # 添加产品编辑权限装饰器
def update_product_api(id):
    """更新产品信息（JSON API）"""
    try:
        logger.debug(f'正在更新产品 ID={id}...')
        
        # 查找产品
        product = Product.query.get(id)
        if not product:
            return jsonify({
                'success': False,
                'message': f'未找到ID为 {id} 的产品'
            }), 404
        
        # 检查所有权
        if product.owner_id != current_user.id and current_user.role not in ['admin', 'product_manager']:
            logger.warning(f'用户 {current_user.username} 尝试编辑不属于他的产品 {id}')
            return jsonify({
                'success': False,
                'message': '您没有权限编辑此产品'
            }), 403
        
        # 获取表单数据
        product_data = {
            'type': request.form.get('type'),
            'category': request.form.get('category'),
            'product_mn': request.form.get('product_mn'),
            'product_name': request.form.get('product_name'),
            'model': request.form.get('model'),
            'specification': request.form.get('specification'),
            'brand': request.form.get('brand'),
            'unit': request.form.get('unit'),
            'retail_price': request.form.get('retail_price'),
            'currency': request.form.get('currency', Config.DEFAULT_CURRENCY)  # 默认货币
        }
        
        # 只有管理员能修改生产状态
        if current_user.role == 'admin':
            # 直接从表单获取status字段，而不是错误的is_active字段
            status_value = request.form.get('status')
            if status_value in ['active', 'discontinued', 'upcoming']:
                product_data['status'] = status_value
                logger.debug(f"管理员更新产品状态: status={status_value}")
            else:
                # 如果状态值无效，保持原状态
                product_data['status'] = product.status
                logger.debug(f"状态值无效({status_value})，保持原状态: {product_data['status']}")
        else:
            # 非管理员不能修改生产状态，保持原状态
            product_data['status'] = product.status
            logger.debug(f"非管理员用户无法修改产品生产状态，保持原状态: {product_data['status']}")
        
        # 验证必填字段
        required_fields = ['product_name', 'model', 'product_mn']
        for field in required_fields:
            if not product_data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'{field} 字段为必填项'
                }), 400
        
        # 验证MN号唯一性 (排除当前产品)
        if product.product_mn != product_data['product_mn']:
            existing_product = Product.query.filter_by(product_mn=product_data['product_mn']).first()
            if existing_product:
                return jsonify({
                    'success': False,
                    'message': f'MN号 {product_data["product_mn"]} 已存在'
                }), 400
        
        # 处理零售价格
        if product_data['retail_price']:
            try:
                product_data['retail_price'] = Decimal(product_data['retail_price'])
            except InvalidOperation:
                return jsonify({
                    'success': False,
                    'message': '零售价格格式不正确'
                }), 400
        else:
            product_data['retail_price'] = Decimal('0.00')
        
        # 设置数据是否有变化的标志，用于决定是否提交到数据库
        data_changed = False
        
        # 处理产品图片上传（智能选择本地或云端）
        image_changed = False
        if 'product_image' in request.files:
            product_image = request.files['product_image']
            if product_image.filename:  # 确保有文件被上传
                logger.debug(f'处理产品图片上传: {product_image.filename}')
                # 使用智能存储系统（NAS 优先，Supabase 备份）
                try:
                    from app.utils.smart_storage_manager import get_smart_product_storage
                    smart_storage = get_smart_product_storage()
                    image_url = smart_storage.upload_product_file(product.id, product_image, 'image', 'product')

                    if image_url:
                        # 如果已有旧图片且是本地路径，删除本地文件
                        if product.image_path and not product.image_path.startswith('http'):
                            old_image_path = os.path.join(UPLOAD_FOLDER, product.image_path)
                            if os.path.exists(old_image_path):
                                try:
                                    os.remove(old_image_path)
                                    logger.info(f"删除旧本地图片: {old_image_path}")
                                except Exception as e:
                                    logger.warning(f"删除旧图片失败: {str(e)}")

                        # 更新图片路径
                        product.image_path = image_url
                        image_changed = True
                        data_changed = True
                        logger.debug(f'产品图片上传成功: {image_url}')
                    else:
                        logger.warning('产品图片上传失败')
                except Exception as e:
                    logger.error(f'产品图片上传异常: {str(e)}')
        
        # 检查是否需要删除图片
        if request.form.get('remove_image') == 'true' and product.image_path:
            logger.debug('删除产品图片')
            try:
                # 使用智能存储系统删除文件（支持本地和云端）
                supabase_client = get_supabase_client()
                
                # 判断文件类型并删除
                if product.image_path.startswith('http'):
                    # 云端文件，使用智能存储删除
                    delete_success = supabase_client.delete_product_file(product.id, 'image', 'product')
                    if delete_success:
                        logger.info(f'云端产品图片删除成功: {product.image_path}')
                    else:
                        logger.warning(f'云端产品图片删除失败: {product.image_path}')
                else:
                    # 本地文件，直接删除
                    image_path = os.path.join(UPLOAD_FOLDER, product.image_path)
                    if os.path.exists(image_path):
                        os.remove(image_path)
                        logger.info(f'本地产品图片删除成功: {image_path}')
                
                # 清空图片路径
                product.image_path = None
                image_changed = True
                data_changed = True
                
            except Exception as e:
                logger.error(f'删除产品图片失败: {str(e)}')
                flash(_('删除图片失败，请重试'), 'warning')
        
        # 处理PDF文件上传（智能选择本地或云端）
        pdf_changed = False
        if 'product_pdf' in request.files:
            product_pdf = request.files['product_pdf']
            if product_pdf.filename:  # 确保有文件被上传
                logger.debug(f'处理产品PDF上传: {product_pdf.filename}')
                # 使用智能存储系统（NAS 优先，Supabase 备份）
                try:
                    from app.utils.smart_storage_manager import get_smart_product_storage
                    smart_storage = get_smart_product_storage()
                    pdf_url = smart_storage.upload_product_file(product.id, product_pdf, 'pdf', 'product')

                    if pdf_url:
                        # 如果已有旧PDF且是本地路径，删除本地文件
                        if product.pdf_path and not product.pdf_path.startswith('http'):
                            old_pdf_path = os.path.join(current_app.static_folder, product.pdf_path)
                            if os.path.exists(old_pdf_path):
                                try:
                                    os.remove(old_pdf_path)
                                    logger.info(f"删除旧本地PDF: {old_pdf_path}")
                                except Exception as e:
                                    logger.warning(f"删除旧PDF失败: {str(e)}")
                        
                        # 更新PDF路径
                        product.pdf_path = pdf_url
                        pdf_changed = True
                        data_changed = True
                        logger.debug(f'产品PDF上传成功: {pdf_url}')
                    else:
                        logger.warning('产品PDF上传失败')
                except Exception as e:
                    logger.error(f'产品PDF上传异常: {str(e)}')
        
        # 检查是否需要删除PDF文件
        if request.form.get('remove_pdf') == 'true' and product.pdf_path:
            logger.debug('删除产品PDF文件')
            try:
                # 使用智能存储系统删除PDF文件（支持本地和云端）
                supabase_client = get_supabase_client()
                
                # 判断文件类型并删除
                if product.pdf_path.startswith('http'):
                    # 云端文件，使用智能存储删除
                    delete_success = supabase_client.delete_product_file(product.id, 'pdf', 'product')
                    if delete_success:
                        logger.info(f'云端产品PDF删除成功: {product.pdf_path}')
                    else:
                        logger.warning(f'云端产品PDF删除失败: {product.pdf_path}')
                else:
                    # 本地文件，直接删除
                    pdf_path = os.path.join(current_app.static_folder, product.pdf_path)
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                        logger.info(f'本地产品PDF删除成功: {pdf_path}')
                
                # 清空PDF路径
                product.pdf_path = None
                pdf_changed = True
                data_changed = True
                
            except Exception as e:
                logger.error(f'删除产品PDF文件失败: {str(e)}')
                flash(_('删除PDF文件失败，请重试'), 'warning')
        
        # 更新产品信息 - 仅当值发生变化时才更新
        if product.type != product_data['type']:
            product.type = product_data['type']
            data_changed = True

        if product.product_mn != product_data['product_mn']:
            product.product_mn = product_data['product_mn']
            data_changed = True

        if product.model != product_data['model']:
            product.model = product_data['model']
            data_changed = True
            
        if product.specification != product_data['specification']:
            product.specification = product_data['specification']
            data_changed = True
            
        if product.brand != product_data['brand']:
            product.brand = product_data['brand']
            data_changed = True
            
        if product.unit != product_data['unit']:
            product.unit = product_data['unit']
            data_changed = True
            
        # 特殊处理Decimal类型的比较
        current_price = float(product.retail_price) if product.retail_price else 0.0
        new_price = float(product_data['retail_price']) if product_data['retail_price'] else 0.0
        if abs(current_price - new_price) > 0.001:  # 使用小数点精度进行比较
            product.retail_price = product_data['retail_price']
            data_changed = True
        
        # 更新货币类型
        if product.currency != product_data['currency']:
            product.currency = product_data['currency']
            data_changed = True
        
        # 获取厂商产品标记并更新
        is_vendor_product = request.form.get('is_vendor_product') == 'on'
        if product.is_vendor_product != is_vendor_product:
            product.is_vendor_product = is_vendor_product
            data_changed = True
            logger.debug(f'厂商产品标记从 {product.is_vendor_product} 更新为 {is_vendor_product}')
        
        # 更新生产状态 - 只有当用户有权限且状态有变化时才更新
        if product.status != product_data['status']:
            product.status = product_data['status']
            data_changed = True
            logger.info(f'产品状态已更新: ID={product.id}, 新状态={product.status}')
        
        # 如果存在图片变更或任何其他数据变更，则保存更新
        if data_changed or image_changed or pdf_changed:
            logger.info(f'产品数据已变更，正在提交更新: ID={product.id}')
            db.session.commit()
            return jsonify({
                'success': True,
                'message': '产品更新成功',
                'product': {
                    'id': product.id,
                    'product_name': product.name,  # 使用智能属性
                    'product_mn': product.product_mn,
                    'image_updated': image_changed,
                    'pdf_updated': pdf_changed
                }
            })
        else:
            logger.info(f'产品数据未变更，不需要更新: ID={product.id}')
            return jsonify({
                'success': True,
                'message': '产品数据未发生变化',
                'product': {
                    'id': product.id,
                    'product_name': product.name,  # 使用智能属性
                    'product_mn': product.product_mn
                }
            })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新产品时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'更新产品失败: {str(e)}'
        }), 500

@bp.route('/api/products/<int:id>', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product(id):
    """获取单个产品详情"""
    try:
        logger.debug(f'正在获取产品详情: ID={id}')
        
        # 查询产品
        product = Product.query.get(id)
        if not product:
            return jsonify({
                'error': '未找到产品',
                'message': f'未找到ID为 {id} 的产品'
            }), 404
        
        # 查询所有者信息
        owner_name = None
        if product.owner_id:
            from app.models.user import User
            owner = User.query.get(product.owner_id)
            if owner:
                # 优先使用真实姓名，如果没有则使用用户名
                owner_name = owner.real_name if owner.real_name else owner.username
        
        # 小数类型转换为浮点数
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        
        # 构建响应数据
        response = {
            'id': product.id,
            'type': product.type,
            'category': product.category_name,  # 使用智能属性
            'category_id': product.category_id,  # 模态框编辑需要
            'subcategory_id': product.subcategory_id,  # 模态框编辑需要
            'region_id': product.region_id,  # 模态框编辑需要
            'product_mn': product.product_mn,
            'product_name': product.product_name,  # 使用真正的产品名称字段
            'name': product.product_name,  # 模态框编辑需要真正的产品名称字段
            'model': product.model,
            'specification': product.specification,
            'brand': product.brand,
            'unit': product.unit,
            'retail_price': decimal_to_float(product.retail_price) if product.retail_price else 0,
            'currency': product.currency if hasattr(product, 'currency') else Config.DEFAULT_CURRENCY,
            'status': product.status,
            'is_vendor_product': product.is_vendor_product if hasattr(product, 'is_vendor_product') else False,
            'created_at': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else None,
            'updated_at': product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else None,
            'owner_id': product.owner_id,
            'owner_name': owner_name,
            'image_path': product.image_path,
            'pdf_path': product.pdf_path,
            # 配置来源信息（用于判断是否为引入产品）
            'source_configuration_id': product.source_configuration_id,
            'source_type': product.source_type
        }

        return jsonify(response)
        
    except Exception as e:
        logger.error(f'获取产品详情时出错: {str(e)}')
        return jsonify({
            'error': '获取产品详情失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/<int:id>/coefficient', methods=['POST'])
@login_required
def update_product_coefficient(id):
    """更新产品积分系数 (仅admin)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': '无权限'}), 403

    product = Product.query.get_or_404(id)
    data = request.get_json(silent=True) or {}
    coeff_str = str(data.get('coefficient', '')).strip()

    if coeff_str:
        try:
            val = float(coeff_str)
            if val < 1.0 or val > 10.0:
                return jsonify({'success': False, 'message': '系数范围 1.0 ~ 10.0'}), 400
            product.points_coefficient_override = val
            product.points_coefficient_override_at = datetime.now()
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '系数格式不正确'}), 400
    else:
        product.points_coefficient_override = None
        product.points_coefficient_override_at = None

    db.session.commit()
    return jsonify({
        'success': True,
        'coefficient': round(float(product.points_coefficient), 1),
        'points': product.points,
        'is_override': product.points_coefficient_override is not None
    })


@bp.route('/api/products/<int:id>/delete', methods=['POST'])
@login_required
# 注意：不使用 @permission_required 装饰器，因为创建者即使没有模块权限也可以删除自己的产品
def delete_product(id):
    """删除产品API - 创建者或有权限的角色可以删除未被引用的产品"""
    try:
        product = Product.query.get_or_404(id)

        # 权限检查：管理员、产品经理或产品创建者可以删除
        can_delete = (
            current_user.role == 'admin' or
            current_user.role in ['product_manager', 'product'] or
            current_user.has_permission('product', 'delete') or
            product.owner_id == current_user.id
        )

        if not can_delete:
            return jsonify({
                'success': False,
                'message': '您没有权限删除此产品'
            }), 403
        
        # 检查产品是否被报价单引用（按MN编号检查，MN是唯一标识）
        from app.models.quotation import QuotationDetail
        referenced_count = QuotationDetail.query.filter(
            QuotationDetail.product_mn == product.product_mn
        ).count()

        if referenced_count > 0:
            return jsonify({
                'success': False,
                'message': f'该产品（MN: {product.product_mn}）已被 {referenced_count} 个报价单引用，不能删除。如需停产，请使用"停产"功能。',
                'code': 'PRODUCT_REFERENCED'
            }), 400
        
        # 如果产品未被引用，可以删除
        product_name = product.name  # 使用智能属性
        
        # 删除产品图片文件（如果存在）
        if product.image_path:
            try:
                image_file_path = os.path.join(current_app.static_folder, product.image_path)
                if os.path.exists(image_file_path):
                    os.remove(image_file_path)
                    logger.debug(f"已删除产品图片文件: {image_file_path}")
            except Exception as e:
                logger.warning(f"删除产品图片文件失败: {str(e)}")
        
        # 删除PDF文件（如果存在）
        if hasattr(product, 'pdf_path') and product.pdf_path:
            try:
                pdf_file_path = os.path.join(current_app.static_folder, product.pdf_path)
                if os.path.exists(pdf_file_path):
                    os.remove(pdf_file_path)
                    logger.debug(f"已删除产品PDF文件: {pdf_file_path}")
            except Exception as e:
                logger.warning(f"删除产品PDF文件失败: {str(e)}")
        
        # 删除产品记录
        db.session.delete(product)
        db.session.commit()
        
        logger.info(f'{current_user.role} {current_user.username} 删除了产品: {product_name} (ID: {id})')
        
        return jsonify({
            'success': True,
            'message': f'产品 {product_name} 删除成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除产品时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'删除产品失败: {str(e)}'
        }), 500

@bp.route('/api/products/<int:id>/toggle-status', methods=['POST'])
@login_required
@permission_required('product', 'edit')  # 添加产品编辑权限装饰器
def toggle_product_status(id):
    """切换产品状态API - 仅管理员可用"""
    try:
        logger.debug(f'正在切换产品状态: ID={id}')
        
        # 检查用户是否为管理员
        if current_user.role != 'admin':
            logger.warning(f'非管理员用户 {current_user.username} 尝试切换产品状态: ID={id}')
            return jsonify({
                'success': False,
                'message': '只有管理员可以切换产品生产状态'
            }), 403
        
        # 解析请求数据
        data = request.json
        if not data or 'status' not in data:
            return jsonify({
                'success': False,
                'message': '请求数据格式不正确'
            }), 400
            
        # 获取目标状态
        target_status = data['status']
        if target_status not in ['active', 'discontinued']:
            return jsonify({
                'success': False,
                'message': '状态值无效，请使用 active 或 discontinued'
            }), 400
        
        # 查找产品
        product = Product.query.get(id)
        if not product:
            return jsonify({
                'success': False,
                'message': f'未找到ID为 {id} 的产品'
            }), 404
        
        # 如果状态没有变化，直接返回成功
        if product.status == target_status:
            return jsonify({
                'success': True,
                'message': '产品状态未变更',
                'status': target_status
            })
            
        # 更新状态
        product.status = target_status
        db.session.commit()
        
        # 记录操作
        status_text = '已停产' if target_status == 'discontinued' else '生产中'
        logger.info(f'管理员 {current_user.username} 将产品状态更新为 {status_text}: ID={product.id}, 名称={product.name}')
        
        return jsonify({
            'success': True,
            'message': f'产品状态已更新为 {status_text}',
            'status': target_status
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'切换产品状态时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'切换产品状态失败: {str(e)}'
        }), 500

@bp.route('/api/products/units', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product_units():
    """获取去重后的产品单位列表（含中英文）"""
    try:
        logger.debug('正在获取产品单位列表...')
        units = db.session.query(Product.unit, Product.unit_en).distinct().filter(
            Product.unit.isnot(None)
        ).all()

        is_ovs = Config.IS_OVS
        unit_list = []
        seen = set()
        for u, u_en in units:
            if not u or u in seen:
                continue
            seen.add(u)
            display = (u_en or u) if is_ovs else u
            unit_list.append({'value': u, 'value_en': u_en or '', 'display': display})

        unit_list.sort(key=lambda x: x['display'])
        logger.debug(f'找到 {len(unit_list)} 个单位')
        return jsonify(unit_list)
        
    except Exception as e:
        logger.error(f'获取产品单位列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品单位列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/types', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product_types():
    """获取产品类型列表，并按规则转换"""
    try:
        logger.debug('正在获取产品类型列表...')
        # 使用 distinct 获取唯一的类型列表
        types = db.session.query(Product.type).distinct().filter(
            Product.type.isnot(None),
            Product.type != ''
        ).all()
        
        # 将结果转换为列表
        type_list = [type_value[0] for type_value in types if type_value[0]]
        
        # 确保列表中有标准类型
        standard_types = {
            '0': '项目产品',
            '1': '渠道产品'
        }
        
        # 处理特殊类型值的转换
        result_types = []
        for type_value in type_list:
            if type_value in standard_types:
                # 如果是特殊值，转换为对应的名称
                result_types.append({
                    'value': type_value,
                    'text': standard_types[type_value]
                })
            else:
                # 保留原始值
                result_types.append({
                    'value': type_value,
                    'text': type_value
                })
        
        logger.debug(f'找到 {len(result_types)} 个产品类型')
        return jsonify(result_types)
        
    except Exception as e:
        logger.error(f'获取产品类型列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品类型列表失败',
            'message': str(e)
        }), 500

@bp.route('/api/products/brands', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def get_product_brands():
    """获取去重后的产品品牌列表"""
    try:
        logger.debug('正在获取产品品牌列表...')
        # 使用 distinct 获取唯一的品牌列表
        brands = db.session.query(Product.brand).distinct().filter(
            Product.brand.isnot(None),
            Product.brand != ''
        ).all()
        
        # 将结果转换为列表
        brand_list = [brand[0] for brand in brands if brand[0]]
        
        # 排序品牌列表
        brand_list.sort()

        logger.debug(f'找到 {len(brand_list)} 个品牌')
        return jsonify({'brands': brand_list})
        
    except Exception as e:
        logger.error(f'获取产品品牌列表时出错: {str(e)}')
        return jsonify({
            'error': '获取产品品牌列表失败',
            'message': str(e)
        }), 500

@bp.route('/products/<int:id>/detail', methods=['GET'])
@login_required
@permission_required('product', 'view')  # 添加产品查看权限装饰器
def view_product_detail(id):
    """查看产品详情页面"""
    try:
        from app.models.product_spec import ProductSpec

        # 获取产品详情
        product = Product.query.get_or_404(id)

        # 检查产品停产状态的权限：只有产品经理、解决方案经理和管理员可以查看停产产品
        if product.status == 'discontinued' and current_user.role not in ['admin', 'product_manager', 'solution_manager']:
            logger.warning(f"用户 {current_user.username} 尝试查看停产产品详情: {id}")
            flash(_('您没有权限查看已停产的产品'), 'danger')
            return redirect(url_for('product.product_list'))

        # 获取产品规格数据（包含缺失的编码规格）
        from app.services.spec_service import SpecService
        product_specs = SpecService.get_specs_with_coded_fields(
            SpecService.TYPE_PRODUCT,
            id,
            product.subcategory_id
        )

        # 将规格按分类分组（与配置矩阵保持一致）
        from app.models.product_code import SpecificationDictionary
        from app.models.spec_template import SpecCategory
        specs_by_category = {}
        spec_categories = []  # 按 display_order 排序的分类列表

        # 获取所有分类（按 display_order 排序）
        all_categories = SpecCategory.query.filter_by(is_active=True).order_by(SpecCategory.display_order).all()
        category_map = {cat.id: cat for cat in all_categories}

        # 通过 field_name 匹配 SpecificationDictionary 获取分类信息、英文名称和排序
        spec_names = [s['field_name'] for s in product_specs]
        definitions = SpecificationDictionary.query.filter(SpecificationDictionary.name.in_(spec_names)).all()
        name_to_category = {d.name: d.category_id for d in definitions}
        name_to_name_en = {d.name: d.name_en for d in definitions}
        name_to_display_order = {d.name: d.display_order for d in definitions}
        # 未分类的规格放入 category_id=0
        uncategorized_specs = []
        for spec in product_specs:
            # 添加英文名称：优先使用存储的英文名称，回退到 SpecificationDictionary
            if not spec.get('field_name_en'):
                spec['field_name_en'] = name_to_name_en.get(spec['field_name'], '')
            cat_id = name_to_category.get(spec['field_name'])
            if cat_id and cat_id in category_map:
                if cat_id not in specs_by_category:
                    specs_by_category[cat_id] = []
                specs_by_category[cat_id].append(spec)
            else:
                uncategorized_specs.append(spec)

        # 构建按 display_order 排序的分类列表
        for cat in all_categories:
            if cat.id in specs_by_category:
                sorted_specs = sorted(
                    specs_by_category[cat.id],
                    key=lambda s: s.get('display_order', name_to_display_order.get(s.get('field_name', ''), 9999))
                )
                spec_categories.append({
                    'id': cat.id,
                    'name': cat.name,
                    'name_en': cat.name_en,
                    'specs': sorted_specs
                })

        # 未分类的规格放在最后
        if uncategorized_specs:
            spec_categories.append({
                'id': 0,
                'name': '其他规格',
                'name_en': 'Other Specs',
                'specs': uncategorized_specs
            })

        # 计算有效图片和PDF路径（三级引用）
        from app.utils.product_helpers import get_effective_image, get_effective_pdf
        effective_image = get_effective_image(product)
        effective_pdf = get_effective_pdf(product)

        # 信任数据库路径，图片由前端 onerror 兜底，PDF 由下载/预览路由处理

        # 计算上一个/下一个产品ID（按列表页排序，用窗口函数避免加载全量ID）
        nav_result = db.session.execute(text("""
            WITH ordered AS (
                SELECT p.id,
                       LAG(p.id) OVER w AS prev_id,
                       LEAD(p.id) OVER w AS next_id
                FROM products p
                LEFT JOIN product_subcategories ps ON p.subcategory_id = ps.id
                LEFT JOIN product_categories pc ON ps.category_id = pc.id
                WINDOW w AS (ORDER BY pc.display_order, pc.id, ps.display_order, ps.name, p.model, p.id)
            )
            SELECT prev_id, next_id FROM ordered WHERE id = :pid
        """), {'pid': id}).fetchone()
        prev_product_id = nav_result.prev_id if nav_result else None
        next_product_id = nav_result.next_id if nav_result else None

        # 统计该产品MN在报价单中的数量（单次条件聚合）
        design_quantity = 0
        order_quantity = 0
        if product.product_mn:
            from app.models.quotation import QuotationDetail, Quotation

            stats = db.session.query(
                func.coalesce(func.sum(QuotationDetail.quantity), 0),
                func.coalesce(func.sum(case(
                    (Quotation.project_stage.in_(['awarded', 'signed']), QuotationDetail.quantity),
                    else_=0
                )), 0)
            ).outerjoin(Quotation, QuotationDetail.quotation_id == Quotation.id)\
             .filter(QuotationDetail.product_mn == product.product_mn)\
             .first()
            design_quantity, order_quantity = int(stats[0]), int(stats[1])

        # 获取MN锁定状态
        is_mn_locked = product.is_mn_locked or False

        # 获取模态框所需数据
        from app.models.product_code import ProductCodeField, ProductCodeFieldOption
        from app.utils.dictionary_helpers import get_currency_type_options

        categories = ProductCategory.query.order_by(ProductCategory.display_order).all()
        modal_categories = [
            {'id': cat.id, 'name': cat.name, 'name_en': cat.name_en or cat.name, 'code_letter': cat.code_letter or ''}
            for cat in categories
        ]

        region_fields = ProductCodeField.query.filter_by(field_type='origin_location')\
                                              .order_by(ProductCodeField.position).all()
        modal_regions = []
        for field in region_fields:
            code = field.code or '0'
            if code == "?":
                option = ProductCodeFieldOption.query.filter_by(field_id=field.id).first()
                code = option.code if option else "0"
            modal_regions.append({
                'id': field.id,
                'name': field.name,
                'name_en': field.name_en or field.name,
                'code': code
            })

        modal_currencies = get_currency_type_options()

        # 计算快照状态
        snapshot = product.code_definition_snapshot
        has_valid_snapshot = False
        if snapshot and isinstance(snapshot, dict):
            code_parts = snapshot.get('code_parts', [])
            for part in code_parts:
                if part.get('use_in_code', True) and part.get('code'):
                    has_valid_snapshot = True
                    break

        return render_template('product/tw_product_detail.html',
                               product=product,
                               product_specs=product_specs,
                               spec_categories=spec_categories,
                               effective_image=effective_image,
                               effective_pdf=effective_pdf,
                               prev_product_id=prev_product_id,
                               next_product_id=next_product_id,
                               order_quantity=order_quantity,
                               design_quantity=design_quantity,
                               is_mn_locked=is_mn_locked,
                               modal_categories=modal_categories,
                               modal_regions=modal_regions,
                               modal_currencies=modal_currencies,
                               has_valid_snapshot=has_valid_snapshot)
    except Exception as e:
        logger.error(f'查看产品详情页面时出错: {str(e)}', exc_info=True)
        flash(_('查看产品详情失败: %s') % str(e), 'danger')
        return redirect(url_for('product.product_list')) 

@bp.route('/api/products/<int:id>/update-status', methods=['POST'])
@login_required
@permission_required('product', 'edit')  # 添加产品编辑权限装饰器
def update_product_status(id):
    """更新产品状态API - 支持三种状态切换"""
    try:
        logger.debug(f'正在更新产品状态: ID={id}')
        
        # 检查用户是否为管理员
        if current_user.role != 'admin':
            logger.warning(f'非管理员用户 {current_user.username} 尝试更新产品状态: ID={id}')
            return jsonify({
                'success': False,
                'message': '只有管理员可以更新产品状态'
            }), 403
        
        # 解析请求数据
        data = request.json
        if not data or 'status' not in data:
            return jsonify({
                'success': False,
                'message': '请求数据格式不正确'
            }), 400
            
        # 获取目标状态
        target_status = data['status']
        if target_status not in ['active', 'discontinued', 'upcoming']:
            return jsonify({
                'success': False,
                'message': '状态值无效，请使用 active、discontinued 或 upcoming'
            }), 400
        
        # 查找产品
        product = Product.query.get(id)
        if not product:
            return jsonify({
                'success': False,
                'message': f'未找到ID为 {id} 的产品'
            }), 404
        
        # 如果状态没有变化，直接返回成功
        if product.status == target_status:
            return jsonify({
                'success': True,
                'message': '产品状态未变更',
                'status': target_status
            })
            
        # 更新状态
        product.status = target_status
        db.session.commit()
        
        # 记录操作
        status_text = ''
        if target_status == 'active':
            status_text = '生产中'
        elif target_status == 'discontinued':
            status_text = '已停产'
        elif target_status == 'upcoming':
            status_text = '待上市'
            
        logger.info(f'管理员 {current_user.username} 将产品状态更新为 {status_text}: ID={product.id}, 名称={product.name}')
        
        return jsonify({
            'success': True,
            'message': f'产品状态已更新为 {status_text}',
            'status': target_status
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新产品状态时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'更新产品状态失败: {str(e)}'
        }), 500

# PDF文件下载
@bp.route('/api/products/<int:id>/download-pdf', methods=['GET'])
@login_required
@permission_required('product', 'view')
def download_pdf(id):
    """下载产品PDF文件"""
    file_service = get_product_file_service()
    return file_service.download_pdf(id)



# PDF预览缩略图专用API（用于产品详情页面的小预览图）
@bp.route('/api/products/<int:id>/pdf-preview', methods=['GET'])
@login_required
@permission_required('product', 'view')
def get_pdf_preview(id):
    """获取PDF文件用于预览缩略图（不强制下载，支持PDF.js加载）"""
    file_service = get_product_file_service()
    return file_service.preview_pdf(id)


# ============= 报价单产品级联选择相关API =============

@bp.route('/api/products/subcategories', methods=['GET'])
@login_required
def get_subcategories_api():
    """
    获取指定分类下的所有子分类
    用于报价单产品选择的级联菜单
    优化：使用 JOIN + GROUP BY 单条 SQL，避免加载 Product 对象和 N+1 查询
    """
    try:
        category = request.args.get('category')

        if not category:
            return jsonify({
                'success': False,
                'message': '缺少分类参数'
            }), 400

        from app.models.product_code import ProductCategory, ProductSubcategory

        category_obj = ProductCategory.query.filter_by(name=category).first()
        if not category_obj:
            return jsonify({
                'success': False,
                'message': f'未找到分类: {category}'
            }), 404

        # 单条 SQL：JOIN + GROUP BY 直接统计子分类产品数量
        rows = db.session.query(
            ProductSubcategory.name,
            ProductSubcategory.display_order,
            func.count(Product.id)
        ).join(Product, Product.subcategory_id == ProductSubcategory.id)\
         .filter(
            Product.category_id == category_obj.id,
            Product.status == 'active'
        ).group_by(ProductSubcategory.id, ProductSubcategory.name, ProductSubcategory.display_order)\
         .order_by(func.coalesce(ProductSubcategory.display_order, 999))\
         .all()

        result = [{'name': name, 'count': count} for name, _, count in rows]

        # 兼容旧数据：没有 subcategory_id 但有 product_name 的产品
        orphan_rows = db.session.query(
            Product.product_name,
            func.count(Product.id)
        ).filter(
            Product.category_id == category_obj.id,
            Product.status == 'active',
            Product.subcategory_id.is_(None),
            Product.product_name.isnot(None)
        ).group_by(Product.product_name).all()

        for pname, cnt in orphan_rows:
            result.append({'name': pname, 'count': cnt})

        return jsonify({
            'success': True,
            'subcategories': result
        })

    except Exception as e:
        logger.error(f'查询子分类时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'查询子分类失败: {str(e)}'
        }), 500


@bp.route('/api/products/by-subcategory', methods=['GET'])
@login_required
def get_products_by_subcategory_api():
    """
    获取指定分类和子分类下的所有产品，按型号分组
    用于报价单产品选择：显示型号列表，每个型号下可能有多个产品
    注意：不返回product_desc字段（按需求隐藏描述）
    优化：SQL 直接过滤 + eager load + 批量关联查询，避免 400+ N+1 查询
    """
    try:
        category = request.args.get('category')
        subcategory = request.args.get('subcategory')

        if not category or not subcategory:
            return jsonify({
                'success': False,
                'message': '缺少分类或子分类参数'
            }), 400

        from app.models.product_code import ProductSubcategory, ProductCategory

        category_obj = ProductCategory.query.filter_by(name=category).first()
        if not category_obj:
            return jsonify({
                'success': False,
                'message': f'未找到分类: {category}'
            }), 404

        # 尝试通过子分类表查找，SQL 直接过滤
        subcategory_obj = ProductSubcategory.query.filter_by(
            category_id=category_obj.id, name=subcategory
        ).first()

        if subcategory_obj:
            # SQL 直接按 subcategory_id 过滤 + eager load subcategory_obj
            products = Product.query.options(joinedload(Product.subcategory_obj))\
                .filter(
                    Product.subcategory_id == subcategory_obj.id,
                    Product.status == 'active'
                ).all()
        else:
            # 兼容旧数据：按 product_name 过滤
            products = Product.query.filter(
                Product.category_id == category_obj.id,
                Product.product_name == subcategory,
                Product.subcategory_id.is_(None),
                Product.status == 'active'
            ).all()

        # 批量查询所有匹配产品的关联配置数量（替代逐个调用 get_relations_for_product）
        product_ids = [p.id for p in products]
        subcategory_ids = list({p.subcategory_id for p in products if p.subcategory_id})

        from app.models.product_relation import ProductRelation
        config_counts = {}
        if product_ids:
            # 产品级关联数量
            product_rel_rows = db.session.query(
                ProductRelation.main_product_id,
                func.count(ProductRelation.id)
            ).filter(
                ProductRelation.main_product_type == ProductRelation.MAIN_TYPE_PRODUCT,
                ProductRelation.main_product_id.in_(product_ids),
                ProductRelation.is_active == True
            ).group_by(ProductRelation.main_product_id).all()
            for pid, cnt in product_rel_rows:
                config_counts[pid] = cnt

            # 子分类级关联数量（所有同子分类产品共享）
            sub_rel_count = 0
            if subcategory_ids:
                sub_rel_count = db.session.query(func.count(ProductRelation.id)).filter(
                    ProductRelation.main_product_type == ProductRelation.MAIN_TYPE_SUBCATEGORY,
                    ProductRelation.main_product_id.in_(subcategory_ids),
                    ProductRelation.is_active == True
                ).scalar() or 0

            for pid in product_ids:
                config_counts[pid] = config_counts.get(pid, 0) + sub_rel_count

        # 批量查询同子分类下有图片的兄弟产品（用于 effective_image 的优先级2回退）
        sibling_images = {}
        if subcategory_obj:
            # 按 product_name 分组，找到每组第一个有图片的产品
            sibling_rows = db.session.query(
                Product.product_name,
                Product.image_path
            ).filter(
                Product.subcategory_id == subcategory_obj.id,
                Product.image_path.isnot(None),
                Product.image_path != ''
            ).all()
            for pname, img in sibling_rows:
                if pname and pname not in sibling_images:
                    sibling_images[pname] = img

        # 子分类图片（用于 effective_image 的优先级3回退）
        subcategory_image = None
        if subcategory_obj and hasattr(subcategory_obj, 'image_path'):
            subcategory_image = subcategory_obj.image_path

        # 构建产品数据 + 按型号(model)分组
        model_groups_dict = {}
        for product in products:
            model_key = product.model or product.product_name or '未命名产品'
            if model_key not in model_groups_dict:
                model_groups_dict[model_key] = []

            config_count = config_counts.get(product.id, 0)

            # 内联 effective_image 逻辑，避免逐个查询
            effective_image = product.image_path
            if not effective_image and product.product_name:
                effective_image = sibling_images.get(product.product_name)
            if not effective_image:
                effective_image = subcategory_image

            model_groups_dict[model_key].append({
                'id': product.id,
                'product_name': product.product_name,
                'model': product.model,
                'product_mn': product.product_mn,
                'spec_mn': product.spec_mn,
                'specification': product.specification,
                'brand': product.brand,
                'unit': product.unit,
                'retail_price': float(product.retail_price) if product.retail_price else None,
                'currency': product.currency,
                'status': product.status,
                'code_definition_snapshot': product.code_definition_snapshot,
                'image_path': product.image_path,
                'effective_image': effective_image,
                'config_count': config_count,
                'has_configurations': config_count > 0,
                'points': product.points,
                'points_tier': product.points_tier,
                'points_coefficient': float(product.points_coefficient) if product.points_coefficient else None
            })

        result = sorted([
            {'product_name': plist[0]['product_name'] if plist else model, 'model': model, 'count': len(plist), 'products': plist}
            for model, plist in model_groups_dict.items()
        ], key=lambda x: x['model'])

        return jsonify({
            'success': True,
            'model_groups': result
        })

    except Exception as e:
        logger.error(f'查询产品按子分类时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'查询产品失败: {str(e)}'
        }), 500


@bp.route('/api/products/search', methods=['GET'])
@login_required
def search_products_api():
    """产品搜索API（供产品选择器使用）"""
    try:
        term = request.args.get('term', '').strip()
        if not term or len(term) < 1:
            return jsonify([])

        from app.models.product_code import ProductSubcategory
        search_term = f'%{term}%'

        query = Product.query.outerjoin(
            ProductSubcategory, Product.subcategory_id == ProductSubcategory.id
        ).filter(
            Product.status == 'active',
            or_(
                ProductSubcategory.name.ilike(search_term),
                Product.product_name.ilike(search_term),
                Product.product_mn.ilike(search_term),
                Product.model.ilike(search_term),
                Product.brand.ilike(search_term)
            )
        ).order_by(Product.product_mn.asc()).limit(20).all()

        results = []
        for p in query:
            results.append({
                'id': p.id,
                'product_name': p.name,
                'model': p.model,
                'product_mn': p.product_mn,
                'specification': p.specification,
                'brand': p.brand,
                'unit': p.unit,
                'retail_price': float(p.retail_price) if p.retail_price else None,
                'currency': p.currency,
                'status': p.status,
                'points': p.points,
                'points_tier': p.points_tier
            })

        return jsonify(results)
    except Exception as e:
        logger.error(f'搜索产品失败: {str(e)}')
        return jsonify([])


@bp.route('/api/v1/user/product-points-summary', methods=['GET'])
@login_required
def get_user_product_points_summary():
    """获取当前用户的产品积分汇总（供导航栏使用）

    统一从 ledger 读取。销售：source_type='quotation'，产品经理：source_type='pm_category'。
    """
    try:
        from app.helpers.product_points import get_points_tier
        from app.models.user_points_ledger import UserPointsLedger

        current_year = datetime.now().year
        total_points = db.session.query(
            db.func.coalesce(db.func.sum(UserPointsLedger.points), 0)
        ).filter(
            UserPointsLedger.user_id == current_user.id,
            UserPointsLedger.year == current_year
        ).scalar()

        # 按来源分类汇总：quotation + pm_category 均归为"产品植入积分"
        categories = []
        if total_points > 0:
            categories.append({
                'name': '产品植入积分',
                'name_en': 'Product Points',
                'points': total_points
            })

        return jsonify({
            'success': True,
            'total_points': total_points,
            'points_tier': get_points_tier(total_points),
            'year': current_year,
            'categories': categories
        })
    except Exception as e:
        logger.error(f'获取用户积分汇总失败: {str(e)}')
        return jsonify({'success': False, 'total_points': 0, 'points_tier': 'none', 'year': datetime.now().year, 'categories': []})


@bp.route('/api/products/<int:product_id>/configurations', methods=['GET'])
@login_required
@permission_required('product', 'view')
def get_product_configurations(product_id):
    """
    获取指定产品的配置列表

    根据产品ID获取其关联的配置产品（配件、附件等）
    按产品名称分组返回

    Args:
        product_id: 产品ID

    Returns:
        JSON格式的配置列表，按产品名称分组
    """
    try:
        from app.models.product import Product
        from app.models.product_relation import ProductRelation

        # 验证产品是否存在
        product = Product.query.get(product_id)
        if not product:
            return jsonify({
                'success': False,
                'message': '产品不存在'
            }), 404

        # 获取产品的关联配置
        relations = ProductRelation.get_relations_for_product(product_id)

        # 按产品名称分组配置
        config_groups = {}
        for relation in relations:
            if relation.related_product and relation.related_product.status == 'active':
                related_prod = relation.related_product
                product_name = related_prod.product_name or '其他配置'

                if product_name not in config_groups:
                    config_groups[product_name] = []

                config_groups[product_name].append({
                    'id': related_prod.id,
                    'model': related_prod.model,
                    'product_mn': related_prod.product_mn,
                    'specification': related_prod.specification,
                    'product_name': related_prod.product_name,
                    'brand': related_prod.brand,
                    'unit': related_prod.unit,
                    'retail_price': float(related_prod.retail_price) if related_prod.retail_price else None,
                    'currency': related_prod.currency,
                    'relation_type': relation.relation_type,
                    'is_required': relation.is_required,
                    'default_quantity': relation.default_quantity
                })

        # 转换为列表格式
        configurations = []
        for product_name, items in config_groups.items():
            configurations.append({
                'product_name': product_name,
                'items': items
            })

        logger.info(f'获取产品 {product_id} 的配置，共 {len(configurations)} 个分组')

        return jsonify({
            'success': True,
            'configurations': configurations
        })

    except Exception as e:
        logger.error(f'获取产品配置时出错: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'获取配置失败: {str(e)}'
        }), 500


@bp.route('/products/export', methods=['GET'])
@login_required
@admin_required
def export_products():
    """导出产品库为Excel文件"""
    try:
        from app.models.product_spec import ProductSpec

        # 辅助函数：获取产品规格的展示文本（优先从 ProductSpec 表，回退到 specification 字段）
        def get_product_spec_display(product, ProductSpec_model):
            """获取产品规格的展示文本"""
            # 优先从 code_definition_snapshot 获取
            if product.code_definition_snapshot:
                snapshot = product.code_definition_snapshot
                code_parts = snapshot.get('code_parts', [])
                spec_parts = []
                for part in code_parts:
                    field_name = part.get('field_name', '')
                    value = part.get('value', '')
                    unit = part.get('unit', '')
                    if field_name and value:
                        display_value = f"{value} {unit}" if unit else value
                        spec_parts.append(f"{field_name}: {display_value}")
                if spec_parts:
                    return '; '.join(spec_parts)

            # 然后从 ProductSpec 表获取
            specs = ProductSpec_model.query.filter_by(product_id=product.id).order_by(ProductSpec_model.display_order).all()
            if specs:
                spec_parts = []
                for spec in specs:
                    if spec.field_name and spec.field_value:
                        unit = getattr(spec, 'unit', '') or ''
                        display_value = f"{spec.field_value} {unit}" if unit else spec.field_value
                        spec_parts.append(f"{spec.field_name}: {display_value}")
                if spec_parts:
                    return '; '.join(spec_parts)

            # 最后回退到 specification 字段
            return product.specification or ''

        # 查询所有产品（按分类体系排序，与产品列表页一致）
        products = Product.query\
            .outerjoin(ProductSubcategory, Product.subcategory_id == ProductSubcategory.id)\
            .outerjoin(ProductCategory, ProductSubcategory.category_id == ProductCategory.id)\
            .order_by(
                ProductCategory.display_order.asc(),
                ProductCategory.id.asc(),
                ProductSubcategory.display_order.asc(),
                ProductSubcategory.id.asc(),
                Product.product_name.asc(),
                Product.id.asc()
            ).all()

        # 创建工作簿
        wb = Workbook()

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        normal_font = Font(name='微软雅黑', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== Sheet 1: 产品列表 ==========
        ws_list = wb.active
        ws_list.title = "产品列表"

        # 产品列表表头（子分类 = 产品系列，产品名称 = 产品本身名称）
        list_headers = ['产品类型', '产品类别', '状态', '子分类', '产品名称', '型号',
                        '规格', '品牌', '单位', '价格', '货币', 'MN号', '创建时间']

        for col_idx, header in enumerate(list_headers, 1):
            cell = ws_list.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # 状态映射
        status_map = {
            'active': '生产中',
            'discontinued': '已停产',
            'upcoming': '待上市'
        }

        # 类型映射（涵盖所有可能的产品类型值，包括英文和中文）
        type_map = {
            # 英文 key
            'standard': '标准产品',
            'channel': '渠道产品',
            'third party': '第三方产品',
            'third_party': '第三方产品',  # 兼容下划线写法
            'project': '项目产品',
            # 中文 key（直接保留）
            '标准产品': '标准产品',
            '渠道产品': '渠道产品',
            '第三方产品': '第三方产品',
            '项目产品': '项目产品',
        }

        # 填充产品数据
        for row_idx, product in enumerate(products, 2):
            # 获取子分类名称（产品系列）
            subcategory_name = ''
            if product.subcategory_obj:
                subcategory_name = product.subcategory_obj.name

            # 获取产品类别（优先使用category_obj）
            category_name = ''
            if product.category_obj:
                category_name = product.category_obj.name
            elif product.category:
                category_name = product.category

            row_data = [
                type_map.get(product.type, product.type or ''),
                category_name,
                status_map.get(product.status, product.status or ''),
                subcategory_name,                   # 子分类（产品系列）
                product.product_name or '',         # 产品名称（产品本身的名称）
                product.model or '',
                get_product_spec_display(product, ProductSpec),  # 从 ProductSpec 表获取规格
                product.brand or '',
                product.unit or '',
                float(product.retail_price) if product.retail_price else '',
                product.currency or Config.DEFAULT_CURRENCY,
                product.product_mn or '',
                product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else ''
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_list.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = left_alignment
                cell.border = thin_border

        # 设置列宽（增加了子分类列，共13列）
        # A:产品类型, B:产品类别, C:状态, D:子分类, E:产品名称, F:型号, G:规格, H:品牌, I:单位, J:价格, K:货币, L:MN号, M:创建时间
        list_column_widths = {'A': 12, 'B': 12, 'C': 10, 'D': 18, 'E': 20, 'F': 15,
                             'G': 40, 'H': 12, 'I': 8, 'J': 12, 'K': 8, 'L': 15, 'M': 18}
        for col, width in list_column_widths.items():
            ws_list.column_dimensions[col].width = width

        # ========== Sheet 2-N: 按子分类分组的规格表 ==========
        # 子分类到Sheet名称的映射（用于产品列表超链接）
        subcategory_to_sheet = {}

        # 按子分类分组
        subcategory_groups = {}
        for product in products:
            # 获取子分类名称
            subcategory_name = product.subcategory_obj.name if product.subcategory_obj else None
            if not subcategory_name:
                continue

            subcategory_id = product.subcategory_id or 0
            key = (subcategory_id, subcategory_name)

            if key not in subcategory_groups:
                subcategory_groups[key] = []
            subcategory_groups[key].append(product)

        used_sheet_names = set(['产品列表'])

        for (subcategory_id, subcategory_name), group_products in subcategory_groups.items():
            # 收集该子分类所有产品的规格数据
            products_specs = []  # [(product, specs_dict), ...]
            all_spec_names = []  # 收集所有规格名称（保持顺序）

            for product in group_products:
                specs_dict = {}  # {field_name: {value, use_in_code, field_code, unit}}

                if product.code_definition_snapshot:
                    snapshot = product.code_definition_snapshot
                    code_parts = snapshot.get('code_parts', [])
                    for part in code_parts:
                        field_name = part.get('field_name', '')
                        if field_name:
                            unit = part.get('unit', '')
                            value = part.get('value', '')
                            # 合并值和单位显示
                            display_value = f"{value} {unit}" if value and unit else value
                            specs_dict[field_name] = {
                                'value': display_value,
                                'use_in_code': '是' if part.get('use_in_code', False) else '否',
                                'field_code': part.get('field_code', '') or part.get('code', '')
                            }
                            if field_name not in all_spec_names:
                                all_spec_names.append(field_name)
                else:
                    specs = ProductSpec.query.filter_by(product_id=product.id).order_by(ProductSpec.id).all()
                    for spec in specs:
                        if spec.field_name:
                            unit = getattr(spec, 'unit', '') or ''
                            value = spec.field_value or ''
                            # 合并值和单位显示
                            display_value = f"{value} {unit}" if value and unit else value
                            specs_dict[spec.field_name] = {
                                'value': display_value,
                                'use_in_code': '是' if spec.field_code else '否',
                                'field_code': spec.field_code or ''
                            }
                            if spec.field_name not in all_spec_names:
                                all_spec_names.append(spec.field_name)

                # 即使没有规格也添加产品（规格留空）
                products_specs.append((product, specs_dict))

            # 如果子分类下没有产品，跳过
            if not products_specs:
                continue

            # 生成Sheet名称（子分类名，截断到31字符）
            sheet_name = subcategory_name
            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('*', '_')
            sheet_name = sheet_name.replace('?', '_').replace('[', '_').replace(']', '_')
            sheet_name = sheet_name.replace(':', '_')

            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # 确保sheet名称唯一
            original_name = sheet_name
            counter = 1
            while sheet_name in used_sheet_names:
                suffix = f"_{counter}"
                sheet_name = original_name[:31-len(suffix)] + suffix
                counter += 1
            used_sheet_names.add(sheet_name)

            # 记录子分类到Sheet名称的映射
            subcategory_to_sheet[subcategory_id] = sheet_name

            # 创建规格Sheet
            ws_spec = wb.create_sheet(title=sheet_name)

            # ===== 第1行：规格名称栏头 + 各产品型号 =====
            # A1: 子分类名（作为行标题）
            cell = ws_spec.cell(row=1, column=1, value=subcategory_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

            # 每个产品占1列，第1行放型号
            for idx, (product, _) in enumerate(products_specs):
                col = 2 + idx  # 从第2列开始
                model_name = product.model or f'产品{idx+1}'
                cell = ws_spec.cell(row=1, column=col, value=model_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # ===== 第2行：MN编号 =====
            cell = ws_spec.cell(row=2, column=1, value='MN编号')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

            for idx, (product, _) in enumerate(products_specs):
                col = 2 + idx
                cell = ws_spec.cell(row=2, column=col, value=product.product_mn or '')
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # ===== 第3行：Spec MN =====
            cell = ws_spec.cell(row=3, column=1, value='Spec MN')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

            for idx, (product, _) in enumerate(products_specs):
                col = 2 + idx
                cell = ws_spec.cell(row=3, column=col, value=product.spec_mn or '')
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # ===== 第4行起：规格数据（规格名 + 各产品的指标值）=====
            for row_idx, spec_name in enumerate(all_spec_names, 4):
                # A列：规格名称
                cell = ws_spec.cell(row=row_idx, column=1, value=spec_name)
                cell.font = normal_font
                cell.alignment = left_alignment
                cell.border = thin_border

                # 各产品的规格值（每个产品1列，只放指标值）
                for p_idx, (product, specs_dict) in enumerate(products_specs):
                    col = 2 + p_idx  # 每个产品1列
                    spec_data = specs_dict.get(spec_name, {'value': '', 'use_in_code': '', 'field_code': ''})
                    cell = ws_spec.cell(row=row_idx, column=col, value=spec_data['value'])
                    cell.font = normal_font
                    cell.alignment = left_alignment
                    cell.border = thin_border

            # 设置列宽
            ws_spec.column_dimensions['A'].width = 15  # 规格名称列
            from openpyxl.utils import get_column_letter
            for idx in range(len(products_specs)):
                col = 2 + idx  # 每个产品1列
                ws_spec.column_dimensions[get_column_letter(col)].width = 18  # 指标值列

        # ========== 为产品列表的产品名称添加超链接 ==========
        link_font = Font(name='微软雅黑', size=10, color='0066CC', underline='single')

        for row_idx, product in enumerate(products, 2):
            # 获取子分类ID
            subcategory_id = product.subcategory_id

            # 如果该子分类有对应的规格Sheet，添加超链接
            if subcategory_id and subcategory_id in subcategory_to_sheet:
                sheet_name = subcategory_to_sheet[subcategory_id]
                cell = ws_list.cell(row=row_idx, column=4)  # D列是子分类
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.font = link_font

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        filename = f"产品库导出-{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        logger.info(f"用户 {current_user.username} 导出产品库，共 {len(products)} 个产品")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"导出产品库失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        flash(f'导出失败：{str(e)}', 'danger')
        return redirect(url_for('product.products_page'))


@bp.route('/p/<string:product_mn>')
def public_product_info(product_mn):
    """公开产品信息页面（无需登录）"""
    from app.models.product_spec import ProductSpec
    from app.models.dictionary import Dictionary

    # 获取语言参数（从URL参数，默认中文）
    lang = request.args.get('lang', 'zh')

    # 根据 product_mn 查找产品
    product = Product.query.filter_by(product_mn=product_mn).first()
    if not product:
        return render_template('product/public_not_found.html', product_mn=product_mn, lang=lang), 404

    # 获取规格数据
    specs = ProductSpec.query.filter_by(product_id=product.id).order_by(ProductSpec.display_order).all()

    # 计算有效图片路径（三级引用）
    from app.utils.product_helpers import get_effective_image
    effective_image = get_effective_image(product)

    # 获取厂商信息（查找标记为 is_vendor=True 的厂商企业）
    vendor_info = None
    vendor = Dictionary.query.filter_by(
        type='company',
        is_vendor=True,
        is_active=True
    ).first()
    if vendor and vendor.website:
        vendor_info = {
            'name': vendor.value,  # 公司全称
            'website': vendor.website
        }

    # 获取单位信息
    specs_with_unit = []
    for spec in specs:
        specs_with_unit.append({
            'name': spec.field_name,
            'value': spec.field_value or '-',
            'unit': getattr(spec, 'unit', '') or '',
            'code': spec.field_code
        })

    return render_template('product/public_info.html',
                           product=product,
                           effective_image=effective_image,
                           specs=specs_with_unit,
                           vendor_info=vendor_info,
                           lang=lang)


@bp.route('/products/<int:id>/qrcode')
@login_required
@permission_required('product', 'view')
def generate_product_qrcode(id):
    """生成产品二维码（仅包含详情页URL，兼容所有扫描器）"""
    import qrcode
    from io import BytesIO

    try:
        product = Product.query.get_or_404(id)

        if not product.product_mn:
            return jsonify({'error': '产品MN编码为空，无法生成二维码'}), 400

        # 获取当前语言设置
        lang = session.get('language', 'zh')

        # 生成详情页URL（附带语言参数）
        # 注：只使用URL作为二维码内容，确保iPhone等扫描器兼容性
        # 产品详细信息在扫描后打开的网页上查看
        qr_content = url_for('product.public_product_info',
                             product_mn=product.product_mn,
                             lang=lang,
                             _external=True)

        # 生成二维码
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(qr_content)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # 返回图片
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='image/png',
            as_attachment=False
        )

    except Exception as e:
        logger.error(f"生成产品二维码失败: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# 配置引入功能 API
# ============================================================

@bp.route('/api/products/configurations-tree', methods=['GET'])
@login_required
@permission_required('product', 'create')
def get_configurations_tree():
    """获取可引入配置的树状数据

    返回按分类/子分类组织的配置产品树，用于产品库引入功能。

    **环境自适应**：
    - SP8D: 查询本地 ProductConfiguration 表
    - OVS: 调用SP8D的远程API

    Returns:
        JSON: 树状配置数据
        {
            "success": true,
            "data": [
                {
                    "id": "cat_1",
                    "type": "category",
                    "name": "基站",
                    "children": [
                        {
                            "id": "sub_3",
                            "type": "subcategory",
                            "name": "宏基站",
                            "children": [
                                {
                                    "id": 123,
                                    "type": "configuration",
                                    "template_model": "BS2800",
                                    "mn_code": "EBS4X33N",
                                    ...
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    """
    from config import Config

    # ============================================================
    # 环境自适应逻辑：OVS调用SP8D API，SP8D查询本地数据库
    # ============================================================
    if Config.IS_OVS:
        # OVS环境：调用SP8D的远程API
        from app.services.sp8d_api_service import sp8d_api_service

        tree_data = sp8d_api_service.get_configurations_tree()

        if tree_data is not None:
            # 查询 OVS 本地已引入的 product_mn 集合，用于标记 already_imported
            local_imported_mns = set(
                mn for (mn,) in db.session.query(Product.product_mn).filter(
                    Product.product_mn.isnot(None),
                    Product.is_deleted == False
                ).all()
            )

            # 标记为远程配置，并根据 OVS 本地 Product 表判断是否已引入
            def mark_remote_configs(nodes):
                for node in nodes:
                    if node.get('type') == 'configuration':
                        node['is_remote'] = True
                        node['already_imported'] = node.get('mn_code', '') in local_imported_mns
                    if 'children' in node:
                        mark_remote_configs(node['children'])

            mark_remote_configs(tree_data)
            return jsonify({'success': True, 'data': tree_data})
        else:
            return jsonify({
                'success': False,
                'message': _('SP8D系统暂时不可用，请稍后重试或手动输入产品配置')
            }), 503

    # SP8D环境：查询本地数据库（保持原有逻辑）
    from app.models.spec_template import ProductConfiguration, SpecTemplate

    # 查询 pilot/production 状态的配置
    configs = ProductConfiguration.query.filter(
        ProductConfiguration.status.in_(['pilot', 'production']),
        ProductConfiguration.deleted_at.is_(None),
        ProductConfiguration.mn_code.isnot(None)
    ).join(SpecTemplate).all()

    # 按分类/子分类组织成树
    category_map = {}

    for config in configs:
        template = config.template
        if not template or not template.category_id:
            continue

        cat_id = template.category_id
        sub_id = template.subcategory_id

        # 检查是否已引入（排除已删除的产品）
        already_imported = Product.query.filter_by(
            source_configuration_id=config.id,
            is_deleted=False
        ).first() is not None

        # 获取状态显示文本
        status_display = _('可生产') if config.status == 'production' else _('小批价')

        # 查找区域ID和区域名称（通过区域编码查找 ProductCodeField）
        # ProductCodeField 有 code 字段或其第一个 option 的 code 来标识区域
        region_id = None
        region_name_display = config.region_name
        region_name_en_display = config.region_name_en if hasattr(config, 'region_name_en') else None
        if config.region:
            from app.models.product_code import ProductCodeField, ProductCodeFieldOption
            # 先尝试通过 ProductCodeField.code 匹配
            region_field = ProductCodeField.query.filter(
                ProductCodeField.field_type == 'origin_location',
                ProductCodeField.code == config.region
            ).first()
            if region_field:
                region_id = region_field.id
                # 始终以 ProductCodeField 为准，修正可能过期的 region_name
                region_name_display = region_field.name
                region_name_en_display = region_field.name_en
            else:
                # 如果 field.code 是 '?' 则通过 option.code 查找
                region_option = ProductCodeFieldOption.query.join(ProductCodeField).filter(
                    ProductCodeField.field_type == 'origin_location',
                    ProductCodeFieldOption.code == config.region
                ).first()
                if region_option:
                    region_id = region_option.field_id

        # 获取产品名称（使用模板名称）
        product_name = template.name or ''

        # 构建配置节点数据
        config_node = {
            'id': config.id,
            'type': 'configuration',
            'template_id': template.id,
            'template_model': template.model,
            'mn_code': config.mn_code,
            'config_code': config.config_code,
            'status': config.status,
            'status_display': status_display,
            'region': config.region,
            'region_id': region_id,
            'region_name': region_name_display or config.region,
            'region_name_en': region_name_en_display,
            'category_id': template.category_id,
            'category_name': template.category.name if template.category else None,
            'category_name_en': template.category.name_en if template.category else None,
            'subcategory_id': template.subcategory_id,
            'subcategory_name': template.subcategory.name if template.subcategory else None,
            'subcategory_name_en': template.subcategory.name_en if template.subcategory else None,
            'product_name': product_name,
            'product_name_en': template.name_en or None,
            'already_imported': already_imported
        }

        # 添加到分类树
        if cat_id not in category_map:
            category = template.category
            category_map[cat_id] = {
                'id': f'cat_{cat_id}',
                'type': 'category',
                'name': category.name if category else f'Category {cat_id}',
                'category_id': cat_id,
                'children': {}
            }

        # 添加到子分类
        if sub_id:
            if sub_id not in category_map[cat_id]['children']:
                subcategory = template.subcategory
                category_map[cat_id]['children'][sub_id] = {
                    'id': f'sub_{sub_id}',
                    'type': 'subcategory',
                    'name': subcategory.name if subcategory else f'Subcategory {sub_id}',
                    'subcategory_id': sub_id,
                    'children': []
                }
            category_map[cat_id]['children'][sub_id]['children'].append(config_node)
        else:
            # 没有子分类的情况（直接放在分类下）
            if '_no_sub' not in category_map[cat_id]['children']:
                category_map[cat_id]['children']['_no_sub'] = {
                    'id': f'sub_none_{cat_id}',
                    'type': 'subcategory',
                    'name': _('未分类'),
                    'subcategory_id': None,
                    'children': []
                }
            category_map[cat_id]['children']['_no_sub']['children'].append(config_node)

    # 转换为列表结构
    tree = []
    for cat_id in sorted(category_map.keys()):
        cat_node = category_map[cat_id].copy()
        cat_node['children'] = list(cat_node['children'].values())
        tree.append(cat_node)

    return jsonify({'success': True, 'data': tree})


@bp.route('/api/products/<int:product_id>/configuration-specs', methods=['GET'])
@login_required
@permission_required('product', 'edit')
def get_configuration_specs(product_id):
    """获取产品关联配置的规格数据

    返回该产品源配置中的所有有值的规格项，用于规格引入选择。
    编码规格（use_in_code=True）将被标记为必选。

    Args:
        product_id: 产品ID

    Returns:
        JSON: {
            "success": true,
            "data": {
                "configuration": {配置信息},
                "specs": [规格列表]
            }
        }
    """
    from app.models.spec_template import ProductConfiguration, SpecTemplate, SpecTemplateItem
    from app.models.spec_template import generate_safe_code_char

    product = Product.query.get_or_404(product_id)

    if not product.source_configuration_id:
        return jsonify({'success': False, 'message': _('该产品没有关联的配置版本')})

    config = ProductConfiguration.query.get(product.source_configuration_id)
    if not config:
        return jsonify({'success': False, 'message': _('关联的配置版本不存在')})

    template = config.template
    if not template:
        return jsonify({'success': False, 'message': _('配置模板不存在')})

    specs = []

    # 遍历模板规格项
    for item in template.items:
        if not item.spec_dict:
            continue

        # 获取配置值（优先）或通用值
        value = config.get_spec_value(item.id)
        if not value:
            continue  # 只返回有值的规格

        # 获取编码字符（如果是编码规格）
        code_char = None
        if item.use_in_code and item.options:
            code_char = generate_safe_code_char(value, item.options)
            # 处理编码长度
            if item.code_length == 2:
                code_char = code_char.ljust(2, 'X')[:2]
            else:
                code_char = code_char[:1]

        specs.append({
            'template_item_id': item.id,
            'name': item.spec_dict.name,
            'name_en': item.spec_dict.name_en,
            'value': value,
            'unit': item.spec_dict.unit or '',
            'category': item.spec_dict.category.name if item.spec_dict.category else '',
            'use_in_code': item.use_in_code,
            'code_char': code_char,
            'is_required': item.use_in_code  # 编码规格强制必选
        })

    # 按编码规格优先排序
    specs.sort(key=lambda x: (not x['use_in_code'], x['category']))

    return jsonify({
        'success': True,
        'data': {
            'configuration': {
                'id': config.id,
                'config_code': config.config_code,
                'region': config.region,
                'region_name': config.region_name or config.region,
                'template_model': template.model,
                'mn_code': config.mn_code
            },
            'specs': specs
        }
    })


@bp.route('/api/sp8d-configuration-specs/<int:config_id>', methods=['GET'])
@login_required
@permission_required('product', 'edit')
def get_sp8d_configuration_specs(config_id):
    """获取SP8D远程配置的规格数据（代理SP8D API）

    不绑定产品，直接根据SP8D配置版本ID获取规格。
    用于SP8D跨NAS导入产品时，在规格引入选择模态框中展示可选规格。
    返回格式与本地 get_configuration_specs 一致。
    """
    try:
        from app.services.sp8d_api_service import sp8d_api_service
        spec_data = sp8d_api_service.get_configuration_specs(config_id)

        if not spec_data:
            return jsonify({'success': False, 'message': _('无法从SP8D获取规格数据')})

        # 转换为与本地 get_configuration_specs 相同的返回格式
        config_info = spec_data.get('configuration', {})
        raw_specs = spec_data.get('specs', [])

        specs = []
        for spec in raw_specs:
            specs.append({
                'template_item_id': spec.get('template_item_id'),
                'name': spec.get('name', ''),
                'name_en': spec.get('name_en', ''),
                'value': spec.get('value', ''),
                'unit': spec.get('unit', ''),
                'category': spec.get('category', ''),
                'use_in_code': spec.get('use_in_code', False),
                'code_char': spec.get('code_char'),
                'is_required': spec.get('use_in_code', False)
            })

        return jsonify({
            'success': True,
            'data': {
                'configuration': config_info,
                'specs': specs
            }
        })

    except Exception as e:
        logger.error(f"获取SP8D配置规格失败: {e}")
        return jsonify({'success': False, 'message': _('获取SP8D规格数据失败')})


@bp.route('/api/products/<int:product_id>/import-specs', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def import_configuration_specs(product_id):
    """从配置导入规格到产品

    清空产品现有规格，导入选中的配置规格项。
    支持本地配置(from_config)和SP8D远程配置(from_sp8d)。
    重新生成产品描述和 spec_mn。

    Args:
        product_id: 产品ID

    Request Body:
        {
            "spec_ids": [1, 2, 5],  // 选中的规格项 template_item_id 列表
            "sp8d_config_id": 123   // (可选) SP8D远程配置版本ID
        }

    Returns:
        JSON: {
            "success": true,
            "message": "成功导入 X 个规格项",
            "redirect": "产品详情页URL"
        }
    """
    from app.models.product_spec import ProductSpec

    product = Product.query.get_or_404(product_id)

    # 检查请求体中是否有 sp8d_config_id，有则走SP8D远程导入逻辑
    data = request.get_json()
    sp8d_config_id = data.get('sp8d_config_id') if data else None
    if sp8d_config_id:
        return _import_sp8d_specs(product, int(sp8d_config_id))

    # 本地配置导入
    from app.models.spec_template import ProductConfiguration, SpecTemplate, SpecTemplateItem
    from app.models.spec_template import generate_safe_code_char

    if not product.source_configuration_id:
        return jsonify({'success': False, 'message': _('该产品没有关联的配置版本')})

    config = ProductConfiguration.query.get(product.source_configuration_id)
    if not config:
        return jsonify({'success': False, 'message': _('关联的配置版本不存在')})

    if not data:
        return jsonify({'success': False, 'message': _('请求数据无效')})

    selected_ids = data.get('spec_ids', [])

    try:
        # 清空现有规格
        ProductSpec.query.filter_by(product_id=product_id).delete()

        template = config.template
        new_specs = []
        display_order = 0

        # 创建新规格
        for item in template.items:
            if item.id not in selected_ids:
                continue

            if not item.spec_dict:
                continue

            value = config.get_spec_value(item.id)
            if not value:
                continue

            # 获取编码字符（如果是编码规格）
            code_char = None
            if item.use_in_code and item.options:
                code_char = generate_safe_code_char(value, item.options)
                # 处理编码长度
                if item.code_length == 2:
                    code_char = code_char.ljust(2, 'X')[:2]
                else:
                    code_char = code_char[:1]

            spec = ProductSpec(
                product_id=product_id,
                field_name=item.spec_dict.name,
                field_value=value,
                field_code=code_char,
                unit=item.spec_dict.unit or None,
                include_in_description=True,
                display_order=display_order
            )
            db.session.add(spec)
            new_specs.append(spec)
            display_order += 1

        db.session.flush()

        # 重新生成产品描述（包含单位）
        description_parts = []
        for spec in new_specs:
            if spec.include_in_description and spec.field_value:
                unit_str = f" {spec.unit}" if spec.unit else ""
                description_parts.append(f"{spec.field_name}: {spec.field_value}{unit_str}")
        product.specification = ", ".join(description_parts) if description_parts else ""

        # 直接使用配置版本的 mn_code 作为 spec_mn（不重新计算，避免 ProductCodeField 配置不一致问题）
        if config.mn_code:
            product.spec_mn = config.mn_code
            product.product_mn = config.mn_code
        else:
            # 如果配置没有 mn_code，回退到计算
            from app.utils.product_helpers import generate_spec_mn
            spec_mn = generate_spec_mn(product)
            if spec_mn:
                product.spec_mn = spec_mn
                product.product_mn = spec_mn

        # 锁定产品MN编码（引入后不可再修改）
        product.is_mn_locked = True

        # 生成编码定义快照
        from app.utils.product_helpers import generate_product_snapshot
        snapshot = generate_product_snapshot(product=product, source="config_import")
        if snapshot:
            product.code_definition_snapshot = snapshot
            logger.info(f'配置引入后生成编码快照成功: 产品ID={product.id}')

        product.updated_at = datetime.now()
        db.session.commit()

        logger.info(f"产品 {product_id} 成功导入 {len(new_specs)} 个规格项")

        return jsonify({
            'success': True,
            'message': _('成功导入 %(count)d 个规格项', count=len(new_specs)),
            'redirect': url_for('product.view_product_detail', id=product_id)
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"导入规格失败: {str(e)}")
        return jsonify({'success': False, 'message': _('导入失败: %(error)s', error=str(e))})


def _import_sp8d_specs(product, sp8d_config_id):
    """从SP8D远程配置导入选中的规格到产品（内部辅助函数）

    Args:
        product: 目标产品对象
        sp8d_config_id: SP8D远程配置版本ID（从请求体传入）
    """
    from app.models.product_spec import ProductSpec
    from app.services.sp8d_api_service import sp8d_api_service

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': _('请求数据无效')})

    selected_ids = set(data.get('spec_ids', []))

    try:
        # 从SP8D获取规格数据
        spec_data = sp8d_api_service.get_configuration_specs(sp8d_config_id)
        if not spec_data or not spec_data.get('specs'):
            return jsonify({'success': False, 'message': _('无法从SP8D获取规格数据')})

        config_info = spec_data.get('configuration', {})
        all_specs = spec_data['specs']

        # 清空现有规格
        ProductSpec.query.filter_by(product_id=product.id).delete()

        new_specs = []
        for idx, spec in enumerate(all_specs):
            if spec.get('template_item_id') not in selected_ids:
                continue

            is_coded = spec.get('use_in_code', False)
            ps = ProductSpec(
                product_id=product.id,
                field_name=spec.get('name', ''),
                field_name_en=spec.get('name_en', ''),
                field_value=spec.get('value', ''),
                field_value_en=spec.get('value_en', '') or None,
                field_code=spec.get('code_char', '') if is_coded else None,
                unit=spec.get('unit', '') or None,
                include_in_description=spec.get('include_in_description', is_coded),
                display_order=spec.get('display_order', idx)
            )
            db.session.add(ps)
            new_specs.append(ps)

        db.session.flush()

        # 使用与 Edit+Apply 相同的描述生成逻辑
        from app.services.spec_service import SpecService
        product.specification = SpecService.generate_description('product', product.id)

        # 使用SP8D配置的 mn_code
        mn_code = config_info.get('mn_code')
        if mn_code:
            product.spec_mn = mn_code
            product.product_mn = mn_code
        else:
            from app.utils.product_helpers import generate_spec_mn
            spec_mn = generate_spec_mn(product)
            if spec_mn:
                product.spec_mn = spec_mn
                product.product_mn = spec_mn

        product.is_mn_locked = True

        # 生成编码定义快照
        from app.utils.product_helpers import generate_product_snapshot
        snapshot = generate_product_snapshot(product=product, source="sp8d_import")
        if snapshot:
            product.code_definition_snapshot = snapshot

        product.updated_at = datetime.now()
        db.session.commit()

        logger.info(f"产品 {product.id} 从SP8D成功导入 {len(new_specs)} 个规格项")

        return jsonify({
            'success': True,
            'message': _('成功导入 %(count)d 个规格项', count=len(new_specs)),
            'redirect': url_for('product.view_product_detail', id=product.id)
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"SP8D规格导入失败: {str(e)}")
        return jsonify({'success': False, 'message': _('导入失败: %(error)s', error=str(e))})


# ============================================================================
# 产品关联配置相关API (从已废弃的product_management.py迁移 2025-12-26)
# ============================================================================

@bp.route('/api/product/<int:product_id>/relations', methods=['GET'])
@login_required
def get_product_relations(product_id):
    """
    获取产品的关联配置列表

    支持互斥组分组显示：
    - 互斥组内的产品会被分组在一起
    - 每个互斥组有自己的显示名称、选择模式、是否必选等属性
    - 普通关联产品单独列出
    """
    try:
        from app.models.product_relation import ProductRelation

        product = Product.query.get_or_404(product_id)

        relations_query = ProductRelation.query.join(
            Product,
            ProductRelation.related_product_id == Product.id
        ).filter(
            ProductRelation.main_product_type == ProductRelation.MAIN_TYPE_PRODUCT,
            ProductRelation.main_product_id == product_id,
            ProductRelation.is_active == True
        ).order_by(
            ProductRelation.display_order.asc(),
            Product.product_name.desc()
        ).all()

        normal_relations = []
        mutual_groups = {}

        lang = session.get('language', 'zh')
        badge_label_map = {
            'required_accessory': {'zh': '必选', 'en': 'Required'},
            'mutual_exclusion_required': {'zh': '必选互斥', 'en': 'Req.Excl'},
            'mutual_exclusion_optional': {'zh': '可选互斥', 'en': 'Opt.Excl'},
            'optional_accessory': {'zh': '可选配件', 'en': 'Optional'},
            'recommended': {'zh': '推荐', 'en': 'Recommend'},
            'alternative': {'zh': '替代产品', 'en': 'Alt'}
        }

        for relation in relations_query:
            if not relation.related_product:
                continue

            badge_class_map = {
                'required_accessory': 'badge relation-type-badge relation-type-required rounded-pill',
                'recommended': 'badge relation-type-badge relation-type-recommended rounded-pill',
                'required_mutual': 'badge relation-type-badge relation-type-required-mutual rounded-pill',
                'optional_mutual': 'badge relation-type-badge relation-type-optional-mutual rounded-pill'
            }
            badge_class = badge_class_map.get(relation.relation_type, 'badge relation-type-badge relation-type-unknown rounded-pill')

            nested_count = ProductRelation.query.join(
                Product,
                ProductRelation.related_product_id == Product.id
            ).filter(
                ProductRelation.main_product_type == ProductRelation.MAIN_TYPE_PRODUCT,
                ProductRelation.main_product_id == relation.related_product_id,
                ProductRelation.is_active == True,
                or_(
                    ProductRelation.relation_type == 'required_accessory',
                    ProductRelation.relation_type == 'recommended',
                    ProductRelation.mutual_exclusion_group.isnot(None)
                )
            ).count()

            related = relation.related_product
            relation_info = {
                'id': relation.id,
                'related_product_id': relation.related_product_id,
                'product_name': related.product_name or related.model or '',
                'product_model': related.model or related.product_name or '',
                'product_mn': related.product_mn,
                'brand': related.brand or '',
                'specification': related.specification or '',
                'retail_price': float(related.retail_price) if related.retail_price else 0,
                'unit': related.unit or 'Set',
                'points': related.points,
                'points_tier': related.points_tier,
                'relation_type': relation.relation_type,
                'relation_type_label': badge_label_map.get(relation.relation_type, {}).get(lang, relation.relation_type),
                'relation_type_label_class': badge_class,
                'default_quantity': relation.default_quantity,
                'is_required': relation.is_required,
                'is_default': relation.is_default or False,
                'has_nested_configs': nested_count > 0,
                'nested_count': nested_count
            }

            if relation.is_in_mutual_exclusion_group():
                group_id = relation.mutual_exclusion_group
                if group_id not in mutual_groups:
                    is_required = relation.is_group_required or False
                    group_badge_class = badge_class_map.get(
                        'required_mutual' if is_required else 'optional_mutual',
                        'badge relation-type-badge relation-type-unknown rounded-pill'
                    )
                    mutual_groups[group_id] = {
                        'group_id': group_id,
                        'group_name': relation.group_display_name or group_id,
                        'is_required': is_required,
                        'selection_mode': relation.group_selection_mode or 'single',
                        'badge_class': group_badge_class,
                        'badge_label': badge_label_map.get('mutual_exclusion_required' if is_required else 'mutual_exclusion_optional', {}).get(lang, '互斥'),
                        'products': []
                    }
                mutual_groups[group_id]['products'].append(relation_info)
            else:
                normal_relations.append(relation_info)

        total_count = len(normal_relations) + sum(len(group['products']) for group in mutual_groups.values())

        return jsonify({
            'success': True,
            'data': normal_relations,
            'groups': mutual_groups,
            'total': total_count
        })

    except Exception as e:
        logger.error(f"获取产品关联失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取产品关联失败: {str(e)}'}), 500


@bp.route('/api/product/<int:product_id>/relations/<int:relation_id>', methods=['DELETE'])
@login_required
@permission_required('product', 'edit')
def delete_product_relation(product_id, relation_id):
    """删除产品关联配置"""
    try:
        from app.models.product_relation import ProductRelation

        relation = ProductRelation.query.filter_by(
            id=relation_id,
            main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
            main_product_id=product_id
        ).first()

        if not relation:
            return jsonify({'success': False, 'message': '关联记录不存在'}), 404

        db.session.delete(relation)
        db.session.commit()

        logger.info(f"用户 {current_user.id} 删除了产品 {product_id} 的关联 {relation_id}")
        return jsonify({'success': True, 'message': '删除成功'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"删除产品关联失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'}), 500


@bp.route('/api/product/<int:product_id>/mutual-exclusion-group', methods=['DELETE'])
@login_required
@permission_required('product', 'edit')
def delete_mutual_exclusion_group(product_id):
    """删除互斥组（删除组内所有产品关联）"""
    try:
        from app.models.product_relation import ProductRelation

        data = request.get_json()
        group_id = data.get('group_id')

        if not group_id:
            return jsonify({'success': False, 'message': '缺少互斥组ID'}), 400

        relations = ProductRelation.query.filter_by(
            main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
            main_product_id=product_id,
            mutual_exclusion_group=group_id
        ).all()

        if not relations:
            return jsonify({'success': False, 'message': '互斥组不存在或已被删除'}), 404

        deleted_count = 0
        for relation in relations:
            db.session.delete(relation)
            deleted_count += 1

        db.session.commit()

        logger.info(f"用户 {current_user.id} 删除了产品 {product_id} 的互斥组 {group_id}")
        return jsonify({
            'success': True,
            'message': f'成功删除互斥组，共删除 {deleted_count} 个产品关联',
            'deleted_count': deleted_count
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"删除互斥组失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'}), 500


@bp.route('/api/product/<int:product_id>/relations/batch', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def batch_add_product_relations(product_id):
    """批量添加产品关联配置"""
    try:
        from app.models.product_relation import ProductRelation

        product = Product.query.get_or_404(product_id)

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '缺少请求数据'}), 400

        product_ids = data.get('product_ids', [])
        relation_type = data.get('relation_type', 'required_accessory')
        default_quantity = data.get('default_quantity', 1)

        if not product_ids:
            return jsonify({'success': False, 'message': '请至少选择一个产品'}), 400

        valid_types = ['required_accessory', 'recommended', 'optional_accessory', 'alternative']
        if relation_type not in valid_types:
            return jsonify({'success': False, 'message': f'无效的关联类型: {relation_type}'}), 400

        is_required = (relation_type == 'required_accessory')

        added_count = 0
        skipped_count = 0
        errors = []

        for related_product_id in product_ids:
            try:
                related_product = Product.query.get(related_product_id)
                if not related_product:
                    errors.append(f'产品ID {related_product_id} 不存在')
                    continue

                existing = ProductRelation.query.filter_by(
                    main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
                    main_product_id=product_id,
                    related_product_id=related_product_id,
                    is_active=True
                ).first()

                if existing:
                    skipped_count += 1
                    continue

                relation = ProductRelation(
                    main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
                    main_product_id=product_id,
                    related_product_id=related_product_id,
                    relation_type=relation_type,
                    is_required=is_required,
                    default_quantity=default_quantity,
                    is_active=True,
                    display_order=0
                )
                db.session.add(relation)
                added_count += 1

            except Exception as e:
                errors.append(f'添加产品ID {related_product_id} 失败: {str(e)}')

        db.session.commit()

        message_parts = []
        if added_count > 0:
            message_parts.append(f'成功添加 {added_count} 个关联产品')
        if skipped_count > 0:
            message_parts.append(f'跳过 {skipped_count} 个已存在的关联')
        if errors:
            message_parts.append(f'失败 {len(errors)} 个')

        message = '；'.join(message_parts)

        logger.info(f"用户 {current_user.id} 批量添加了产品 {product_id} 的关联: {message}")

        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'added': added_count,
                'skipped': skipped_count,
                'failed': len(errors),
                'errors': errors
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"批量添加产品关联失败: {str(e)}")
        return jsonify({'success': False, 'message': f'批量添加失败: {str(e)}'}), 500


@bp.route('/api/product/<int:product_id>/relations/batch-mutual-group', methods=['POST'])
@login_required
@permission_required('product', 'edit')
def batch_add_mutual_exclusion_group(product_id):
    """批量添加互斥组配置"""
    try:
        from app.models.product_relation import ProductRelation

        product = Product.query.get_or_404(product_id)

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '缺少请求数据'}), 400

        product_ids = data.get('product_ids', [])
        group_name = data.get('group_name', '').strip()
        is_required = data.get('is_required', False)
        default_product_id = data.get('default_product_id')
        default_quantity = data.get('default_quantity', 1)

        if not product_ids or len(product_ids) < 2:
            return jsonify({'success': False, 'message': '互斥组至少需要2个产品'}), 400

        if not group_name:
            return jsonify({'success': False, 'message': '请输入互斥组名称'}), 400

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        group_id = f"{group_name.lower().replace(' ', '_')}_{timestamp}"

        added_count = 0
        skipped_count = 0
        errors = []

        for pid in product_ids:
            try:
                related_product = Product.query.get(pid)
                if not related_product:
                    errors.append(f'产品ID {pid} 不存在')
                    continue

                existing = ProductRelation.query.filter_by(
                    main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
                    main_product_id=product_id,
                    related_product_id=pid,
                    is_active=True
                ).first()

                if existing:
                    skipped_count += 1
                    continue

                relation = ProductRelation(
                    main_product_type=ProductRelation.MAIN_TYPE_PRODUCT,
                    main_product_id=product_id,
                    related_product_id=pid,
                    relation_type='recommended',
                    default_quantity=default_quantity,
                    is_required=False,
                    mutual_exclusion_group=group_id,
                    group_display_name=group_name,
                    is_group_required=is_required,
                    group_selection_mode='single',
                    is_default=(pid == default_product_id)
                )

                db.session.add(relation)
                added_count += 1

            except Exception as e:
                errors.append(f'添加产品ID {pid} 失败: {str(e)}')

        db.session.commit()

        message = f'成功创建互斥组"{group_name}"'
        if skipped_count > 0:
            message += f'，跳过 {skipped_count} 个已存在的关联'

        logger.info(f"用户 {current_user.id} 为产品 {product_id} 创建了互斥组: {group_name}")

        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'group_id': group_id,
                'added': added_count,
                'skipped': skipped_count,
                'failed': len(errors),
                'errors': errors
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"创建互斥组失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建失败: {str(e)}'}), 500


@bp.route('/api/product-tree', methods=['GET'])
@login_required
def get_product_tree():
    """获取产品树数据 - 用于树状选择器"""
    try:
        tree_data = []

        categories = ProductCategory.get_ordered_list()

        for category in categories:
            category_node = {
                'id': f'cat_{category.id}',
                'category_id': category.id,
                'name': category.name,
                'code_letter': category.code_letter,
                'type': 'category',
                'icon': 'fa-layer-group',
                'children': []
            }

            subcategories = ProductSubcategory.query.filter_by(
                category_id=category.id
            ).order_by(ProductSubcategory.display_order).all()

            for subcategory in subcategories:
                subcategory_node = {
                    'id': f'sub_{subcategory.id}',
                    'subcategory_id': subcategory.id,
                    'name': subcategory.name,
                    'code_letter': subcategory.code_letter,
                    'type': 'subcategory',
                    'icon': 'fa-boxes',
                    'children': []
                }

                products = Product.query.filter_by(
                    subcategory_id=subcategory.id,
                    status='active'
                ).order_by(Product.product_mn).all()

                for product in products:
                    product_node = {
                        'id': f'prod_{product.id}',
                        'product_id': product.id,
                        'mn': product.product_mn,
                        'model': product.model or product.name or '',
                        'name': product.name or '',
                        'description': product.specification or '',
                        'type': 'product',
                        'icon': 'fa-microchip',
                        'status': product.status
                    }
                    subcategory_node['children'].append(product_node)

                if subcategory_node['children']:
                    category_node['children'].append(subcategory_node)

            if category_node['children']:
                tree_data.append(category_node)

        return jsonify({
            'success': True,
            'data': tree_data,
            'total_categories': len(tree_data)
        })

    except Exception as e:
        logger.error(f"获取产品树数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取产品树数据失败: {str(e)}'
        }), 500