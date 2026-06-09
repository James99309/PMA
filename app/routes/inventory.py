from flask import Blueprint, render_template, render_template_string, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from app import db
from app.models.inventory import Inventory, InventoryTransaction, Settlement, SettlementDetail, PurchaseOrder, PurchaseOrderDetail
from app.models.customer import Company
from app.models.product import Product
from app.models.product_code import ProductSubcategory
from app.utils.product_helpers import find_product_by_name
from app.utils.inventory_helpers import update_inventory, process_settlement, generate_order_number, get_inventory_status, calculate_order_totals
from app.decorators import permission_required, permission_required_with_approval_context
from datetime import datetime, date
import logging
from app.models.pricing_order import SettlementOrder, SettlementOrderDetail
from sqlalchemy import select, or_
from sqlalchemy.sql import func
import io
import pandas as pd
from app.helpers.approval_helpers import get_object_approval_instance, get_available_templates
from config import Config

logger = logging.getLogger(__name__)

inventory = Blueprint('inventory', __name__, url_prefix='/inventory')

@inventory.route('/settlement')
@login_required
@permission_required('settlement', 'view')
def settlement_list():
    """结算明细列表 - 使用通用列表组件架构"""
    try:
        # 确保数据库连接正常，如果有失败的事务则回滚
        try:
            db.session.rollback()
        except:
            pass
        
        # 获取查询参数
        search = request.args.get('search', '').strip()
        company_filter = request.args.get('company_filter')
        status_filter = request.args.get('status_filter')
        settlement_company_filter = request.args.get('settlement_company_filter')
        
        # 构建基础查询 - 只获取已审批批价单的结算单明细
        from app.models.pricing_order import SettlementOrderDetail, SettlementOrder, PricingOrder
        query = db.session.query(SettlementOrderDetail).join(SettlementOrder).join(PricingOrder)
        
        # 关键过滤：只显示已审批批价单的结算明细
        query = query.filter(PricingOrder.status == 'approved')
        
        # 应用筛选条件
        if search:
            search_filter = db.or_(
                SettlementOrder.order_number.contains(search),
                SettlementOrderDetail.product_name.contains(search),
                SettlementOrderDetail.product_mn.contains(search)
            )
            query = query.filter(search_filter)
        
        if company_filter:
            query = query.filter(SettlementOrder.distributor_id == company_filter)
        
        if settlement_company_filter:
            query = query.filter(SettlementOrderDetail.settlement_company_id == settlement_company_filter)
        
        if status_filter:
            if status_filter == 'completed':
                query = query.filter(SettlementOrderDetail.settlement_status == 'settled')
            elif status_filter == 'pending':
                query = query.filter(SettlementOrderDetail.settlement_status == 'pending')
        
        # 计算统计数据
        all_settlement_details = query.all()
        
        total_count = len(all_settlement_details)
        settled_count = len([d for d in all_settlement_details if d.settlement_status == 'settled'])
        pending_count = len([d for d in all_settlement_details if d.settlement_status == 'pending'])
        draft_count = total_count - settled_count - pending_count
        
        total_amount = sum(float(d.total_price or 0) for d in all_settlement_details) / 10000
        settled_amount = sum(float(d.total_price or 0) for d in all_settlement_details if d.settlement_status == 'settled') / 10000
        pending_amount = sum(float(d.total_price or 0) for d in all_settlement_details if d.settlement_status == 'pending') / 10000
        draft_amount = total_amount - settled_amount - pending_amount
        
        # 本月结算统计
        current_month = datetime.now().strftime('%Y-%m')
        thismonth_details = [d for d in all_settlement_details if d.settlement_status == 'settled' and d.settlement_date and d.settlement_date.strftime('%Y-%m') == current_month]
        thismonth_count = len(thismonth_details)
        thismonth_amount = sum(float(d.total_price or 0) for d in thismonth_details) / 10000
        
        # 获取公司列表用于筛选
        settlement_order_company_ids = db.session.query(
            SettlementOrder.distributor_id
        ).distinct().subquery()
        
        settlement_companies_all = db.session.query(Company).filter(
            Company.id.in_(
                db.session.query(settlement_order_company_ids.c.distributor_id)
            ),
            Company.is_deleted == False
        ).order_by(Company.company_name).all()
        
        settlement_company_ids = db.session.query(
            SettlementOrderDetail.settlement_company_id
        ).filter(
            SettlementOrderDetail.settlement_company_id.isnot(None)
        ).distinct().subquery()
        
        settlement_companies = db.session.query(Company).filter(
            Company.id.in_(
                db.session.query(settlement_company_ids.c.settlement_company_id)
            )
        ).order_by(Company.company_name).all()
        
        # 构建筛选搜索配置
        filter_config = {
            'action_url': url_for('inventory.settlement_list'),
            'form_id': 'settlementFilterForm',
            'reset_url': url_for('inventory.settlement_list'),
            'auto_submit': True,
            'ajax_mode': True,
            'ajax_endpoint': url_for('inventory.settlement_list_ajax'),
            'ajax_target': 'settlementTableBody',
            'ajax_columns': 12,
            'dynamic_reset_button': True,
            'search_field_id': 'search',
            
            'search_field': {
                'name': 'search',
                'label': '搜索',
                'placeholder': '结算单号、项目名称或产品名称',
                'value': search,
                'col_width': 3
            },
            
            'filter_fields': [
                {
                    'name': 'company_filter',
                    'label': '结算单公司',
                    'all_option_text': '全部公司',
                    'current_value': company_filter if company_filter and request.args else '',
                    'col_width': 2,
                    'options': [
                        {'value': company.id, 'label': company.company_name, 'translate': False} 
                        for company in settlement_companies_all
                    ]
                },
                {
                    'name': 'settlement_company_filter',
                    'label': '结算目标公司',
                    'all_option_text': '全部目标公司',
                    'current_value': settlement_company_filter if settlement_company_filter and request.args else '',
                    'col_width': 2,
                    'options': [
                        {'value': company.id, 'label': company.company_name, 'translate': False} 
                        for company in settlement_companies
                    ]
                },
                {
                    'name': 'status_filter',
                    'label': '结算状态',
                    'all_option_text': '全部状态',
                    'current_value': status_filter if status_filter and request.args else '',
                    'col_width': 2,
                    'options': [
                        {'value': 'completed', 'label': '已结算', 'translate': True},
                        {'value': 'pending', 'label': '待结算', 'translate': True}
                    ]
                }
            ],
            
            'search_button_text': '搜索',
            'reset_button_text': '重置'
        }
        
        # 构建通用列表配置
        list_config = {
            'module_name': 'settlement',
            'title': '结算明细管理',  # 设置为与页面标题一致
            'ajax_mode': True,
            
            # 无限滚动配置
            'infinite_scroll': {
                'enabled': True,
                'page_size': 50,
                'scroll_threshold': 100,
                'container_selector': '.table-responsive',
                'scroll_mode': 'container'  # 明确设置为容器滚动模式
            },
            
            # 统计卡片配置
            'stats': {
                'cards': [
                    {
                        'id': 'total',
                        'title': '全部明细',
                        'icon': 'fas fa-list',
                        'value': total_count,
                        'amount': total_amount,
                        'unit': '条',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'primary',
                        'data_key': 'total'
                    },
                    {
                        'id': 'settled',
                        'title': '已结算',
                        'icon': 'fas fa-check-circle',
                        'value': settled_count,
                        'amount': settled_amount,
                        'unit': '条',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'success',
                        'data_key': 'settled'
                    },
                    {
                        'id': 'pending',
                        'title': '待结算',
                        'icon': 'fas fa-clock',
                        'value': pending_count,
                        'amount': pending_amount,
                        'unit': '条',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'warning',
                        'data_key': 'pending'
                    },
                    {
                        'id': 'thismonth',
                        'title': '本月结算',
                        'icon': 'fas fa-calendar-check',
                        'value': thismonth_count,
                        'amount': thismonth_amount,
                        'unit': '条',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'info',
                        'data_key': 'thismonth'
                    }
                ]
            },
            
            # 筛选配置
            'filter': filter_config,
            
            # 表格配置
            'table': {
                'ajax_target': 'settlementTableBody',
                'title': '结算明细列表',
                'icon': 'fas fa-table',
                'show_header': True,
                'fixed_height_scroll': True,     # 启用固定高度滚动
                'enhanced_striping': True,       # 启用增强斑马纹效果
                'columns': [
                    {
                        'key': 'settlement_order.order_number',
                        'label': '结算单号',
                        'type': 'link',
                        'url_template': '/inventory/settlement_detail/{settlement_order.id}',
                        'width': '140px'
                    },
                    {
                        'key': 'settlement_order.project.project_name',
                        'label': '项目名称',
                        'type': 'text',
                        'width': '200px'
                    },
                    {
                        'key': 'product_name',
                        'label': '产品名称',
                        'type': 'text',
                        'width': '180px'
                    },
                    {
                        'key': 'product_model',
                        'label': '产品型号',
                        'type': 'text',
                        'width': '120px'
                    },
                    {
                        'key': 'brand',
                        'label': '品牌',
                        'type': 'text',
                        'width': '100px'
                    },
                    {
                        'key': 'product_mn',
                        'label': '产品MN',
                        'type': 'text',
                        'width': '120px'
                    },
                    {
                        'key': 'quantity',
                        'label': '数量',
                        'type': 'number',
                        'align': 'center',
                        'width': '80px'
                    },
                    {
                        'key': 'unit_price',
                        'label': '单价',
                        'type': 'number',
                        'format': 'currency',
                        'align': 'end',
                        'width': '100px'
                    },
                    {
                        'key': 'total_price',
                        'label': '总价',
                        'type': 'number',
                        'format': 'currency',
                        'align': 'end',
                        'width': '120px'
                    },
                    {
                        'key': 'settlement_company',
                        'label': '结算目标公司',
                        'type': 'text',
                        'width': '150px'
                    },
                    {
                        'key': 'settlement_status',
                        'label': '结算状态',
                        'type': 'badge',
                        'render': 'render_settlement_status_badge',
                        'width': '100px'
                    },
                    {
                        'key': 'settlement_date',
                        'label': '结算时间',
                        'type': 'date',
                        'format': '%Y-%m-%d',
                        'width': '120px'
                    }
                ]
            }
        }
        
        return render_template('inventory/settlement_list.html', list_config=list_config)
                             
    except Exception as e:
        logger.error(f"获取结算明细列表失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        flash(f'加载结算明细列表失败：{str(e)}', 'danger')
        
        # 错误时的默认list_config
        error_list_config = {
            'module_name': 'settlement',
            'title': '结算明细',  # 设置为与页面标题一致
            'ajax_mode': True,
            'stats': {'cards': []},
            'filter': {
                'action_url': url_for('inventory.settlement_list'),
                'form_id': 'settlementFilterForm',
                'reset_url': url_for('inventory.settlement_list'),
                'search_field': {
                    'name': 'search',
                    'label': '搜索',
                    'placeholder': '结算单号、项目名称或产品名称',
                    'value': '',
                    'col_width': 3
                },
                'filter_fields': [],
                'search_button_text': '搜索',
                'reset_button_text': '重置'
            },
            'table': {
                'ajax_target': 'settlementTableBody',
                'columns': []
            }
        }
        
        return render_template('inventory/settlement_list.html', list_config=error_list_config)

@inventory.route('/settlement/export')
@login_required
@permission_required('inventory', 'view')
def export_settlement_list():
    """导出结算明细列表为Excel"""
    try:
        # 获取查询参数（和settlement_list函数相同的逻辑）
        search = request.args.get('search', '').strip()
        company_filter = request.args.get('company_filter')
        status_filter = request.args.get('status_filter')
        settlement_company_filter = request.args.get('settlement_company_filter')
        
        # 构建基础查询 - 只获取已审批批价单的结算单明细
        from app.models.pricing_order import SettlementOrderDetail, SettlementOrder, PricingOrder
        query = db.session.query(SettlementOrderDetail).join(SettlementOrder).join(PricingOrder)
        
        # 关键过滤：只显示已审批批价单的结算明细
        query = query.filter(PricingOrder.status == 'approved')
        
        # 搜索条件
        if search:
            search_filter = db.or_(
                SettlementOrder.order_number.contains(search),
                SettlementOrderDetail.product_name.contains(search),
                SettlementOrderDetail.product_mn.contains(search)
            )
            query = query.filter(search_filter)
        
        # 结算单公司过滤（分销商）
        if company_filter:
            query = query.filter(SettlementOrder.distributor_id == company_filter)
        
        # 结算目标公司过滤
        if settlement_company_filter:
            query = query.filter(SettlementOrderDetail.settlement_company_id == settlement_company_filter)
        
        # 结算状态过滤
        if status_filter:
            if status_filter == 'completed':
                query = query.filter(SettlementOrderDetail.settlement_status == 'settled')
            elif status_filter == 'pending':
                query = query.filter(SettlementOrderDetail.settlement_status == 'pending')
        
        # 排序
        query = query.order_by(SettlementOrder.created_at.desc(), SettlementOrderDetail.id.desc())
        
        # 获取所有数据（不分页）
        settlement_details = query.all()
        
        # 准备筛选条件信息
        filter_info = []
        filter_info.append(['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        filter_info.append(['数据总数', f'{len(settlement_details)} 条记录'])
        filter_info.append(['', ''])  # 空行
        filter_info.append(['筛选条件', ''])
        
        # 添加具体筛选条件
        if search:
            filter_info.append(['搜索关键词', search])
        
        if company_filter:
            # 获取公司名称
            from app.models.customer import Company
            company = Company.query.get(company_filter)
            company_name = company.company_name if company else f'ID:{company_filter}'
            filter_info.append(['结算单公司', company_name])
        
        if settlement_company_filter:
            from app.models.customer import Company
            company = Company.query.get(settlement_company_filter)
            company_name = company.company_name if company else f'ID:{settlement_company_filter}'
            filter_info.append(['结算目标公司', company_name])
        
        if status_filter:
            status_map = {
                'completed': '已结算',
                'pending': '待结算'
            }
            filter_info.append(['结算状态', status_map.get(status_filter, status_filter)])
        
        # 如果没有任何筛选条件
        if not any([search, company_filter, settlement_company_filter, status_filter]):
            filter_info.append(['筛选状态', '未应用筛选条件，显示全部数据'])
        
        filter_info.append(['', ''])  # 空行
        
        # 准备导出数据
        export_data = []
        for detail in settlement_details:
            export_data.append({
                '结算单号': detail.settlement_order.order_number,
                '项目名称': detail.settlement_order.project.project_name if detail.settlement_order.project else '无项目',
                '产品名称': detail.product_name if detail.product_name else '无产品',
                '产品型号': detail.product_model if detail.product_model else '-',
                '品牌': detail.brand if detail.brand else '-',
                '产品MN': detail.product_mn if detail.product_mn else '-',
                '数量': detail.quantity,
                '单价(元)': round(detail.unit_price, 2),
                '总价(元)': round(detail.total_price, 2),
                '结算目标公司': detail.settlement_company.company_name if detail.settlement_company else '未指定',
                '结算状态': '已结算' if detail.settlement_status == 'settled' else '待结算',
                '结算时间': detail.settlement_date.strftime('%Y-%m-%d %H:%M') if detail.settlement_date else '-'
            })
        
        # 创建DataFrame
        df = pd.DataFrame(export_data)
        
        # 创建筛选条件DataFrame
        filter_df = pd.DataFrame(filter_info, columns=['项目', '值'])
        
        # 生成文件名
        current_time = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'结算明细表-{current_time}.xlsx'
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 先写入筛选条件信息
            filter_df.to_excel(writer, sheet_name='结算明细', index=False, header=False, startrow=0)
            
            # 计算数据开始行（筛选条件行数 + 2行间距）
            data_start_row = len(filter_info) + 2
            
            # 写入数据表格
            df.to_excel(writer, sheet_name='结算明细', index=False, startrow=data_start_row)
            
            # 获取工作表并设置格式
            worksheet = writer.sheets['结算明细']
            
            # 设置筛选条件区域的格式
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 标题行格式（导出时间、数据总数等）
            title_font = Font(bold=True, size=12)
            filter_font = Font(bold=True, size=10, color='0066CC')
            
            for row in range(1, len(filter_info) + 1):
                cell_a = worksheet.cell(row=row, column=1)
                cell_b = worksheet.cell(row=row, column=2)
                
                # 设置关键信息的格式
                if row <= 2:  # 导出时间和数据总数
                    cell_a.font = title_font
                    cell_b.font = title_font
                elif cell_a.value == '筛选条件':  # 筛选条件标题
                    cell_a.font = filter_font
                elif cell_a.value and cell_a.value not in ['', '筛选条件']:  # 具体筛选项
                    cell_a.font = Font(bold=True, size=9)
            
            # 设置数据表格标题行格式
            header_row = data_start_row + 1
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            column_widths = {
                'A': 18,  # 项目/结算单号
                'B': 25,  # 值/项目名称
                'C': 20,  # 产品名称
                'D': 15,  # 产品型号
                'E': 12,  # 品牌
                'F': 15,  # 产品MN
                'G': 10,  # 数量
                'H': 12,  # 单价
                'I': 12,  # 总价
                'J': 20,  # 结算目标公司
                'K': 12,  # 结算状态
                'L': 18,  # 结算时间
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置数据区域的边框
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为数据表格添加边框
            for row in range(header_row, header_row + len(df) + 1):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
        
        output.seek(0)
        
        logger.info(f"用户 {current_user.username} 导出结算明细列表，共 {len(settlement_details)} 条记录")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"导出结算明细列表失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        flash(f'导出失败：{str(e)}', 'danger')
        return redirect(url_for('inventory.settlement_list'))

@inventory.route('/settlement_orders')
@login_required
@permission_required('settlement', 'view')
def settlement_order_list():
    """结算单列表"""
    try:
        # 确保数据库事务状态干净
        try:
            db.session.rollback()
        except Exception:
            pass
            
        print("=== 执行了 settlement_order_list 函数 ===")
        logger.info("=== 执行了 settlement_order_list 函数 ===")
        
        # 获取搜索和筛选参数
        search = request.args.get('search', '').strip()
        settlement_company_id = request.args.get('settlement_company', '')
        settlement_status = request.args.get('settlement_status', '')
        
        # 调试信息：检查参数获取情况
        logger.info(f"=== 结算单列表参数调试 ===")
        logger.info(f"URL参数: {dict(request.args)}")
        logger.info(f"搜索关键词: '{search}'")
        logger.info(f"结算公司ID: '{settlement_company_id}'")
        logger.info(f"结算状态: '{settlement_status}'")
        
        # 构建基础查询：只获取来自已审批批价单的结算单
        from app.models.pricing_order import PricingOrder
        from app.models.project import Project
        
        query = SettlementOrder.query.join(
            PricingOrder, SettlementOrder.pricing_order_id == PricingOrder.id
        ).join(
            Project, SettlementOrder.project_id == Project.id
        ).filter(
            PricingOrder.status == 'approved'
        )
        
        # 应用搜索条件
        if search:
            query = query.filter(
                db.or_(
                    SettlementOrder.order_number.ilike(f'%{search}%'),
                    Project.project_name.ilike(f'%{search}%')
                )
            )
        
        # 应用公司筛选
        if settlement_company_id:
            query = query.filter(SettlementOrder.dealer_id == settlement_company_id)
        
        # 应用状态筛选
        if settlement_status:
            query = query.filter(SettlementOrder.settlement_status == settlement_status)
        
        # 执行查询，添加事务保护
        try:
            settlement_orders = query.order_by(SettlementOrder.created_at.desc()).all()
        except Exception as e:
            logger.error(f"结算单查询失败: {e}")
            # 回滚失败的事务
            try:
                db.session.rollback()
            except Exception:
                pass
            # 返回空结果，避免页面崩溃
            settlement_orders = []
        
        # 初始化统计变量
        fully_settled_count = 0
        partially_settled_count = 0
        pending_count = 0  # 正确初始化
        
        fully_settled_amount = 0.0
        partially_settled_amount = 0.0
        pending_amount = 0.0  # 正确初始化
        
        # 先统计数据库中实际的状态分布
        status_distribution = {}
        for order in settlement_orders:
            status = order.status
            if status not in status_distribution:
                status_distribution[status] = {'count': 0, 'amount': 0.0}
            status_distribution[status]['count'] += 1
            status_distribution[status]['amount'] += float(order.total_amount or 0.0)
        
        logger.info("=== 数据库中结算单实际状态分布 ===")
        for status, data in status_distribution.items():
            logger.info(f"状态 '{status}': {data['count']} 单, {data['amount']:.2f} 元")
        
        # 统计每个结算单的结算状态（只统计已审批批价单的结算单）
        for order in settlement_orders:
            order_amount = order.total_amount or 0.0
            # 确认是已审批批价单的结算单
            pricing_order = order.pricing_order_ref
            if pricing_order and pricing_order.status == 'approved':
                settlement_status = order.settlement_status
                
                logger.info(f"结算单 {order.order_number}: 批价单状态='{pricing_order.status}', 结算状态='{settlement_status}', 金额={order_amount}")
                
                # 根据settlement_status字段统计数量和金额
                if settlement_status == 'pending':
                    pending_count += 1
                    pending_amount += order_amount
                    logger.info(f"  -> 计入待结算: 当前待结算数量={pending_count}")
                elif settlement_status == 'fully_settled':
                    fully_settled_count += 1
                    fully_settled_amount += order_amount
                    logger.info(f"  -> 计入完全结算: 当前完全结算数量={fully_settled_count}")
                elif settlement_status == 'partially_settled':
                    partially_settled_count += 1
                    partially_settled_amount += order_amount
                    logger.info(f"  -> 计入部分结算: 当前部分结算数量={partially_settled_count}")
                else:
                    # 如果有其他状态，记录并按待结算处理
                    logger.warning(f"结算单 {order.order_number} 有未知结算状态: '{settlement_status}', 按待结算处理")
                    pending_count += 1
                    pending_amount += order_amount
                    logger.info(f"  -> 未知状态计入待结算: 当前待结算数量={pending_count}")
            else:
                logger.info(f"跳过非已审批批价单的结算单: {order.order_number} (批价单状态: {pricing_order.status if pricing_order else 'None'})")
        
        # 计算总数和总金额
        total_count = len(settlement_orders)
        total_amount = sum(float(order.total_amount or 0.0) for order in settlement_orders)
        
        # 输出调试信息
        logger.info(f"=== 结算单列表统计 ===")
        logger.info(f"总结算单数: {total_count}, 完全结算: {fully_settled_count}, 部分结算: {partially_settled_count}, 待结算: {pending_count}")
        logger.info(f"总金额: {total_amount}, 完全结算金额: {fully_settled_amount}, 部分结算金额: {partially_settled_amount}, 待结算金额: {pending_amount}")
        logger.info(f"传递给模板的待结算数量: {pending_count}, 待结算金额万元: {float(pending_amount) / 10000}")
        
        # 验证数学关系
        calculated_total = fully_settled_count + partially_settled_count + pending_count
        if calculated_total != total_count:
            logger.warning(f"结算单数量统计不匹配！计算值: {calculated_total}, 实际值: {total_count}")
        
        # 获取有结算单的公司用于筛选下拉框
        company_ids_with_settlements = db.session.query(Company.id.distinct()).join(
            SettlementOrder, Company.id == SettlementOrder.dealer_id
        ).join(
            PricingOrder, SettlementOrder.pricing_order_id == PricingOrder.id
        ).filter(
            PricingOrder.status == 'approved',
            Company.is_deleted == False
        ).all()
        
        company_ids = [row[0] for row in company_ids_with_settlements]
        companies = Company.query.filter(
            Company.id.in_(company_ids),
            Company.is_deleted == False
        ).order_by(Company.company_name).all()
        
        # 调试筛选配置构建
        logger.info(f"=== 构建筛选配置 ===")
        logger.info(f"即将传递给模板的 settlement_status: '{settlement_status}'")
        logger.info(f"settlement_status 是否为空: {settlement_status == ''}")
        logger.info(f"settlement_status 布尔值: {bool(settlement_status)}")
        
        # 构建筛选搜索配置
        filter_config = {
            'action_url': url_for('inventory.settlement_order_list'),
            'form_id': 'settlementOrderFilterForm',
            'reset_url': url_for('inventory.settlement_order_list'),
            'auto_submit': True,                # 启用自动筛选（关键配置）
            'ajax_mode': True,                  # 启用AJAX模式
            'ajax_endpoint': url_for('inventory.settlement_order_list_ajax'),
            'ajax_target': 'settlementTableBody',
            'ajax_columns': 7,
            'dynamic_reset_button': True,       # 启用动态重置按钮
            'search_field_id': 'search',        # 搜索字段ID（修复搜索功能）
            
            'search_field': {
                'name': 'search',
                'label': '搜索',
                'placeholder': '结算单编号或项目名称',
                'value': search,
                'col_width': 4
            },
            
            'filter_fields': [
                {
                    'name': 'settlement_company',
                    'label': '结算公司',
                    'all_option_text': '全部公司',
                    'current_value': settlement_company_id if settlement_company_id and request.args else '',
                    'col_width': 3,
                    'options': [
                        {'value': company.id, 'label': company.company_name, 'translate': False} 
                        for company in companies
                    ]
                },
                {
                    'name': 'settlement_status',
                    'label': '结算状态',
                    'all_option_text': '全部状态',
                    'current_value': settlement_status if settlement_status and request.args else '',
                    'col_width': 3,
                    'options': [
                        {'value': 'pending', 'label': '待结算', 'translate': True},
                        {'value': 'partially_settled', 'label': '部分结算', 'translate': True},
                        {'value': 'fully_settled', 'label': '已结算', 'translate': True}
                    ]
                }
            ],
            
            'search_button_text': '搜索',
            'reset_button_text': '重置'
        }
        
        # 统计数据
        stats = {
            'total': total_count,
            'total_amount': float(total_amount) / 10000,  # 转换为万元
            'fully_settled': fully_settled_count,
            'fully_settled_amount': float(fully_settled_amount) / 10000,
            'partially_settled': partially_settled_count,
            'partially_settled_amount': float(partially_settled_amount) / 10000,
            'pending': pending_count,  # 使用真实统计值
            'pending_amount': float(pending_amount) / 10000  # 使用真实统计值
        }
        
        # 构建通用列表配置
        list_config = {
            'module_name': 'settlement',
            'title': '结算管理',  # 设置为与页面标题一致
            'ajax_mode': True,
            
            # 无限滚动配置
            'infinite_scroll': {
                'enabled': True,
                'page_size': 60,
                'scroll_threshold': 100,
                'container_selector': '.table-responsive',
                'scroll_mode': 'container'  # 明确设置为容器滚动模式
            },
            
            # 统计卡片配置
            'stats': {
                'cards': [
                    {
                        'id': 'total',
                        'title': '总结算',
                        'icon': 'fas fa-list',
                        'value': stats['total'],
                        'amount': f"{stats['total_amount']:.2f}",
                        'unit': '单',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'primary',
                        'data_key': 'total'
                    },
                    {
                        'id': 'fullySettled',
                        'title': '已结算',
                        'icon': 'fas fa-check-circle',
                        'value': stats['fully_settled'],
                        'amount': f"{stats['fully_settled_amount']:.2f}",
                        'unit': '单',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'success',
                        'data_key': 'fully_settled'
                    },
                    {
                        'id': 'partiallySettled',
                        'title': '部分结算',
                        'icon': 'fas fa-exclamation-triangle',
                        'value': stats['partially_settled'],
                        'amount': f"{stats['partially_settled_amount']:.2f}",
                        'unit': '单',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'warning',
                        'data_key': 'partially_settled'
                    },
                    {
                        'id': 'pending',
                        'title': '待结算',
                        'icon': 'fas fa-clock',
                        'value': stats['pending'],
                        'amount': f"{stats['pending_amount']:.2f}",
                        'unit': '单',
                        'amount_unit': Config.AMOUNT_UNIT,
                        'color': 'danger',
                        'data_key': 'pending'
                    }
                ]
            },
            
            # 筛选配置（复用现有筛选组件）
            'filter': filter_config,
            
            # 表格配置
            'table': {
                'ajax_target': 'settlementTableBody',
                'title': '结算单列表',
                'icon': 'fas fa-table',
                'show_batch_actions': False,
                'fixed_height_scroll': True,     # 启用固定高度滚动
                'enhanced_striping': True,       # 启用增强斑马纹效果
                'columns': [
                    {
                        'key': 'order_number',
                        'field': 'order_number',
                        'label': '结算单编号',
                        'type': 'link',
                        'width': '140px',
                        'sort_type': 'string'
                    },
                    {
                        'key': 'project_name',
                        'field': 'project_name',
                        'label': '关联项目',
                        'type': 'text',
                        'width': '200px',
                        'sort_type': 'string'
                    },
                    {
                        'key': 'dealer_name',
                        'field': 'dealer_name',
                        'label': '结算公司',
                        'type': 'text',
                        'width': '150px',
                        'sort_type': 'string'
                    },
                    {
                        'key': 'product_count',
                        'field': 'product_count',
                        'label': '产品数量',
                        'type': 'number',
                        'align': 'end',
                        'width': '80px',
                        'sort_type': 'number'
                    },
                    {
                        'key': 'total_amount',
                        'field': 'total_amount',
                        'label': '总金额',
                        'type': 'text',  # 已格式化的金额字符串
                        'align': 'end',
                        'width': '120px',
                        'sort_type': 'number'
                    },
                    {
                        'key': 'settlement_status',
                        'field': 'settlement_status',
                        'label': '结算情况',
                        'type': 'badge',
                        'render': 'render_settlement_situation_badge',
                        'width': '100px',
                        'sort_type': 'string'
                    },
                    {
                        'key': 'created_time',
                        'field': 'created_time',
                        'label': '创建时间',
                        'type': 'text',
                        'width': '120px',
                        'sort_type': 'date'
                    }
                ]
            }
        }
        
        logger.info(f"即将传递给模板的数据: pending_count={stats['pending']}, pending_amount={stats['pending_amount']}")
        
        return render_template('inventory/settlement_order_list.html', 
                             settlement_orders=settlement_orders,
                             list_config=list_config,
                             companies=companies)
                             
    except Exception as e:
        logger.error(f"获取结算单列表失败：{str(e)}")
        flash(f'获取结算单列表失败：{str(e)}', 'danger')
        # 获取有结算单的公司以供筛选使用
        try:
            companies = db.session.query(Company).join(
                SettlementOrder, Company.id == SettlementOrder.dealer_id
            ).join(
                PricingOrder, SettlementOrder.pricing_order_id == PricingOrder.id
            ).filter(
                PricingOrder.status == 'approved',
                Company.is_deleted == False
            ).distinct().order_by(Company.company_name).all()
        except:
            companies = []
        
        # 错误处理时的默认列表配置
        error_list_config = {
            'module_name': 'settlement',
            'title': '结算单列表',
            'ajax_mode': True,
            'stats': {
                'cards': [
                    {'id': 'total', 'title': '总结算', 'icon': 'fas fa-list', 'value': 0, 'amount': '0.00', 'unit': '单', 'amount_unit': Config.AMOUNT_UNIT, 'color': 'primary', 'data_key': 'total'},
                    {'id': 'fullySettled', 'title': '已结算', 'icon': 'fas fa-check-circle', 'value': 0, 'amount': '0.00', 'unit': '单', 'amount_unit': Config.AMOUNT_UNIT, 'color': 'success', 'data_key': 'fully_settled'},
                    {'id': 'partiallySettled', 'title': '部分结算', 'icon': 'fas fa-exclamation-triangle', 'value': 0, 'amount': '0.00', 'unit': '单', 'amount_unit': Config.AMOUNT_UNIT, 'color': 'warning', 'data_key': 'partially_settled'},
                    {'id': 'pending', 'title': '待结算', 'icon': 'fas fa-clock', 'value': 0, 'amount': '0.00', 'unit': '单', 'amount_unit': Config.AMOUNT_UNIT, 'color': 'danger', 'data_key': 'pending'}
                ]
            },
            'filter': {
                'action_url': url_for('inventory.settlement_order_list'),
                'form_id': 'settlementOrderFilterForm',
                'reset_url': url_for('inventory.settlement_order_list'),
                'search_field': {'name': 'search', 'label': '搜索', 'placeholder': '结算单编号或项目名称', 'value': '', 'col_width': 4},
                'filter_fields': [],
                'search_button_text': '搜索',
                'reset_button_text': '重置'
            },
            'table': {
                'ajax_target': 'settlementTableBody',
                'title': '结算单列表',
                'icon': 'fas fa-table',
                'show_batch_actions': False,
                'columns': []
            }
        }
        
        return render_template('inventory/settlement_order_list.html',
                             settlement_orders=[],
                             companies=companies,
                             list_config=error_list_config)

@inventory.route('/settlement_process/<order_number>')
@login_required
def settlement_process(order_number):
    """结算处理页面"""
    try:
        settlement_order = SettlementOrder.query.filter_by(order_number=order_number).first_or_404()
        # 只获取 company_type 为 'dealer' 且未删除的公司
        companies = Company.query.filter(
            Company.company_type == 'dealer',
            Company.is_deleted == False
        ).order_by(Company.company_name).all()
        
        # 获取URL参数中的选中公司ID
        selected_company_id = request.args.get('selected_company')
        
        return render_template('inventory/settlement_process.html',
                             settlement_order=settlement_order,
                             companies=companies,
                             selected_company_id=selected_company_id)
                             
    except Exception as e:
        logger.error(f"获取结算处理页面失败：{str(e)}")
        flash(f'获取结算处理页面失败：{str(e)}', 'danger')
        return redirect(url_for('inventory.settlement_order_list'))

@inventory.route('/settlement_orders/export')
@login_required
@permission_required('settlement', 'view')
def export_settlement_orders():
    """导出结算单列表为Excel"""
    try:
        # 获取筛选参数
        search = request.args.get('search', '').strip()
        settlement_company_id = request.args.get('settlement_company', '')
        settlement_status = request.args.get('settlement_status', '')
        
        # 构建查询（复用列表页面的逻辑）
        from app.models.pricing_order import PricingOrder
        from app.models.project import Project
        
        query = SettlementOrder.query.join(
            PricingOrder, SettlementOrder.pricing_order_id == PricingOrder.id
        ).join(
            Project, SettlementOrder.project_id == Project.id
        ).filter(
            PricingOrder.status == 'approved'
        )
        
        # 应用搜索条件
        if search:
            query = query.filter(
                db.or_(
                    SettlementOrder.order_number.ilike(f'%{search}%'),
                    Project.project_name.ilike(f'%{search}%')
                )
            )
        
        # 应用公司筛选
        if settlement_company_id:
            query = query.filter(SettlementOrder.dealer_id == settlement_company_id)
        
        # 应用状态筛选
        if settlement_status:
            query = query.filter(SettlementOrder.settlement_status == settlement_status)
        
        # 执行查询
        settlement_orders = query.order_by(SettlementOrder.created_at.desc()).all()
        
        # 准备筛选条件信息
        filter_info = []
        filter_info.append(['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        filter_info.append(['数据总数', f'{len(settlement_orders)} 条记录'])
        filter_info.append(['', ''])  # 空行
        filter_info.append(['筛选条件', ''])
        
        # 添加具体筛选条件
        if search:
            filter_info.append(['搜索关键词', search])
        
        if settlement_company_id:
            company = Company.query.get(settlement_company_id)
            company_name = company.company_name if company else f'ID:{settlement_company_id}'
            filter_info.append(['结算公司', company_name])
        
        if settlement_status:
            status_map = {
                'pending': '待结算',
                'partially_settled': '部分结算',
                'fully_settled': '已结算'
            }
            filter_info.append(['结算状态', status_map.get(settlement_status, settlement_status)])
        
        # 如果没有任何筛选条件
        if not any([search, settlement_company_id, settlement_status]):
            filter_info.append(['筛选状态', '未应用筛选条件，显示全部数据'])
        
        filter_info.append(['', ''])  # 空行
        
        # 准备导出数据
        export_data = []
        for order in settlement_orders:
            status_map = {
                'fully_settled': '完全结算',
                'partially_settled': '部分结算',
                'pending': '待结算'
            }
            
            export_data.append({
                '结算单编号': order.order_number,
                '关联项目': order.project.project_name if order.project else '无项目',
                '结算公司': order.dealer.company_name if order.dealer else '无公司',
                '产品数量': len(order.details),
                '总金额(万元)': round(float(order.total_amount or 0.0) / 10000, 2),
                '结算情况': status_map.get(order.settlement_status, order.settlement_status),
                '创建时间': order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '-'
            })
        
        # 创建DataFrame
        df = pd.DataFrame(export_data)
        
        # 创建筛选条件DataFrame
        filter_df = pd.DataFrame(filter_info, columns=['项目', '值'])
        
        # 生成文件名
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'结算单统计-{current_time}.xlsx'
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 先写入筛选条件信息
            filter_df.to_excel(writer, sheet_name='结算单统计', index=False, header=False, startrow=0)
            
            # 计算数据开始行（筛选条件行数 + 2行间距）
            data_start_row = len(filter_info) + 2
            
            # 写入数据表格
            df.to_excel(writer, sheet_name='结算单统计', index=False, startrow=data_start_row)
            
            # 获取工作表并设置格式
            worksheet = writer.sheets['结算单统计']
            
            # 设置筛选条件区域的格式
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 标题行格式（导出时间、数据总数等）
            title_font = Font(bold=True, size=12)
            filter_font = Font(bold=True, size=10, color='0066CC')
            
            for row in range(1, len(filter_info) + 1):
                cell_a = worksheet.cell(row=row, column=1)
                cell_b = worksheet.cell(row=row, column=2)
                
                # 设置关键信息的格式
                if row <= 2:  # 导出时间和数据总数
                    cell_a.font = title_font
                    cell_b.font = title_font
                elif cell_a.value == '筛选条件':  # 筛选条件标题
                    cell_a.font = filter_font
                elif cell_a.value and cell_a.value not in ['', '筛选条件']:  # 具体筛选项
                    cell_a.font = Font(bold=True, size=9)
            
            # 设置数据表格标题行格式
            header_row = data_start_row + 1
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            column_widths = {
                'A': 18,  # 结算单编号
                'B': 25,  # 关联项目
                'C': 20,  # 结算公司
                'D': 12,  # 产品数量
                'E': 15,  # 总金额
                'F': 12,  # 结算情况
                'G': 18,  # 创建时间
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置数据区域的边框
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为数据表格添加边框
            for row in range(header_row, header_row + len(df) + 1):
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
        
        output.seek(0)
        
        logger.info(f"用户 {current_user.username} 导出结算单列表，共 {len(settlement_orders)} 条记录")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"导出结算单列表失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        flash(f'导出失败：{str(e)}', 'danger')
        return redirect(url_for('inventory.settlement_order_list'))

@inventory.route('/settlement/create', methods=['GET', 'POST'])
@login_required
@permission_required('settlement', 'create')
def create_settlement():
    """创建结算 - 重定向到结算单列表"""
    return redirect(url_for('inventory.settlement_order_list'))

@inventory.route('/settlement/<int:id>')
@login_required
@permission_required('settlement', 'view')
def settlement_detail(id):
    """结算详情"""
    settlement_order = SettlementOrder.query.get_or_404(id)
    return render_template('inventory/settlement_detail.html', settlement_order=settlement_order)

@inventory.route('/inventory_settlement/<int:id>')
@login_required
@permission_required('settlement', 'view')
def inventory_settlement_detail(id):
    """库存结算详情"""
    settlement = Settlement.query.get_or_404(id)
    return render_template('inventory/inventory_settlement_detail.html', settlement=settlement)

@inventory.route('/settlement/<int:id>/execute', methods=['POST'])
@login_required
@permission_required('settlement', 'create')
def execute_settlement(id):
    """执行结算 - 将结算单与库存进行关联并扣减库存"""
    try:
        settlement_order = SettlementOrder.query.get_or_404(id)
        
        # 检查结算单状态
        if settlement_order.status != 'approved':
            return jsonify({'success': False, 'message': '只有已批准的结算单才能执行结算'})
        
        # 检查是否已经执行过结算
        existing_settlement = Settlement.query.filter_by(
            settlement_number=f"INV-{settlement_order.order_number}"
        ).first()
        if existing_settlement:
            return jsonify({'success': False, 'message': '该结算单已经执行过库存结算'})
        
        # 准备结算项目
        settlement_items = []
        for detail in settlement_order.details:
            # 根据产品名称和MN查找对应的产品
            product = None
            if detail.product_mn:
                product = Product.query.filter_by(product_mn=detail.product_mn).first()
            if not product and detail.product_name:
                # 根据产品名称查找（使用公共辅助函数）
                product = find_product_by_name(detail.product_name)

            if product:
                settlement_items.append({
                    'product_id': product.id,
                    'quantity': detail.quantity,
                    'notes': f'结算单{settlement_order.order_number}执行'
                })
            else:
                logger.warning(f"未找到产品: {detail.product_name} (MN: {detail.product_mn})")
        
        if not settlement_items:
            return jsonify({'success': False, 'message': '未找到可结算的产品库存'})
        
        # 执行库存结算
        success, message, settlement = process_settlement(
            company_id=settlement_order.distributor_id,
            settlement_items=settlement_items,
            description=f'执行结算单 {settlement_order.order_number}',
            user_id=current_user.id
        )
        
        if success:
            # 更新结算单号以关联库存结算
            settlement.settlement_number = f"INV-{settlement_order.order_number}"
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': '结算执行成功，库存已更新',
                'settlement_id': settlement.id
            })
        else:
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        logger.error(f"执行结算失败：{str(e)}")
        return jsonify({'success': False, 'message': f'执行结算失败：{str(e)}'})

@inventory.route('/api/settlement/<int:id>')
@login_required
@permission_required('settlement', 'view')
def get_settlement_info(id):
    """获取结算单详情API"""
    try:
        settlement_order = SettlementOrder.query.get_or_404(id)
        
        return jsonify({
            'success': True,
            'settlement': {
                'order_number': settlement_order.order_number,
                'distributor_name': settlement_order.dealer.company_name if settlement_order.dealer else '无经销商',
                'details_count': len(settlement_order.details),
                'total_amount': settlement_order.formatted_total_amount,
                'status': settlement_order.status
            }
        })
    except Exception as e:
        logger.error(f"获取结算详情失败：{str(e)}")
        return jsonify({'success': False, 'message': f'获取结算详情失败：{str(e)}'})

@inventory.route('/api/company/<int:company_id>/products')
@login_required
# @permission_required('inventory', 'view')  # 临时注释掉权限检查
def get_company_products(company_id):
    """获取公司库存产品API"""
    try:
        # 获取该公司的库存产品
        inventories = Inventory.query.filter_by(company_id=company_id).filter(Inventory.quantity > 0).all()
        
        products = []
        for inventory in inventories:
            products.append({
                'id': inventory.product_id,
                'product_name': inventory.product.name,
                'product_model': inventory.product.product_model,
                'quantity': inventory.quantity,
                'unit': inventory.unit
            })
        
        return jsonify({
            'success': True,
            'products': products
        })
    except Exception as e:
        logger.error(f"获取公司库存产品失败：{str(e)}")
        return jsonify({'success': False, 'message': f'获取库存产品失败：{str(e)}'})

@inventory.route('/api/company/<int:company_id>/inventory_details')
@login_required
def get_company_inventory_details(company_id):
    """获取公司库存详情API - 用于结算处理"""
    try:
        # 获取该公司的所有库存产品
        inventories = Inventory.query.filter_by(company_id=company_id).all()
        
        inventory_dict = {}
        for inventory in inventories:
            inventory_dict[inventory.product.name] = {
                'product_id': inventory.product_id,
                'product_name': inventory.product.name,
                'product_model': inventory.product.model or '',
                'product_mn': inventory.product.product_mn or '',
                'quantity': inventory.quantity,
                'unit': inventory.unit or '件',
                'min_stock': inventory.min_stock,
                'max_stock': inventory.max_stock
            }
        
        return jsonify({
            'success': True,
            'inventory': inventory_dict
        })
    except Exception as e:
        logger.error(f"获取公司库存详情失败：{str(e)}")
        return jsonify({'success': False, 'message': f'获取库存详情失败：{str(e)}'})

@inventory.route('/api/settlement_orders/filter', methods=['GET'])
@login_required
@permission_required('settlement', 'view')
def settlement_order_list_ajax():
    """结算单列表AJAX筛选API"""
    try:
        # 确保数据库连接正常，如果有失败的事务则回滚
        try:
            db.session.rollback()
        except:
            pass
    except:
        pass
    
    # 获取搜索和筛选参数
    search = request.args.get('search', '')
    settlement_company = request.args.get('settlement_company', '')
    settlement_status = request.args.get('settlement_status', '')
    
    # 分页参数
    offset = request.args.get('offset', 0, type=int)
    limit = request.args.get('limit', 20, type=int)
    
    # 排序参数
    sort_field = request.args.get('sort_field', '')
    sort_direction = request.args.get('sort_direction', 'asc')
    
    # 限制每次加载数量的范围
    if limit not in [10, 20, 30, 50]:
        limit = 20
    
    # 构建查询：只获取来自已审批批价单的结算单
    from app.models.pricing_order import PricingOrder
    from app.models.project import Project
    
    query = SettlementOrder.query.join(
        PricingOrder, SettlementOrder.pricing_order_id == PricingOrder.id
    ).join(
        Project, SettlementOrder.project_id == Project.id
    ).filter(
        PricingOrder.status == 'approved'
    )
    
    # 搜索条件
    if search:
        query = query.filter(
            db.or_(
                SettlementOrder.order_number.ilike(f'%{search}%'),
                Project.project_name.ilike(f'%{search}%')
            )
        )
    
    # 筛选条件
    if settlement_company:
        try:
            company_id = int(settlement_company)
            query = query.filter(SettlementOrder.dealer_id == company_id)
        except (ValueError, TypeError):
            pass
    
    if settlement_status:
        query = query.filter(SettlementOrder.settlement_status == settlement_status)
    
    # 应用排序
    from app.utils.sorting_service import SortingService, create_basic_field_mappings
    
    # 创建排序配置
    sorting_config = {
        'field_mappings': create_basic_field_mappings(SettlementOrder, [
            'order_number', 'total_amount', 'settlement_status', 'created_at', 'updated_at'
        ]),
        'relation_mappings': {},
        'default_sort': {'field': 'created_at', 'direction': 'desc'}
    }
    
    # 创建排序服务并应用排序
    sorting_service = SortingService(SettlementOrder, sorting_config)
    query = sorting_service.apply_sort(query, sort_field, sort_direction)
    
    # 执行查询
    total_count = query.count()
    settlement_orders = query.offset(offset).limit(limit).all()
    has_more = (offset + limit) < total_count
    
    # 计算统计数据（用于更新统计卡片）
    all_orders_for_stats = query.all()
    
    # 分类统计
    total_stats_count = len(all_orders_for_stats)
    total_stats_amount = sum(float(order.total_amount or 0) for order in all_orders_for_stats) / 10000  # 转换为万元
    
    fully_settled_orders = [o for o in all_orders_for_stats if o.settlement_status == 'fully_settled']
    fully_settled_stats_count = len(fully_settled_orders)
    fully_settled_stats_amount = sum(float(order.total_amount or 0) for order in fully_settled_orders) / 10000
    
    partially_settled_orders = [o for o in all_orders_for_stats if o.settlement_status == 'partially_settled']
    partially_settled_stats_count = len(partially_settled_orders)
    partially_settled_stats_amount = sum(float(order.total_amount or 0) for order in partially_settled_orders) / 10000
    
    pending_orders = [o for o in all_orders_for_stats if o.settlement_status == 'pending']
    pending_stats_count = len(pending_orders)
    pending_stats_amount = sum(float(order.total_amount or 0) for order in pending_orders) / 10000
    
    # 构建统计数据
    statistics = {
        'total_count': total_stats_count,
        'total_amount': total_stats_amount,
        'fully_settled_count': fully_settled_stats_count,
        'fully_settled_amount': fully_settled_stats_amount,
        'partially_settled_count': partially_settled_stats_count,
        'partially_settled_amount': partially_settled_stats_amount,
        'pending_count': pending_stats_count,
        'pending_amount': pending_stats_amount
    }
    
    # 检测移动端并使用智能移动卡片
    from app.utils.mobile_helpers import is_mobile_request
    from types import SimpleNamespace
    
    if settlement_orders:
        # 格式化结算单数据为标准结构
        formatted_results = []
        for order in settlement_orders:
            project_name = order.project.project_name if order.project else '未知项目'
            dealer_name = order.dealer.company_name if order.dealer else '未知经销商'
            
            # 获取结算状态标签
            status_labels = {
                'pending': '待结算',
                'partially_settled': '部分结算',
                'fully_settled': '已完成结算'
            }
            status_label = status_labels.get(order.settlement_status, order.settlement_status)
            
            formatted_order = SimpleNamespace(
                id=order.id,
                order_number=order.order_number,
                project_name=project_name,
                dealer_name=dealer_name,
                total_amount=order.total_amount or 0,
                settlement_status=order.settlement_status,
                settlement_status_label=status_label,
                created_at=order.created_at,
                updated_at=order.updated_at
            )
            formatted_results.append(formatted_order)
        
        if is_mobile_request():
            # 智能移动卡片配置 - 结算单管理
            smart_mobile_card = {
                'module': 'settlement_order',
                'title_field': {'field': 'order_number'},
                'link_url': '/inventory/settlement/{id}',
                'badges': [
                    {'field': 'settlement_status_label', 'renderer': 'settlement_status'}
                ],
                'details': [
                    {'field': 'project_name', 'label': '关联项目'},
                    {'field': 'dealer_name', 'label': '经销商'},
                    {'field': 'total_amount', 'label': '结算金额', 'format': 'currency'},
                    {'field': 'created_at', 'label': '创建时间', 'format': 'date'},
                    {'field': 'updated_at', 'label': '更新时间', 'format': 'date'}
                ]
            }
            
            # 使用智能移动卡片模板渲染
            html = render_template_string('''
            {% from 'macros/ui_helpers.html' import render_smart_mobile_cards %}
            {{ render_smart_mobile_cards(items, card_config) }}
            ''', items=formatted_results, card_config=smart_mobile_card)
        else:
            # 桌面端使用传统表格行渲染
            html = render_template('inventory/settlement_order_rows.html', settlement_orders=settlement_orders)
    else:
        if is_mobile_request():
            html = '<div class="text-center py-4">暂无符合条件的数据</div>'
        else:
            html = render_template('inventory/settlement_order_rows.html', settlement_orders=settlement_orders)
    
    return jsonify({
        'success': True,
        'html': html,
        'has_more': has_more,
        'total_count': total_count,
        'loaded_count': len(settlement_orders),
        'statistics': statistics  # 用于更新统计卡片
    })

@inventory.route('/api/settlement/filter', methods=['GET'])
@login_required
def settlement_list_ajax():
    """结算明细列表AJAX筛选API"""
    # 获取查询参数
    search = request.args.get('search', '').strip()
    company_filter = request.args.get('company_filter')
    status_filter = request.args.get('status_filter')
    settlement_company_filter = request.args.get('settlement_company_filter')
    
    # 分页参数
    offset = request.args.get('offset', 0, type=int)
    limit = request.args.get('limit', 20, type=int)
    
    # 限制每次加载数量的范围
    if limit not in [10, 20, 30, 50]:
        limit = 20
    
    # 构建基础查询 - 只获取已审批批价单的结算单明细
    from app.models.pricing_order import SettlementOrderDetail, SettlementOrder, PricingOrder
    query = db.session.query(SettlementOrderDetail).join(SettlementOrder).join(PricingOrder)
    
    # 关键过滤：只显示已审批批价单的结算明细
    query = query.filter(PricingOrder.status == 'approved')
    
    # 搜索条件
    if search:
        search_filter = db.or_(
            SettlementOrder.order_number.contains(search),
            SettlementOrderDetail.product_name.contains(search),
            SettlementOrderDetail.product_mn.contains(search)
        )
        query = query.filter(search_filter)
    
    # 结算单公司过滤（分销商）
    if company_filter:
        query = query.filter(SettlementOrder.distributor_id == company_filter)
    
    # 结算目标公司过滤
    if settlement_company_filter:
        query = query.filter(SettlementOrderDetail.settlement_company_id == settlement_company_filter)
    
    # 结算状态过滤
    if status_filter:
        if status_filter == 'completed':
            query = query.filter(SettlementOrderDetail.settlement_status == 'settled')
        elif status_filter == 'pending':
            query = query.filter(SettlementOrderDetail.settlement_status == 'pending')
    
    # 排序
    query = query.order_by(SettlementOrder.created_at.desc(), SettlementOrderDetail.id.desc())
    
    # 执行查询（按创建时间倒序）
    total_count = query.count()
    settlement_details_raw = query.offset(offset).limit(limit).all()
    has_more = (offset + limit) < total_count
    
    # 处理结算明细数据用于显示
    settlement_details = []
    
    for detail in settlement_details_raw:
        # 获取产品信息（按MN号精确匹配）
        product = None
        if detail.product_mn:
            product = Product.query.filter_by(product_mn=detail.product_mn).first()
        
        # 构建明细信息
        detail_info = {
            'id': detail.id,
            'settlement_order': detail.settlement_order,
            'product': product,
            'product_name': detail.product_name,
            'product_model': detail.product_model,
            'product_mn': detail.product_mn,
            'brand': detail.brand,
            'quantity': detail.quantity,
            'unit': detail.unit,
            'unit_price': detail.unit_price,
            'total_price': detail.total_price,
            'settlement_status': detail.settlement_status,
            'settlement_company': detail.settlement_company,
            'settlement_date': detail.settlement_date,
            'settlement_notes': detail.settlement_notes,
            'is_settled': detail.settlement_status == 'settled'  # 保持兼容性
        }
        
        settlement_details.append(detail_info)
    
    # 计算统计数据（用于更新统计卡片）
    all_details_for_stats = query.all()
    
    # 分类统计
    total_stats_count = len(all_details_for_stats)
    total_stats_amount = sum(float(detail.total_price or 0) for detail in all_details_for_stats) / 10000  # 转换为万元
    
    settled_details = [d for d in all_details_for_stats if d.settlement_status == 'settled']
    settled_stats_count = len(settled_details)
    settled_stats_amount = sum(float(detail.total_price or 0) for detail in settled_details) / 10000
    
    pending_details = [d for d in all_details_for_stats if d.settlement_status == 'pending']
    pending_stats_count = len(pending_details)
    pending_stats_amount = sum(float(detail.total_price or 0) for detail in pending_details) / 10000
    
    # 本月结算统计
    from datetime import datetime
    current_month = datetime.now().strftime('%Y-%m')
    thismonth_details = [d for d in settled_details if d.settlement_date and d.settlement_date.strftime('%Y-%m') == current_month]
    thismonth_stats_count = len(thismonth_details)
    thismonth_stats_amount = sum(float(detail.total_price or 0) for detail in thismonth_details) / 10000
    
    # 构建统计数据
    statistics = {
        'total_count': total_stats_count,
        'total_amount': total_stats_amount,
        'settled_count': settled_stats_count,
        'settled_amount': settled_stats_amount,
        'pending_count': pending_stats_count,
        'pending_amount': pending_stats_amount,
        'thismonth_count': thismonth_stats_count,
        'thismonth_amount': thismonth_stats_amount
    }
    
    # 渲染HTML片段
    html = render_template('inventory/settlement_rows.html', 
                          settlement_details=settlement_details)
    
    return jsonify({
        'success': True,
        'html': html,
        'has_more': has_more,
        'total_count': total_count,
        'loaded_count': len(settlement_details),
        'statistics': statistics  # 用于更新统计卡片
    })

@inventory.route('/api/company/<int:company_id>/product/<int:product_id>/stock')
@login_required
def get_product_stock(company_id, product_id):
    """获取指定公司的指定产品库存"""
    try:
        inventory = Inventory.query.filter_by(
            company_id=company_id,
            product_id=product_id
        ).first()
        
        stock = inventory.quantity if inventory else 0
        
        return jsonify({
            'success': True,
            'stock': stock
        })
        
    except Exception as e:
        logger.error(f"获取产品库存失败：{str(e)}")
        return jsonify({'success': False, 'message': f'获取库存失败：{str(e)}'})

@inventory.route('/api/company_inventory/<int:company_id>')
@login_required
@permission_required('inventory', 'view')
def get_company_inventory(company_id):
    """获取公司库存"""
    inventories = Inventory.query.filter_by(company_id=company_id).join(Product).all()
    
    result = []
    for inv in inventories:
        status = get_inventory_status(company_id, inv.product_id)
        result.append({
            'product_id': inv.product_id,
            'product_name': inv.product.name,
            'product_mn': inv.product.product_mn,
            'quantity': inv.quantity,
            'unit': inv.unit,
            'status': status['status'],
            'warning': status['warning']
        })
    
    return jsonify(result)

@inventory.route('/api/product_info/<int:product_id>')
@login_required
@permission_required('inventory', 'view')
def get_product_info(product_id):
    """获取产品信息"""
    try:
        product = Product.query.get_or_404(product_id)
        return jsonify({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'model': product.product_model,
                'desc': product.product_desc,
                'unit': product.unit,
                'mn': product.product_mn
            }
        })
    except Exception as e:
        logger.error(f"获取产品信息失败：{str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@inventory.route('/api/settle_product', methods=['POST'])
@login_required
@permission_required('settlement', 'create')
def settle_product():
    """将结算单明细中的产品结算到指定公司的库存"""
    try:
        data = request.get_json()
        detail_id = data.get('detail_id')
        company_id = data.get('company_id')
        notes = data.get('notes', '')
        
        if not detail_id or not company_id:
            return jsonify({'success': False, 'message': '缺少必要参数'})
        
        # 获取结算单明细
        from app.models.pricing_order import SettlementOrderDetail
        detail = SettlementOrderDetail.query.get_or_404(detail_id)
        
        # 检查是否已经结算过
        order_number = detail.settlement_order.order_number
        existing_settlement = Settlement.query.filter_by(
            settlement_number=f"INV-{order_number}"
        ).first()
        
        if existing_settlement:
            # 检查该产品是否已经在结算记录中
            existing_detail = SettlementDetail.query.filter_by(
                settlement_id=existing_settlement.id
            ).join(Product).filter(
                Product.product_name == detail.product_name
            ).first()
            
            if existing_detail:
                return jsonify({'success': False, 'message': '该产品已经结算过了'})
        
        # 获取或创建产品
        product = None
        # SettlementOrderDetail没有product_id字段，需要根据产品名称和MN查找
        if detail.product_mn:
            product = Product.query.filter_by(product_mn=detail.product_mn).first()
        if not product and detail.product_name:
            # 根据产品名称查找（使用公共辅助函数）
            product = find_product_by_name(detail.product_name)

        if not product:
            return jsonify({'success': False, 'message': f'未找到产品: {detail.product_name}'})
        
        # 检查或创建库存记录
        inventory = Inventory.query.filter_by(
            company_id=company_id,
            product_id=product.id
        ).first()
        
        if not inventory:
            # 创建新的库存记录
            inventory = Inventory(
                company_id=company_id,
                product_id=product.id,
                quantity=0,
                unit=detail.unit,
                created_by_id=current_user.id
            )
            db.session.add(inventory)
            db.session.flush()  # 获取ID
        
        # 记录变动前的库存
        quantity_before = inventory.quantity
        
        # 检查库存是否充足
        if inventory.quantity < detail.quantity:
            return jsonify({'success': False, 'message': f'库存不足：当前库存 {inventory.quantity}，需要结算 {detail.quantity}'})
        
        # 扣减库存数量（结算是出库操作）
        inventory.quantity -= detail.quantity
        inventory.updated_at = datetime.now()
        
        # 创建或更新结算记录
        if not existing_settlement:
            # 创建新的结算记录
            settlement = Settlement(
                settlement_number=f"INV-{order_number}",
                company_id=company_id,
                settlement_date=datetime.now(),
                status='completed',
                total_items=1,
                description=f'结算单 {order_number} 产品结算',
                created_by_id=current_user.id,
                approved_by_id=current_user.id,
                approved_at=datetime.now()
            )
            db.session.add(settlement)
            db.session.flush()
        else:
            settlement = existing_settlement
            settlement.total_items += 1
            settlement.updated_at = datetime.now()
        
        # 创建结算明细记录
        settlement_detail = SettlementDetail(
            settlement_id=settlement.id,
            inventory_id=inventory.id,
            product_id=product.id,
            quantity_settled=detail.quantity,
            quantity_before=quantity_before,
            quantity_after=inventory.quantity,
            unit=detail.unit,
            notes=notes or f'结算单{order_number}产品结算'
        )
        db.session.add(settlement_detail)
        
        # 创建库存变动记录
        transaction = InventoryTransaction(
            inventory_id=inventory.id,
            transaction_type='out',  # 改为出库
            quantity=detail.quantity,
            quantity_before=quantity_before,
            quantity_after=inventory.quantity,
            reference_type='settlement',
            reference_id=settlement.id,
            description=f'结算出库: {detail.product_name}',
            created_by_id=current_user.id
        )
        db.session.add(transaction)
        
        # 提交事务
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '产品结算成功',
            'settlement_id': settlement.id,
            'inventory_id': inventory.id,
            'new_quantity': inventory.quantity
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"产品结算失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'结算失败: {str(e)}'})

@inventory.route('/api/settlement_order/<int:settlement_order_id>')
@login_required
@permission_required('settlement', 'view')
def get_settlement_order_detail(settlement_order_id):
    """获取结算单详情（用于模态框显示）"""
    try:
        settlement_order = SettlementOrder.query.get_or_404(settlement_order_id)
        
        # 获取结算状态信息
        settled_products = {}
        settlement_companies = {}
        
        existing_settlement = Settlement.query.filter_by(
            settlement_number=f"INV-{settlement_order.order_number}"
        ).first()
        
        if existing_settlement:
            for detail in existing_settlement.details:
                if detail.product and detail.product.product_name:
                    key = detail.product.product_name
                    settled_products[key] = existing_settlement.settlement_date.strftime('%Y-%m-%d %H:%M') if existing_settlement.settlement_date else ''
                    settlement_companies[key] = existing_settlement.company.company_name if existing_settlement.company else ''
        
        # 构建HTML内容
        html_content = f"""
        <div class="row mb-4">
            <div class="col-md-6">
                <h6>结算单信息</h6>
                <table class="table table-sm">
                    <tr><td><strong>结算单号：</strong></td><td>{settlement_order.order_number}</td></tr>
                    <tr><td><strong>经销商：</strong></td><td>{settlement_order.dealer.company_name if settlement_order.dealer else '无经销商'}</td></tr>
                    <tr><td><strong>关联项目：</strong></td><td>{settlement_order.project.project_name if settlement_order.project else '无关联项目'}</td></tr>
                    <tr><td><strong>状态：</strong></td><td>
                        <span class="badge {'bg-success' if settlement_order.status == 'approved' else 'bg-warning' if settlement_order.status == 'pending' else 'bg-secondary'}">
                            {'已批准' if settlement_order.status == 'approved' else '审批中' if settlement_order.status == 'pending' else '草稿'}
                        </span>
                    </td></tr>
                    <tr><td><strong>创建时间：</strong></td><td>{settlement_order.created_at.strftime('%Y-%m-%d %H:%M') if settlement_order.created_at else '-'}</td></tr>
                </table>
            </div>
            <div class="col-md-6">
                <h6>金额信息</h6>
                <table class="table table-sm">
                    <tr><td><strong>结算总金额：</strong></td><td class="text-success">{Config.CURRENCY_SYMBOL}{settlement_order.formatted_total_amount}</td></tr>
                    <tr><td><strong>产品项数：</strong></td><td>{len(settlement_order.details)}</td></tr>
                    <tr><td><strong>总折扣率：</strong></td><td>{settlement_order.discount_percentage}%</td></tr>
                </table>
            </div>
        </div>
        
        <h6>产品明细</h6>
        <div class="table-responsive">
            <table class="table table-hover">
                <thead class="table-light">
                    <tr>
                        <th>产品名称</th>
                        <th>型号</th>
                        <th>品牌</th>
                        <th>数量</th>
                        <th>单价</th>
                        <th>小计</th>
                        <th>结算状态</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for detail in settlement_order.details:
            is_settled = detail.product_name in settled_products
            settlement_date = settled_products.get(detail.product_name, '')
            settlement_company = settlement_companies.get(detail.product_name, '')
            
            status_html = ''
            if is_settled:
                status_html = f'''
                    <span class="badge bg-success">已结算</span><br>
                    <small class="text-muted">{settlement_company}</small><br>
                    <small class="text-muted">{settlement_date}</small>
                '''
            else:
                status_html = '<span class="badge bg-warning">待结算</span>'
            
            html_content += f"""
                <tr>
                    <td><strong>{detail.product_name}</strong></td>
                    <td>{detail.product_model or '-'}</td>
                    <td>{detail.brand or '-'}</td>
                    <td>{detail.quantity} {detail.unit or '件'}</td>
                    <td>{Config.CURRENCY_SYMBOL}{detail.unit_price:.2f}</td>
                    <td class="text-success">{Config.CURRENCY_SYMBOL}{detail.total_price:.2f}</td>
                    <td>{status_html}</td>
                </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </div>
        """
        
        return jsonify({
            'success': True,
            'html': html_content
        })
        
    except Exception as e:
        logger.error(f"获取结算单详情失败：{str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@inventory.route('/settlement_detail/<order_number>')
@login_required
def settlement_detail_api(order_number):
    """获取结算单详情API"""
    try:
        from app.models.pricing_order import SettlementOrder, SettlementOrderDetail
        
        # 获取结算单
        settlement_order = SettlementOrder.query.filter_by(order_number=order_number).first()
        if not settlement_order:
            return jsonify({'success': False, 'message': '结算单不存在'})
        
        # 获取结算单明细
        details = SettlementOrderDetail.query.filter_by(settlement_order_id=settlement_order.id).all()
        
        # 构建HTML内容
        html_content = f"""
        <div class="row mb-3">
            <div class="col-md-6">
                <h6>结算单信息</h6>
                <table class="table table-sm">
                    <tr><td>结算单号:</td><td>{settlement_order.order_number}</td></tr>
                    <tr><td>项目名称:</td><td>{settlement_order.project.project_name if settlement_order.project else '无项目'}</td></tr>
                    <tr><td>结算公司:</td><td>{settlement_order.dealer.company_name if settlement_order.dealer else '无公司'}</td></tr>
                    <tr><td>创建时间:</td><td>{settlement_order.created_at.strftime('%Y-%m-%d %H:%M') if settlement_order.created_at else '-'}</td></tr>
                </table>
            </div>
            <div class="col-md-6">
                <h6>统计信息</h6>
                <table class="table table-sm">
                    <tr><td>产品数量:</td><td>{len(details)} 项</td></tr>
                    <tr><td>总金额:</td><td>{Config.CURRENCY_SYMBOL}{sum(d.total_price or 0 for d in details):,.2f}</td></tr>
                </table>
            </div>
        </div>
        
        <h6>产品明细</h6>
        <div class="table-responsive">
            <table class="table table-sm table-striped">
                <thead>
                    <tr>
                        <th>产品名称</th>
                        <th>型号</th>
                        <th>品牌</th>
                        <th>数量</th>
                        <th>单价</th>
                        <th>总价</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for detail in details:
            html_content += f"""
                    <tr>
                        <td>{detail.product_name or '-'}</td>
                        <td>{detail.product_model or '-'}</td>
                        <td>{detail.brand or '-'}</td>
                        <td>{detail.quantity}</td>
                        <td>{Config.CURRENCY_SYMBOL}{detail.unit_price or 0:,.2f}</td>
                        <td>{Config.CURRENCY_SYMBOL}{detail.total_price or 0:,.2f}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </div>
        """
        
        return jsonify({'success': True, 'html': html_content})
        
    except Exception as e:
        logger.error(f"获取结算单详情失败：{str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@inventory.route('/settle_product/<int:detail_id>', methods=['POST'])
@login_required
def settle_single_product(detail_id):
    """结算单个产品到库存"""
    try:
        from app.models.pricing_order import SettlementOrderDetail
        
        # 获取结算明细
        detail = SettlementOrderDetail.query.get_or_404(detail_id)
        
        # 检查是否已经结算
        settlement_key = f"{detail.settlement_order.order_number}_{detail.product_name}"
        existing_settlement = Settlement.query.filter(
            Settlement.settlement_number == f"INV-{detail.settlement_order.order_number}"
        ).first()
        
        if existing_settlement:
            # 检查该产品是否已在结算中
            for settlement_detail in existing_settlement.details:
                if (settlement_detail.product and 
                    settlement_detail.product.product_name == detail.product_name):
                    return jsonify({'success': False, 'message': '该产品已经结算过了'})
        
        # 获取或创建产品
        product = None
        # SettlementOrderDetail没有product_id字段，需要根据产品名称和MN查找
        if detail.product_mn:
            product = Product.query.filter_by(product_mn=detail.product_mn).first()
        if not product and detail.product_name:
            # 根据产品名称查找（使用公共辅助函数）
            product = find_product_by_name(detail.product_name)

        if not product:
            return jsonify({'success': False, 'message': '找不到对应的产品信息'})
        
        # 获取结算公司
        settlement_company = detail.settlement_order.dealer
        if not settlement_company:
            return jsonify({'success': False, 'message': '结算单没有指定结算公司'})
        
        # 创建或更新结算记录
        if not existing_settlement:
            existing_settlement = Settlement(
                settlement_number=f"INV-{detail.settlement_order.order_number}",
                company_id=settlement_company.id,
                settlement_date=datetime.now(),
                status='completed',
                notes=f"从结算单 {detail.settlement_order.order_number} 结算"
            )
            db.session.add(existing_settlement)
            db.session.flush()  # 获取ID
        
        # 创建结算明细
        settlement_detail = SettlementDetail(
            settlement_id=existing_settlement.id,
            product_id=product.id,
            quantity=detail.quantity,
            unit_price=detail.unit_price or 0,
            total_price=detail.total_price or 0,
            notes=f"从结算单明细 {detail_id} 结算"
        )
        db.session.add(settlement_detail)
        
        # 更新库存 - 结算应该是出库操作，减少库存
        inventory = Inventory.query.filter_by(
            product_id=product.id,
            company_id=settlement_company.id
        ).first()
        
        if not inventory:
            return jsonify({'success': False, 'message': f'公司 {settlement_company.company_name} 没有产品 {product.product_name} 的库存记录'})
        
        # 检查库存是否充足
        if inventory.quantity < detail.quantity:
            return jsonify({'success': False, 'message': f'库存不足：当前库存 {inventory.quantity}，需要结算 {detail.quantity}'})
        
        # 记录变动前后数量
        quantity_before = inventory.quantity
        quantity_after = inventory.quantity - detail.quantity
        
        # 扣减库存
        inventory.quantity -= detail.quantity
        inventory.updated_at = datetime.now()
        
        # 更新结算明细以包含库存变动信息
        settlement_detail.quantity_settled = detail.quantity
        settlement_detail.quantity_before = quantity_before
        settlement_detail.quantity_after = quantity_after
        settlement_detail.inventory_id = inventory.id
        
        # 创建库存变动记录
        transaction = InventoryTransaction(
            inventory_id=inventory.id,
            transaction_type='out',
            quantity=detail.quantity,
            quantity_before=quantity_before,
            quantity_after=quantity_after,
            unit_price=detail.unit_price or 0,
            total_price=detail.total_price or 0,
            transaction_date=datetime.now(),
            reference_type='settlement',
            reference_id=existing_settlement.id,
            description=f"结算出库：{detail.settlement_order.order_number}",
            created_by_id=current_user.id
        )
        db.session.add(transaction)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '产品结算成功'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"结算产品失败：{str(e)}")
        import traceback
        logger.error(f"错误详情: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'结算失败：{str(e)}'})

@inventory.route('/api/settle_product_to_company', methods=['POST'])
@login_required
@permission_required('settlement', 'create')
def settle_product_to_company():
    """将结算单明细中的产品结算到指定公司（新版本，支持MN号精确匹配和记录结算目标公司）"""
    try:
        data = request.get_json()
        detail_id = data.get('detail_id')
        company_id = data.get('company_id')
        notes = data.get('notes', '')
        
        if not detail_id or not company_id:
            return jsonify({'success': False, 'message': '缺少必要参数'})
        
        # 获取结算单明细
        from app.models.pricing_order import SettlementOrderDetail
        detail = SettlementOrderDetail.query.get_or_404(detail_id)
        
        # 检查是否已经结算过
        if detail.settlement_status == 'settled':
            return jsonify({'success': False, 'message': '该产品已经结算过了'})
        
        # 检查产品MN号
        if not detail.product_mn:
            return jsonify({'success': False, 'message': '该产品没有MN号，无法进行精确匹配结算'})
        
        # 获取或创建产品（按MN号精确匹配）
        product = Product.query.filter_by(product_mn=detail.product_mn).first()
        if not product:
            return jsonify({'success': False, 'message': f'未找到MN号为 {detail.product_mn} 的产品'})
        
        # 获取结算目标公司
        settlement_company = Company.query.get_or_404(company_id)
        
        # 检查或创建库存记录
        inventory = Inventory.query.filter_by(
            company_id=company_id,
            product_id=product.id
        ).first()
        
        if not inventory:
            return jsonify({'success': False, 'message': f'公司 {settlement_company.company_name} 没有产品 {product.product_name} (MN: {detail.product_mn}) 的库存记录'})
        
        # 检查库存数量，支持部分结算
        if inventory.quantity == 0:
            return jsonify({'success': False, 'message': f'库存为0，无法结算'})
        
        # 计算实际可结算数量
        actual_settle_quantity = min(inventory.quantity, detail.quantity)
        is_partial = actual_settle_quantity < detail.quantity
        
        # 记录变动前的库存
        quantity_before = inventory.quantity
        quantity_after = inventory.quantity - actual_settle_quantity
        
        # 扣减库存数量（结算是出库操作）
        inventory.quantity -= actual_settle_quantity
        inventory.updated_at = datetime.now()
        
        # 更新结算明细状态
        detail.settlement_company_id = company_id
        detail.settlement_status = 'settled'
        detail.settlement_date = datetime.now()
        detail.settlement_notes = notes or f'结算到 {settlement_company.company_name}'
        
        # 如果是部分结算，需要创建新的明细记录保留未结算部分
        if is_partial:
            remaining_quantity = detail.quantity - actual_settle_quantity
            detail.quantity = actual_settle_quantity  # 当前明细改为已结算数量
            
            # 创建新的明细记录保留未结算部分
            new_detail = SettlementOrderDetail(
                settlement_order_id=detail.settlement_order_id,
                product_name=detail.product_name,
                product_model=detail.product_model,
                product_desc=detail.product_desc,
                brand=detail.brand,
                product_mn=detail.product_mn,
                quantity=remaining_quantity,
                unit=detail.unit,
                unit_price=detail.unit_price,
                total_price=detail.unit_price * remaining_quantity if detail.unit_price else 0,
                settlement_status='pending'
            )
            db.session.add(new_detail)
        
        # 创建库存变动记录
        transaction = InventoryTransaction(
            inventory_id=inventory.id,
            transaction_type='out',
            quantity=actual_settle_quantity,
            quantity_before=quantity_before,
            quantity_after=quantity_after,
            description=f'结算出库 - {detail.settlement_order.order_number}',
            reference_type='settlement',
            reference_id=detail.id,
            created_by_id=current_user.id
        )
        db.session.add(transaction)
        
        # 更新结算单的settlement_status字段
        settlement_order = detail.settlement_order
        settlement_order.update_settlement_status()
        
        db.session.commit()
        
        message = f'产品结算成功，已从 {settlement_company.company_name} 扣减库存 {actual_settle_quantity} 件'
        if is_partial:
            remaining_quantity = detail.quantity - actual_settle_quantity
            message += f'，剩余 {remaining_quantity} 件未结算'
            
        return jsonify({
            'success': True, 
            'message': message,
            'settlement_info': {
                'company_name': settlement_company.company_name,
                'quantity_before': quantity_before,
                'quantity_after': quantity_after,
                'settled_quantity': actual_settle_quantity,
                'settlement_date': detail.settlement_date.strftime('%Y-%m-%d %H:%M:%S'),
                'is_partial': is_partial,
                'quantity_remaining': detail.quantity - actual_settle_quantity if is_partial else 0
            }
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"结算产品到指定公司失败：{str(e)}")
        return jsonify({'success': False, 'message': f'结算失败：{str(e)}'})

def _resolve_default_company(user):
    """决定当前用户默认看哪家公司的库存。

    新规则(2026-05-24):
    - 任何用户都优先看自己 linked_company_id 对应的公司(锁定,不可切换)
    - admin 默认看平台所属公司(vendor),并可在切换器里选其他公司 / 全局视图
    - 既不是 admin、又没 linked_company_id → 返回 None(无可看库存,显示空态)

    返回 (company_id | None, locked: bool)
    """
    if user.linked_company_id:
        return user.linked_company_id, True  # 所有用户(含 admin 若有绑定)默认锁在自己公司

    # admin 无绑定 → 默认指向第一家有库存的客户公司(厂商自营仓库是独立 scope,见
    # at_stock_list 的 scope='vendor';这里只处理 customer 仓库的 fallback)
    if getattr(user, 'role', None) == 'admin':
        fallback = db.session.query(Company.id).join(
            Inventory, Inventory.company_id == Company.id
        ).filter(
            Company.is_deleted == False,
            Inventory.is_vendor_warehouse == False
        ).order_by(Company.company_name).first()
        return (fallback[0] if fallback else None), False

    # 非 admin 且无 linked_company_id → 无可看库存
    return None, False


def _user_can_view_company(user, company_id):
    """权限校验:用户能否查看指定公司的库存。
    - admin:所有公司
    - 其他用户:仅自己 linked_company_id 对应的公司
    """
    if getattr(user, 'role', None) == 'admin':
        return True
    return user.linked_company_id == company_id


def _switchable_companies():
    """返回当前所有有库存的客户公司(供 switcher 下拉用)。

    厂商自营仓库是独立 scope(scope='vendor'),由模板单独渲染,不在此列表。
    """
    rows = db.session.query(
        Company.id, Company.company_name, Company.company_type,
        func.count(Inventory.id).label('inv_count'),
        func.sum(Inventory.quantity).label('total_qty')
    ).join(Inventory, Inventory.company_id == Company.id).filter(
        Company.is_deleted == False,
        Inventory.is_vendor_warehouse == False
    ).group_by(Company.id, Company.company_name, Company.company_type).order_by(
        Company.company_name
    ).all()
    out = [{
        'id': r.id,
        'name': r.company_name,
        'type': r.company_type or 'user',
        'inv_count': r.inv_count or 0,
        'total_qty': int(r.total_qty or 0),
    } for r in rows]

    # 厂商自营仓库聚合作为独立条目(id='vendor' 字符串区分)
    vendor_agg = db.session.query(
        func.count(Inventory.id).label('inv_count'),
        func.sum(Inventory.quantity).label('total_qty')
    ).filter(Inventory.is_vendor_warehouse == True).first()
    if vendor_agg and (vendor_agg.inv_count or 0) > 0:
        from app.utils.inventory_helpers import get_vendor_warehouse_label
        out.insert(0, {
            'id': 'vendor',
            'name': get_vendor_warehouse_label(),
            'type': 'vendor',
            'inv_count': vendor_agg.inv_count or 0,
            'total_qty': int(vendor_agg.total_qty or 0),
        })
    return out


@inventory.route('/')
@inventory.route('/tw_stock')
@inventory.route('/tw_stock/<int:_legacy_company_id>')
@inventory.route('/stock')
@inventory.route('/index')
@login_required
def _legacy_inventory_redirect(_legacy_company_id=None):
    """老 URL(浏览器缓存 / 收藏夹遗留)兜底跳转到 AT 库存页。"""
    if _legacy_company_id is not None:
        return redirect(url_for('inventory.at_stock_list', scope=str(_legacy_company_id)), code=301)
    return redirect(url_for('inventory.at_stock_list'), code=301)


@inventory.route('/at_stock')
@inventory.route('/at_stock/<scope>')
@login_required
@permission_required('inventory', 'view')
def at_stock_list(scope=None):
    """AT 风格库存管理 — 支持单公司视图 + 全局聚合视图 + 厂商自营仓视图。

    scope:
      None / 缺省 → 默认公司
      'global'    → 全局聚合(仅厂商管理员)
      'vendor'    → 厂商自营仓库(系统级,不在 companies 表)
      数字字符串  → 指定客户公司 id
    """
    if scope == 'global':
        if not _is_vendor_admin(current_user):
            flash('您没有权限查看全局库存', 'danger')
            return redirect(url_for('inventory.at_stock_list'))
        ctx = _build_global_stock_context()
    elif scope == 'vendor':
        ctx = _build_vendor_stock_context()
    else:
        try:
            company_id = int(scope) if scope else None
        except (TypeError, ValueError):
            company_id = None
        ctx = _build_stock_list_context(company_id)
        if ctx is None:
            return redirect(url_for('inventory.at_stock_list'))
    return render_template('inventory/at_stock_list.html', **ctx)


def _build_vendor_stock_context():
    """厂商自营仓库视图(is_vendor_warehouse=true 的库存,company_id=NULL)。

    厂商在 dictionaries 表里(type='company', is_vendor=true),不在 companies 表,
    所以 current_company 构造为 SimpleNamespace,id='vendor',name 取字典。
    """
    from types import SimpleNamespace
    from app.utils.inventory_helpers import get_vendor_warehouse_label

    vendor_label = get_vendor_warehouse_label()
    current_company = SimpleNamespace(
        id='vendor', company_name=vendor_label, company_type='vendor'
    )

    inventories = Inventory.query.filter_by(is_vendor_warehouse=True).join(
        Product, Product.id == Inventory.product_id
    ).order_by(Product.product_name).all()

    inv_ids = [i.id for i in inventories]
    last_tx_map = {}
    if inv_ids:
        subq = db.session.query(
            InventoryTransaction.inventory_id,
            func.max(InventoryTransaction.transaction_date).label('max_date')
        ).filter(InventoryTransaction.inventory_id.in_(inv_ids)).group_by(
            InventoryTransaction.inventory_id
        ).subquery()
        latest_txs = db.session.query(InventoryTransaction).join(
            subq,
            (InventoryTransaction.inventory_id == subq.c.inventory_id) &
            (InventoryTransaction.transaction_date == subq.c.max_date)
        ).all()
        last_tx_map = {tx.inventory_id: tx for tx in latest_txs}

    from datetime import datetime as _dt
    month_start = _dt.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_tx = db.session.query(func.count(InventoryTransaction.id)).filter(
        InventoryTransaction.inventory_id.in_(inv_ids) if inv_ids else False,
        InventoryTransaction.transaction_date >= month_start
    ).scalar() or 0

    stats = {
        'products': len(inventories),
        'total_qty': sum(i.quantity for i in inventories),
        'monthly_tx': int(monthly_tx),
    }

    tx_rows = []
    if inv_ids:
        recent_txs = db.session.query(InventoryTransaction).filter(
            InventoryTransaction.inventory_id.in_(inv_ids)
        ).order_by(InventoryTransaction.id.desc()).limit(100).all()
        inv_by_id = {i.id: i for i in inventories}
        for tx in recent_txs:
            inv = inv_by_id.get(tx.inventory_id)
            tx_rows.append({
                'tx': tx,
                'product': inv.product if inv else None,
                'ref_url': _build_ref_url(tx.reference_type, tx.reference_id),
            })

    return {
        'is_global': False,
        'is_vendor_warehouse': True,
        'current_company': current_company,
        'inventories': inventories,
        'global_inventories': [],
        'last_tx_map': last_tx_map,
        'switchable': _switchable_companies(),
        'locked': False,
        'stats': stats,
        'tx_rows': tx_rows,
        'is_vendor_admin': _is_vendor_admin(current_user),
        'vendor_warehouse_label': vendor_label,
    }


def _build_global_stock_context():
    """全局聚合视图:跨公司汇总 + 流水合并(仅厂商管理员)。"""
    # 1. 全部库存 → 按 product_id 聚合
    rows = db.session.query(Inventory).join(
        Product, Product.id == Inventory.product_id
    ).all()
    grouped = {}  # product_id → {product, total_qty, unit, companies:[{company, qty, last_tx}], inv_ids:[]}
    for inv in rows:
        pid = inv.product_id
        if pid not in grouped:
            grouped[pid] = {
                'product': inv.product,
                'total_qty': 0,
                'unit': inv.unit or '件',
                'companies': [],
                'inv_ids': [],
            }
        g = grouped[pid]
        g['total_qty'] += inv.quantity or 0
        g['companies'].append({
            'company': inv.company,  # 厂商仓时为 None
            'is_vendor_warehouse': bool(inv.is_vendor_warehouse),
            'inv_id': inv.id,
            'quantity': inv.quantity or 0,
            'unit': inv.unit or '件',
        })
        g['inv_ids'].append(inv.id)

    # 2. 取所有 inv 的最近一次流水(供子行显示"最后变动")
    all_inv_ids = [i.id for i in rows]
    last_tx_by_inv = {}
    if all_inv_ids:
        subq = db.session.query(
            InventoryTransaction.inventory_id,
            func.max(InventoryTransaction.transaction_date).label('max_date')
        ).filter(InventoryTransaction.inventory_id.in_(all_inv_ids)).group_by(
            InventoryTransaction.inventory_id
        ).subquery()
        latest = db.session.query(InventoryTransaction).join(
            subq,
            (InventoryTransaction.inventory_id == subq.c.inventory_id) &
            (InventoryTransaction.transaction_date == subq.c.max_date)
        ).all()
        last_tx_by_inv = {tx.inventory_id: tx for tx in latest}

    # 把 last_tx 串到各 company 子行
    for g in grouped.values():
        for sub in g['companies']:
            sub['last_tx'] = last_tx_by_inv.get(sub['inv_id'])
        # 子行按数量降序排,大库存的公司在前
        g['companies'].sort(key=lambda s: s['quantity'], reverse=True)

    # 按产品名排序
    global_inventories = sorted(grouped.values(), key=lambda g: g['product'].product_name or '')

    # 3. 全局 KPI
    from datetime import datetime as _dt
    month_start = _dt.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_tx = db.session.query(func.count(InventoryTransaction.id)).filter(
        InventoryTransaction.transaction_date >= month_start
    ).scalar() or 0
    stats = {
        'products': len(global_inventories),
        'total_qty': sum(g['total_qty'] for g in global_inventories),
        'monthly_tx': int(monthly_tx),
    }

    # 4. 全局最近 100 条流水(带公司列)
    tx_rows = []
    if all_inv_ids:
        recent = db.session.query(InventoryTransaction).order_by(
            InventoryTransaction.id.desc()
        ).limit(100).all()
        # 预加载 inv → company / product 映射
        inv_by_id = {i.id: i for i in rows}
        for tx in recent:
            inv = inv_by_id.get(tx.inventory_id)
            tx_rows.append({
                'tx': tx,
                'company': inv.company if inv else None,
                'product': inv.product if inv else None,
                'ref_url': _build_ref_url(tx.reference_type, tx.reference_id),
            })

    switchable = _switchable_companies()

    from app.utils.inventory_helpers import get_vendor_warehouse_label
    return {
        'is_global': True,
        'current_company': None,
        'global_inventories': global_inventories,
        'inventories': [],          # 模板兼容
        'last_tx_map': {},          # 模板兼容
        'switchable': switchable,
        'locked': False,
        'stats': stats,
        'tx_rows': tx_rows,
        'is_vendor_admin': True,
        'vendor_warehouse_label': get_vendor_warehouse_label(),
    }


def _build_stock_list_context(company_id=None):
    """构建库存主页的渲染上下文(供 tw_stock_list / at_stock_list 共用)。

    返回 dict | None。None 表示权限被拒,调用方应 redirect。
    空 company 时返回完整 ctx(模板内自行处理空态)。
    """
    # 1. Resolve target company
    locked = False
    if company_id is None:
        company_id, locked = _resolve_default_company(current_user)
        if company_id is None:
            from app.utils.inventory_helpers import get_vendor_warehouse_label
            return {
                'is_global': False,
                'current_company': None, 'inventories': [],
                'global_inventories': [],
                'switchable': _switchable_companies(),
                'locked': False,
                'last_tx_map': {}, 'tx_rows': [],
                'stats': {'products': 0, 'total_qty': 0, 'monthly_tx': 0},
                'is_vendor_admin': _is_vendor_admin(current_user),
                'vendor_warehouse_label': get_vendor_warehouse_label(),
            }
    else:
        if not _user_can_view_company(current_user, company_id):
            flash('您没有权限查看该公司的库存', 'danger')
            return None
        # admin 可切换;其他人锁在自己公司
        locked = not _is_vendor_admin(current_user)

    company = Company.query.get_or_404(company_id)

    inventories = Inventory.query.filter_by(company_id=company_id).join(
        Product, Product.id == Inventory.product_id
    ).order_by(Product.product_name).all()

    inv_ids = [i.id for i in inventories]
    last_tx_map = {}
    if inv_ids:
        subq = db.session.query(
            InventoryTransaction.inventory_id,
            func.max(InventoryTransaction.transaction_date).label('max_date')
        ).filter(InventoryTransaction.inventory_id.in_(inv_ids)).group_by(
            InventoryTransaction.inventory_id
        ).subquery()
        latest_txs = db.session.query(InventoryTransaction).join(
            subq,
            (InventoryTransaction.inventory_id == subq.c.inventory_id) &
            (InventoryTransaction.transaction_date == subq.c.max_date)
        ).all()
        last_tx_map = {tx.inventory_id: tx for tx in latest_txs}

    from datetime import datetime as _dt
    month_start = _dt.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_tx = db.session.query(func.count(InventoryTransaction.id)).filter(
        InventoryTransaction.inventory_id.in_(inv_ids) if inv_ids else False,
        InventoryTransaction.transaction_date >= month_start
    ).scalar() or 0

    stats = {
        'products': len(inventories),
        'total_qty': sum(i.quantity for i in inventories),
        'monthly_tx': int(monthly_tx),
    }

    switchable = _switchable_companies() if not locked else []

    tx_rows = []
    if inv_ids:
        recent_txs = db.session.query(InventoryTransaction).filter(
            InventoryTransaction.inventory_id.in_(inv_ids)
        ).order_by(InventoryTransaction.id.desc()).limit(100).all()
        inv_by_id = {i.id: i for i in inventories}
        for tx in recent_txs:
            inv = inv_by_id.get(tx.inventory_id)
            tx_rows.append({
                'tx': tx,
                'product': inv.product if inv else None,
                'ref_url': _build_ref_url(tx.reference_type, tx.reference_id),
            })

    from app.utils.inventory_helpers import get_vendor_warehouse_label
    return {
        'is_global': False,
        'current_company': company,
        'inventories': inventories,
        'global_inventories': [],
        'last_tx_map': last_tx_map,
        'switchable': switchable,
        'locked': locked,
        'stats': stats,
        'tx_rows': tx_rows,
        'is_vendor_admin': _is_vendor_admin(current_user),
        'vendor_warehouse_label': get_vendor_warehouse_label(),
    }


@inventory.route('/api/tw_stock_adjust', methods=['POST'])
@login_required
@permission_required('inventory', 'edit')
def tw_stock_adjust():
    """Tailwind 调整模态的后端 API.

    Accepts: inventory_id (required for 'out'/'set'; for 'in' on new product handled separately later)
             action_type ('in' | 'out' | 'set')
             quantity (positive integer)
             reason (label string, mandatory)
             notes (free text, optional)
    """
    try:
        data = request.get_json() or request.form
        inventory_id = data.get('inventory_id')
        action_type = data.get('action_type')
        quantity = data.get('quantity')
        reason = (data.get('reason') or '').strip()
        notes = (data.get('notes') or '').strip()

        if not inventory_id or not action_type or not quantity:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400
        if not reason:
            return jsonify({'success': False, 'message': '原因必填'}), 400
        try:
            qty = int(quantity)
            if qty <= 0:
                return jsonify({'success': False, 'message': '数量必须大于 0'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '数量必须是有效整数'}), 400

        inv = Inventory.query.get_or_404(int(inventory_id))

        # Permission check: external user can only adjust own company's inventory
        if not _user_can_view_company(current_user, inv.company_id):
            return jsonify({'success': False, 'message': '没有权限'}), 403

        # Resolve quantity_change and transaction_type
        if action_type == 'in':
            quantity_change = qty
            trans_type = 'in'
        elif action_type == 'out':
            quantity_change = -qty
            trans_type = 'out'
        elif action_type == 'set':
            quantity_change = qty - inv.quantity
            trans_type = 'adjustment'
        else:
            return jsonify({'success': False, 'message': '无效的操作类型'}), 400

        # Reject out-of-stock
        if quantity_change < 0 and inv.quantity + quantity_change < 0:
            return jsonify({'success': False, 'message': f'库存不足,当前 {inv.quantity}'}), 400

        # Build description: reason · notes
        description = reason if not notes else f'{reason} · {notes}'

        success, message, updated = update_inventory(
            company_id=inv.company_id,
            product_id=inv.product_id,
            quantity_change=quantity_change,
            transaction_type=trans_type,
            description=description,
            reference_type='manual',
            user_id=current_user.id,
        )
        if not success:
            return jsonify({'success': False, 'message': message}), 400

        return jsonify({
            'success': True,
            'message': '调整成功',
            'new_quantity': updated.quantity,
            'change': quantity_change,
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"tw_stock_adjust failed: {e}")
        return jsonify({'success': False, 'message': f'操作失败: {e}'}), 500


# =============================================================================
# Task 3: 全局流水(厂商管理员专属)
# =============================================================================

def _is_vendor_admin(user):
    """是否有"全局库存"+ 切换公司的权限。
    规则:仅角色为 admin 的用户;其他人(无论是否有 linked_company_id)都只能看自己公司。
    """
    return getattr(user, 'role', None) == 'admin'




def _build_ref_url(ref_type, ref_id):
    """流水关联单据 → 可点击 URL。"""
    if not ref_id:
        return None
    if ref_type == 'order':
        # PurchaseOrder
        try:
            return url_for('purchase_order.at_detail_view', order_id=ref_id)
        except Exception:
            return None
    if ref_type == 'shipment':
        # 独立发货详情页已移除，发货流水不再可点
        return None
    if ref_type == 'settlement':
        # 跳到批价单(settlement_order.id → 找其 pricing_order_id)
        try:
            from app.models.pricing_order import SettlementOrderDetail
            d = SettlementOrderDetail.query.get(ref_id)
            if d and d.settlement_order_id:
                so = d.settlement_order
                if so and so.pricing_order_id:
                    try:
                        return url_for('pricing_order.detail_view', order_id=so.pricing_order_id, _anchor='settlement')
                    except Exception:
                        pass
        except Exception:
            pass
        return None
    return None


# =============================================================================
# Task 4: Excel 导入工具
# =============================================================================

@inventory.route('/api/import/template', methods=['GET'])
@login_required
@permission_required('inventory', 'create')
def tw_import_template():
    """下载 Excel 导入模板(带示例行)。仅厂商管理员。"""
    if not _is_vendor_admin(current_user):
        flash('没有权限', 'danger')
        return redirect(url_for('inventory.at_stock_list'))
    bio = io.BytesIO()
    example_data = [
        {'MN号': 'PNR2100-001', '公司名': '北京 ABC 经销商', '数量': 100, '单位': '台', '存储位置': 'A 区 1 排', '最低库存': 10, '最高库存': 500, '备注': '示例行,可删除'},
        {'MN号': 'CMP2600-002', '公司名': '上海 XYZ 经销商', '数量': 50, '单位': '台', '存储位置': '', '最低库存': '', '最高库存': '', '备注': ''},
    ]
    df = pd.DataFrame(example_data, columns=['MN号', '公司名', '数量', '单位', '存储位置', '最低库存', '最高库存', '备注'])
    with pd.ExcelWriter(bio, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='库存')
    bio.seek(0)
    from flask import send_file as _send
    return _send(bio,
                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                 as_attachment=True,
                 download_name='库存导入模板.xlsx')


@inventory.route('/api/import/preview', methods=['POST'])
@login_required
@permission_required('inventory', 'create')
def tw_import_preview():
    """上传 Excel,解析并返回预览 JSON,不写库。"""
    if not _is_vendor_admin(current_user):
        return jsonify({'success': False, 'message': '没有权限'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'success': False, 'message': '请选择文件'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传 .xlsx 或 .xls 文件'}), 400

    try:
        df = pd.read_excel(file, dtype={'MN号': str, '公司名': str})
    except Exception as e:
        return jsonify({'success': False, 'message': f'文件解析失败: {e}'}), 400

    required = ['MN号', '公司名', '数量']
    for col in required:
        if col not in df.columns:
            return jsonify({'success': False, 'message': f'缺少必需列: {col}'}), 400

    # 预加载 lookups,避免 N+1
    product_map = {p.product_mn: p for p in Product.query.filter(Product.product_mn.isnot(None)).all()}
    company_map = {c.company_name: c for c in Company.query.filter(Company.is_deleted == False).all()}

    rows = []
    counts = {'success': 0, 'warning': 0, 'failed': 0}

    for idx, r in df.iterrows():
        line_no = int(idx) + 2  # Excel 行号(表头是 1)
        mn = str(r.get('MN号', '') or '').strip()
        company_name = str(r.get('公司名', '') or '').strip()
        qty_raw = r.get('数量')
        try:
            qty = int(qty_raw) if pd.notna(qty_raw) else None
        except (ValueError, TypeError):
            qty = None

        product = product_map.get(mn) if mn else None
        company = company_map.get(company_name) if company_name else None

        # 分类
        status = None
        reason = ''
        current_qty = None
        if not mn:
            status, reason = 'failed', 'MN号为空'
        elif not product:
            status, reason = 'failed', f'MN号 {mn} 在产品库找不到'
        elif not company_name:
            status, reason = 'failed', '公司名为空'
        elif not company:
            status, reason = 'failed', f'公司 {company_name} 找不到'
        elif qty is None or qty < 0:
            status, reason = 'failed', '数量非法'
        else:
            existing = Inventory.query.filter_by(company_id=company.id, product_id=product.id).first()
            if existing:
                status = 'warning'
                current_qty = existing.quantity
                reason = f'覆盖 {existing.quantity} → {qty}'
            else:
                status = 'success'
                reason = '新建'

        counts[status] += 1
        rows.append({
            'line_no': line_no,
            'mn': mn,
            'company_name': company_name,
            'product_name': product.product_name if product else '',
            'product_id': product.id if product else None,
            'company_id': company.id if company else None,
            'quantity': qty,
            'unit': str(r.get('单位') or '').strip(),
            'location': str(r.get('存储位置') or '').strip(),
            'min_stock': (int(r['最低库存']) if pd.notna(r.get('最低库存')) else None),
            'max_stock': (int(r['最高库存']) if pd.notna(r.get('最高库存')) else None),
            'current_qty': current_qty,
            'status': status,
            'reason': reason,
        })

    return jsonify({
        'success': True,
        'rows': rows,
        'summary': {
            'total': len(rows),
            **counts,
        },
    })


@inventory.route('/api/import/commit', methods=['POST'])
@login_required
@permission_required('inventory', 'create')
def tw_import_commit():
    """提交导入:对每行非 failed 状态调 update_inventory 写入。单事务。"""
    if not _is_vendor_admin(current_user):
        return jsonify({'success': False, 'message': '没有权限'}), 403

    data = request.get_json() or {}
    rows = data.get('rows') or []
    if not rows:
        return jsonify({'success': False, 'message': '没有可导入的行'}), 400

    batch_id = int(datetime.now().timestamp())
    written = 0
    skipped = 0
    errors = []

    try:
        for row in rows:
            if row.get('status') == 'failed':
                skipped += 1
                continue
            company_id = row.get('company_id')
            product_id = row.get('product_id')
            target_qty = row.get('quantity')

            if not company_id or not product_id or target_qty is None:
                skipped += 1
                errors.append(f"第 {row.get('line_no')} 行: 缺少字段")
                continue

            inv = Inventory.query.filter_by(company_id=company_id, product_id=product_id).first()
            current = inv.quantity if inv else 0
            delta = int(target_qty) - current
            if delta == 0:
                # 数量没变,但更新元数据(unit/location)如果填了
                if inv and (row.get('unit') or row.get('location')):
                    if row.get('unit'):
                        inv.unit = row.get('unit')
                    if row.get('location'):
                        inv.location = row.get('location')
                continue

            success, msg, updated_inv = update_inventory(
                company_id=company_id,
                product_id=product_id,
                quantity_change=delta,
                transaction_type='adjustment',
                description=f'期初导入 batch #{batch_id}',
                reference_type='import',
                reference_id=batch_id,
                user_id=current_user.id,
            )
            if success:
                # 元数据补充
                if updated_inv:
                    if row.get('unit'):
                        updated_inv.unit = row.get('unit')
                    if row.get('location'):
                        updated_inv.location = row.get('location')
                    if row.get('min_stock') is not None:
                        updated_inv.min_stock = row.get('min_stock')
                    if row.get('max_stock') is not None:
                        updated_inv.max_stock = row.get('max_stock')
                written += 1
            else:
                skipped += 1
                errors.append(f"第 {row.get('line_no')} 行: {msg}")
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"导入提交失败: {e}")
        return jsonify({'success': False, 'message': f'导入失败: {e}'}), 500

    return jsonify({
        'success': True,
        'batch_id': batch_id,
        'written': written,
        'skipped': skipped,
        'errors': errors[:10],
    })