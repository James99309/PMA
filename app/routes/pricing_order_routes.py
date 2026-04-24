from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models.pricing_order import PricingOrder, PricingOrderDetail, SettlementOrder, SettlementOrderDetail, PricingOrderApprovalRecord
from app.models.project import Project
from app.models.quotation import Quotation
from app.models.customer import Company
from app.services.pricing_order_service import PricingOrderService
from app.services.discount_permission_service import DiscountPermissionService
from app.permissions import check_permission, permission_required, is_admin_or_ceo
import os

# 条件导入 PDF 生成器（本地开发时可跳过 weasyprint）
if os.environ.get('SKIP_WEASYPRINT', '').lower() in ('1', 'true', 'yes'):
    PDFGenerator = None
else:
    from app.services.pdf_generator import PDFGenerator
import logging
import json
from datetime import datetime

logger = logging.getLogger(__name__)

pricing_order_bp = Blueprint('pricing_order', __name__)


@pricing_order_bp.route('/api/companies/<company_type>')
@login_required
def api_companies_for_pricing(company_type):
    """API端点 - 为批价单获取企业列表（经销商/分销商）"""
    try:
        from app.utils.access_control import get_viewable_data, can_view_company
        
        # 获取用户可查看的企业
        query = get_viewable_data(Company, current_user)
        query = query.filter(Company.is_deleted == False)
        
        # 根据类型筛选企业
        type_mapping = {
            'dealer': ['dealer'],  # 经销商
            'distributor': ['distributor']  # 分销商
        }
        
        if company_type in type_mapping:
            company_types = type_mapping[company_type]
            query = query.filter(Company.company_type.in_(company_types))
        
        companies = query.order_by(Company.company_name).all()
        
        # 格式化返回数据
        result = []
        for company in companies:
            result.append({
                'id': company.id,
                'name': company.company_name,
                'type': company.company_type,
                'owner_name': company.owner.real_name if company.owner else '未指定',
                'is_readable': can_view_company(current_user, company)
            })
        
        return jsonify(result)
    except Exception as e:
        logger.error(f"获取批价单企业列表失败: {str(e)}")
        return jsonify([])


def check_pricing_edit_permission(pricing_order, current_user):
    """
    检查批价单编辑权限，支持审批上下文

    Returns:
        tuple: (can_edit_pricing, can_edit_settlement, is_approval_context,
                can_edit_quantity, can_edit_discount_price, can_edit_basic_info)
    """
    # 检查是否在审批上下文中
    is_approval_context = False

    if pricing_order.status == 'pending':
        # 方式1: 检查通用审批系统（优先，支持分支流程等复杂场景）
        from app.helpers.approval_helpers import is_current_approver
        if is_current_approver('pricing_order', pricing_order.id, current_user.id):
            is_approval_context = True

    # 根据上下文选择权限检查方式
    can_edit_pricing = PricingOrderService.can_edit_pricing_details(
        pricing_order, current_user, is_approval_context=is_approval_context
    )
    can_edit_settlement = PricingOrderService.can_edit_settlement_details(
        pricing_order, current_user, is_approval_context=is_approval_context
    )
    
    # 细粒度权限检查
    can_edit_quantity = PricingOrderService.can_edit_quantity(
        pricing_order, current_user, is_approval_context=is_approval_context
    )
    can_edit_discount_price = PricingOrderService.can_edit_discount_and_price(
        pricing_order, current_user, is_approval_context=is_approval_context
    )
    can_edit_basic_info = PricingOrderService.can_edit_basic_info(
        pricing_order, current_user, is_approval_context=is_approval_context
    )
    
    return (can_edit_pricing, can_edit_settlement, is_approval_context,
            can_edit_quantity, can_edit_discount_price, can_edit_basic_info)


@pricing_order_bp.route('/project/<int:project_id>/start_pricing_process', methods=['POST'])
@login_required
def start_pricing_process(project_id):
    """启动批价流程（从项目页面的签约按钮触发）"""
    try:
        project = Project.query.get_or_404(project_id)
        
        # 检查项目是否在批价或签约阶段（允许签约后再次创建批价单）
        if project.current_stage not in ['quoted', 'signed']:
            return jsonify({
                'success': False,
                'message': '项目必须在批价或签约阶段才能发起批价流程'
            })
        
        # 支持从报价单页面指定报价单ID，或从项目页面使用最新报价单
        data = request.get_json() or {}
        quotation_id = data.get('quotation_id')

        if quotation_id:
            # 从报价单页面：使用指定的报价单
            quotation = Quotation.query.get(quotation_id)
            if not quotation or quotation.project_id != project_id:
                return jsonify({
                    'success': False,
                    'message': '报价单不存在或不属于该项目'
                })
        else:
            # 从项目页面：使用最新报价单
            quotation = Quotation.query.filter_by(project_id=project_id).order_by(
                Quotation.created_at.desc()
            ).first()

            if not quotation:
                return jsonify({
                    'success': False,
                    'message': '项目没有关联的报价单，无法发起批价流程'
                })
        
        # 检查报价单是否有审核标记
        has_approval = (
            # 传统审核流程：有审核状态且不是pending/rejected，且有已审核阶段
            (quotation.approval_status and 
             quotation.approval_status != 'pending' and
             quotation.approval_status != 'rejected' and
             quotation.approved_stages) or
            # 或者有确认徽章（产品明细已确认）
            (quotation.confirmation_badge_status == 'confirmed')
        )
        
        if not has_approval:
            return jsonify({
                'success': False,
                'message': f'报价单 {quotation.quotation_number} 尚未完成审核，无法发起批价流程。请先完成报价单审批。'
            })
        
        # 检查该报价单下是否有未完成的批价单（草稿、审批中、已拒绝）
        existing_pending_order = PricingOrder.query.filter_by(
            project_id=project_id,
            quotation_id=quotation.id
        ).filter(
            PricingOrder.status.in_(['draft', 'pending', 'rejected'])
        ).first()

        # 如果有未完成的批价单，同步报价单备注后跳转到该批价单继续编辑
        if existing_pending_order:
            # 同步报价单整体备注（始终以报价单为准）
            existing_pending_order.notes = quotation.notes or ''
            # 同步各行备注：只更新 item_note 为空的行（避免覆盖用户在批价单里手动填写的备注）
            from app.models.quotation import QuotationDetail as _QD
            # 构建当前报价单明细的 id→note 和 product_name→note 映射
            current_qd_by_id = {}
            current_qd_by_name = {}
            for qd in _QD.query.filter_by(quotation_id=quotation.id).all():
                if qd.item_note:
                    current_qd_by_id[qd.id] = qd.item_note
                    current_qd_by_name[qd.product_name] = qd.item_note
            for pod in existing_pending_order.pricing_details:
                if not pod.item_note:
                    note = (current_qd_by_id.get(pod.source_quotation_detail_id)
                            or current_qd_by_name.get(pod.product_name))
                    if note:
                        pod.item_note = note
            db.session.commit()
            return jsonify({
                'success': True,
                'pricing_order_id': existing_pending_order.id,
                'message': f'该报价单已有未完成的批价单 {existing_pending_order.order_number}，将跳转到该批价单',
                'redirect_url': url_for('pricing_order.edit_pricing_order', order_id=existing_pending_order.id)
            })
        
        # 创建新的批价单 - V2逻辑支持
        if PricingOrderService.should_use_v2_flow():
            # V2版本：只创建草稿，不生成审批流程
            pricing_order, error = PricingOrderService.create_pricing_order(
                project_id=project_id,
                quotation_id=quotation.id,
                current_user_id=current_user.id
            )
            # V2版本下，检查并规范化厂商直签状态
            if pricing_order and not error:
                normalized, message = PricingOrderService.normalize_direct_contract_status(pricing_order, current_user)
                if not normalized:
                    logger.warning(f"规范化厂商直签状态失败: {message}")
                elif message:
                    logger.info(f"厂商直签状态已规范化: {message}")
                db.session.commit()
        else:
            # V1版本：使用原有逻辑
            pricing_order, error = PricingOrderService.create_pricing_order(
                project_id=project_id,
                quotation_id=quotation.id,
                current_user_id=current_user.id
            )
        
        if error:
            return jsonify({
                'success': False,
                'message': error
            })

        return jsonify({
            'success': True,
            'pricing_order_id': pricing_order.id,
            'redirect_url': url_for('pricing_order.edit_pricing_order', order_id=pricing_order.id)
        })
        
    except Exception as e:
        logger.error(f"启动批价流程失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'系统错误: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>')
@login_required
def edit_pricing_order(order_id):
    """批价单编辑页面"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限管理
        if not PricingOrderService.can_view_pricing_order(pricing_order, current_user):
            flash('您没有权限查看该批价单', 'danger')
            return redirect(url_for('project.list_projects'))
        
        # 检查编辑权限 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        can_view_settlement = PricingOrderService.can_view_settlement_tab(current_user)
        
        # 检查是否有审批历史（用于决定是否显示流程图）
        has_approval_history = False
        from app.helpers.approval_helpers import get_object_approval_instance
        approval_instance = get_object_approval_instance('pricing_order', order_id)
        if approval_instance:
            has_approval_history = True
        
        # 获取客户数据（分销商和经销商）- 应用数据所有权过滤
        from app.utils.access_control import get_viewable_data
        
        # 获取用户有权限查看的经销商类型公司（分销商下拉框也显示经销商类型的公司）
        dealers = get_viewable_data(Company, current_user, [Company.company_type.in_(['经销商', 'dealer'])]).all()
        
        # 分销商下拉框显示的也是经销商类型的公司（因为系统中没有单独的分销商类型）
        distributors = dealers

        # 获取当前审批步骤信息（V2统一审批系统）
        current_approval_step = None
        editable_fields = []
        if pricing_order.status == 'pending':
            from app.helpers.approval_helpers import get_object_approval_instance
            from app.models.approval import ApprovalStatus

            # 优先使用V2系统
            approval_instance = get_object_approval_instance('pricing_order', pricing_order.id)
            if approval_instance and approval_instance.status == ApprovalStatus.PENDING:
                # 🔥 修复：使用 get_current_step_info() 获取步骤信息（支持动态步骤）
                current_approval_step = approval_instance.get_current_step_info()
                # 获取 editable_fields（字典或对象都支持）
                step_editable_fields = None
                if current_approval_step:
                    if isinstance(current_approval_step, dict):
                        step_editable_fields = current_approval_step.get('editable_fields')
                    else:
                        step_editable_fields = current_approval_step.editable_fields
                if step_editable_fields:
                    import json
                    try:
                        # step_editable_fields 可能是字符串或已解析的列表
                        if isinstance(step_editable_fields, str):
                            editable_fields = json.loads(step_editable_fields)
                        else:
                            editable_fields = step_editable_fields
                    except:
                        editable_fields = []
        
        current_approval_record = None
        
        # 获取用户的折扣权限
        discount_limits = DiscountPermissionService.get_user_discount_limits(current_user)
        
        # 获取审批步骤的折扣权限状态
        from app.helpers.approval_helpers import get_approval_step_discount_status
        step_discount_statuses = get_approval_step_discount_status(pricing_order)

        return render_template('pricing_order/edit_pricing_order.html',
                             pricing_order=pricing_order,
                             can_edit_pricing=can_edit_pricing,
                             can_edit_settlement=can_edit_settlement,
                             can_view_settlement=can_view_settlement,
                             can_edit_quantity=can_edit_quantity,
                             can_edit_discount_price=can_edit_discount_price,
                             can_edit_basic_info=can_edit_basic_info,
                             distributors=distributors,
                             dealers=dealers,
                             current_approval_record=current_approval_record,
                             current_approval_step=current_approval_step,
                             editable_fields=editable_fields,
                             discount_limits=discount_limits,
                             step_discount_statuses=step_discount_statuses,
                             has_approval_history=has_approval_history)
        
    except Exception as e:
        logger.error(f"访问批价单编辑页面失败: {str(e)}")
        flash(f'访问失败: {str(e)}', 'danger')
        return redirect(url_for('project.list_projects'))


@pricing_order_bp.route('/<int:order_id>/update_basic_info', methods=['POST'])
@login_required
def update_basic_info(order_id):
    """更新批价单基本信息"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        logger.info(f"开始更新批价单{order_id}基本信息，当前用户: {current_user.id}, 创建者: {pricing_order.created_by}, 状态: {pricing_order.status}")
        
        # 权限检查 - 使用统一的权限检查函数
        (_, _, _, _, _, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        logger.info(f"基本信息编辑权限检查结果: {can_edit_basic_info}")
        
        if not can_edit_basic_info:
            logger.warning(f"用户{current_user.id}没有权限编辑批价单{order_id}")
            return jsonify({
                'success': False,
                'message': '您没有权限编辑该批价单'
            }), 403
        
        data = request.get_json()
        logger.info(f"更新基本信息请求数据: {data}")
        
        if not data:
            logger.warning("请求数据为空")
            return jsonify({
                'success': False,
                'message': '请求数据不能为空'
            }), 400
        
        # 处理厂商直签和厂家提货字段
        is_direct_contract = data.get('is_direct_contract', False)
        is_factory_pickup = data.get('is_factory_pickup', False)
        
        # V2版本的厂商直签权限检查
        if 'is_direct_contract' in data and PricingOrderService.should_use_v2_flow(pricing_order):
            # 只有厂家账户可以设置厂商直签
            if not current_user.is_vendor_user():
                return jsonify({
                    'success': False,
                    'message': '只有厂家账户可以设置厂商直签状态'
                }), 403
            
            # 审批中不可修改
            if pricing_order.status in ['pending', 'approved']:
                return jsonify({
                    'success': False,
                    'message': '审批过程中和审批通过后不可修改厂商直签状态'
                }), 403
        
        pricing_order.is_direct_contract = is_direct_contract
        pricing_order.is_factory_pickup = is_factory_pickup
        logger.info(f"设置厂商直签: {is_direct_contract}, 厂家提货: {is_factory_pickup}")
        
        # 根据厂商直签状态处理经销商和分销商
        if is_direct_contract:
            # 厂商直签时，清空经销商和分销商
            pricing_order.dealer_id = None
            pricing_order.distributor_id = None
            logger.info("厂商直签开启，清空经销商和分销商")
        else:
            # 非厂商直签时，正常处理经销商和分销商
            if 'dealer_id' in data:
                dealer_id = data['dealer_id']
                logger.info(f"处理经销商ID: {dealer_id}, 类型: {type(dealer_id)}")
                if dealer_id and str(dealer_id).strip():
                    try:
                        dealer_id = int(dealer_id)
                        pricing_order.dealer_id = dealer_id
                        logger.info(f"设置经销商ID为: {dealer_id}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"无效的经销商ID: {dealer_id}, 错误: {e}")
                        pricing_order.dealer_id = None
                else:
                    pricing_order.dealer_id = None
                    logger.info("清空经销商ID")
            
            # 处理分销商：如果厂家提货开启，清空分销商
            if is_factory_pickup:
                pricing_order.distributor_id = None
                logger.info("厂家提货开启，清空分销商")
            elif 'distributor_id' in data:
                distributor_id = data['distributor_id']
                logger.info(f"处理分销商ID: {distributor_id}, 类型: {type(distributor_id)}")
                if distributor_id and str(distributor_id).strip():
                    try:
                        distributor_id = int(distributor_id)
                        pricing_order.distributor_id = distributor_id
                        logger.info(f"设置分销商ID为: {distributor_id}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"无效的分销商ID: {distributor_id}, 错误: {e}")
                        pricing_order.distributor_id = None
                else:
                    pricing_order.distributor_id = None
                    logger.info("清空分销商ID")

        # 同步业务类型到结算单
        PricingOrderService.sync_business_type_to_settlements(pricing_order)

        db.session.commit()
        logger.info(f"成功更新批价单 {order_id} 基本信息")
        
        return jsonify({
            'success': True,
            'message': '基本信息更新成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"更新基本信息失败: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'更新失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/<int:order_id>/update_pricing_detail', methods=['POST'])
@login_required
def update_pricing_detail(order_id):
    """更新批价单明细"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑批价单明细'
            })
        
        data = request.get_json()
        detail_id = data.get('detail_id')
        quantity = data.get('quantity')
        discount_rate = data.get('discount_rate')
        unit_price = data.get('unit_price')  # 新增单价更新
        item_note = data.get('item_note')  # None means "no change", '' means "clear"

        success, error = PricingOrderService.update_pricing_detail(
            order_id, detail_id, quantity=quantity, discount_rate=discount_rate, unit_price=unit_price, item_note=item_note
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': error
            })
        
        # 重新获取更新后的数据
        pricing_order = PricingOrder.query.get(order_id)

        # 获取更新后的明细数据
        updated_detail = PricingOrderDetail.query.get(detail_id)

        # 构建结算单明细数据
        settlement_details_data = []
        for sd in pricing_order.settlement_details:
            settlement_details_data.append({
                'id': sd.id,
                'product_name': sd.product_name,
                'product_model': sd.product_model,
                'product_desc': sd.product_desc,
                'brand': sd.brand,
                'market_price': sd.market_price,
                'discount_rate': sd.discount_rate,
                'unit_price': sd.unit_price,
                'quantity': sd.quantity,
                'total_price': sd.total_price
            })

        return jsonify({
            'success': True,
            'message': '明细更新成功',
            'pricing_total_amount': pricing_order.formatted_pricing_total_amount,
            'pricing_discount_percentage': pricing_order.pricing_discount_percentage,
            'settlement_total_amount': pricing_order.formatted_settlement_total_amount,
            'settlement_discount_percentage': pricing_order.settlement_discount_percentage,
            'updated_detail': {
                'id': updated_detail.id,
                'discount_rate': updated_detail.discount_rate,
                'unit_price': updated_detail.unit_price,
                'total_price': updated_detail.total_price,
                'item_note': updated_detail.item_note or '',
            } if updated_detail else None,
            'settlement_details': settlement_details_data
        })
        
    except Exception as e:
        logger.error(f"更新批价单明细失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'更新失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/update_notes', methods=['POST'])
@login_required
def update_pricing_notes(order_id):
    """更新批价单整体备注"""
    pricing_order = PricingOrder.query.get_or_404(order_id)
    (can_edit_pricing, can_edit_settlement, is_approval_context,
     can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
    if not can_edit_pricing:
        return jsonify({'success': False, 'message': '没有权限编辑批价单'})
    data = request.get_json()
    pricing_order.notes = data.get('notes', '') or ''
    db.session.commit()
    return jsonify({'success': True})


@pricing_order_bp.route('/<int:order_id>/update_settlement_detail', methods=['POST'])
@login_required
def update_settlement_detail(order_id):
    """更新结算单明细"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限检查函数
        can_edit_settlement = PricingOrderService.can_edit_settlement_details(
            pricing_order, current_user, is_approval_context=True
        )
        if not can_edit_settlement:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑结算单明细'
            })
        
        data = request.get_json()
        detail_id = data.get('detail_id')
        discount_rate = data.get('discount_rate')
        unit_price = data.get('unit_price')
        
        success, error = PricingOrderService.update_settlement_detail(
            order_id, detail_id, discount_rate=discount_rate, unit_price=unit_price
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': error
            })
        
        # 重新获取更新后的数据
        pricing_order = PricingOrder.query.get(order_id)
        
        # 获取更新后的明细数据
        updated_detail = SettlementOrderDetail.query.get(detail_id)
        
        return jsonify({
            'success': True,
            'message': '结算单明细更新成功',
            'settlement_total_amount': pricing_order.formatted_settlement_total_amount,
            'settlement_discount_percentage': pricing_order.settlement_discount_percentage,
            'updated_detail': {
                'id': updated_detail.id,
                'discount_rate': updated_detail.discount_rate,
                'unit_price': updated_detail.unit_price,
                'total_price': updated_detail.total_price
            } if updated_detail else None
        })
        
    except Exception as e:
        logger.error(f"更新结算单明细失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'更新失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/update_total_discount', methods=['POST'])
@login_required
def update_total_discount_rate(order_id):
    """更新批价单或结算单的总折扣率"""
    try:
        data = request.get_json()
        tab_type = data.get('tab_type', 'pricing')  # pricing 或 settlement
        total_discount_rate = data.get('total_discount_rate')

        if total_discount_rate is None:
            return jsonify({'success': False, 'message': '缺少折扣率参数'})

        # 转换为小数形式
        discount_rate_decimal = float(total_discount_rate) / 100
        
        # 获取批价单
        from app.utils.access_control import get_viewable_data
        viewable_orders = get_viewable_data(PricingOrder, current_user)
        pricing_order = viewable_orders.filter(PricingOrder.id == order_id).first_or_404()
        
        # 检查编辑权限 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({'success': False, 'message': '无权限编辑此批价单'})
        
        # 根据tab类型获取相应的明细列表
        if tab_type == 'pricing':
            details = pricing_order.pricing_details
        else:  # settlement
            details = pricing_order.settlement_details
        
        # 更新所有明细的折扣率和价格
        updated_details = []
        for detail in details:
            if detail.market_price and detail.market_price > 0:
                detail.discount_rate = discount_rate_decimal
                detail.unit_price = detail.market_price * discount_rate_decimal
                detail.total_price = detail.unit_price * detail.quantity
            updated_details.append({
                'id': detail.id,
                'discount_rate': detail.discount_rate,
                'unit_price': detail.unit_price,
                'total_price': detail.total_price
            })

        # 同步更新总折扣率字段
        if tab_type == 'pricing':
            pricing_order.pricing_total_discount_rate = discount_rate_decimal
            pricing_order.pricing_total_amount = sum(d.total_price or 0 for d in details)

            # 注意：修改批价单折扣率时，不再同步覆盖结算单折扣率
            # 结算单有独立的折扣率，由用户在结算 tab 中单独设置
        else:
            # 结算单更新不影响批价单
            pricing_order.settlement_total_discount_rate = discount_rate_decimal
            pricing_order.settlement_total_amount = sum(d.total_price or 0 for d in details)

        # 标记批价单已修改
        pricing_order.updated_at = datetime.utcnow()

        # 保存到数据库
        db.session.commit()

        # 构建结算单明细数据
        settlement_details_data = []
        for sd in pricing_order.settlement_details:
            settlement_details_data.append({
                'id': sd.id,
                'product_name': sd.product_name,
                'product_model': sd.product_model,
                'product_desc': sd.product_desc,
                'brand': sd.brand,
                'market_price': sd.market_price,
                'discount_rate': sd.discount_rate,
                'unit_price': sd.unit_price,
                'quantity': sd.quantity,
                'total_price': sd.total_price
            })

        # 返回与明细更新相同格式的响应
        return jsonify({
            'success': True,
            'pricing_total_amount': pricing_order.formatted_pricing_total_amount,
            'pricing_discount_percentage': pricing_order.pricing_discount_percentage,
            'settlement_total_amount': pricing_order.formatted_settlement_total_amount,
            'settlement_discount_percentage': pricing_order.settlement_discount_percentage,
            'updated_details': updated_details,
            'settlement_details': settlement_details_data
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"更新总折扣率失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})


@pricing_order_bp.route('/<int:order_id>/submit', methods=['POST'])
@login_required
def submit_pricing_order(order_id):
    """提交批价单审批"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查
        if pricing_order.created_by != current_user.id and current_user.role != 'admin':
            return jsonify({
                'success': False,
                'message': '您没有权限提交该批价单'
            })
        
        success, error = PricingOrderService.submit_for_approval(order_id, current_user.id)
        
        if not success:
            return jsonify({
                'success': False,
                'message': error
            })
        
        return jsonify({
            'success': True,
            'message': '批价单已提交审批'
        })
        
    except Exception as e:
        logger.error(f"提交批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'提交失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/approve', methods=['POST'])
@login_required
def approve_pricing_order(order_id):
    """审批批价单 - 已弃用，建议使用统一审批API"""
    # 注意：此路由已弃用，前端现在统一使用 /approval/approve/{instanceId}
    # 数据保存逻辑已移到批结算审批动作内部
    
    try:
        current_app.logger.warning(f"⚠️ 使用了已弃用的批价单审批路由: {order_id}")
        
        pricing_order = PricingOrder.query.get_or_404(order_id)
        data = request.get_json()
        
        action = data.get('action')  # 'approve' 或 'reject'
        comment = data.get('comment', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({
                'success': False,
                'message': '无效的审批动作'
            })
        
        # 获取当前审批步骤
        current_step = pricing_order.current_approval_step
        
        # 简化调用，不传递前端数据
        success, error = PricingOrderService.approve_step(
            order_id, current_step, current_user.id, action, comment, None, None
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': error
            })
        
        action_text = '通过' if action == 'approve' else '拒绝'
        return jsonify({
            'success': True,
            'message': f'审批{action_text}成功'
        })
        
    except Exception as e:
        logger.error(f"审批批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'审批失败: {str(e)}'
        })


def evaluate_branch_condition_for_display(branch_condition, pricing_order):
    """
    为前端显示评估分支条件，确定具体的审批人
    直接实现字段获取和条件评估逻辑，避免依赖实例方法
    """
    try:
        from app.models.user import User
        
        # 获取分支条件配置
        field = branch_condition.get('field')
        conditions = branch_condition.get('conditions', [])
        default_branch = branch_condition.get('default_branch')
        
        
        # 获取对象字段值（实现特殊处理逻辑）
        field_value = None
        
        # 特殊处理：批价单的 project_type 字段 — 从关联项目获取（报价单创建时项目可能未审批，project_type 为空）
        if field == 'project_type':
            if hasattr(pricing_order, 'project') and pricing_order.project:
                field_value = pricing_order.project.project_type
        
        # 通用字段获取
        if field_value is None:
            try:
                # 支持点分隔的嵌套字段访问，如 project.project_type
                if '.' in field:
                    parts = field.split('.')
                    value = pricing_order
                    for part in parts:
                        if hasattr(value, part):
                            value = getattr(value, part)
                        else:
                            value = None
                            break
                    field_value = value
                else:
                    # 直接字段访问
                    if hasattr(pricing_order, field):
                        field_value = getattr(pricing_order, field)
                    else:
                        pass
            except Exception as e:
                logger.error(f"🔍 [显示评估] 获取字段值失败: field={field}, error={str(e)}")
        
        
        # 查找匹配的条件
        for condition in conditions:
            condition_value = condition.get('value')
            operator = condition.get('operator', 'equals')
            
            
            # 实现条件评估逻辑
            match = False
            if field_value is None:
                match = operator in ['is_null', 'is_empty']
            else:
                obj_str = str(field_value)
                cond_str = str(condition_value)
                
                if operator == 'equals':
                    match = obj_str == cond_str
                elif operator == 'in':
                    # 处理多值匹配
                    if isinstance(condition_value, list):
                        match = obj_str in [str(v) for v in condition_value]
                    else:
                        # 处理逗号分隔的字符串
                        cond_values = [v.strip() for v in cond_str.split(',')]
                        match = obj_str in cond_values
                elif operator == 'not_equals':
                    match = obj_str != cond_str
                elif operator == 'contains':
                    match = cond_str.lower() in obj_str.lower()
                elif operator == 'not_contains':
                    match = cond_str.lower() not in obj_str.lower()
                # 可以根据需要添加更多操作符
            
            
            if match:
                approver_id = condition.get('approver_id')
                if approver_id:
                    user = User.query.get(approver_id)
                    if user:
                        return {
                            'user_id': approver_id,
                            'username': user.username,
                            'real_name': user.real_name
                        }
                break
        
        # 如果没有匹配的条件，使用默认分支
        if default_branch and default_branch.get('approver_id'):
            approver_id = default_branch.get('approver_id')
            user = User.query.get(approver_id)
            if user:
                return {
                    'user_id': approver_id,
                    'username': user.username,
                    'real_name': user.real_name
                }
        
        return None
        
    except Exception as e:
        logger.error(f"🔍 [显示评估] 评估分支条件失败: {e}")
        return None


@pricing_order_bp.route('/api/approval/<int:order_id>/flow')
@login_required
def get_pricing_order_approval_flow(order_id):
    """获取批价单审批流程信息 - 统一API"""
    try:
        # 获取批价单（使用数据归属机制验证权限）
        from app.utils.access_control import get_viewable_data

        from app.utils.access_control import can_view_pricing_order
        pricing_order = PricingOrder.query.get(order_id)
        if not pricing_order or not can_view_pricing_order(current_user, pricing_order):
            return jsonify({
                'success': False,
                'message': '批价单不存在或您没有权限查看'
            }), 404

        # 获取统一审批流程数据
        from app.helpers.approval_helpers import get_object_approval_instance
        
        # 查找审批实例（包括已召回的实例）
        approval_instance = get_object_approval_instance('pricing_order', order_id)
        
        # 🔍 调试：打印找到的审批实例信息
        if approval_instance:
            pass
        else:
            pass
        
        if not approval_instance:
            # 如果使用新系统但未找到实例，尝试检查旧系统
            if pricing_order.status in ['pending', 'approved', 'rejected']:
                # 有可能是V1系统的数据，尝试调用旧API获取数据
                return get_approval_flow(order_id)
            
            # 返回控制信息，即使没有审批实例
            control_info = {
                'status': pricing_order.status,
                'can_submit': pricing_order.status == 'draft' and pricing_order.created_by == current_user.id,
                'can_recall': False,
                'can_resubmit': False,
                'is_creator': pricing_order.created_by == current_user.id
            }
            
            return jsonify({
                'success': False,
                'message': '未找到审批流程实例',
                'control_info': control_info,
                'data': {
                    'order_number': pricing_order.order_number,
                    'status': pricing_order.status,
                    'flow_data': []
                }
            })
        
        # 获取审批步骤和记录
        from app.models.approval import ApprovalRecord, ApprovalStep, ApprovalStatus
        
        # 构建流程数据
        flow_data = []
        
        # 获取模板步骤（从实例的快照或模板中获取）
        template_steps = []
        if approval_instance.template_snapshot and 'steps' in approval_instance.template_snapshot:
            # 从快照中获取步骤，转换为统一格式
            snapshot_steps = approval_instance.template_snapshot.get('steps', [])
            template_steps = []
            for i, step in enumerate(snapshot_steps):
                step_data = {
                    'id': step.get('step_id'),
                    'step_order': step.get('step_order'),
                    'step_name': step.get('step_name'),
                    'approver_user_id': step.get('approver_user_id'),
                    'action_type': step.get('action_type')
                }
                template_steps.append(step_data)
                # 🔍 调试：特别关注第一步
                if i == 0:
                    pass
        else:
            # 回退到实际模板
            from app.models.approval import ApprovalProcessTemplate
            template = ApprovalProcessTemplate.query.get(approval_instance.process_id)
            if template:
                steps = ApprovalStep.query.filter_by(process_id=template.id).order_by(ApprovalStep.step_order).all()
                template_steps = [{'id': s.id, 'step_order': s.step_order, 'step_name': s.step_name, 
                                 'approver_user_id': s.approver_user_id, 'action_type': s.action_type} for s in steps]
        
        # 获取审批记录
        approval_records = ApprovalRecord.query.filter_by(
            instance_id=approval_instance.id
        ).order_by(ApprovalRecord.timestamp).all()

        # 🔥 修复：current_step 可能是 step_order 或 step_id，需要智能匹配
        current_step_value = approval_instance.current_step
        matched_step_order = None

        # 先尝试用 step_order 匹配
        for step in template_steps:
            if step.get('step_order') == current_step_value:
                matched_step_order = current_step_value
                break

        # 如果 step_order 匹配不到，尝试用 step_id 匹配
        if matched_step_order is None:
            for step in template_steps:
                if step.get('id') == current_step_value:
                    matched_step_order = step.get('step_order')
                    logger.info(f"[审批流程] current_step={current_step_value} 通过 step_id 匹配到 step_order={matched_step_order}")
                    break

        current_step_order = matched_step_order

        # 构建流程数据
        for step in template_steps:
            step_data = {
                # 前端期望的字段名
                'id': step['id'],
                'stage_order': step['step_order'],
                'stage_name': step['step_name'],
                'action_type': step.get('action_type', ''),
                'status': 'pending',
                'approver_name': '',
                'approver_id': step.get('approver_user_id'),
                'completed_at': None,
                'comment': '',
                'is_current': False,
                'can_approve': False,  # 前端期望的审批权限字段
                # 兼容字段名
                'step_order': step['step_order'],
                'step_name': step['step_name']
            }
            
            # 🔍 调试：打印步骤信息
            
            # 查找对应的审批记录
            step_record = next((r for r in approval_records if r.step_id == step['id']), None)

            # 兜底：模板快照模式（动态生成的step_id写入记录时为NULL）
            if not step_record and approval_records:
                approve_records = [r for r in approval_records if r.action in ('approve', 'reject')]
                if approve_records and all(r.step_id is None for r in approve_records):
                    step_order = step['step_order']
                    if step_order <= len(approve_records):
                        step_record = approve_records[step_order - 1]

            if step_record:
                # 🔍 调试：检查审批记录
                
                step_data.update({
                    'status': 'approved' if step_record.action == 'approve' else 'rejected' if step_record.action == 'reject' else 'pending',
                    'completed_at': step_record.timestamp.isoformat() if step_record.timestamp else None,
                    'comment': step_record.comment or ''
                })
                
                # 🔥 修复：只有在非召回操作时才使用记录中的审批人姓名
                # 召回操作的记录中approver_id是召回发起人，不是步骤的真正审批人
                if step_record.action != 'recall':
                    step_data['approver_name'] = step_record.approver.real_name if step_record.approver else ''
            
            # 判断是否当前步骤
            # 🔥 修复：使用已计算的 current_step_order（已兼容 step_id 和 step_order）
            if current_step_order == step['step_order']:
                step_data['is_current'] = True
                # 🔍 前端JavaScript期望current状态标记在status字段中
                if approval_instance.status == ApprovalStatus.PENDING:
                    step_data['status'] = 'current'
            
            # 获取审批人姓名
            if step_data['approver_id']:
                from app.models.user import User
                approver = User.query.get(step_data['approver_id'])
                if approver and not step_data['approver_name']:
                    step_data['approver_name'] = approver.real_name or approver.username
                    # 🔍 调试：打印审批人信息
            else:
                # 对于分支决策等特殊步骤，尝试通过分支条件确定具体审批人
                if step.get('action_type') == 'branch_decision':
                    # 尝试通过分支条件确定具体审批人
                    snapshot_step = None
                    if approval_instance.template_snapshot and 'steps' in approval_instance.template_snapshot:
                        snapshot_steps = approval_instance.template_snapshot.get('steps', [])
                        snapshot_step = next((s for s in snapshot_steps if s.get('step_id') == step['id']), None)
                    
                    if snapshot_step and snapshot_step.get('branch_condition'):
                        try:
                            # 调用分支条件评估逻辑确定审批人
                            determined_approver = evaluate_branch_condition_for_display(
                                snapshot_step['branch_condition'], 
                                pricing_order
                            )
                            if determined_approver:
                                step_data['approver_name'] = determined_approver['real_name']
                                step_data['approver_id'] = determined_approver['user_id']
                            else:
                                step_data['approver_name'] = '分支决策'
                        except Exception as e:
                            logger.error(f"🔍 分支条件评估失败: {e}")
                            step_data['approver_name'] = '分支决策'
                    else:
                        step_data['approver_name'] = '分支决策'
            
            flow_data.append(step_data)
        
        # 检查当前用户是否可以审批
        can_approve = False
        current_stage = None

        # 查找当前步骤和用户权限
        for i, step in enumerate(flow_data):
            if step['is_current']:
                current_stage = step['stage_order']
                # 检查当前用户是否是当前步骤的审批人
                if step['approver_id'] == current_user.id:
                    can_approve = True
                    # 重要：设置当前步骤的 can_approve 字段，前端组件依赖此字段显示审批操作
                    step['can_approve'] = True
                break
        
        # 确定流程状态（用于前端显示）
        flow_status = pricing_order.status
        
        # 🔍 调试：打印状态比较信息
        
        # 使用枚举值比较而不是字符串比较
        from app.models.approval import ApprovalStatus
        if approval_instance.status == ApprovalStatus.RECALLED:
            flow_status = 'recalled'
        elif approval_instance.status == ApprovalStatus.APPROVED:
            flow_status = 'approved'
        elif approval_instance.status == ApprovalStatus.REJECTED:
            flow_status = 'rejected'
        elif approval_instance.status == ApprovalStatus.PENDING:
            flow_status = 'pending'
        
        
        return jsonify({
            'success': True,
            'approval_flow': {
                'order_number': pricing_order.order_number,
                'status': flow_status,  # 使用流程状态
                'pricing_order_status': pricing_order.status,  # 批价单状态
                'instance_id': approval_instance.id,
                'instance_status': str(approval_instance.status),
                'current_step': approval_instance.current_step,
                'current_stage': current_stage,
                'can_approve': can_approve,
                'stages': flow_data,  # 重命名为stages以匹配前端期望
                'flow_data': flow_data  # 保留原字段名作为兼容
            }
        })
        
    except Exception as e:
        import traceback
        logger.error(f"获取批价单审批流程失败: {str(e)}")
        logger.error(f"错误详情: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'获取失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/<int:order_id>/approval_flow')
@login_required
def get_approval_flow(order_id):
    """获取审批流程信息 - 旧版兼容API"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限管理
        if not PricingOrderService.can_view_pricing_order(pricing_order, current_user):
            return jsonify({
                'success': False,
                'message': '您没有权限查看该批价单'
            })
        
        # 构建审批流程信息
        flow_data = []
        for record in pricing_order.approval_records:
            flow_data.append({
                'step_order': record.step_order,
                'step_name': record.step_name,
                'approver_role': record.approver_role,
                'approver_name': record.approver.real_name or record.approver.username if record.approver else '未指定',
                'action': record.action,
                'comment': record.comment,
                'approved_at': record.approved_at.strftime('%Y-%m-%d %H:%M:%S') if record.approved_at else None,
                'is_current': record.step_order == pricing_order.current_approval_step,
                'is_fast_approval': record.is_fast_approval,
                'fast_approval_reason': record.fast_approval_reason
            })
        
        return jsonify({
            'success': True,
            'data': {
                'order_number': pricing_order.order_number,
                'status': pricing_order.status,
                'current_step': pricing_order.current_approval_step,
                'flow_data': flow_data
            }
        })
        
    except Exception as e:
        logger.error(f"获取审批流程失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取失败: {str(e)}'
        })


def _build_pricing_order_query(current_user):
    """构建批价单查询（含权限过滤），供列表和AJAX共用"""
    from app.utils.access_control import get_viewable_data
    from sqlalchemy.orm import joinedload

    query = get_viewable_data(PricingOrder, current_user)
    query = query.options(
        joinedload(PricingOrder.project),
        joinedload(PricingOrder.creator)
    )

    return query


def _apply_pricing_order_filters(query, search='', status_filter='', owner_filter=''):
    """对批价单查询应用筛选条件"""
    if search:
        query = query.outerjoin(Project, PricingOrder.project_id == Project.id).filter(
            db.or_(
                PricingOrder.order_number.ilike(f'%{search}%'),
                Project.project_name.ilike(f'%{search}%')
            )
        )

    if status_filter:
        query = query.filter(PricingOrder.status == status_filter)

    if owner_filter:
        query = query.filter(PricingOrder.created_by == int(owner_filter))

    return query


def _apply_pricing_order_sort(query, sort_field, sort_order):
    """对批价单查询应用排序"""
    sort_map = {
        'order_number': PricingOrder.order_number,
        'pricing_total_amount': PricingOrder.pricing_total_amount,
        'created_at': PricingOrder.created_at,
        'updated_at': PricingOrder.updated_at,
        'status': PricingOrder.status,
    }
    order_attr = sort_map.get(sort_field, PricingOrder.created_at)
    return query.order_by(order_attr.desc() if sort_order == 'desc' else order_attr.asc())


def _get_pricing_order_stats(base_query):
    """计算批价单统计数据（count 用 SQL，amount 跨货币换算到系统默认货币）"""
    from flask_babel import gettext as _
    from sqlalchemy import case, func
    from config import Config
    from app.utils.i18n import get_current_language
    from app.services.multi_currency_aggregation import MultiCurrencyAggregationService

    # 数量统计（不涉及货币，直接 SQL）
    count_result = base_query.with_entities(
        func.count(PricingOrder.id).label('total'),
        func.count(case((PricingOrder.status == 'approved', PricingOrder.id))).label('approved'),
        func.count(case((PricingOrder.status == 'pending', PricingOrder.id))).label('pending'),
        func.count(case((PricingOrder.status == 'draft', PricingOrder.id))).label('draft'),
    ).first()

    # 金额统计（跨货币换算到 Config.DEFAULT_CURRENCY）
    amount_stats = MultiCurrencyAggregationService.sum_converted_with_conditions(
        base_query,
        PricingOrder.pricing_total_amount,
        PricingOrder.currency,
        {
            'total': None,
            'approved': PricingOrder.status == 'approved',
            'pending': PricingOrder.status == 'pending',
            'draft': PricingOrder.status == 'draft',
        }
    )

    # 语言感知的金额单位
    current_lang = get_current_language()
    if current_lang == 'en':
        amount_unit = 'M'
        amount_divisor = 1000000
    else:
        amount_unit = Config.AMOUNT_UNIT
        amount_divisor = Config.AMOUNT_DIVISOR

    total_count = count_result.total or 0
    approved_count = count_result.approved or 0
    pending_count = count_result.pending or 0
    draft_count = count_result.draft or 0

    stats_cards = [
        {
            'id': 'total',
            'title': _('全部批价单'),
            'icon': 'fas fa-file-invoice-dollar',
            'value': total_count,
            'amount': round(amount_stats['total'] / amount_divisor, 2),
            'unit': _('份'),
            'amount_unit': amount_unit,
            'color': 'primary',
        },
        {
            'id': 'approved',
            'title': _('已批准'),
            'icon': 'fas fa-check-circle',
            'value': approved_count,
            'amount': round(amount_stats['approved'] / amount_divisor, 2),
            'unit': _('份'),
            'amount_unit': amount_unit,
            'color': 'success',
        },
        {
            'id': 'pending',
            'title': _('审批中'),
            'icon': 'fas fa-clock',
            'value': pending_count,
            'amount': round(amount_stats['pending'] / amount_divisor, 2),
            'unit': _('份'),
            'amount_unit': amount_unit,
            'color': 'warning',
        },
        {
            'id': 'draft',
            'title': _('草稿'),
            'icon': 'fas fa-edit',
            'value': draft_count,
            'amount': round(amount_stats['draft'] / amount_divisor, 2),
            'unit': _('份'),
            'amount_unit': amount_unit,
            'color': 'secondary',
        },
    ]
    return stats_cards, total_count


@pricing_order_bp.route('/list')
@login_required
def list_pricing_orders():
    """批价单列表页面"""
    from flask_babel import gettext as _

    from app.models.user import User

    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status_filter', '').strip()
    owner_filter = request.args.get('owner_filter', '').strip()
    sort_field = request.args.get('sort', 'created_at')
    sort_order = request.args.get('order', 'desc')
    offset = request.args.get('offset', 0, type=int)
    limit = request.args.get('limit', 30, type=int)
    limit = min(limit, 100)

    # 构建查询（仅1次 get_viewable_data）
    base_query = _build_pricing_order_query(current_user)

    # 获取负责人列表（从 base_query 提取，无需额外构建）
    unique_owner_ids = {row[0] for row in base_query.with_entities(
        PricingOrder.created_by.distinct()
    ).filter(PricingOrder.created_by.isnot(None)).all()}
    available_users = User.query.filter(
        User.id.in_(unique_owner_ids)
    ).order_by(User.real_name, User.username).all()

    # 应用筛选
    filtered_query = _apply_pricing_order_filters(base_query, search, status_filter, owner_filter)

    # 统计（1次聚合，同时拿到 total_count）
    stats_cards, total_count = _get_pricing_order_stats(filtered_query)

    # 排序和分页
    query = _apply_pricing_order_sort(filtered_query, sort_field, sort_order)
    pricing_orders = query.offset(offset).limit(limit).all()
    has_more = (offset + limit) < total_count

    # 筛选配置
    filter_config = {
        'action_url': url_for('pricing_order.list_pricing_orders'),
        'form_id': 'filterForm',
        'reset_url': url_for('pricing_order.list_pricing_orders'),
        'search_field': {
            'name': 'search',
            'label': _('搜索'),
            'placeholder': _('批价单号或项目名称'),
            'value': search,
        },
        'filter_fields': [
            {
                'name': 'owner_filter',
                'label': _('负责人'),
                'all_option_text': _('全部负责人'),
                'current_value': owner_filter,
                'col_width': 2,
                'options': [
                    {'value': str(u.id), 'label': u.real_name or u.username}
                    for u in available_users
                ],
            },
            {
                'name': 'status_filter',
                'label': _('状态'),
                'all_option_text': _('全部状态'),
                'current_value': status_filter,
                'col_width': 2,
                'options': [
                    {'value': 'draft', 'label': _('草稿')},
                    {'value': 'pending', 'label': _('审批中')},
                    {'value': 'approved', 'label': _('已批准')},
                    {'value': 'rejected', 'label': _('已拒绝')},
                ],
            },
        ],
        'auto_submit': True,
        'ajax_mode': True,
        'dynamic_reset_button': True,
    }

    # 列表配置
    list_config = {
        'stats': {'cards': stats_cards},
        'table': {
            'columns': [
                {'key': 'owner', 'field': 'created_by', 'label': _('负责人'), 'type': 'text', 'width': '100px', 'sort_type': 'string'},
                {'key': 'order_number', 'field': 'order_number', 'label': _('批价单号'), 'type': 'link', 'width': '180px', 'min_width': '160px', 'sort_type': 'string'},
                {'key': 'project_name', 'field': 'project_id', 'label': _('项目名称'), 'type': 'text', 'width': '240px', 'min_width': '200px', 'sort_type': 'string'},
                {'key': 'status', 'field': 'status', 'label': _('状态'), 'type': 'text', 'width': '100px', 'sort_type': 'string'},
                {'key': 'pricing_total_amount', 'field': 'pricing_total_amount', 'label': _('批价总额'), 'type': 'number', 'align': 'end', 'width': '140px', 'sort_type': 'currency'},
                {'key': 'created_at', 'field': 'created_at', 'label': _('创建时间'), 'type': 'date', 'width': '120px', 'sort_type': 'date'},
            ],
        },
    }

    # 查询结算Tab权限（用于批价单模态框）
    can_view_settlement = PricingOrderService.can_view_settlement_tab(current_user)

    return render_template('pricing_order/tw_list.html',
                         pricing_orders=pricing_orders,
                         sort_field=sort_field,
                         sort_order=sort_order,
                         offset=offset,
                         limit=limit,
                         has_more=has_more,
                         total_count=total_count,
                         filter_config=filter_config,
                         list_config=list_config,
                         can_view_settlement=can_view_settlement,
                         now_year=datetime.now().year)


@pricing_order_bp.route('/api/list')
@login_required
def pricing_order_list_ajax():
    """批价单列表AJAX分页加载"""
    from flask_babel import gettext as _

    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status_filter', '').strip()
    owner_filter = request.args.get('owner_filter', '').strip()
    sort_field = request.args.get('sort_field', request.args.get('sort', 'created_at'))
    sort_order = request.args.get('sort_direction', request.args.get('order', 'desc'))
    offset = request.args.get('offset', 0, type=int)
    limit = request.args.get('limit', 30, type=int)
    limit = min(limit, 100)

    query = _build_pricing_order_query(current_user)
    query = _apply_pricing_order_filters(query, search, status_filter, owner_filter)

    total_count = query.count()
    query = _apply_pricing_order_sort(query, sort_field, sort_order)
    pricing_orders = query.offset(offset).limit(limit).all()

    html = render_template('pricing_order/tw_list_rows.html', pricing_orders=pricing_orders)

    return jsonify({
        'success': True,
        'html': html,
        'total_count': total_count,
        'loaded_count': offset + len(pricing_orders),
        'has_more': (offset + len(pricing_orders)) < total_count,
    })


@pricing_order_bp.route('/<int:order_id>/add_product', methods=['POST'])
@login_required
def add_product_to_pricing(order_id):
    """添加产品到批价单"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑批价单明细'
            })
        
        data = request.get_json()
        
        # 验证必填字段
        required_fields = ['product_name', 'market_price', 'quantity']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'字段 {field} 不能为空'
                })
        
        # 创建批价单明细
        pricing_detail = PricingOrderDetail(
            pricing_order_id=order_id,
            product_name=data['product_name'],
            product_model=data.get('product_model', ''),
            product_desc=data.get('product_desc', ''),
            brand=data.get('brand', ''),
            unit=data.get('unit', '台'),
            market_price=float(data['market_price']),
            quantity=int(data['quantity']),
            discount_rate=float(data.get('discount_rate', 0.8)),
            source_type='manual'
        )
        
        # 计算价格
        pricing_detail.calculate_prices()
        
        db.session.add(pricing_detail)
        db.session.flush()

        # 获取关联的结算单
        settlement_order = pricing_order.settlement_orders[0] if pricing_order.settlement_orders else None

        # 同时创建结算单明细
        settlement_detail = SettlementOrderDetail(
            pricing_order_id=order_id,
            settlement_order_id=settlement_order.id if settlement_order else None,
            product_name=pricing_detail.product_name,
            product_model=pricing_detail.product_model,
            product_desc=pricing_detail.product_desc,
            brand=pricing_detail.brand,
            unit=pricing_detail.unit,
            product_mn=pricing_detail.product_mn,
            market_price=pricing_detail.market_price,
            unit_price=pricing_detail.unit_price,
            quantity=pricing_detail.quantity,
            discount_rate=pricing_detail.discount_rate,
            pricing_detail_id=pricing_detail.id
        )
        settlement_detail.calculate_prices()
        db.session.add(settlement_detail)
        
        # 重新计算总额和总折扣率（基于明细数据）
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '产品添加成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"添加产品失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'添加失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/delete_product/<int:detail_id>', methods=['DELETE'])
@login_required
def delete_product_from_pricing(order_id, detail_id):
    """从批价单删除产品"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑批价单明细'
            })
        
        pricing_detail = PricingOrderDetail.query.filter_by(
            pricing_order_id=order_id, id=detail_id
        ).first()
        
        if not pricing_detail:
            return jsonify({
                'success': False,
                'message': '产品明细不存在'
            })
        
        # 删除对应的结算单明细
        settlement_detail = SettlementOrderDetail.query.filter_by(
            pricing_detail_id=detail_id
        ).first()
        if settlement_detail:
            db.session.delete(settlement_detail)
        
        # 删除批价单明细
        db.session.delete(pricing_detail)
        
        # 重新计算总额和总折扣率（基于明细数据）
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '产品删除成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"删除产品失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/batch_delete_products', methods=['DELETE'])
@login_required
def batch_delete_products_from_pricing(order_id):
    """批量删除批价单产品明细"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑批价单明细'
            })
        
        data = request.get_json()
        detail_ids = data.get('detail_ids', [])
        
        if not detail_ids:
            return jsonify({
                'success': False,
                'message': '请选择要删除的产品明细'
            })
        
        deleted_count = 0
        for detail_id in detail_ids:
            pricing_detail = PricingOrderDetail.query.filter_by(
                pricing_order_id=order_id, id=detail_id
            ).first()
            
            if pricing_detail:
                # 删除对应的结算单明细
                settlement_detail = SettlementOrderDetail.query.filter_by(
                    pricing_detail_id=detail_id
                ).first()
                if settlement_detail:
                    db.session.delete(settlement_detail)
                
                # 删除批价单明细
                db.session.delete(pricing_detail)
                deleted_count += 1
        
        # 重新计算总额和总折扣率
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'成功删除 {deleted_count} 个产品',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"批量删除产品失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'批量删除失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/save_pricing_details', methods=['POST'])
@login_required
def save_pricing_details(order_id):
    """保存批价单明细（批量保存）"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 记录当前项目阶段状态用于调试
        project_stage_before = pricing_order.project.current_stage if pricing_order.project else None
        logger.info(f"保存批价单 {pricing_order.order_number} 明细前，项目阶段: {project_stage_before}")
        
        # 权限检查 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑批价单明细'
            })
        
        data = request.get_json()
        details_data = data.get('details', [])
        
        if not details_data:
            return jsonify({
                'success': False,
                'message': '请添加至少一个产品明细'
            })
        
        # 删除现有明细
        existing_details = PricingOrderDetail.query.filter_by(pricing_order_id=order_id).all()
        for detail in existing_details:
            # 同时删除对应的结算单明细
            settlement_detail = SettlementOrderDetail.query.filter_by(
                pricing_detail_id=detail.id
            ).first()
            if settlement_detail:
                db.session.delete(settlement_detail)
            db.session.delete(detail)
        
        # 创建新明细
        for detail_data in details_data:
            # 验证必填字段
            if not detail_data.get('product_name'):
                continue
            
            # 创建批价单明细
            pricing_detail = PricingOrderDetail(
                pricing_order_id=order_id,
                product_name=detail_data['product_name'],
                product_model=detail_data.get('product_model', ''),
                product_desc=detail_data.get('product_desc', ''),
                brand=detail_data.get('brand', ''),
                unit=detail_data.get('unit', '台'),
                product_mn=detail_data.get('product_mn', ''),
                market_price=float(detail_data.get('market_price', 0)),
                quantity=int(detail_data.get('quantity', 1)),
                discount_rate=float(detail_data.get('discount_rate', 100)) / 100,
                source_type='manual'
            )
            
            # 计算价格
            pricing_detail.calculate_prices()
            db.session.add(pricing_detail)
            db.session.flush()  # 获取ID

            # 获取关联的结算单
            settlement_order = pricing_order.settlement_orders[0] if pricing_order.settlement_orders else None

            # 同时创建结算单明细
            settlement_detail = SettlementOrderDetail(
                pricing_order_id=order_id,
                settlement_order_id=settlement_order.id if settlement_order else None,
                product_name=pricing_detail.product_name,
                product_model=pricing_detail.product_model,
                product_desc=pricing_detail.product_desc,
                brand=pricing_detail.brand,
                unit=pricing_detail.unit,
                product_mn=pricing_detail.product_mn,
                market_price=pricing_detail.market_price,
                unit_price=pricing_detail.unit_price,
                quantity=pricing_detail.quantity,
                discount_rate=pricing_detail.discount_rate,
                pricing_detail_id=pricing_detail.id
            )
            settlement_detail.calculate_prices()
            db.session.add(settlement_detail)
        
        # 重新计算总额和总折扣率（基于明细数据）
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)
        
        db.session.commit()
        
        # 检查项目阶段是否被意外修改
        project_stage_after = pricing_order.project.current_stage if pricing_order.project else None
        if project_stage_before != project_stage_after:
            logger.warning(f"警告：保存批价单 {pricing_order.order_number} 明细时，项目阶段发生了意外变化: {project_stage_before} -> {project_stage_after}")
        else:
            logger.info(f"保存批价单 {pricing_order.order_number} 明细后，项目阶段保持不变: {project_stage_after}")
        
        return jsonify({
            'success': True,
            'message': '批价单明细保存成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"保存批价单明细失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'保存失败: {str(e)}'
        })


@pricing_order_bp.route('/test')
@login_required
def test_pricing():
    """测试页面 - 仅用于开发调试"""
    return "<h1>批价单功能测试页面</h1><p>当前系统运行正常</p>"


@pricing_order_bp.route('/<int:order_id>/save_all', methods=['POST'])
@login_required
def save_all_pricing_data(order_id):
    """保存批价单所有数据（基本信息和明细）"""
    try:
        
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 记录当前项目阶段状态用于调试
        project_stage_before = pricing_order.project.current_stage if pricing_order.project else None
        
        # 权限检查 - 使用统一的权限检查函数
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
        
        
        if not can_edit_pricing:
            return jsonify({
                'success': False,
                'message': '您没有权限编辑该批价单'
            })
        
        
        # 获取请求数据
        data = request.get_json()
        basic_info = data.get('basic_info', {})
        pricing_details = data.get('pricing_details', [])
        settlement_details = data.get('settlement_details', [])
        
        
        if basic_info:
            pass
        
        if pricing_details:
            pass
        
        if settlement_details:
            pass
        
        # 更新基本信息
        # 处理厂商直签和厂家提货字段
        is_direct_contract = basic_info.get('is_direct_contract', False)
        is_factory_pickup = basic_info.get('is_factory_pickup', False)
        
        pricing_order.is_direct_contract = is_direct_contract
        pricing_order.is_factory_pickup = is_factory_pickup
        
        # 根据厂商直签状态处理经销商和分销商
        if is_direct_contract:
            # 厂商直签时，清空经销商和分销商
            pricing_order.dealer_id = None
            pricing_order.distributor_id = None
        else:
            # 非厂商直签时，正常处理经销商和分销商
            if 'dealer_id' in basic_info:
                dealer_id = basic_info['dealer_id']
                if dealer_id and str(dealer_id).strip():
                    try:
                        pricing_order.dealer_id = int(dealer_id)
                    except (ValueError, TypeError):
                        pricing_order.dealer_id = None
                else:
                    pricing_order.dealer_id = None
            
            # 处理分销商：如果厂家提货开启，清空分销商
            if is_factory_pickup:
                pricing_order.distributor_id = None
            elif 'distributor_id' in basic_info:
                distributor_id = basic_info['distributor_id']
                if distributor_id and str(distributor_id).strip():
                    try:
                        pricing_order.distributor_id = int(distributor_id)
                    except (ValueError, TypeError):
                        pricing_order.distributor_id = None
                else:
                    pricing_order.distributor_id = None
        
        # 处理总折扣率
        if 'pricing_total_discount_rate' in basic_info:
            pricing_total_discount_rate = basic_info['pricing_total_discount_rate']
            if pricing_total_discount_rate and str(pricing_total_discount_rate).strip():
                try:
                    # 将百分比转换为小数
                    discount_rate = float(pricing_total_discount_rate) / 100
                    pricing_order.pricing_total_discount_rate = discount_rate
                except (ValueError, TypeError):
                    pass  # 保留原值
        
        # 处理币种
        if 'currency' in basic_info:
            currency = basic_info['currency']
            if currency and str(currency).strip():
                pricing_order.currency = currency
        
        # 保存批价单明细（如果提供）
        if pricing_details:
            # 删除现有批价单明细
            existing_pricing_details = PricingOrderDetail.query.filter_by(pricing_order_id=order_id).all()
            for detail in existing_pricing_details:
                # 同时删除对应的结算单明细
                settlement_detail = SettlementOrderDetail.query.filter_by(
                    pricing_detail_id=detail.id
                ).first()
                if settlement_detail:
                    db.session.delete(settlement_detail)
                db.session.delete(detail)
            
            # 创建新的批价单明细
            for detail_data in pricing_details:
                if not detail_data.get('product_name'):
                    continue
                
                # 获取前端传递的数据
                market_price = float(detail_data.get('market_price', 0))
                quantity = int(detail_data.get('quantity', 1))
                discount_rate_percent = float(detail_data.get('discount_rate', 100))
                unit_price = float(detail_data.get('unit_price', 0))
                
                # 转换折扣率为小数形式
                discount_rate = discount_rate_percent / 100
                
                # 如果前端没有传递单价，则根据折扣率计算
                if unit_price == 0 and market_price > 0:
                    unit_price = market_price * discount_rate
                
                logger.info(f"保存批价单明细 - 产品: {detail_data['product_name']}, 市场价: {market_price}, 单价: {unit_price}, 数量: {quantity}, 折扣率: {discount_rate_percent}%")
                
                pricing_detail = PricingOrderDetail(
                    pricing_order_id=order_id,
                    product_name=detail_data['product_name'],
                    product_model=detail_data.get('product_model', ''),
                    product_desc=detail_data.get('product_desc', ''),
                    brand=detail_data.get('brand', ''),
                    unit=detail_data.get('unit', '台'),
                    product_mn=detail_data.get('product_mn', ''),
                    market_price=market_price,
                    unit_price=unit_price,  # 直接使用计算好的单价
                    quantity=quantity,
                    discount_rate=discount_rate,
                    source_type='manual'
                )
                # 重新计算总价以确保一致性
                pricing_detail.total_price = unit_price * quantity
                db.session.add(pricing_detail)
                db.session.flush()
                
                # 查找对应的结算单明细数据
                settlement_data = None
                for s_detail in settlement_details:
                    if s_detail.get('product_name') == detail_data['product_name']:
                        settlement_data = s_detail
                        break
                
                # 创建结算单明细，优先使用前端传递的结算单数据
                if settlement_data:
                    # 使用前端传递的结算单明细数据
                    settlement_market_price = float(settlement_data.get('market_price', market_price))
                    settlement_quantity = int(settlement_data.get('quantity', quantity))
                    settlement_discount_rate_percent = float(settlement_data.get('discount_rate', discount_rate_percent))
                    settlement_unit_price = float(settlement_data.get('unit_price', unit_price))
                    
                    # 转换折扣率为小数形式
                    settlement_discount_rate = settlement_discount_rate_percent / 100
                    
                    # 如果前端没有传递单价，则根据折扣率计算
                    if settlement_unit_price == 0 and settlement_market_price > 0:
                        settlement_unit_price = settlement_market_price * settlement_discount_rate
                    
                    logger.info(f"保存结算单明细 - 产品: {settlement_data['product_name']}, 市场价: {settlement_market_price}, 单价: {settlement_unit_price}, 数量: {settlement_quantity}, 折扣率: {settlement_discount_rate_percent}%")
                else:
                    # 如果没有对应的结算单数据，使用批价单数据作为默认值
                    settlement_market_price = market_price
                    settlement_quantity = quantity
                    settlement_discount_rate = discount_rate
                    settlement_unit_price = unit_price
                    logger.info(f"未找到对应结算单明细，使用批价单数据作为默认值 - 产品: {detail_data['product_name']}")

                # 获取关联的结算单
                settlement_order = pricing_order.settlement_orders[0] if pricing_order.settlement_orders else None

                settlement_detail = SettlementOrderDetail(
                    pricing_order_id=order_id,
                    settlement_order_id=settlement_order.id if settlement_order else None,
                    product_name=pricing_detail.product_name,
                    product_model=pricing_detail.product_model,
                    product_desc=pricing_detail.product_desc,
                    brand=pricing_detail.brand,
                    unit=pricing_detail.unit,
                    product_mn=pricing_detail.product_mn,
                    market_price=settlement_market_price,
                    unit_price=settlement_unit_price,
                    quantity=settlement_quantity,
                    discount_rate=settlement_discount_rate,
                    pricing_detail_id=pricing_detail.id
                )
                # 重新计算总价以确保一致性
                settlement_detail.total_price = settlement_unit_price * settlement_quantity
                db.session.add(settlement_detail)
        
        # 保存结算单明细（如果提供且有权限）
        # can_edit_settlement 已经在上面获取了，直接使用
        if settlement_details and can_edit_settlement:
            # 更新现有结算单明细
            for detail_data in settlement_details:
                if not detail_data.get('id'):
                    continue
                
                settlement_detail = SettlementOrderDetail.query.get(detail_data['id'])
                if settlement_detail and settlement_detail.pricing_order_id == order_id:
                    if 'discount_rate' in detail_data:
                        discount_rate_percent = float(detail_data['discount_rate'])
                        settlement_detail.discount_rate = discount_rate_percent / 100
                        logger.info(f"保存时更新结算单明细 {detail_data['id']}: 折扣率从前端 {discount_rate_percent}% 转换为 {settlement_detail.discount_rate:.3f}")
                    if 'unit_price' in detail_data:
                        settlement_detail.unit_price = float(detail_data['unit_price'])
                    settlement_detail.calculate_prices()
        
        
        # 重新计算总额和总折扣率（基于明细数据）
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)

        # 同步业务类型到结算单
        PricingOrderService.sync_business_type_to_settlements(pricing_order)

        try:
            db.session.commit()
        except Exception as e:
            logger.error(f"🔥 [SAVE_ALL_DEBUG] ❌ 数据库事务提交失败: {str(e)}")
            raise
        
        # 检查项目阶段是否被意外修改
        project_stage_after = pricing_order.project.current_stage if pricing_order.project else None
        if project_stage_before != project_stage_after:
            logger.warning(f"警告：保存批价单 {pricing_order.order_number} 数据时，项目阶段发生了意外变化: {project_stage_before} -> {project_stage_after}")
        else:
            logger.info(f"保存批价单 {pricing_order.order_number} 数据后，项目阶段保持不变: {project_stage_after}")
        
        return jsonify({
            'success': True,
            'message': '批价单保存成功'
        })
        
    except Exception as e:
        logger.error(f"🔥 [SAVE_ALL_DEBUG] ❌ 保存过程中发生异常: {str(e)}")
        logger.error(f"🔥 [SAVE_ALL_DEBUG] 异常类型: {type(e).__name__}")
        import traceback
        logger.error(f"🔥 [SAVE_ALL_DEBUG] 异常堆栈:\n{traceback.format_exc()}")
        
        try:
            db.session.rollback()
        except Exception as rollback_e:
            logger.error(f"🔥 [SAVE_ALL_DEBUG] 数据库回滚失败: {str(rollback_e)}")
        
        return jsonify({
            'success': False,
            'message': f'保存失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/save_and_submit', methods=['POST'])
@login_required
def save_and_submit_pricing_order(order_id):
    """保存并提交批价单审批"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查
        if pricing_order.created_by != current_user.id and current_user.role != 'admin':
            return jsonify({
                'success': False,
                'message': '您没有权限提交该批价单'
            })
        
        # 先保存所有数据
        data = request.get_json()
        basic_info = data.get('basic_info', {})
        pricing_details = data.get('pricing_details', [])
        settlement_details = data.get('settlement_details', [])

        # 更新基本信息
        # 处理厂商直签和厂家提货字段
        is_direct_contract = basic_info.get('is_direct_contract', False)
        is_factory_pickup = basic_info.get('is_factory_pickup', False)

        pricing_order.is_direct_contract = is_direct_contract
        pricing_order.is_factory_pickup = is_factory_pickup

        # 根据厂商直签状态处理经销商和分销商
        if is_direct_contract:
            # 厂商直签时，清空经销商和分销商
            pricing_order.dealer_id = None
            pricing_order.distributor_id = None
        else:
            # 非厂商直签时，正常处理经销商和分销商
            if 'dealer_id' in basic_info:
                dealer_id = basic_info['dealer_id']
                if dealer_id and str(dealer_id).strip():
                    try:
                        pricing_order.dealer_id = int(dealer_id)
                    except (ValueError, TypeError) as e:
                        pricing_order.dealer_id = None
                        logger.error(f"dealer_id转换失败: {e}")
                else:
                    pricing_order.dealer_id = None

            # 处理分销商：如果厂家提货开启，清空分销商
            if is_factory_pickup:
                pricing_order.distributor_id = None
            elif 'distributor_id' in basic_info:
                distributor_id = basic_info['distributor_id']
                if distributor_id and str(distributor_id).strip():
                    try:
                        pricing_order.distributor_id = int(distributor_id)
                    except (ValueError, TypeError) as e:
                        pricing_order.distributor_id = None
                else:
                    pricing_order.distributor_id = None
        
        # 处理总折扣率
        if 'pricing_total_discount_rate' in basic_info:
            pricing_total_discount_rate = basic_info['pricing_total_discount_rate']
            
            if pricing_total_discount_rate is not None:
                try:
                    pricing_order.pricing_total_discount_rate = float(pricing_total_discount_rate)
                except (ValueError, TypeError) as e:
                    pass
            else:
                pass
        else:
            pass
        
        # 保存明细（如果提供）
        if pricing_details:
            # 删除现有明细
            existing_details = PricingOrderDetail.query.filter_by(pricing_order_id=order_id).all()
            for detail in existing_details:
                settlement_detail = SettlementOrderDetail.query.filter_by(
                    pricing_detail_id=detail.id
                ).first()
                if settlement_detail:
                    db.session.delete(settlement_detail)
                db.session.delete(detail)
            
            # 创建新明细
            for detail_data in pricing_details:
                if not detail_data.get('product_name'):
                    continue
                
                # 获取前端传递的数据
                market_price = float(detail_data.get('market_price', 0))
                quantity = int(detail_data.get('quantity', 1))
                discount_rate_percent = float(detail_data.get('discount_rate', 100))
                unit_price = float(detail_data.get('unit_price', 0))
                
                # 转换折扣率为小数形式
                discount_rate = discount_rate_percent / 100
                
                # 如果前端没有传递单价，则根据折扣率计算
                if unit_price == 0 and market_price > 0:
                    unit_price = market_price * discount_rate
                
                logger.info(f"提交批价单明细 - 产品: {detail_data['product_name']}, 市场价: {market_price}, 单价: {unit_price}, 数量: {quantity}, 折扣率: {discount_rate_percent}%")
                
                pricing_detail = PricingOrderDetail(
                    pricing_order_id=order_id,
                    product_name=detail_data['product_name'],
                    product_model=detail_data.get('product_model', ''),
                    product_desc=detail_data.get('product_desc', ''),
                    brand=detail_data.get('brand', ''),
                    unit=detail_data.get('unit', '台'),
                    product_mn=detail_data.get('product_mn', ''),
                    market_price=market_price,
                    unit_price=unit_price,  # 直接使用计算好的单价
                    quantity=quantity,
                    discount_rate=discount_rate,
                    source_type='manual'
                )
                # 重新计算总价以确保一致性
                pricing_detail.total_price = unit_price * quantity
                db.session.add(pricing_detail)
                db.session.flush()
                
                # 同时创建结算单明细，使用对应的结算单数据
                settlement_data = None
                # 查找对应的结算单明细数据
                for s_detail in settlement_details:
                    if s_detail.get('product_name') == detail_data['product_name']:
                        settlement_data = s_detail
                        break
                
                if settlement_data:
                    # 使用前端传递的结算单明细数据
                    settlement_market_price = float(settlement_data.get('market_price', market_price))
                    settlement_quantity = int(settlement_data.get('quantity', quantity))
                    settlement_discount_rate_percent = float(settlement_data.get('discount_rate', discount_rate_percent))
                    settlement_unit_price = float(settlement_data.get('unit_price', unit_price))
                    
                    # 转换折扣率为小数形式
                    settlement_discount_rate = settlement_discount_rate_percent / 100
                    
                    # 如果前端没有传递单价，则根据折扣率计算
                    if settlement_unit_price == 0 and settlement_market_price > 0:
                        settlement_unit_price = settlement_market_price * settlement_discount_rate
                    
                    logger.info(f"提交结算单明细 - 产品: {settlement_data['product_name']}, 市场价: {settlement_market_price}, 单价: {settlement_unit_price}, 数量: {settlement_quantity}, 折扣率: {settlement_discount_rate_percent}%")
                else:
                    # 如果没有对应的结算单数据，使用批价单数据
                    settlement_market_price = market_price
                    settlement_quantity = quantity
                    settlement_discount_rate = discount_rate
                    settlement_unit_price = unit_price
                    logger.info(f"未找到对应结算单明细，使用批价单数据 - 产品: {detail_data['product_name']}")

                # 获取关联的结算单
                settlement_order = pricing_order.settlement_orders[0] if pricing_order.settlement_orders else None

                settlement_detail = SettlementOrderDetail(
                    pricing_order_id=order_id,
                    settlement_order_id=settlement_order.id if settlement_order else None,
                    product_name=pricing_detail.product_name,
                    product_model=pricing_detail.product_model,
                    product_desc=pricing_detail.product_desc,
                    brand=pricing_detail.brand,
                    unit=pricing_detail.unit,
                    product_mn=pricing_detail.product_mn,
                    market_price=settlement_market_price,
                    unit_price=settlement_unit_price,
                    quantity=settlement_quantity,
                    discount_rate=settlement_discount_rate,
                    pricing_detail_id=pricing_detail.id
                )
                # 重新计算总价以确保一致性
                settlement_detail.total_price = settlement_unit_price * settlement_quantity
                db.session.add(settlement_detail)
        
        # 重新计算总额和总折扣率（基于明细数据）
        pricing_order.calculate_pricing_totals(recalculate_discount_rate=True)
        pricing_order.calculate_settlement_totals(recalculate_discount_rate=True)

        # 同步业务类型到结算单
        PricingOrderService.sync_business_type_to_settlements(pricing_order)

        # 在审批提交前强制保存数据到数据库
        db.session.commit()

        # 提交审批
        success, error = PricingOrderService.submit_for_approval(order_id, current_user.id)

        if not success:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': error
            })

        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '批价单已保存并提交审批'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"保存并提交批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'保存并提交失败: {str(e)}'
        })


@pricing_order_bp.route('/<int:order_id>/recall', methods=['POST'])
@login_required
def recall_pricing_order(order_id):
    """召回批价单"""
    try:
        data = request.get_json(silent=True) or {}
        reason = data.get('reason', '')
        
        success, error = PricingOrderService.recall_pricing_order(
            order_id, current_user.id, reason
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': error
            }), 400
        
        logger.info(f"用户 {current_user.username} 召回了批价单 {order_id}")
        
        return jsonify({
            'success': True,
            'message': '批价单已成功召回'
        })
        
    except Exception as e:
        logger.error(f"召回批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'召回失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/<int:order_id>/delete', methods=['DELETE'])
@login_required
def delete_pricing_order(order_id):
    """删除批价单"""
    try:
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查：创建人或管理员/CEO可删除
        if pricing_order.created_by != current_user.id and not is_admin_or_ceo():
            return jsonify({
                'success': False,
                'message': '您没有权限删除该批价单'
            }), 403

        # 状态检查：创建人可删除draft/rejected，管理员可删除任意状态
        deletable_statuses = ['draft', 'rejected']
        if pricing_order.status not in deletable_statuses and not is_admin_or_ceo():
            return jsonify({
                'success': False,
                'message': '只有草稿或已驳回状态的批价单才能删除'
            }), 400

        data = request.get_json(silent=True) or {}
        reason = data.get('reason', '')

        # 删除相关的结算单明细
        settlement_details = SettlementOrderDetail.query.filter_by(
            pricing_order_id=order_id
        ).all()
        for detail in settlement_details:
            db.session.delete(detail)
        
        # 删除结算单主记录
        settlement_orders = SettlementOrder.query.filter_by(
            pricing_order_id=order_id
        ).all()
        for settlement_order in settlement_orders:
            db.session.delete(settlement_order)
        
        # 获取批价单明细ID列表，用于清理通过pricing_detail_id关联的结算明细
        pricing_details = PricingOrderDetail.query.filter_by(
            pricing_order_id=order_id
        ).all()
        pricing_detail_ids = [d.id for d in pricing_details]
        if pricing_detail_ids:
            extra_settlement_details = SettlementOrderDetail.query.filter(
                SettlementOrderDetail.pricing_detail_id.in_(pricing_detail_ids)
            ).all()
            for detail in extra_settlement_details:
                db.session.delete(detail)

        # 删除批价单明细
        for detail in pricing_details:
            db.session.delete(detail)
        
        # 删除审批记录（如果有）
        approval_records = PricingOrderApprovalRecord.query.filter_by(
            pricing_order_id=order_id
        ).all()
        for record in approval_records:
            db.session.delete(record)
        
        # 删除V2审批系统相关记录
        from app.models.approval import ApprovalInstance, ApprovalRecord
        
        # 查找并删除V2审批实例
        v2_approval_instances = ApprovalInstance.query.filter_by(
            object_type='pricing_order',
            object_id=order_id
        ).all()
        
        for instance in v2_approval_instances:
            # 先删除审批记录
            instance_records = ApprovalRecord.query.filter_by(
                instance_id=instance.id
            ).all()
            for record in instance_records:
                db.session.delete(record)
            
            # 删除审批实例
            db.session.delete(instance)
            logger.info(f"删除批价单关联的V2审批实例: instance_id={instance.id}")
        
        # 删除批价单
        db.session.delete(pricing_order)
        db.session.commit()
        
        logger.info(f"用户 {current_user.username} 删除了批价单 {pricing_order.order_number}，原因: {reason}")
        
        return jsonify({
            'success': True,
            'message': '批价单已成功删除',
            'project_id': pricing_order.project_id
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"删除批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/<int:order_id>/export_pdf/<pdf_type>')
@login_required
def export_pdf(order_id, pdf_type):
    """导出批价单/结算单PDF（优先使用Word模板，失败则回退到HTML模板）"""
    try:
        from io import BytesIO

        # 获取批价单
        pricing_order = PricingOrder.query.get_or_404(order_id)

        # 权限检查
        from app.utils.access_control import can_view_pricing_order
        if not can_view_pricing_order(current_user, pricing_order):
            flash('您没有权限查看该批价单', 'danger')
            return redirect(url_for('pricing_order.list_pricing_orders'))

        # 优先尝试使用Word模板生成PDF
        use_word_template = request.args.get('template', 'word') == 'word'
        include_notes = request.args.get('include_notes') == '1'
        result = None

        if use_word_template:
            try:
                from app.services.word_generator import WordGenerator
                word_generator = WordGenerator()

                if pdf_type == 'pricing':
                    result = word_generator.generate_pricing_order_pdf(pricing_order, include_notes=include_notes)
                elif pdf_type == 'settlement':
                    result = word_generator.generate_settlement_order_pdf(pricing_order, include_notes=include_notes)
                else:
                    flash('无效的PDF类型', 'danger')
                    return redirect(url_for('pricing_order.edit_pricing_order', order_id=order_id))

            except Exception as word_error:
                logger.warning(f"Word模板生成PDF失败，回退到HTML模板: {str(word_error)}")
                result = None

        # 如果Word模板失败或未使用，回退到HTML模板
        if result is None:
            pdf_generator = PDFGenerator()

            if pdf_type == 'pricing':
                result = pdf_generator.generate_pricing_order_pdf(pricing_order, include_notes=include_notes)
            elif pdf_type == 'settlement':
                result = pdf_generator.generate_settlement_order_pdf(pricing_order, include_notes=include_notes)
            else:
                flash('无效的PDF类型', 'danger')
                return redirect(url_for('pricing_order.edit_pricing_order', order_id=order_id))

        # 返回PDF文件
        pdf_io = BytesIO(result['content'])
        pdf_io.seek(0)

        return send_file(
            pdf_io,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=result['filename']
        )

    except Exception as e:
        logger.error(f"导出PDF失败: {str(e)}")
        flash(f'PDF导出失败: {str(e)}', 'danger')
        return redirect(url_for('pricing_order.edit_pricing_order', order_id=order_id))


@pricing_order_bp.route('/<int:order_id>/export_word/<doc_type>')
@login_required
def export_word(order_id, doc_type):
    """导出批价单/结算单Word文档"""
    try:
        from io import BytesIO
        logger.info(f"[Word导出] 开始导出, order_id={order_id}, doc_type={doc_type}")

        logger.info("[Word导出] 尝试导入 WordGenerator...")
        from app.services.word_generator import WordGenerator
        logger.info("[Word导出] WordGenerator 导入成功")

        # 获取批价单
        pricing_order = PricingOrder.query.get_or_404(order_id)
        logger.info(f"[Word导出] 批价单: {pricing_order.order_number}")

        # 权限检查
        from app.utils.access_control import can_view_pricing_order
        if not can_view_pricing_order(current_user, pricing_order):
            flash('您没有权限查看该批价单', 'danger')
            return redirect(url_for('pricing_order.list_pricing_orders'))

        # 生成Word文档
        include_notes = request.args.get('include_notes') == '1'
        word_generator = WordGenerator()
        logger.info(f"[Word导出] WordGenerator 实例化成功, 模板目录: {word_generator.template_dir}")

        if doc_type == 'pricing':
            logger.info("[Word导出] 调用 generate_pricing_order_word_v2...")
            result = word_generator.generate_pricing_order_word_v2(pricing_order, include_notes=include_notes)
        elif doc_type == 'settlement':
            logger.info("[Word导出] 调用 generate_settlement_order_word_v2...")
            result = word_generator.generate_settlement_order_word_v2(pricing_order, include_notes=include_notes)
        else:
            flash('无效的文档类型', 'danger')
            return redirect(url_for('pricing_order.edit_pricing_order', order_id=order_id))

        logger.info(f"[Word导出] ✅ 生成成功: {result['filename']}, 大小: {len(result['content'])} 字节")

        # 返回Word文件
        doc_io = BytesIO(result['content'])
        doc_io.seek(0)

        return send_file(
            doc_io,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=result['filename']
        )

    except Exception as e:
        import traceback
        logger.error(f"[Word导出] ❌ 导出失败: {str(e)}")
        logger.error(f"[Word导出] 详细错误: {traceback.format_exc()}")
        flash(f'Word导出失败: {str(e)}', 'danger')
        return redirect(url_for('pricing_order.edit_pricing_order', order_id=order_id))


@pricing_order_bp.route('/<int:order_id>/admin_rollback', methods=['POST'])
@login_required
def admin_rollback_pricing_order(order_id):
    """管理员退回已通过的批价单"""
    try:
        from app.permissions import is_admin_or_ceo
        from flask import abort
        
        # 检查管理员或CEO权限
        if not is_admin_or_ceo():
            abort(403)
        
        # 检查是否可以退回
        if not PricingOrderService.can_admin_rollback_pricing_order(order_id, current_user.id):
            return jsonify({
                'success': False,
                'message': '权限不足或批价单状态不允许退回'
            }), 403
        
        # 获取退回原因
        data = request.get_json(silent=True) or {}
        reason = data.get('reason', '')
        
        # 执行退回操作
        success, message = PricingOrderService.admin_rollback_pricing_order(
            order_id, current_user.id, reason
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': message
            }), 400
        
        logger.info(f"管理员 {current_user.username} 退回了批价单 {order_id}")
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        logger.error(f"管理员退回批价单失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'退回失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/api/<int:order_id>/status')
@login_required
def get_pricing_order_status(order_id):
    """获取批价单状态信息 - 用于审批后页面状态更新"""
    try:
        from app.utils.access_control import get_viewable_data
        
        # 获取批价单信息
        viewable_orders = get_viewable_data(PricingOrder, current_user)
        
        pricing_order = viewable_orders.filter(PricingOrder.id == order_id).first()
        if not pricing_order:
            logger.warning(f"❌ [状态API] 批价单 {order_id} 不存在或用户 {current_user.username} 无权访问")
            return jsonify({
                'success': False,
                'message': '批价单不存在或您没有权限访问'
            }), 404
        
        logger.debug(f"✅ [状态API] 成功获取批价单: ID={pricing_order.id}, 状态={pricing_order.status}")
        
        # 检查权限 - 统一权限检查
        try:
            (can_edit_pricing, can_edit_settlement, is_approval_context,
             can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)
            logger.debug(f"✅ [状态API] 权限检查完成: 批价={can_edit_pricing}, 结算={can_edit_settlement}, 审批中={is_approval_context}")
        except Exception as perm_error:
            logger.error(f"❌ [状态API] 权限检查失败: {str(perm_error)}")
            raise perm_error
        
        # 构建状态响应
        response_data = {
            'success': True,
            'status': pricing_order.status,
            'is_locked': pricing_order.status in ['approved', 'paid'],
            'permissions': {
                'can_edit_pricing': can_edit_pricing,
                'can_edit_settlement': can_edit_settlement,
                'can_edit_quantity': can_edit_quantity,
                'can_edit_discount_price': can_edit_discount_price,
                'can_edit_basic_info': can_edit_basic_info,
                'is_approval_context': is_approval_context
            },
            'order_info': {
                'order_number': pricing_order.order_number,
                'current_approver': None,  # 可以后续扩展
                'last_updated': pricing_order.updated_at.isoformat() if pricing_order.updated_at else None
            }
        }
        
        logger.info(f"✅ [状态API] 成功返回批价单 {order_id} 状态, 锁定={response_data['is_locked']}")
        return jsonify(response_data)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"❌ [状态API] 获取批价单 {order_id} 状态失败: {str(e)}")
        logger.error(f"❌ [状态API] 详细错误信息:\n{error_details}")
        return jsonify({
            'success': False,
            'message': f'获取状态失败: {str(e)}'
        }), 500


# ========== V2 统一审批系统 API ==========

@pricing_order_bp.route('/api/approval/<int:order_id>/submit', methods=['POST'])
@login_required
@permission_required('pricing_order', 'view')
def submit_pricing_order_approval(order_id):
    """提交批价单审批 - V2统一审批系统"""
    try:
        logger.info(f"[V2审批] 提交批价单审批: order_id={order_id}, user_id={current_user.id}")
        
        # 获取批价单
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查 - 只有创建人可以提交
        if pricing_order.created_by != current_user.id and current_user.role != 'admin':
            logger.warning(f"[V2审批] 无权限提交: order_id={order_id}, user_id={current_user.id}")
            return jsonify({
                'success': False,
                'message': '只有创建人可以提交审批'
            }), 403
        
        # 检查状态
        if pricing_order.status not in ['draft', 'rejected']:
            logger.warning(f"[V2审批] 状态不允许提交: status={pricing_order.status}")
            return jsonify({
                'success': False,
                'message': '只有草稿或被拒绝状态的批价单才能提交审批'
            })
        
        # 校验关联项目是否有类型（分支审批依赖 project_type）
        if pricing_order.project and not pricing_order.project.project_type:
            return jsonify({
                'success': False,
                'message': '关联项目未设置类型，无法提交批价单审批'
            }), 400
        if not pricing_order.project:
            return jsonify({
                'success': False,
                'message': '批价单未关联项目，无法提交审批'
            }), 400

        # 获取请求数据并保存（使用 silent=True 避免空body时抛出异常）
        data = request.get_json(silent=True) or {}
        logger.info(f"[V2审批] 收到请求数据: {data}")
        
        # 使用专门的审批保存函数
        success, message = PricingOrderService.save_pricing_order_core_data(
            order_id, data, current_user
        )
        
        if not success:
            logger.error(f"[V2审批] 数据保存失败: {message}")
            return jsonify({
                'success': False,
                'message': f'数据保存失败: {message}'
            })
        
        # 导入V2审批系统
        from app.helpers.approval_helpers import start_approval_process, get_available_templates
        from app.models.approval import ApprovalProcessTemplate
        
        # 获取批价单审批模板
        template = ApprovalProcessTemplate.query.filter_by(
            object_type='pricing_order',
            is_active=True
        ).first()
        
        if not template:
            logger.error(f"[V2审批] 未找到批价单审批模板")
            return jsonify({
                'success': False,
                'message': '未找到可用的审批模板'
            }), 400
        
        # 启动V2审批流程（使用 auto_commit=False 确保与批价单状态更新在同一事务中）
        approval_instance = start_approval_process(
            'pricing_order',
            order_id,
            template.id,
            current_user.id,
            auto_commit=False
        )

        if not approval_instance:
            db.session.rollback()
            logger.error(f"[V2审批] 启动审批流程失败")
            return jsonify({
                'success': False,
                'message': '启动审批流程失败'
            })

        # 更新批价单状态为待审批
        pricing_order.status = 'pending'
        # 统一提交：审批实例创建 + 批价单状态更新
        db.session.commit()
        
        logger.info(f"[V2审批] 成功提交: order_id={order_id}, approval_instance={approval_instance.id}")
        return jsonify({
            'success': True,
            'message': '批价单已提交审批',
            'approval_instance_id': approval_instance.id
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"[V2审批] 提交失败: order_id={order_id}, error={str(e)}")
        return jsonify({
            'success': False,
            'message': f'提交审批失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/api/approval/<int:order_id>/resubmit', methods=['POST'])
@login_required
@permission_required('pricing_order', 'view')
def resubmit_pricing_order_approval(order_id):
    """重新提交批价单审批 - V2统一审批系统"""
    try:
        logger.info(f"[V2审批] 重新提交批价单审批: order_id={order_id}, user_id={current_user.id}")
        
        # 获取批价单
        pricing_order = PricingOrder.query.get_or_404(order_id)
        
        # 权限检查
        if pricing_order.created_by != current_user.id and current_user.role != 'admin':
            return jsonify({
                'success': False,
                'message': '只有创建人可以重新提交审批'
            }), 403
        
        # 校验关联项目是否有类型（分支审批依赖 project_type）
        if pricing_order.project and not pricing_order.project.project_type:
            return jsonify({
                'success': False,
                'message': '关联项目未设置类型，无法提交批价单审批'
            }), 400
        if not pricing_order.project:
            return jsonify({
                'success': False,
                'message': '批价单未关联项目，无法提交审批'
            }), 400

        # 获取请求数据并保存（使用 silent=True 避免空body时抛出异常）
        data = request.get_json(silent=True) or {}

        # 使用专门的审批保存函数
        success, message = PricingOrderService.save_pricing_order_core_data(
            order_id, data, current_user
        )

        if not success:
            return jsonify({
                'success': False,
                'message': f'数据保存失败: {message}'
            })

        # 重新提交逻辑与提交相同
        return submit_pricing_order_approval(order_id)
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"[V2审批] 重新提交失败: order_id={order_id}, error={str(e)}")
        return jsonify({
            'success': False,
            'message': f'重新提交失败: {str(e)}'
        }), 500


# ========== 批价单模态框 API ==========

@pricing_order_bp.route('/api/<int:quotation_id>/by_quotation')
@login_required
def get_pricing_order_by_quotation(quotation_id):
    """根据报价单ID获取批价单"""
    try:
        from app.utils.access_control import get_viewable_data

        # 获取用户可查看的批价单
        viewable_orders = get_viewable_data(PricingOrder, current_user)

        # 查找该报价单关联的批价单（取最新的一个）
        pricing_order = viewable_orders.filter(
            PricingOrder.quotation_id == quotation_id
        ).order_by(PricingOrder.created_at.desc()).first()

        if not pricing_order:
            return jsonify({
                'success': False,
                'pricing_order': None,
                'message': '未找到关联的批价单'
            })

        return jsonify({
            'success': True,
            'pricing_order': {
                'id': pricing_order.id,
                'order_number': pricing_order.order_number,
                'status': pricing_order.status
            }
        })

    except Exception as e:
        logger.error(f"获取批价单失败: quotation_id={quotation_id}, error={str(e)}")
        return jsonify({
            'success': False,
            'message': f'获取批价单失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/api/<int:order_id>/detail')
@login_required
def get_pricing_order_detail_api(order_id):
    """获取批价单详细信息 - 用于模态框展示"""
    try:
        from app.utils.access_control import can_view_pricing_order
        pricing_order = PricingOrder.query.get(order_id)
        if not pricing_order or not can_view_pricing_order(current_user, pricing_order):
            return jsonify({
                'success': False,
                'message': '批价单不存在或您没有权限访问'
            }), 404

        # 检查权限
        (can_edit_pricing, can_edit_settlement, is_approval_context,
         can_edit_quantity, can_edit_discount_price, can_edit_basic_info) = check_pricing_edit_permission(pricing_order, current_user)

        # 获取当前审批步骤的可编辑字段列表（用于前端字段控制）
        editable_fields = []
        if pricing_order.status == 'pending' and is_approval_context:
            from app.helpers.approval_helpers import get_object_approval_instance
            from app.models.approval import ApprovalStatus

            approval_instance = get_object_approval_instance('pricing_order', pricing_order.id)
            if approval_instance and approval_instance.status == ApprovalStatus.PENDING:
                # 🔥 修复：使用 get_current_step_info() 获取步骤信息（支持动态步骤）
                current_step = approval_instance.get_current_step_info()
                if current_step:
                    # 获取 editable_fields（字典或对象都支持）
                    step_editable_fields = current_step.get('editable_fields') if isinstance(current_step, dict) else current_step.editable_fields
                    if step_editable_fields:
                        import json
                        try:
                            editable_fields = json.loads(step_editable_fields) if isinstance(step_editable_fields, str) else step_editable_fields
                        except:
                            editable_fields = []

        # 判断是否可以提交/召回/重新提交（创建人或管理员）
        # - draft: 可提交
        # - pending: 可召回
        # - rejected/recalled: 可重新提交
        submit_statuses = ['draft', 'pending', 'rejected', 'recalled']
        can_submit = pricing_order.status in submit_statuses and (
            pricing_order.created_by == current_user.id or current_user.role == 'admin'
        )

        # 构建批价明细数据
        pricing_details = []
        pricing_total = 0
        pricing_market_total = 0
        for detail in pricing_order.pricing_details:
            pricing_details.append({
                'id': detail.id,
                'product_name': detail.product_name,
                'product_model': detail.product_model,
                'product_desc': detail.product_desc,
                'brand': detail.brand,
                'unit': detail.unit,
                'product_mn': detail.product_mn,
                'market_price': detail.market_price,
                'discount_rate': detail.discount_rate,
                'unit_price': detail.unit_price,
                'quantity': detail.quantity,
                'total_price': detail.total_price,
                'item_note': detail.item_note or '',
            })
            pricing_total += detail.total_price or 0
            pricing_market_total += (detail.market_price or 0) * (detail.quantity or 0)

        # 构建结算明细数据
        settlement_details = []
        settlement_total = 0
        settlement_market_total = 0
        for detail in pricing_order.settlement_details:
            settlement_details.append({
                'id': detail.id,
                'product_name': detail.product_name,
                'product_model': detail.product_model,
                'product_desc': detail.product_desc,
                'brand': detail.brand,
                'unit': detail.unit,
                'product_mn': detail.product_mn,
                'market_price': detail.market_price,
                'discount_rate': detail.discount_rate,
                'unit_price': detail.unit_price,
                'quantity': detail.quantity,
                'total_price': detail.total_price
            })
            settlement_total += detail.total_price or 0
            settlement_market_total += (detail.market_price or 0) * (detail.quantity or 0)

        # 计算总折扣率
        pricing_discount_rate = (pricing_total / pricing_market_total) if pricing_market_total > 0 else 0
        settlement_discount_rate = (settlement_total / settlement_market_total) if settlement_market_total > 0 else 0

        # 计算分销利润
        distributor_profit = pricing_total - settlement_total if pricing_total > settlement_total else 0

        # 判断是否可编辑（草稿、已驳回、已召回状态且是创建者或管理员）
        editable_statuses = ['draft', 'rejected', 'recalled']
        can_edit = pricing_order.status in editable_statuses and (
            pricing_order.created_by == current_user.id or current_user.role == 'admin'
        )

        # 判断是否可删除（草稿、已驳回、已召回状态且是创建者或管理员）
        can_delete = pricing_order.status in editable_statuses and (
            pricing_order.created_by == current_user.id or current_user.role == 'admin'
        )

        # 判断是否可导出PDF（仅审批通过后）
        # Tailwind 模态框权限：仅管理员/商务助理/财务总监有导出和创建订单权限
        # 注：Bootstrap 详情页保持原有逻辑（创建人/销售负责人也可导出批价单）
        can_export_pricing = False
        can_export_settlement = False
        can_create_sales_order = False
        if pricing_order.status == 'approved':
            # 导出和创建订单权限：仅管理员/商务助理/财务总监
            allowed_roles = ['admin', 'business_admin', 'finance_director']
            can_export_pricing = current_user.role in allowed_roles
            can_export_settlement = current_user.role in allowed_roles
            can_create_sales_order = current_user.role in allowed_roles

        # 构建响应
        response_data = {
            'success': True,
            'id': pricing_order.id,
            'order_number': pricing_order.order_number,
            'status': pricing_order.status,
            'project_id': pricing_order.project_id,
            'project_name': pricing_order.project.project_name if pricing_order.project else None,
            'quotation_id': pricing_order.quotation_id,
            'quotation_number': pricing_order.quotation.quotation_number if pricing_order.quotation else None,
            'creator_name': pricing_order.creator.real_name or pricing_order.creator.username if pricing_order.creator else None,
            'created_at': pricing_order.created_at.strftime('%Y-%m-%d %H:%M') if pricing_order.created_at else None,
            # 渠道信息
            'dealer_id': pricing_order.dealer_id,
            'dealer_name': pricing_order.dealer.company_name if pricing_order.dealer else None,
            'distributor_id': pricing_order.distributor_id,
            'distributor_name': pricing_order.distributor.company_name if pricing_order.distributor else None,
            'is_direct_contract': pricing_order.is_direct_contract or False,
            'is_factory_pickup': pricing_order.is_factory_pickup or False,
            'notes': pricing_order.notes or '',
            # 明细数据
            'pricing_details': pricing_details,
            'settlement_details': settlement_details,
            'pricing_total': pricing_total,
            'settlement_total': settlement_total,
            'pricing_discount_rate': pricing_discount_rate,
            'settlement_discount_rate': settlement_discount_rate,
            'distributor_profit': distributor_profit,
            # 权限
            'can_submit': can_submit,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'can_edit_pricing': can_edit_pricing,
            'can_edit_settlement': can_edit_settlement,
            # 细粒度权限
            'can_edit_quantity': can_edit_quantity,
            'can_edit_discount_price': can_edit_discount_price,
            'can_edit_basic_info': can_edit_basic_info,
            'is_approval_context': is_approval_context,
            # 当前审批步骤可编辑字段列表（用于 ApprovalFieldControl）
            'editable_fields': editable_fields,
            # 导出权限（仅审批通过后有效）
            'can_export_pricing': can_export_pricing,
            'can_export_settlement': can_export_settlement,
            # 创建客户订单权限
            'can_create_sales_order': can_create_sales_order
        }

        return jsonify(response_data)

    except Exception as e:
        import traceback
        logger.error(f"获取批价单详情失败: order_id={order_id}, error={str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': f'获取批价单详情失败: {str(e)}'
        }), 500


@pricing_order_bp.route('/export_excel')
@login_required
def export_excel_list():
    """导出批价单/结算单Excel"""
    from flask_babel import gettext as _
    from io import BytesIO
    from sqlalchemy.orm import joinedload, subqueryload
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # 角色权限检查
    allowed_roles = ['admin', 'business_admin', 'finance_director']
    if current_user.role not in allowed_roles:
        return jsonify({'success': False, 'message': '您没有权限执行此操作'}), 403

    # 解析参数
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)  # 0 或空 = 全部月份
    exclude_statuses = request.args.getlist('exclude_statuses')
    include_settlement = request.args.get('include_settlement', '1') == '1'

    if not year:
        year = datetime.now().year

    # 构建查询 - 复用权限过滤
    query = _build_pricing_order_query(current_user)
    query = query.options(
        subqueryload(PricingOrder.pricing_details),          # 1对多：子查询避免笛卡尔积
        subqueryload(PricingOrder.settlement_details),       # 1对多：子查询避免笛卡尔积
        joinedload(PricingOrder.dealer),                     # 1对1：JOIN高效
        joinedload(PricingOrder.distributor),                 # 1对1：JOIN高效
        subqueryload(PricingOrder.settlement_orders)         # 1对多：子查询避免笛卡尔积
            .subqueryload(SettlementOrder.details)
            .joinedload(SettlementOrderDetail.settlement_company),
        subqueryload(PricingOrder.settlement_orders)
            .joinedload(SettlementOrder.creator),            # 修复N+1查询
        subqueryload(PricingOrder.settlement_orders)
            .joinedload(SettlementOrder.project),            # 修复N+1查询
        subqueryload(PricingOrder.settlement_orders)
            .joinedload(SettlementOrder.dealer),             # 修复N+1查询
        subqueryload(PricingOrder.settlement_orders)
            .joinedload(SettlementOrder.distributor),        # 修复N+1查询
    )

    # 按年月筛选
    start_date = datetime(year, month if month else 1, 1)
    if month:
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)
    query = query.filter(PricingOrder.created_at >= start_date, PricingOrder.created_at < end_date)

    # 排除指定状态
    if exclude_statuses:
        query = query.filter(PricingOrder.status.notin_(exclude_statuses))

    query = query.order_by(PricingOrder.created_at.desc())
    pricing_orders = query.all()

    # ---- 创建 Excel ----
    wb = Workbook()

    # 公共样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    normal_font = Font(name='微软雅黑', size=10)
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def write_cell(ws, row, col, value, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = align or left_align
        return cell

    status_labels = {
        'draft': '草稿', 'pending': '审批中',
        'approved': '已批准', 'rejected': '已拒绝',
    }
    flow_type_labels = {
        'channel_follow': '渠道跟进类',
        'sales_focus': '销售重点类',
        'sales_key': '销售重点类',
        'sales_opportunity': '销售机会类',
    }
    settlement_status_labels = {
        'pending': '待结算', 'partially_settled': '部分结算',
        'fully_settled': '全部结算',
    }
    detail_settlement_labels = {
        'draft': '草稿', 'pending': '待结算', 'settled': '已结算',
    }

    # ===== Sheet1: 批价单 =====
    ws1 = wb.active
    ws1.title = '批价单'
    pricing_headers = [
        '批价单号', '销售负责人', '项目名称', '创建时间', '状态',
        '审批流类型', '批价单总金额', '整体折扣率', '币种',
        '经销商', '分销商',
        '产品名称', '产品型号', '品牌', '单位', '数量',
        '市场单价', '折扣率', '折扣后价格', '小计'
    ]
    write_header(ws1, pricing_headers)

    row = 2
    for po in pricing_orders:
        creator_name = (po.creator.real_name or po.creator.username) if po.creator else ''
        project_name = po.project.project_name if po.project else ''
        created_str = po.created_at.strftime('%Y-%m-%d') if po.created_at else ''
        status_text = status_labels.get(po.status, po.status)
        flow_text = flow_type_labels.get(po.approval_flow_type, po.approval_flow_type or '')
        dealer_name = po.dealer.company_name if po.dealer else ''
        distributor_name = po.distributor.company_name if po.distributor else ''
        discount_pct = f"{(po.pricing_total_discount_rate or 1) * 100:.2f}%"

        details = po.pricing_details
        if not details:
            # 无明细时输出一行基本信息
            write_cell(ws1, row, 1, po.order_number)
            write_cell(ws1, row, 2, creator_name)
            write_cell(ws1, row, 3, project_name)
            write_cell(ws1, row, 4, created_str)
            write_cell(ws1, row, 5, status_text)
            write_cell(ws1, row, 6, flow_text)
            write_cell(ws1, row, 7, po.pricing_total_amount or 0, right_align)
            write_cell(ws1, row, 8, discount_pct, right_align)
            write_cell(ws1, row, 9, po.currency or 'CNY')
            write_cell(ws1, row, 10, dealer_name)
            write_cell(ws1, row, 11, distributor_name)
            row += 1
        else:
            for d in details:
                write_cell(ws1, row, 1, po.order_number)
                write_cell(ws1, row, 2, creator_name)
                write_cell(ws1, row, 3, project_name)
                write_cell(ws1, row, 4, created_str)
                write_cell(ws1, row, 5, status_text)
                write_cell(ws1, row, 6, flow_text)
                write_cell(ws1, row, 7, po.pricing_total_amount or 0, right_align)
                write_cell(ws1, row, 8, discount_pct, right_align)
                write_cell(ws1, row, 9, po.currency or 'CNY')
                write_cell(ws1, row, 10, dealer_name)
                write_cell(ws1, row, 11, distributor_name)
                write_cell(ws1, row, 12, d.product_name or '')
                write_cell(ws1, row, 13, d.product_model or '')
                write_cell(ws1, row, 14, d.brand or '')
                write_cell(ws1, row, 15, d.unit or '')
                write_cell(ws1, row, 16, d.quantity or 0, right_align)
                write_cell(ws1, row, 17, d.market_price or 0, right_align)
                write_cell(ws1, row, 18, f"{(d.discount_rate or 1) * 100:.2f}%", right_align)
                write_cell(ws1, row, 19, d.unit_price or 0, right_align)
                write_cell(ws1, row, 20, d.total_price or 0, right_align)
                row += 1

    # 设置列宽
    from openpyxl.utils import get_column_letter
    col_widths_1 = [16, 12, 25, 12, 10, 14, 14, 12, 8, 20, 20, 25, 18, 12, 8, 8, 12, 10, 14, 14]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet2: 结算单（可选）=====
    if include_settlement:
        ws2 = wb.create_sheet(title='结算单')
        settlement_headers = [
            '结算单号', '销售负责人', '项目名称', '创建时间',
            '审批状态', '结算状态', '总金额', '整体折扣率', '币种',
            '经销商', '分销商', '关联批价单号',
            '产品名称', '产品型号', '品牌', '单位', '数量',
            '市场单价', '折扣率', '折扣后价格', '小计',
            '明细结算状态', '结算对象公司'
        ]
        write_header(ws2, settlement_headers)

        row2 = 2
        for po in pricing_orders:
            for so in po.settlement_orders:
                creator_name = (so.creator.real_name or so.creator.username) if so.creator else ''
                project_name = so.project.project_name if so.project else ''
                created_str = so.created_at.strftime('%Y-%m-%d') if so.created_at else ''
                status_text = status_labels.get(so.status, so.status)
                settle_status_text = settlement_status_labels.get(
                    so.get_dynamic_settlement_status(), so.settlement_status or ''
                )
                dealer_name = so.dealer.company_name if so.dealer else ''
                distributor_name = so.distributor.company_name if so.distributor else ''
                discount_pct = f"{(so.total_discount_rate or 1) * 100:.2f}%"

                details = so.details
                if not details:
                    write_cell(ws2, row2, 1, so.order_number)
                    write_cell(ws2, row2, 2, creator_name)
                    write_cell(ws2, row2, 3, project_name)
                    write_cell(ws2, row2, 4, created_str)
                    write_cell(ws2, row2, 5, status_text)
                    write_cell(ws2, row2, 6, settle_status_text)
                    write_cell(ws2, row2, 7, so.total_amount or 0, right_align)
                    write_cell(ws2, row2, 8, discount_pct, right_align)
                    write_cell(ws2, row2, 9, so.currency or 'CNY')
                    write_cell(ws2, row2, 10, dealer_name)
                    write_cell(ws2, row2, 11, distributor_name)
                    write_cell(ws2, row2, 12, po.order_number)
                    row2 += 1
                else:
                    for d in details:
                        write_cell(ws2, row2, 1, so.order_number)
                        write_cell(ws2, row2, 2, creator_name)
                        write_cell(ws2, row2, 3, project_name)
                        write_cell(ws2, row2, 4, created_str)
                        write_cell(ws2, row2, 5, status_text)
                        write_cell(ws2, row2, 6, settle_status_text)
                        write_cell(ws2, row2, 7, so.total_amount or 0, right_align)
                        write_cell(ws2, row2, 8, discount_pct, right_align)
                        write_cell(ws2, row2, 9, so.currency or 'CNY')
                        write_cell(ws2, row2, 10, dealer_name)
                        write_cell(ws2, row2, 11, distributor_name)
                        write_cell(ws2, row2, 12, po.order_number)
                        write_cell(ws2, row2, 13, d.product_name or '')
                        write_cell(ws2, row2, 14, d.product_model or '')
                        write_cell(ws2, row2, 15, d.brand or '')
                        write_cell(ws2, row2, 16, d.unit or '')
                        write_cell(ws2, row2, 17, d.quantity or 0, right_align)
                        write_cell(ws2, row2, 18, d.market_price or 0, right_align)
                        write_cell(ws2, row2, 19, f"{(d.discount_rate or 1) * 100:.2f}%", right_align)
                        write_cell(ws2, row2, 20, d.unit_price or 0, right_align)
                        write_cell(ws2, row2, 21, d.total_price or 0, right_align)
                        write_cell(ws2, row2, 22, detail_settlement_labels.get(d.settlement_status, d.settlement_status or ''))
                        write_cell(ws2, row2, 23, d.settlement_company.company_name if d.settlement_company else '')
                        row2 += 1

        col_widths_2 = [16, 12, 25, 12, 10, 10, 14, 12, 8, 20, 20, 16, 25, 18, 12, 8, 8, 12, 10, 14, 14, 12, 20]
        for i, w in enumerate(col_widths_2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    # 输出文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_str = f"-{month:02d}" if month else ""
    filename = f"批价单导出_{year}{month_str}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.document',
        as_attachment=True,
        download_name=filename
    )