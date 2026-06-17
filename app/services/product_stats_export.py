# -*- coding: utf-8 -*-
"""产品统计电子表格通用构建器(透视友好)。

报价单批量导出 / 批价单(批价·结算)批量导出共用此核心,避免重复。
两个 sheet:
  统计      — 按型号聚合(求和数量/合计),供直接看汇总
  产品统计  — 扁平明细,每个产品行一条,供自建透视

调用方负责把各自数据整理成 rows(每行 10 列,见 ROW_SPEC),并给出
单据编号列 / 单据状态列的本地化表头(如「报价单编号」「批价单状态」)。
"""
from io import BytesIO

from flask_babel import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

_HFILL = PatternFill('solid', fgColor='2F2E2B')

# 每行(明细)10 列顺序约定:
#   [产品MN编号, 产品名称, 型号, 数量, 单价, 合计, 单据编号, 项目名称, 项目阶段, 单据状态]
ROW_SPEC = ('mn', 'name', 'model', 'qty', 'unit_price', 'total', 'doc_no', 'project', 'stage', 'status')


def _style_header(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = _HFILL
        c.alignment = Alignment(vertical='center')


def _make_table(ws, name, ncols, last_row):
    """把 A1:<ncols>x<last_row> 套成 Excel 表格(边框+隔行底纹+表头筛选)。"""
    if last_row < 2:
        return
    ref = 'A1:%s%d' % (get_column_letter(ncols), last_row)
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tbl)


def build_product_stats_xlsx(rows, doc_no_label, status_label, table_prefix='t'):
    """构建产品统计 xlsx(BytesIO)。

    rows:        list[list],每行 10 列,顺序见 ROW_SPEC。
    doc_no_label:单据编号列表头(已本地化),如 _('报价单编号') / _('批价单编号')。
    status_label:单据状态列表头(已本地化),如 _('报价单状态') / _('批价单状态')。
    table_prefix:Excel 表对象名前缀(同一工作簿内需唯一)。
    """
    agg = {}  # model -> {name, qty, total}
    for r in rows:
        model, name, qty, total = r[2], r[1], r[3] or 0, r[5] or 0
        key = model or (r[0] or '')
        a = agg.setdefault(key, {'name': name or '', 'qty': 0, 'total': 0.0})
        a['qty'] += qty
        a['total'] += total
        if not a['name'] and name:
            a['name'] = name

    wb = Workbook()

    # sheet 1:统计(按型号聚合)
    ws1 = wb.active
    ws1.title = _('统计')
    ws1.append([_('型号'), _('产品名称'), _('求和项:数量'), _('求和项:合计')])
    _style_header(ws1)
    for model in sorted(agg.keys()):
        a = agg[model]
        ws1.append([model, a['name'], a['qty'], round(a['total'], 2)])
    _make_table(ws1, table_prefix + '_summary', 4, 1 + len(agg))
    ws1.append([_('总计'), '', sum(a['qty'] for a in agg.values()),
                round(sum(a['total'] for a in agg.values()), 2)])
    for c in ws1[ws1.max_row]:
        c.font = Font(bold=True)
    for i, w in enumerate([20, 30, 14, 16], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    # sheet 2:产品统计(扁平明细)
    ws = wb.create_sheet(_('产品统计'))
    ws.append([_('产品MN编号'), _('产品名称'), _('型号'), _('数量'), _('单价'),
               _('合计'), doc_no_label, _('项目名称'), _('项目阶段'), status_label])
    _style_header(ws)
    for r in rows:
        ws.append(r)
    _make_table(ws, table_prefix + '_detail', 10, 1 + len(rows))
    for i, w in enumerate([20, 28, 18, 8, 12, 14, 18, 28, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
