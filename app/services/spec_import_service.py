# -*- coding: utf-8 -*-
"""
规格模板导入导出服务

提供规格模板的Excel导出和导入功能：
- 导出空白模板（含产品分类下拉、所有规格定义）
- 导入Excel创建新规格模板
"""

import io
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
import re
from flask_babel import _

from app import db
from app.models.spec_template import (
    SpecCategory, SpecTemplate, SpecTemplateItem,
    ProductConfiguration, ProductConfigValue, REGIONS, CONFIG_STATUS
)
from app.models.product_code import ProductCategory, ProductSubcategory, ProductCodeField

logger = logging.getLogger(__name__)


class SpecImportService:
    """规格模板导入导出服务"""

    # 样式定义
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="137FEC", end_color="137FEC", fill_type="solid")
    READONLY_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    EDITABLE_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    @classmethod
    def generate_import_template(cls):
        """
        生成空白导入模板 Excel

        Returns:
            bytes: Excel文件的二进制内容
        """
        wb = Workbook()

        # ========== Sheet 1: 模板基本信息 ==========
        ws1 = wb.active
        ws1.title = "模板信息"

        # 获取产品分类和子分类数据
        product_categories = ProductCategory.query.order_by(ProductCategory.display_order).all()
        subcategories = ProductSubcategory.query.order_by(ProductSubcategory.display_order).all()

        # 构建分类名称列表
        category_names = [cat.name for cat in product_categories]

        # 按分类组织子分类
        subcategories_by_category = {}
        for subcat in subcategories:
            if subcat.category_id not in subcategories_by_category:
                subcategories_by_category[subcat.category_id] = []
            subcategories_by_category[subcat.category_id].append(subcat.name)

        # ========== 创建隐藏的数据Sheet（用于分类-子分类联动） ==========
        ws_data = wb.create_sheet(title="_数据")

        # 写入分类列表到第一列
        ws_data['A1'] = "分类列表"
        for idx, cat_name in enumerate(category_names, start=2):
            ws_data[f'A{idx}'] = cat_name

        # 为每个分类创建命名范围（子分类列表）
        # 将子分类写入后续列，每个分类一列
        col_idx = 2
        category_to_range = {}  # 分类名 -> 命名范围名

        for cat in product_categories:
            subcat_names = subcategories_by_category.get(cat.id, [])
            if subcat_names:
                col_letter = get_column_letter(col_idx)
                # 写入分类名称作为列标题
                ws_data[f'{col_letter}1'] = cat.name
                # 写入子分类
                for row_idx, subcat_name in enumerate(subcat_names, start=2):
                    ws_data[f'{col_letter}{row_idx}'] = subcat_name

                # 创建命名范围（将分类名中的特殊字符替换为下划线）
                safe_name = re.sub(r'[^\w]', '_', cat.name)
                range_name = f'子分类_{safe_name}'
                range_ref = f'_数据!${col_letter}$2:${col_letter}${len(subcat_names) + 1}'

                try:
                    from openpyxl.workbook.defined_name import DefinedName
                    defn = DefinedName(range_name, attr_text=range_ref)
                    wb.defined_names[range_name] = defn
                    category_to_range[cat.name] = range_name
                except Exception as e:
                    logger.warning(f"创建命名范围失败: {e}")

                col_idx += 1

        # 隐藏数据Sheet
        ws_data.sheet_state = 'hidden'

        # 定义模板信息字段（去掉模板名称和版本号）
        info_fields = [
            ("型号名称", "", True, "必填，产品型号（如：ABC-2024）"),
            ("产品分类", "", True, "必填，从下拉列表选择"),
            ("产品子分类", "", True, "必填，先选择产品分类，再选择子分类"),
            ("备注", "", False, "可选，模板描述信息"),
        ]

        # 写入表头
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 8
        ws1.column_dimensions['D'].width = 40

        ws1['A1'] = "字段"
        ws1['B1'] = "值"
        ws1['C1'] = "必填"
        ws1['D1'] = "说明"

        # 定义保护样式
        locked_protection = Protection(locked=True)
        unlocked_protection = Protection(locked=False)

        for col in ['A', 'B', 'C', 'D']:
            ws1[f'{col}1'].font = cls.HEADER_FONT
            ws1[f'{col}1'].fill = cls.HEADER_FILL
            ws1[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
            ws1[f'{col}1'].border = cls.THIN_BORDER
            ws1[f'{col}1'].protection = locked_protection

        # 写入字段
        for idx, (field_name, default_value, required, description) in enumerate(info_fields, start=2):
            ws1[f'A{idx}'] = field_name
            ws1[f'B{idx}'] = default_value
            ws1[f'C{idx}'] = "是" if required else ""
            ws1[f'D{idx}'] = description

            # 样式和保护
            ws1[f'A{idx}'].fill = cls.READONLY_FILL
            ws1[f'A{idx}'].protection = locked_protection
            ws1[f'B{idx}'].fill = cls.EDITABLE_FILL
            ws1[f'B{idx}'].protection = unlocked_protection  # 可编辑
            ws1[f'C{idx}'].fill = cls.READONLY_FILL
            ws1[f'C{idx}'].protection = locked_protection
            ws1[f'D{idx}'].fill = cls.READONLY_FILL
            ws1[f'D{idx}'].protection = locked_protection

            for col in ['A', 'B', 'C', 'D']:
                ws1[f'{col}{idx}'].border = cls.THIN_BORDER
                ws1[f'{col}{idx}'].alignment = Alignment(vertical='center')

        # 添加产品分类下拉验证（B3）
        if category_names:
            dv_category = DataValidation(
                type="list",
                formula1=f'"{",".join(category_names)}"',
                allow_blank=True
            )
            dv_category.error = "请从列表中选择产品分类"
            dv_category.errorTitle = "无效选择"
            ws1.add_data_validation(dv_category)
            dv_category.add(ws1['B3'])  # 产品分类行

        # 添加子分类联动下拉验证（B4）
        # 使用 INDIRECT 函数实现联动
        if category_to_range:
            # 构建 INDIRECT 公式
            # 格式: INDIRECT("子分类_" & SUBSTITUTE(B3, " ", "_"))
            dv_subcategory = DataValidation(
                type="list",
                formula1='INDIRECT("子分类_"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(B3," ","_"),"/","_"),"-","_"))',
                allow_blank=True
            )
            dv_subcategory.error = "请先选择产品分类，再选择子分类"
            dv_subcategory.errorTitle = "无效选择"
            ws1.add_data_validation(dv_subcategory)
            dv_subcategory.add(ws1['B4'])  # 产品子分类行

        # 启用工作表保护（只允许编辑未锁定的单元格）
        ws1.protection = SheetProtection(
            sheet=True,
            objects=True,
            scenarios=True,
            selectLockedCells=False,  # 允许选择锁定单元格
            selectUnlockedCells=False  # 允许选择未锁定单元格
        )

        # ========== Sheet 2: 规格项选择 ==========
        ws2 = wb.create_sheet(title="规格项选择")

        # 获取所有规格分类和定义（统一主表）
        from app.models.product_code import SpecificationDictionary
        spec_categories = SpecCategory.query.filter_by(is_active=True).order_by(SpecCategory.display_order).all()
        spec_definitions = SpecificationDictionary.query.filter_by(is_active=True).order_by(
            SpecificationDictionary.category_id, SpecificationDictionary.display_order
        ).all()

        # 构建分类名称映射
        category_id_to_name = {cat.id: cat.name for cat in spec_categories}

        # 定义列（简化版：去掉英文名、值类型、测试条件、测试方法，增加必填）
        # 列顺序：选择、规格ID、分类、规格名称、单位、通用值、必填、用于编码
        spec_headers = [
            ("选择", 6, True, "打√选择此规格"),
            ("规格ID", 10, False, "系统ID（只读）"),
            ("分类", 15, False, "规格分类（只读）"),
            ("规格名称", 30, False, "规格名称（只读）"),
            ("单位", 10, False, "单位（只读）"),
            ("通用值", 25, True, "填写该型号的通用规格值"),
            ("必填", 6, True, "打√表示必填"),
            ("用于编码", 8, True, "打√参与编码"),
        ]

        # 定义保护样式
        locked_protection = Protection(locked=True)
        unlocked_protection = Protection(locked=False)

        # 写入表头
        for col_idx, (header, width, editable, description) in enumerate(spec_headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws2.column_dimensions[col_letter].width = width
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cls.THIN_BORDER
            cell.protection = locked_protection

        # 添加 √/空 下拉验证（用于选择、必填、用于编码列）- 使用打勾符号更直观
        dv_checkbox = DataValidation(type="list", formula1='"√,"', allow_blank=True)
        dv_checkbox.error = "请选择 √ 或留空"
        dv_checkbox.errorTitle = "无效选择"
        ws2.add_data_validation(dv_checkbox)

        # 写入规格定义数据
        # 列顺序：选择(1)、规格ID(2)、分类(3)、规格名称(4)、单位(5)、通用值(6)、必填(7)、用于编码(8)
        row_idx = 2
        for definition in spec_definitions:
            category_name = category_id_to_name.get(definition.category_id, "")

            # 写入数据
            ws2.cell(row=row_idx, column=1, value="")  # 选择
            ws2.cell(row=row_idx, column=2, value=definition.id)  # 规格ID
            ws2.cell(row=row_idx, column=3, value=category_name)  # 分类
            ws2.cell(row=row_idx, column=4, value=definition.name)  # 规格名称
            ws2.cell(row=row_idx, column=5, value=definition.unit or "")  # 单位
            ws2.cell(row=row_idx, column=6, value="")  # 通用值
            ws2.cell(row=row_idx, column=7, value="")  # 必填
            ws2.cell(row=row_idx, column=8, value="")  # 用于编码

            # 应用样式和保护
            for col_idx, (_, _, editable, _) in enumerate(spec_headers, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.border = cls.THIN_BORDER
                if editable:
                    cell.fill = cls.EDITABLE_FILL
                    cell.protection = unlocked_protection  # 可编辑
                else:
                    cell.fill = cls.READONLY_FILL
                    cell.protection = locked_protection  # 只读

            # 添加下拉验证并居中对齐（模拟复选框效果）
            for col in [1, 7, 8]:  # 选择、必填、用于编码列
                cell = ws2.cell(row=row_idx, column=col)
                dv_checkbox.add(cell)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

        # 冻结首行
        ws2.freeze_panes = 'A2'

        # 启用工作表保护（只允许编辑未锁定的单元格）
        ws2.protection = SheetProtection(
            sheet=True,
            objects=True,
            scenarios=True,
            selectLockedCells=False,  # 允许选择锁定单元格
            selectUnlockedCells=False  # 允许选择未锁定单元格
        )

        # 导出为二进制
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def import_template(cls, file, user_id):
        """
        导入Excel创建规格模板

        Args:
            file: 上传的文件对象
            user_id: 当前用户ID

        Returns:
            dict: {
                'success': bool,
                'message': str,
                'template_id': int (成功时),
                'errors': list (失败时),
                'warnings': list
            }
        """
        errors = []
        warnings = []

        try:
            # 加载Excel文件
            wb = load_workbook(file)

            # 验证Sheet存在
            if "模板信息" not in wb.sheetnames:
                return {'success': False, 'message': _('缺少"模板信息"工作表'), 'errors': []}
            if "规格项选择" not in wb.sheetnames:
                return {'success': False, 'message': _('缺少"规格项选择"工作表'), 'errors': []}

            ws1 = wb["模板信息"]
            ws2 = wb["规格项选择"]

            # ========== 读取模板基本信息 ==========
            template_info = cls._read_template_info(ws1)

            if not template_info.get('model'):
                errors.append(_('型号名称不能为空'))
            if not template_info.get('category_name'):
                errors.append(_('产品分类不能为空'))
            if not template_info.get('subcategory_name'):
                errors.append(_('产品子分类不能为空'))

            if errors:
                return {'success': False, 'message': _('模板信息验证失败'), 'errors': errors}

            # 查找产品分类ID
            category = ProductCategory.query.filter_by(name=template_info['category_name']).first()
            if not category:
                errors.append(_('未找到产品分类: %(name)s', name=template_info['category_name']))
                return {'success': False, 'message': _('产品分类不存在'), 'errors': errors}

            # 查找产品子分类ID
            subcategory = ProductSubcategory.query.filter_by(
                name=template_info['subcategory_name'],
                category_id=category.id
            ).first()
            if not subcategory:
                # 尝试不限制分类查找
                subcategory = ProductSubcategory.query.filter_by(name=template_info['subcategory_name']).first()
                if not subcategory:
                    errors.append(_('未找到产品子分类: %(name)s', name=template_info['subcategory_name']))
                    return {'success': False, 'message': _('产品子分类不存在'), 'errors': errors}
                else:
                    warnings.append(_('子分类 "%(subcat)s" 不属于分类 "%(cat)s"，已自动关联',
                                      subcat=template_info['subcategory_name'],
                                      cat=template_info['category_name']))

            # 检查模板是否已存在
            existing = SpecTemplate.query.filter_by(model=template_info['model']).first()
            if existing:
                return {
                    'success': False,
                    'message': _('型号 "%(model)s" 的规格模板已存在', model=template_info['model']),
                    'errors': [_('模板已存在，如需修改请在网页上编辑')]
                }

            # ========== 读取规格项 ==========
            spec_items, spec_errors, spec_warnings = cls._read_spec_items(ws2)
            errors.extend(spec_errors)
            warnings.extend(spec_warnings)

            if errors:
                return {'success': False, 'message': _('规格项验证失败'), 'errors': errors, 'warnings': warnings}

            # 筛选选中的规格项
            selected_items = [item for item in spec_items if item.get('selected')]

            if not selected_items:
                return {
                    'success': False,
                    'message': _('请至少选择一个规格项'),
                    'errors': [_('在"规格项选择"工作表的"选择"列填写Y来选择规格项')],
                    'warnings': warnings
                }

            # ========== 创建模板 ==========
            template = SpecTemplate(
                model=template_info['model'],
                name=template_info.get('name'),
                description=template_info.get('description'),
                version=template_info.get('version', 'V1.0'),
                category_id=category.id,
                subcategory_id=subcategory.id,
                created_by=user_id
            )
            db.session.add(template)
            db.session.flush()  # 获取ID

            # 创建规格项
            for idx, item in enumerate(selected_items):
                dict_id = item['definition_id']
                template_item = SpecTemplateItem(
                    template_id=template.id,
                    spec_dict_id=dict_id,
                    definition_id=None,
                    general_value=item.get('general_value'),
                    is_required=item.get('is_required', False),
                    display_order=idx,
                    use_in_code=item.get('use_in_code', False)
                )
                db.session.add(template_item)

            db.session.commit()

            return {
                'success': True,
                'message': _('规格模板创建成功'),
                'template_id': template.id,
                'template_model': template.model,
                'item_count': len(selected_items),
                'warnings': warnings
            }

        except Exception as e:
            db.session.rollback()
            logger.error(f"导入规格模板失败: {str(e)}")
            return {
                'success': False,
                'message': _('导入失败: %(error)s', error=str(e)),
                'errors': [str(e)]
            }

    @classmethod
    def _read_template_info(cls, ws):
        """读取模板基本信息"""
        info = {}

        # 读取各行数据（第2列是值）
        # 新结构：型号名称、产品分类、产品子分类、备注
        field_mapping = {
            '型号名称': 'model',
            '产品分类': 'category_name',
            '产品子分类': 'subcategory_name',
            '备注': 'description',
        }

        for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=2):
            field_name = row[0].value
            field_value = row[1].value

            if field_name and field_name in field_mapping:
                key = field_mapping[field_name]
                info[key] = str(field_value).strip() if field_value else None

        # 设置默认版本号
        info['version'] = 'V1.0'

        return info

    @classmethod
    def _read_spec_items(cls, ws):
        """读取规格项数据"""
        items = []
        errors = []
        warnings = []

        def is_checked(value):
            """检查单元格值是否为"选中"状态（支持 √ 和 Y）"""
            if not value:
                return False
            val = str(value).strip().upper()
            return val in ('Y', '√', '✓', '✔')

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 新列顺序：选择(0), 规格ID(1), 分类(2), 规格名称(3), 单位(4), 通用值(5), 必填(6), 用于编码(7)
            select_value = row[0].value
            spec_id = row[1].value
            spec_name = row[3].value
            general_value = row[5].value
            is_required_value = row[6].value
            use_in_code = row[7].value

            # 检查是否选中（支持 √ 和 Y）
            is_selected = is_checked(select_value)

            if not is_selected:
                continue

            # 验证规格ID
            if not spec_id:
                errors.append(_('第%(row)d行: 规格ID为空', row=row_idx))
                continue

            # 查找规格定义（统一主表）
            from app.models.product_code import SpecificationDictionary
            try:
                definition = SpecificationDictionary.query.get(int(spec_id))
            except (ValueError, TypeError):
                errors.append(_('第%(row)d行: 规格ID格式错误', row=row_idx))
                continue

            if not definition:
                errors.append(_('第%(row)d行: 未找到规格ID %(id)s', row=row_idx, id=spec_id))
                continue

            # 验证值类型
            if definition.value_type == 'number' and general_value:
                try:
                    float(general_value)
                except (ValueError, TypeError):
                    warnings.append(_('第%(row)d行: 通用值"%(value)s"不是有效数字',
                                      row=row_idx, value=general_value))

            # 检查必填和用于编码（支持 √ 和 Y）
            is_required = is_checked(is_required_value)
            use_code = is_checked(use_in_code)

            items.append({
                'definition_id': definition.id,
                'selected': True,
                'general_value': str(general_value).strip() if general_value else None,
                'is_required': is_required,
                'use_in_code': use_code,
            })

        return items, errors, warnings

    # ========== Phase 2: 配置导出导入 ==========

    @classmethod
    def export_configurations(cls, template_id):
        """
        导出模板的未锁定配置（单Sheet矩阵）

        Args:
            template_id: 规格模板ID

        Returns:
            bytes: Excel文件二进制内容

        Raises:
            ValueError: 模板不存在或无可导出配置
        """
        template = SpecTemplate.query.get(template_id)
        if not template:
            raise ValueError(_('模板不存在'))

        # 筛选未锁定的配置（排除 production 和 discontinued 状态）
        unlocked_configs = [
            c for c in template.configurations
            if c.is_active and not c.mn_locked
        ]

        if not unlocked_configs:
            raise ValueError(_('没有可导出的未锁定配置'))

        wb = Workbook()
        ws = wb.active
        ws.title = "配置矩阵"

        # 写入配置矩阵
        cls._write_config_matrix_sheet(ws, template, unlocked_configs)

        # 导出为二进制
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 预留的空白配置列数量
    EMPTY_CONFIG_COLUMNS = 5

    @classmethod
    def _write_config_matrix_sheet(cls, ws, template, configs):
        """
        写入配置矩阵Sheet

        结构：
        - 行1: 模板型号
        - 行2-5: 配置信息（编码、区域、状态、MN编码）
        - 行6: 表头（规格分类、规格名称、单位、通用值、各配置值）
        - 行7+: 规格值数据

        特点：
        - 已有配置的编码不可修改（锁定）
        - 区域和状态下拉显示中文名称
        - 预留空白列用于新增配置
        """
        locked_protection = Protection(locked=True)
        unlocked_protection = Protection(locked=False)

        # 区域选项 - 从 ProductCodeField 获取（与页面保持一致）
        sales_regions = db.session.query(
            ProductCodeField.code, ProductCodeField.name
        ).filter_by(field_type='origin_location').distinct().all()
        # 显示格式：名称 (代码)，如 "亚太 (AP)"
        region_display_list = [f'{r.name} ({r.code})' for r in sales_regions]
        region_code_to_display = {r.code: f'{r.name} ({r.code})' for r in sales_regions}
        region_display_to_code = {f'{r.name} ({r.code})': r.code for r in sales_regions}

        # 状态选项 - 显示中文名称，只允许非锁定状态
        editable_status_items = [(s[0], s[1]) for s in CONFIG_STATUS if s[0] not in ('production', 'discontinued')]
        status_names = [s[1] for s in editable_status_items]  # 显示名称列表
        status_code_to_name = {s[0]: s[1] for s in CONFIG_STATUS}
        status_name_to_code = {s[1]: s[0] for s in CONFIG_STATUS}

        # 获取模板规格项（按分类和显示顺序排序）
        items = sorted(template.items, key=lambda x: (
            x.spec_dict.category.display_order if x.spec_dict and x.spec_dict.category else 999,
            x.display_order
        ))

        # 构建配置值映射 {config_id: {item_id: value}}
        config_values_map = {}
        for config in configs:
            config_values_map[config.id] = {}
            for cv in config.config_values:
                if cv.template_item_id:
                    config_values_map[config.id][cv.template_item_id] = cv.value

        # ========== 行1: 模板型号 ==========
        ws.cell(row=1, column=1, value="模板型号").fill = cls.READONLY_FILL
        ws.cell(row=1, column=1).protection = locked_protection
        ws.cell(row=1, column=1).border = cls.THIN_BORDER
        ws.cell(row=1, column=2, value=template.model).fill = cls.READONLY_FILL
        ws.cell(row=1, column=2).protection = locked_protection
        ws.cell(row=1, column=2).border = cls.THIN_BORDER

        # ========== 行2-5: 配置信息区（去掉区域名称行） ==========
        info_rows = [
            ("配置编码", 2),
            ("区域", 3),
            ("状态", 4),
            ("MN编码", 5),  # 只读
        ]

        for label, row_num in info_rows:
            ws.cell(row=row_num, column=1, value=label).fill = cls.READONLY_FILL
            ws.cell(row=row_num, column=1).protection = locked_protection
            ws.cell(row=row_num, column=1).border = cls.THIN_BORDER

        # ========== 行6: 表头 ==========
        headers = ["规格分类", "规格名称", "单位", "通用值"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cls.THIN_BORDER
            cell.protection = locked_protection

        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 15

        # ========== 写入各配置列 ==========
        config_start_col = 5  # E列开始

        # 区域下拉验证（显示：名称 (代码)）
        dv_region = DataValidation(
            type="list",
            formula1=f'"{",".join(region_display_list)}"',
            allow_blank=True
        )
        ws.add_data_validation(dv_region)

        # 状态下拉验证（显示中文名称）
        dv_status = DataValidation(
            type="list",
            formula1=f'"{",".join(status_names)}"',
            allow_blank=True
        )
        ws.add_data_validation(dv_status)

        # 写入已有配置
        for config_idx, config in enumerate(configs):
            col = config_start_col + config_idx
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 18

            # 表头 - 配置编号作为列标题
            header_cell = ws.cell(row=6, column=col, value=f"配置{config_idx + 1}")
            header_cell.font = cls.HEADER_FONT
            header_cell.fill = cls.HEADER_FILL
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = cls.THIN_BORDER
            header_cell.protection = locked_protection

            # 行2: 配置编码（已有配置不可修改）
            cell = ws.cell(row=2, column=col, value=config.config_code)
            cell.fill = cls.READONLY_FILL  # 已有配置编码锁定
            cell.protection = locked_protection
            cell.border = cls.THIN_BORDER

            # 行3: 区域（显示：名称 (代码)）
            region_display = region_code_to_display.get(config.region, config.region) if config.region else ""
            cell = ws.cell(row=3, column=col, value=region_display)
            cell.fill = cls.EDITABLE_FILL
            cell.protection = unlocked_protection
            cell.border = cls.THIN_BORDER
            dv_region.add(cell)

            # 行4: 状态（显示中文名称）
            status_display = status_code_to_name.get(config.status, config.status) if config.status else "研发中"
            cell = ws.cell(row=4, column=col, value=status_display)
            cell.fill = cls.EDITABLE_FILL
            cell.protection = unlocked_protection
            cell.border = cls.THIN_BORDER
            dv_status.add(cell)

            # 行5: MN编码（只读）
            cell = ws.cell(row=5, column=col, value=config.mn_code or "")
            cell.fill = cls.READONLY_FILL
            cell.protection = locked_protection
            cell.border = cls.THIN_BORDER

        # ========== 预留空白配置列 ==========
        empty_start_col = config_start_col + len(configs)
        for empty_idx in range(cls.EMPTY_CONFIG_COLUMNS):
            col = empty_start_col + empty_idx
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 18

            # 表头
            header_cell = ws.cell(row=6, column=col, value=f"新配置{empty_idx + 1}")
            header_cell.font = cls.HEADER_FONT
            header_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # 绿色标识新配置
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = cls.THIN_BORDER
            header_cell.protection = locked_protection

            # 行2: 配置编码（可编辑）
            cell = ws.cell(row=2, column=col, value="")
            cell.fill = cls.EDITABLE_FILL
            cell.protection = unlocked_protection
            cell.border = cls.THIN_BORDER

            # 行3: 区域（可编辑，带下拉）
            cell = ws.cell(row=3, column=col, value="")
            cell.fill = cls.EDITABLE_FILL
            cell.protection = unlocked_protection
            cell.border = cls.THIN_BORDER
            dv_region.add(cell)

            # 行4: 状态（可编辑，带下拉）
            cell = ws.cell(row=4, column=col, value="")
            cell.fill = cls.EDITABLE_FILL
            cell.protection = unlocked_protection
            cell.border = cls.THIN_BORDER
            dv_status.add(cell)

            # 行5: MN编码（只读，空）
            cell = ws.cell(row=5, column=col, value="")
            cell.fill = cls.READONLY_FILL
            cell.protection = locked_protection
            cell.border = cls.THIN_BORDER

        # ========== 行7+: 规格值数据 ==========
        current_category_id = None
        total_config_cols = len(configs) + cls.EMPTY_CONFIG_COLUMNS

        for item_idx, item in enumerate(items):
            row = 7 + item_idx
            definition = item.spec_dict
            if not definition:
                continue

            category = definition.category

            # A列: 规格分类（合并同分类的单元格）
            if category and category.id != current_category_id:
                ws.cell(row=row, column=1, value=category.name if category else "")
                current_category_id = category.id if category else None
            ws.cell(row=row, column=1).fill = cls.READONLY_FILL
            ws.cell(row=row, column=1).protection = locked_protection
            ws.cell(row=row, column=1).border = cls.THIN_BORDER

            # B列: 规格名称
            ws.cell(row=row, column=2, value=definition.name).fill = cls.READONLY_FILL
            ws.cell(row=row, column=2).protection = locked_protection
            ws.cell(row=row, column=2).border = cls.THIN_BORDER

            # C列: 单位
            ws.cell(row=row, column=3, value=definition.unit or "").fill = cls.READONLY_FILL
            ws.cell(row=row, column=3).protection = locked_protection
            ws.cell(row=row, column=3).border = cls.THIN_BORDER

            # D列: 通用值
            ws.cell(row=row, column=4, value=item.general_value or "").fill = cls.READONLY_FILL
            ws.cell(row=row, column=4).protection = locked_protection
            ws.cell(row=row, column=4).border = cls.THIN_BORDER

            # E列起: 各配置的规格值（已有配置）
            for config_idx, config in enumerate(configs):
                col = config_start_col + config_idx
                # 获取配置值，如果没有则使用通用值
                config_value = config_values_map.get(config.id, {}).get(item.id)
                display_value = config_value if config_value is not None else (item.general_value or "")

                cell = ws.cell(row=row, column=col, value=display_value)
                cell.fill = cls.EDITABLE_FILL
                cell.protection = unlocked_protection
                cell.border = cls.THIN_BORDER

            # 预留空白列的规格值
            for empty_idx in range(cls.EMPTY_CONFIG_COLUMNS):
                col = empty_start_col + empty_idx
                cell = ws.cell(row=row, column=col, value="")
                cell.fill = cls.EDITABLE_FILL
                cell.protection = unlocked_protection
                cell.border = cls.THIN_BORDER

        # 冻结首6行和前4列
        ws.freeze_panes = 'E7'

        # 启用工作表保护
        ws.protection = SheetProtection(
            sheet=True,
            objects=True,
            scenarios=True,
            selectLockedCells=False,
            selectUnlockedCells=False
        )

    @classmethod
    def import_configurations(cls, file, template_id, user_id):
        """
        导入配置

        Args:
            file: 上传的Excel文件
            template_id: 规格模板ID
            user_id: 当前用户ID

        Returns:
            dict: {
                'success': bool,
                'message': str,
                'created_count': int,
                'updated_count': int,
                'skipped_count': int,
                'warnings': list,
                'errors': list
            }
        """
        template = SpecTemplate.query.get(template_id)
        if not template:
            return {'success': False, 'message': _('模板不存在'), 'errors': []}

        errors = []
        warnings = []

        try:
            wb = load_workbook(file)

            # 查找配置矩阵Sheet
            if "配置矩阵" not in wb.sheetnames:
                return {'success': False, 'message': _('缺少"配置矩阵"工作表'), 'errors': []}

            ws = wb["配置矩阵"]

            # 验证模板型号
            model_in_file = ws.cell(row=1, column=2).value
            if model_in_file and str(model_in_file).strip() != template.model:
                return {
                    'success': False,
                    'message': _('Excel中的模板型号与当前模板不匹配'),
                    'errors': [_('文件中型号: %(file_model)s, 当前模板: %(template_model)s',
                                file_model=model_in_file, template_model=template.model)]
                }

            # 解析配置矩阵
            configs_data, parse_errors, parse_warnings = cls._parse_config_matrix(ws, template)
            errors.extend(parse_errors)
            warnings.extend(parse_warnings)

            if errors:
                return {'success': False, 'message': _('配置数据验证失败'), 'errors': errors, 'warnings': warnings}

            if not configs_data:
                return {'success': False, 'message': _('未找到有效的配置数据'), 'errors': [], 'warnings': warnings}

            # 处理每个配置
            created, updated, skipped = 0, 0, 0

            for config_data in configs_data:
                config_code = config_data.get('config_code')
                if not config_code:
                    warnings.append(_('跳过空配置编码的列'))
                    continue

                # 查找现有配置
                existing = ProductConfiguration.query.filter_by(
                    template_id=template_id,
                    config_code=config_code
                ).first()

                if existing:
                    if existing.mn_locked:
                        skipped += 1
                        warnings.append(_('配置 %(code)s 已锁定，跳过更新', code=config_code))
                        continue

                    # 更新配置
                    cls._update_configuration(existing, config_data)
                    updated += 1
                else:
                    # 创建新配置
                    cls._create_configuration(template, config_data, user_id)
                    created += 1

            db.session.commit()

            return {
                'success': True,
                'message': _('导入完成：新建%(created)d个，更新%(updated)d个，跳过%(skipped)d个',
                            created=created, updated=updated, skipped=skipped),
                'created_count': created,
                'updated_count': updated,
                'skipped_count': skipped,
                'warnings': warnings
            }

        except Exception as e:
            db.session.rollback()
            logger.error(f"导入配置失败: {str(e)}")
            return {
                'success': False,
                'message': _('导入失败: %(error)s', error=str(e)),
                'errors': [str(e)]
            }

    @classmethod
    def _parse_config_matrix(cls, ws, template):
        """
        解析配置矩阵

        结构（更新后）：
        - 行1: 模板型号
        - 行2: 配置编码
        - 行3: 区域（中文名称）
        - 行4: 状态（中文名称）
        - 行5: MN编码
        - 行6: 表头
        - 行7+: 规格值

        Returns:
            tuple: (configs_data, errors, warnings)
        """
        configs_data = []
        errors = []
        warnings = []

        # 区域选项 - 从 ProductCodeField 获取（与导出保持一致）
        sales_regions = db.session.query(
            ProductCodeField.code, ProductCodeField.name
        ).filter_by(field_type='origin_location').distinct().all()
        # 构建显示格式到代码的映射，如 "亚太 (AP)" -> "AP"
        region_display_to_code = {f'{r.name} ({r.code})': r.code for r in sales_regions}
        # 也支持直接的代码
        valid_region_codes = [r.code for r in sales_regions]

        # 状态名称到代码的映射
        status_name_to_code = {s[1]: s[0] for s in CONFIG_STATUS}

        # 构建规格项映射 {name: item}
        item_by_name = {}
        for item in template.items:
            if item.spec_dict:
                item_by_name[item.spec_dict.name] = item

        # 读取配置列（从E列开始，即第5列）
        config_start_col = 5
        max_col = ws.max_column

        for col in range(config_start_col, max_col + 1):
            # 读取配置信息
            config_code = ws.cell(row=2, column=col).value
            if not config_code:
                continue  # 跳过空列

            config_code = str(config_code).strip()

            # 行3: 区域（可能是 "名称 (代码)" 格式或直接代码）
            region_value = ws.cell(row=3, column=col).value
            region_value = str(region_value).strip() if region_value else None
            if region_value:
                # 先尝试从显示格式映射，如 "亚太 (AP)" -> "AP"
                region = region_display_to_code.get(region_value)
                if not region:
                    # 如果不是显示格式，检查是否是有效的代码
                    if region_value in valid_region_codes:
                        region = region_value
                    else:
                        warnings.append(_('配置 %(code)s 的区域 "%(region)s" 无效',
                                         code=config_code, region=region_value))
                        region = None
            else:
                region = None

            # 行4: 状态（可能是中文名称，需要映射回代码）
            status_value = ws.cell(row=4, column=col).value
            status_value = str(status_value).strip() if status_value else None
            # 尝试从名称映射到代码，如果已经是代码则保持不变
            if status_value:
                status = status_name_to_code.get(status_value, status_value)
            else:
                status = 'development'

            # 验证状态值
            valid_statuses = [s[0] for s in CONFIG_STATUS]
            if status not in valid_statuses:
                warnings.append(_('配置 %(code)s 的状态 "%(status)s" 无效，使用默认值 development',
                                 code=config_code, status=status_value))
                status = 'development'

            # 读取规格值（从第7行开始，因为去掉了区域名称行）
            values = []
            for row in range(7, ws.max_row + 1):
                spec_name = ws.cell(row=row, column=2).value  # B列是规格名称
                if not spec_name:
                    continue

                spec_name = str(spec_name).strip()
                item = item_by_name.get(spec_name)
                if not item:
                    continue  # 跳过未知规格

                value = ws.cell(row=row, column=col).value
                value_str = str(value).strip() if value is not None else None

                # 如果值与通用值相同，不需要保存
                if value_str and value_str != (item.general_value or ""):
                    values.append({
                        'template_item_id': item.id,
                        'value': value_str
                    })

            configs_data.append({
                'config_code': config_code,
                'region': region,
                'status': status,
                'values': values
            })

        return configs_data, errors, warnings

    @classmethod
    def _update_configuration(cls, config, data):
        """更新配置"""
        # 更新基本信息（区域名称从代码自动获取，不再单独存储）
        if data.get('region'):
            config.region = data['region']
        if data.get('status'):
            config.status = data['status']

        # 更新配置值
        existing_values = {cv.template_item_id: cv for cv in config.config_values}

        for value_data in data.get('values', []):
            item_id = value_data['template_item_id']
            new_value = value_data['value']

            if item_id in existing_values:
                # 更新现有值
                existing_values[item_id].value = new_value
            else:
                # 创建新值
                cv = ProductConfigValue(
                    configuration_id=config.id,
                    template_item_id=item_id,
                    value=new_value
                )
                db.session.add(cv)

        config.updated_at = datetime.utcnow()

    @classmethod
    def _create_configuration(cls, template, data, user_id):
        """创建新配置"""
        config = ProductConfiguration(
            template_id=template.id,
            config_code=data['config_code'],
            region=data.get('region'),
            status=data.get('status', 'development'),
            created_by=user_id
        )
        db.session.add(config)
        db.session.flush()  # 获取ID

        # 创建配置值
        for value_data in data.get('values', []):
            cv = ProductConfigValue(
                configuration_id=config.id,
                template_item_id=value_data['template_item_id'],
                value=value_data['value']
            )
            db.session.add(cv)

        return config
