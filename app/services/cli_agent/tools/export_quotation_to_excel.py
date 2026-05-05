# -*- coding: utf-8 -*-
"""
export_quotation_to_excel 工具：按模板格式生成报价单 Excel。

直接在 Flask 上下文内查询数据库，使用 openpyxl 生成 .xlsx，
样式严格对照「销售报价单表模版.xlsx」。
"""
from __future__ import annotations

import logging
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from app.services.cli_agent.tools import BaseTool

logger = logging.getLogger(__name__)

_STORAGE_DIR = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', '..', '..', 'storage', 'exports'
))

# ── 大写金额转换 ─────────────────────────────────────────────────────────────

_CN_DIGITS = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
_CN_UPPER  = ['', '万', '亿']


def _num_to_chinese(amount: float) -> str:
    """将金额（元）转换为人民币大写，如 12345.60 → 壹万贰仟叁佰肆拾伍元陆角整"""
    if not amount or amount == 0:
        return '零元整'
    negative = amount < 0
    amount = abs(amount)
    int_part = int(amount)
    dec = round((amount - int_part) * 100)
    jiao, fen = dec // 10, dec % 10

    def _group(n: int) -> str:
        if n == 0:
            return ''
        q, r3 = divmod(n, 1000)
        r2, r1 = divmod(r3, 100)
        r1d, r1u = divmod(r1, 10)
        s = ''
        if q:   s += _CN_DIGITS[q] + '仟'
        if r2:  s += _CN_DIGITS[r2] + '佰'
        elif q and (r1d or r1u): s += '零'
        if r1d: s += _CN_DIGITS[r1d] + '拾'
        elif r2 and r1u: s += '零'
        if r1u: s += _CN_DIGITS[r1u]
        return s

    groups, n = [], int_part
    while n > 0:
        groups.append(n % 10000)
        n //= 10000
    if not groups:
        groups = [0]

    result = ''
    for i, g in enumerate(reversed(groups)):
        seg = _group(g)
        idx = len(groups) - 1 - i
        if seg:
            result += seg + _CN_UPPER[idx]
        elif result and not result.endswith('零'):
            result += '零'
    result = result.rstrip('零') or '零'

    result = ('负' if negative else '') + result + '元'
    if jiao == 0 and fen == 0:
        result += '整'
    elif jiao == 0:
        result += '零' + _CN_DIGITS[fen] + '分'
    elif fen == 0:
        result += _CN_DIGITS[jiao] + '角整'
    else:
        result += _CN_DIGITS[jiao] + '角' + _CN_DIGITS[fen] + '分'
    return result


# ── Excel 样式常量 ───────────────────────────────────────────────────────────

_HEADER_BG  = 'F3F3F3'   # 表头灰底（来自模板）
_BORDER_CLR = '000000'
_GST_BG     = 'FFF8DC'   # OVS Total after GST 高亮

# Logo 资源（PMA 静态资源）+ 真实比例
_LOGO_DIR = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', '..', 'static', 'img', 'company_logos'
))
_LOGO_OVS  = os.path.join(_LOGO_DIR, 'evertac_solutions.png')  # 1913x519, ratio 3.686
_LOGO_SP8D = os.path.join(_LOGO_DIR, 'evertac_cn.png')         # 1678x568, ratio 2.954
_LOGO_OVS_RATIO  = 3.686
_LOGO_SP8D_RATIO = 2.954

# OVS 海外英文版抬头（3 行平铺，对应网页报价单）
_OVS_HEADER_LINES = [
    'EVERTAC SOLUTIONS SINGAPORE PTE. LTD.',
    '18 Boon Lay Way, #03-117 Tradehub 21, Singapore 609966',
    'UEN No/GST Reg. No.: 202230146C    Website: www.evertac-solutions.com',
]
_SP8D_HEADER_LINES = [
    '和源通信(上海)股份有限公司',
    '上海市普陀区武威路88弄中鑫企业广场 19号楼6层',
    '电话: 021-62596028',
]

# 兼容旧调用（其他模块可能引用）
_SP8D_HEADER = '\n'.join(_SP8D_HEADER_LINES)
_OVS_HEADER  = '\n'.join(_OVS_HEADER_LINES[:2])

# SP8D 列宽（中文版，11 列 B-L）
_COL_WIDTHS = {
    'A': 5.5,   'B': 5.0,   'C': 15.5,  'D': 10.33,
    'E': 34.5,  'F': 8.16,  'G': 6.33,  'H': 5.33,
    'I': 7.5,   'J': 8.0,   'K': 9.83,  'L': 12.66,
}

# OVS 列宽（英文版，9 个数据列 B-J，K 留边距）
_OVS_COL_WIDTHS = {
    'A': 1.5,  'B': 6,   'C': 14, 'D': 16, 'E': 38,
    'F': 7,    'G': 10,  'H': 11, 'I': 6,  'J': 12, 'K': 1.5,
}


def _thin(color=_BORDER_CLR):
    from openpyxl.styles import Border, Side
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill_borders(ws, r1, c1, r2, c2):
    """对范围内每个单元格都设置细边框（修复合并单元格外框缺失问题）。"""
    b = _thin()
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b


def _set_cell(ws, row, col, value, *, font_name='微软雅黑', font_size=8,
              bold=False, color=None, h_align=None, v_align='center',
              wrap=False, fill_color=None, num_format=None):
    from openpyxl.styles import Font, Alignment, PatternFill
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=font_name, size=font_size, bold=bold,
                     color=color or '000000')
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align,
                                wrap_text=wrap)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color,
                                fill_type='solid')
    if num_format:
        cell.number_format = num_format
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _row_h(ws, row, height):
    ws.row_dimensions[row].height = height


# ── 主生成函数 ───────────────────────────────────────────────────────────────

def _build_excel(data: dict, db_type: str) -> bytes:
    """根据 plain dict 数据生成报价单 Excel bytes。

    单一布局，视觉对照网页 tw_quotation_edit.html。
    OVS（海外）/ SP8D（中国）只在 logo / 抬头 / 字体 / 标签三处不同，
    通过 _VARIANT 字典参数化，避免双份代码漂移。

    data 结构:
        quotation_number, project_name, company_name, company_address,
        contact_name, contact_phone, quote_date (str), total_amount (float),
        owner_name, currency, payment_terms, shipping_terms, validity,
        ref_no, gst_percent, remarks, details (list of dict)
    detail dict:
        mn (料号), product_name, product_model, product_desc, brand, unit,
        quantity, discount, market_price, unit_price, total_price
    """
    cfg = _VARIANT.get((db_type or '').upper(), _VARIANT['SP8D'])
    return _build_excel_unified(data, cfg)


# 标签字典：OVS 英文 / SP8D 中文（对应网页 i18n）
_LABELS_OVS = {
    'title':           'QUOTATION',
    'company_name':    'Company Name:',
    'company_address': 'Company Address:',
    'contact_name':    'Contact Person:',
    'contact_phone':   'Contact No.:',
    'project':         'Project:',
    'currency':        'Currency:',
    'quotation_no':    'Quotation No.:',
    'quote_date':      'Quotation Date:',
    'payment_terms':   'Payment Terms:',
    'shipping_terms':  'Shipping Terms:',
    'validity':        'Validity:',
    'ref_no':          'Ref No.:',
    'col_sn':          'S/N',
    'col_item_no':     'Item No.',
    'col_brand':       'Brand',
    'col_desc':        'Description',
    'col_disc':        'Disc%',
    'col_market':      'Market',
    'col_unit_price':  'Unit Price',
    'col_qty':         'Qty',
    'col_amount':      'Amount',
    'subtotal':        'Total before GST',
    'tax':             'GST',
    'grand_total':     'Total after GST',
    'remarks':         'Remarks:',
    'sign':            'Signed and Accepted by Customer:',
}

_LABELS_SP8D = {
    'title':           '销 售 报 价 单',
    'company_name':    '客户名称：',
    'company_address': '客户地址：',
    'contact_name':    '联系人：',
    'contact_phone':   '联系电话：',
    'project':         '项目名称：',
    'currency':        '货币：',
    'quotation_no':    '报价编号：',
    'quote_date':      '报价日期：',
    'payment_terms':   '付款条件：',
    'shipping_terms':  '交付条件：',
    'validity':        '有效期：',
    'ref_no':          '参考编号：',
    'col_sn':          '序号',
    'col_item_no':     '产品编码',
    'col_brand':       '品牌',
    'col_desc':        '型号规格',
    'col_disc':        '折扣%',
    'col_market':      '市场价',
    'col_unit_price':  '单价',
    'col_qty':         '数量',
    'col_amount':      '小计',
    'subtotal':        '不含税合计',
    'tax':             '增值税',
    'grand_total':     '含税合计',
    'remarks':         '备注：',
    'sign':            '签收：',
}

# 版式配置：变量 = logo + 抬头 + 字体 + 标签字典 + 货币格式 + 表名
_VARIANT = {
    'OVS': {
        'logo':           _LOGO_OVS,
        'logo_ratio':     _LOGO_OVS_RATIO,
        'header_lines':   _OVS_HEADER_LINES,
        'labels':         _LABELS_OVS,
        'font':           'Arial',
        'currency_fmt':   '#,##0.00',
        'sheet_title':    'Quotation',
        'default_currency': '',
    },
    'SP8D': {
        'logo':           _LOGO_SP8D,
        'logo_ratio':     _LOGO_SP8D_RATIO,
        'header_lines':   _SP8D_HEADER_LINES,
        'labels':         _LABELS_SP8D,
        'font':           '微软雅黑',
        'currency_fmt':   '¥#,##0.00',
        'sheet_title':    '销售报价单',
        'default_currency': 'CNY',
    },
}


def _build_excel_unified(data: dict, cfg: dict) -> bytes:
    """统一渲染逻辑 — 布局 100% 相同，仅 logo/抬头/标签/字体按 cfg 切换。"""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    L    = cfg['labels']
    FONT = cfg['font']
    FMT  = cfg['currency_fmt']

    wb = Workbook()
    ws = wb.active
    ws.title = cfg['sheet_title']

    # 列宽（与网页 9 数据列布局一致）
    for col, w in _OVS_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Row 1-3: logo + 抬头三行
    for r in (1, 2, 3):
        _row_h(ws, r, 22)
    if os.path.exists(cfg['logo']):
        try:
            img = XLImage(cfg['logo'])
            img.width  = 160
            img.height = int(160 / cfg['logo_ratio'])
            ws.add_image(img, 'B1')
        except Exception as e:
            logger.warning(f'[export_quotation_to_excel] logo 插入失败: {e}')

    head = cfg['header_lines']
    _merge(ws, 1, 4, 1, 10)
    _set_cell(ws, 1, 4, head[0] if len(head) > 0 else '',
              font_name=FONT, font_size=14, bold=True, h_align='left')
    _merge(ws, 2, 4, 2, 10)
    _set_cell(ws, 2, 4, head[1] if len(head) > 1 else '',
              font_name=FONT, font_size=10, h_align='left')
    _merge(ws, 3, 4, 3, 10)
    _set_cell(ws, 3, 4, head[2] if len(head) > 2 else '',
              font_name=FONT, font_size=10, h_align='left')

    # Row 4: 空白
    _row_h(ws, 4, 18)

    # Row 5: 大标题
    _row_h(ws, 5, 38)
    _merge(ws, 5, 2, 5, 10)
    _set_cell(ws, 5, 2, L['title'], font_name=FONT, font_size=22,
              bold=True, h_align='center')

    # Row 6-11: metadata 双列（左右对称，对应网页布局）
    info_left = [
        (L['company_name'],    data.get('company_name', '')),
        (L['company_address'], data.get('company_address', '')),
        (L['contact_name'],    data.get('contact_name', '')),
        (L['contact_phone'],   data.get('contact_phone', '')),
        (L['project'],         data.get('project_name', '')),
        (L['currency'],        data.get('currency') or cfg['default_currency']),
    ]
    info_right = [
        (L['quotation_no'],    data.get('quotation_number', '')),
        (L['quote_date'],      data.get('quote_date', '')),
        (L['payment_terms'],   data.get('payment_terms', '')),
        (L['shipping_terms'],  data.get('shipping_terms', '')),
        (L['validity'],        data.get('validity', '')),
        (L['ref_no'],          data.get('ref_no', '')),
    ]
    info_start = 6
    for i, ((lL, vL), (lR, vR)) in enumerate(zip(info_left, info_right)):
        r = info_start + i
        _row_h(ws, r, 22)
        _merge(ws, r, 2, r, 3)
        _set_cell(ws, r, 2, lL, font_name=FONT, font_size=10, bold=True, h_align='left')
        _merge(ws, r, 4, r, 6)
        _set_cell(ws, r, 4, vL, font_name=FONT, font_size=10, h_align='left')
        _merge(ws, r, 7, r, 8)
        _set_cell(ws, r, 7, lR, font_name=FONT, font_size=10, bold=True, h_align='left')
        _merge(ws, r, 9, r, 10)
        _set_cell(ws, r, 9, vR, font_name=FONT, font_size=10, h_align='left')

    # 表头
    header_row = info_start + len(info_left)
    _row_h(ws, header_row, 28)
    headers = [L['col_sn'], L['col_item_no'], L['col_brand'], L['col_desc'],
               L['col_disc'], L['col_market'], L['col_unit_price'], L['col_qty'],
               L['col_amount']]
    for ci, h in enumerate(headers):
        _set_cell(ws, header_row, ci + 2, h, font_name=FONT, font_size=10,
                  bold=True, h_align='center', fill_color=_HEADER_BG)

    # 明细行
    details = data.get('details', [])
    data_start = header_row + 1
    for i, d in enumerate(details):
        r = data_start + i
        _row_h(ws, r, 80)
        discount = d.get('discount')
        disc_pct = f"{int((discount or 1) * 100)} %" if discount is not None else '100 %'
        unit_price = d.get('unit_price') or 0
        qty        = d.get('quantity') or 0
        amount     = d.get('total_price') or (unit_price * qty)
        market     = d.get('market_price') or unit_price

        _set_cell(ws, r, 2, i + 1, font_name=FONT, font_size=10, h_align='center')
        _set_cell(ws, r, 3, d.get('mn') or '', font_name=FONT, font_size=10, h_align='left')
        _set_cell(ws, r, 4, d.get('brand') or '', font_name=FONT, font_size=10, h_align='left')
        # 型号规格 = product_name + product_model + product_desc
        desc_lines = []
        if d.get('product_name'):  desc_lines.append(d['product_name'])
        if d.get('product_model'): desc_lines.append(d['product_model'])
        if d.get('product_desc'):  desc_lines.append(d['product_desc'])
        _set_cell(ws, r, 5, '\n'.join(desc_lines), font_name=FONT, font_size=10,
                  h_align='left', v_align='top', wrap=True)
        _set_cell(ws, r, 6, disc_pct, font_name=FONT, font_size=10, h_align='center')
        _set_cell(ws, r, 7, market, font_name=FONT, font_size=10, h_align='right',
                  num_format=FMT)
        _set_cell(ws, r, 8, unit_price, font_name=FONT, font_size=10, h_align='right',
                  num_format=FMT)
        _set_cell(ws, r, 9, qty, font_name=FONT, font_size=10, h_align='center')
        _set_cell(ws, r, 10, amount, font_name=FONT, font_size=10, h_align='right',
                  num_format=FMT)

    # 汇总三行：subtotal / tax / grand_total（结构对应网页 colspan=4 + 1）
    sub_row = data_start + len(details)
    subtotal = sum((d.get('total_price') or 0) for d in details)
    tax_pct  = data.get('gst_percent', 0) or 0
    tax_amt  = subtotal * tax_pct / 100
    grand    = subtotal + tax_amt

    _row_h(ws, sub_row, 26)
    _merge(ws, sub_row, 6, sub_row, 9)
    _set_cell(ws, sub_row, 6, L['subtotal'], font_name=FONT, font_size=10,
              bold=True, h_align='right')
    _set_cell(ws, sub_row, 10, subtotal, font_name=FONT, font_size=10,
              bold=True, h_align='right', num_format=FMT)

    tax_row = sub_row + 1
    _row_h(ws, tax_row, 26)
    _merge(ws, tax_row, 6, tax_row, 7)
    _set_cell(ws, tax_row, 6, L['tax'], font_name=FONT, font_size=10,
              bold=True, h_align='right')
    _merge(ws, tax_row, 8, tax_row, 9)
    _set_cell(ws, tax_row, 8, f'{tax_pct} %', font_name=FONT, font_size=10, h_align='right')
    _set_cell(ws, tax_row, 10, tax_amt, font_name=FONT, font_size=10,
              h_align='right', num_format=FMT)

    final_row = tax_row + 1
    _row_h(ws, final_row, 26)
    _merge(ws, final_row, 6, final_row, 9)
    _set_cell(ws, final_row, 6, L['grand_total'], font_name=FONT, font_size=10,
              bold=True, h_align='right', fill_color=_GST_BG)
    _set_cell(ws, final_row, 10, grand, font_name=FONT, font_size=10,
              bold=True, h_align='right', fill_color=_GST_BG, num_format=FMT)

    # Remarks
    remark_row = final_row + 1
    _row_h(ws, remark_row, 60)
    _merge(ws, remark_row, 2, remark_row, 3)
    _set_cell(ws, remark_row, 2, L['remarks'], font_name=FONT, font_size=10,
              bold=True, h_align='left', v_align='top')
    _merge(ws, remark_row, 4, remark_row, 10)
    _set_cell(ws, remark_row, 4, data.get('remarks', '') or '', font_name=FONT,
              font_size=10, h_align='left', v_align='top', wrap=True)

    # 签字栏
    sig_row = remark_row + 2
    _row_h(ws, sig_row, 30)
    _merge(ws, sig_row, 7, sig_row, 10)
    _set_cell(ws, sig_row, 7, L['sign'], font_name=FONT, font_size=10,
              bold=True, h_align='right')

    # 边框
    _fill_borders(ws, header_row, 2, remark_row, 10)

    # 打印
    ws.print_area = f'A1:K{sig_row}'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ── 工具类 ────────────────────────────────────────────────────────────────────

class ExportQuotationToExcelTool(BaseTool):
    name = 'export_quotation_to_excel'
    description = (
        '生成报价单 Excel（.xlsx）。两种入参模式：\n\n'
        '【模式 A — 现有报价单导出】当用户说"导出报价单"、"导出报价单Excel"时\n'
        '  必须先获取真实的 quotation_number 才能调用。\n'
        '  若当前上下文没有从 quotations 表直接读到 quotation_number，请先执行：\n'
        '    query_pma_database → "SELECT id, quotation_number FROM quotations WHERE ..."\n'
        '  再用查出的 quotation_number 调用本工具。\n'
        '  **禁止**根据明细行数据或上下文猜测 quotation_number 或 quotation_id。\n'
        '  参数：quotation_number（优先）或 quotation_id\n\n'
        '【模式 B — 直接渲染 BOM 草稿】skill 推导出方案 BOM 后直接出 Excel\n'
        '  不写入数据库，仅生成方案草稿文件。\n'
        '  参数：bom_data（dict，结构见下）\n\n'
        '返回结果包含 download_url，回复"报价单Excel已生成"即可。'
    )
    input_schema = {
        'type': 'object',
        'properties': {
            'quotation_number': {
                'type': 'string',
                'description': '【模式 A】报价单编号，如 QU202604-036',
            },
            'quotation_id': {
                'type': 'integer',
                'description': '【模式 A】报价单 ID（整数）',
            },
            'bom_data': {
                'type': 'object',
                'description': (
                    '【模式 B】BOM 草稿数据，包含：\n'
                    '  region: "CN" 或 "SG"（决定中文/英文版式 + logo）\n'
                    '  project_name: 项目名 / company_name: 客户公司 / contact_name: 联系人\n'
                    '  contact_phone: 电话 / quote_date: YYYY-MM-DD\n'
                    '  currency: 货币(SGD/USD/CNY) / payment_terms / shipping_terms / validity / ref_no\n'
                    '  gst_percent: GST 百分比(海外用,默认 0) / remarks: 备注\n'
                    '  items: [{mn, product_name, product_model, product_desc,\n'
                    '           brand, unit, quantity, discount, market_price, unit_price, total_price?}]\n'
                    '  注：total_price 不传时按 unit_price * quantity 计算'
                ),
            },
        },
    }

    def execute(self, tool_input: dict, context: dict) -> Any:
        tool_input = tool_input or {}
        # ── 模式 B: 直接从 BOM 数据渲染（skill 草稿场景） ──
        bom_data = tool_input.get('bom_data')
        if bom_data:
            return self._execute_from_bom(bom_data, context)

        # ── 模式 A: 按 quotation_number/quotation_id 从数据库导出（原有逻辑） ──
        qnum = tool_input.get('quotation_number', '').strip()
        qid  = tool_input.get('quotation_id')
        user = context.get('user')

        if not qnum and not qid:
            return {'error': '请提供 quotation_number 或 quotation_id'}

        # ── 查询报价单（raw SQL 第一步，与 query_pma_database 用同一 engine）────
        # 避免 Flask-SQLAlchemy scoped session 的事务状态干扰 ORM 查询结果
        try:
            from sqlalchemy import text as _text
            from app.services.chat_db_query import _get_readonly_engine
            engine = _get_readonly_engine()
            real_id: int | None = None
            real_number: str | None = None

            with engine.connect() as conn:
                if qnum:
                    # 精确匹配优先，再 ILIKE 模糊（容忍 LLM 年月序号小偏差）
                    fuzzy = f"%{qnum.replace('-', '%')}%"
                    row = conn.execute(_text(
                        'SELECT id, quotation_number FROM quotations '
                        'WHERE quotation_number = :n OR quotation_number ILIKE :fuzzy '
                        'ORDER BY id DESC LIMIT 1'
                    ), {'n': qnum, 'fuzzy': fuzzy}).fetchone()
                    if row:
                        real_id, real_number = row[0], row[1]
                        if real_number != qnum:
                            logger.info(f'[export_quotation_to_excel] 模糊匹配 {qnum!r} → {real_number}')

                if real_id is None and qid:
                    # qid 可能是 quotations.id、quotation_details.quotation_id 或 quotation_details.id
                    row = conn.execute(_text(
                        'SELECT id, quotation_number FROM quotations WHERE id = :qid '
                        'UNION '
                        'SELECT q.id, q.quotation_number FROM quotations q '
                        'JOIN quotation_details qd ON qd.quotation_id = q.id '
                        'WHERE qd.quotation_id = :qid OR qd.id = :qid '
                        'LIMIT 1'
                    ), {'qid': int(qid)}).fetchone()
                    if row:
                        real_id, real_number = row[0], row[1]
                        logger.info(f'[export_quotation_to_excel] qid={qid} → 报价单 {real_number}(id={real_id})')

        except Exception as e:
            logger.exception('[export_quotation_to_excel] raw SQL 查询异常')
            return {'error': f'查询报价单失败: {e}'}

        if real_id is None:
            ident = qnum or str(qid)
            return {
                'error': (
                    f'未找到报价单（{ident}）。'
                    '请先用 query_pma_database 执行 '
                    '"SELECT id, quotation_number FROM quotations WHERE owner_id=... ORDER BY id DESC LIMIT 5" '
                    '获取准确编号后重试。'
                )
            }

        # ── 第二步：用 raw SQL 加载报价单完整数据（与第一步同一 engine）────────
        try:
            with engine.connect() as conn:
                h = conn.execute(_text(
                    'SELECT q.amount, q.created_at, q.owner_id,'
                    " COALESCE(p.project_name, '') AS project_name,"
                    " COALESCE(c.company_name, '') AS company_name,"
                    " COALESCE(ct.name, '') AS contact_name,"
                    " COALESCE(ct.phone, '') AS contact_phone,"
                    " COALESCE(u.real_name, u.username, '') AS owner_name"
                    ' FROM quotations q'
                    ' LEFT JOIN projects p ON p.id = q.project_id'
                    ' LEFT JOIN companies c ON c.id = q.customer_id'
                    ' LEFT JOIN contacts ct ON ct.id = q.contact_id'
                    ' LEFT JOIN users u ON u.id = q.owner_id'
                    ' WHERE q.id = :real_id'
                ), {'real_id': real_id}).fetchone()

                detail_rows = conn.execute(_text(
                    'SELECT product_name, product_model, product_desc, brand, unit,'
                    ' quantity, discount, unit_price, total_price'
                    ' FROM quotation_details'
                    ' WHERE quotation_id = :real_id'
                    '   AND (is_accessory = FALSE OR is_accessory IS NULL)'
                    ' ORDER BY id'
                ), {'real_id': real_id}).fetchall()

        except Exception as e:
            logger.exception('[export_quotation_to_excel] 数据加载异常')
            return {'error': f'加载报价单数据失败: {e}'}

        if not h:
            return {'error': f'加载报价单 {real_number} 失败，请重试'}

        # ── 权限检查 ────────────────────────────────────────
        if user:
            user_role = getattr(user, 'role', 'user')
            owner_id = h[2]
            admin_roles = {'admin', 'ceo', 'finance', 'finance_director', 'finace_director', 'system'}
            if user_role not in admin_roles and owner_id != user.id:
                try:
                    from app.utils.access_control import get_viewable_data
                    from app.models.quotation import Quotation as _Q
                    accessible = get_viewable_data(_Q, user, [_Q.id == real_id]).first()
                    if not accessible:
                        return {'error': f'您没有权限查看报价单 {real_number}'}
                except Exception as e:
                    logger.warning(f'[export_quotation_to_excel] 权限检查跳过: {e}')

        # ── 构建 data dict ──────────────────────────────────
        quote_date = h[1].strftime('%Y-%m-%d') if h[1] else ''
        data = {
            'quotation_number': real_number,
            'project_name': h[3],
            'company_name': h[4],
            'contact_name': h[5],
            'contact_phone': h[6],
            'quote_date': quote_date,
            'total_amount': float(h[0]) if h[0] else 0.0,
            'owner_name': h[7],
            'details': [
                {
                    'product_name': r[0] or '',
                    'product_model': r[1] or '',
                    'product_desc': r[2] or '',
                    'brand': r[3] or '',
                    'unit': r[4] or '',
                    'quantity': r[5] or 0,
                    'discount': float(r[6]) if r[6] is not None else None,
                    'unit_price': float(r[7]) if r[7] is not None else 0.0,
                    'total_price': float(r[8]) if r[8] is not None else 0.0,
                }
                for r in detail_rows
            ],
        }

        # ── 生成 Excel ────────────────────────────────────────
        from flask import current_app
        db_type = 'OVS' if current_app.config.get('IS_OVS') else 'SP8D'

        try:
            xlsx_bytes = _build_excel(data, db_type)
        except Exception as e:
            logger.exception('[export_quotation_to_excel] 生成 Excel 异常')
            return {'error': f'Excel 生成失败: {e}'}

        # ── 保存文件 ────────────────────────────────────────
        os.makedirs(_STORAGE_DIR, exist_ok=True)
        safe_num = re.sub(r'[^\w\u4e00-\u9fff]+', '_', real_number)
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'报价单_{safe_num}_{date_str}.xlsx'
        file_path = os.path.join(_STORAGE_DIR, filename)

        Path(file_path).write_bytes(xlsx_bytes)
        self._try_upload_nas(file_path, filename, user)

        logger.info(f'[export_quotation_to_excel] 生成: {filename}')
        return {
            'success': True,
            'filename': filename,
            'quotation_number': real_number,
            'download_url': f'/cli/api/exports/{filename}',
            'note': '下载按钮已显示在终端中，只需简短回复"报价单Excel已生成"即可。',
        }


    def _execute_from_bom(self, bom_data: dict, context: dict) -> Any:
        """模式 B：直接从 BOM 数据生成 Excel（不写库，方案草稿场景）。

        skill 推完 BOM → 调本方法 → 拿 download_url。无需 quotation_number。
        """
        if not isinstance(bom_data, dict):
            return {'error': 'bom_data 必须是 dict'}

        items = bom_data.get('items') or []
        if not items:
            return {'error': 'bom_data.items 不能为空'}

        # region → db_type 映射（决定中文/英文版式 + logo）
        region = (bom_data.get('region') or '').upper()
        if region in ('SG', 'OVS'):
            db_type = 'OVS'
        elif region in ('CN', 'SP8D'):
            db_type = 'SP8D'
        else:
            # 没传 region，按 Flask 当前实例配置兜底
            try:
                from flask import current_app
                db_type = 'OVS' if current_app.config.get('IS_OVS') else 'SP8D'
            except Exception:
                db_type = 'SP8D'

        # 字段映射：skill items → _build_excel data['details']
        details = []
        for it in items:
            qty   = it.get('quantity') or it.get('qty') or 0
            price = it.get('unit_price') or it.get('price') or 0
            total = it.get('total_price')
            if total is None and price:
                total = price * qty
            details.append({
                'mn':            it.get('mn') or it.get('product_mn') or '',
                'product_name':  it.get('product_name') or it.get('name') or it.get('name_cn') or '',
                'product_model': it.get('product_model') or it.get('model') or '',
                'product_desc':  it.get('product_desc') or it.get('spec') or it.get('spec_bilingual') or '',
                'brand':         it.get('brand') or '',
                'unit':          it.get('unit') or '',
                'quantity':      qty,
                'discount':      it.get('discount'),
                'market_price':  it.get('market_price'),
                'unit_price':    price or 0,
                'total_price':   total or 0,
            })

        # 组装 data dict
        user = context.get('user') if context else None
        owner_name = ''
        if user:
            owner_name = getattr(user, 'real_name', None) or getattr(user, 'username', '') or ''

        quote_date = bom_data.get('quote_date') or datetime.now().strftime('%Y-%m-%d')
        project_name = bom_data.get('project_name', '') or '方案草稿'
        # 草稿编号：DRAFT-时间戳，避免与正式报价单冲突
        draft_no = bom_data.get('quotation_number') or f'DRAFT-{datetime.now().strftime("%Y%m%d-%H%M")}'

        data = {
            'quotation_number': draft_no,
            'project_name':     project_name,
            'company_name':     bom_data.get('company_name', ''),
            'company_address':  bom_data.get('company_address', ''),
            'contact_name':     bom_data.get('contact_name', ''),
            'contact_phone':    bom_data.get('contact_phone', ''),
            'quote_date':       quote_date,
            'currency':         bom_data.get('currency', ''),
            'payment_terms':    bom_data.get('payment_terms', ''),
            'shipping_terms':   bom_data.get('shipping_terms', ''),
            'validity':         bom_data.get('validity', ''),
            'ref_no':           bom_data.get('ref_no', ''),
            'gst_percent':      bom_data.get('gst_percent', 0) or 0,
            'remarks':          bom_data.get('remarks', ''),
            'total_amount':     sum((d.get('total_price') or 0) for d in details),
            'owner_name':       owner_name,
            'details':          details,
        }

        # 生成 Excel
        try:
            xlsx_bytes = _build_excel(data, db_type)
        except Exception as e:
            logger.exception('[export_quotation_to_excel] BOM 模式 Excel 生成异常')
            return {'error': f'Excel 生成失败: {e}'}

        # 保存到本地 storage/exports(供 PMA Web /cli 终端下载,以及 cli 模式向后兼容)
        os.makedirs(_STORAGE_DIR, exist_ok=True)
        safe_proj = re.sub(r'[^\w一-鿿]+', '_', project_name)[:40]
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'方案草稿_{safe_proj}_{date_str}.xlsx'
        file_path = os.path.join(_STORAGE_DIR, filename)
        Path(file_path).write_bytes(xlsx_bytes)
        self._try_upload_nas(file_path, filename, user)

        logger.info(
            f'[export_quotation_to_excel] BOM 模式生成: {filename}'
            f'（{db_type}, {len(xlsx_bytes)/1024:.1f}KB）'
        )

        # 返回 base64 内容（让 MCP server 透传给 Claude 桌面写到本地目录）
        # PMA Web /cli 终端会忽略 content_base64,继续走 download_url
        import base64 as _b64
        return {
            'success': True,
            'filename': filename,
            'mode': 'bom_draft',
            'region': db_type,
            'size_bytes': len(xlsx_bytes),
            'content_base64': _b64.b64encode(xlsx_bytes).decode('ascii'),
            'download_url': f'/cli/api/exports/{filename}',
            'note': (
                '方案草稿 Excel 已生成（未写入 PMA）。'
                '使用方:Claude 桌面端拿 content_base64 解码写到本地文件;'
                'PMA /cli 终端用 download_url 浏览器下载。'
            ),
        }


    @staticmethod
    def _try_upload_nas(file_path: str, filename: str, user) -> None:
        try:
            from app.utils.synology_webdav_client import get_synology_webdav_client, is_nas_available
            if not is_nas_available():
                return
            client = get_synology_webdav_client()
            remote_dir = f'/exports/cli/{datetime.now().strftime("%Y/%m")}'
            client.ensure_directory(remote_dir)
            with open(file_path, 'rb') as f:
                client.upload(f'{remote_dir}/{filename}', f.read())
        except Exception as e:
            logger.debug(f'[export_quotation_to_excel] NAS 上传跳过: {e}')

