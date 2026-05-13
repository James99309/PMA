# -*- coding: utf-8 -*-
"""
export_to_excel 工具：生成格式化 Excel 文档。

主路径：Claude 使用 PMA xlsx 样式库写完整 openpyxl 代码，
PMA 在临时目录执行，输出 .xlsx。

存储：本地 storage/exports/，可选 NAS WebDAV 备份。
"""
from __future__ import annotations

import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from app.services.cli_agent.tools import BaseTool

logger = logging.getLogger(__name__)

_STORAGE_DIR = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', '..', '..', 'storage', 'exports'
))

# 内置 PMA Excel 样式库（当 DB 中无 pma_xlsx_style skill 时使用）
_FALLBACK_XLSX_STYLE = '''
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ALT_ROW_BG = "EBF3FB"
BORDER_CLR = "B8CCE4"

COMPANY_NAMES = {
    "SP8D": "和源通信（上海）股份有限公司",
    "OVS":  "Evertac Solutions",
}

def _thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def init_workbook(db_type="SP8D", sheet_title="Sheet1"):
    """初始化 PMA 样式 Workbook，返回 (wb, ws)"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws

def write_title(ws, title, row=1, col_span=8):
    """写标题行（深蓝色大字，跨列居中），返回下一可用行号"""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(name="Microsoft YaHei", bold=True, size=14, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=col_span
        )
    return row + 1

def write_header_row(ws, headers, row=2):
    """写表头行（深蓝底白字），返回下一可用行号"""
    fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill      = fill
        c.font      = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin_border()
    ws.row_dimensions[row].height = 22
    return row + 1

def write_data_row(ws, row, values, alt=False):
    """写数据行，alt=True 时隔行浅蓝背景，返回下一可用行号"""
    fill = (
        PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
        if alt else None
    )
    font = Font(name="Microsoft YaHei", size=10)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border    = _thin_border()
        if fill:
            c.fill = fill
    return row + 1

def write_subheader_row(ws, label, row, col_span=8):
    """写分组小标题行（中蓝色背景），返回下一可用行号"""
    fill = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill      = fill
    cell.font      = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    cell.alignment = Alignment(vertical="center")
    cell.border    = _thin_border()
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=col_span
        )
    return row + 1

def auto_column_width(ws, min_w=8, max_w=40):
    """根据内容自动调整列宽（中文字符按2倍计算）"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                    max_len = max(max_len, length)
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def freeze_header(ws, freeze_row=2):
    """冻结表头（freeze_row 行以上）"""
    ws.freeze_panes = f"A{freeze_row + 1}"
'''


class ExportToExcelTool(BaseTool):
    name = 'export_to_excel'
    description = (
        '生成 Excel 文档（.xlsx）。\n'
        '当用户说"生成Excel"、"导出表格"、"导出Excel"、"下载Excel"时使用。\n\n'
        '⚠️ 重要：**禁止**把 query_pma_database 返回的数据手抄到 python_code 里（会被 max_tokens 截断！）。\n'
        '系统会自动注入 `query_results` 变量，结构：\n'
        '  query_results = [\n'
        '      {"sql": "...", "columns": ["c1","c2"], "rows": [[v1,v2], ...]},\n'
        '      ...   # 按本轮查询时间先后排列, [-1] 是最近一次\n'
        '  ]\n'
        '直接用 `query_results[-1]["rows"]` / `query_results[-2]["rows"]` 等引用即可。\n\n'
        '参数：\n'
        '- python_code（必填）：使用 [PMA Excel 样式库] 写完整 openpyxl 代码。\n'
        '  代码结尾必须是 wb.save("__OUTPUT__")，工具会自动替换为实际路径。\n'
        '  DB_TYPE 从系统 [运行环境] 获取（SP8D 或 OVS）。\n\n'
        '[PMA Excel 样式库] 常用函数：\n'
        '  from pma_xlsx_style import (\n'
        '      init_workbook, write_title, write_header_row, write_data_row,\n'
        '      write_subheader_row, auto_column_width, freeze_header,\n'
        '      DARK_BLUE, MID_BLUE, ALT_ROW_BG, COMPANY_NAMES\n'
        '  )\n'
        '  wb, ws = init_workbook(db_type=DB_TYPE, sheet_title="总览")\n'
        '  row = write_title(ws, "报表标题", row=1, col_span=列数)\n'
        '  row = write_header_row(ws, ["列1", "列2", ...], row=row)\n'
        '  data = query_results[-1]["rows"]   # 直接引用查询结果，不要内联\n'
        '  for i, item in enumerate(data):\n'
        '      row = write_data_row(ws, row, list(item), alt=(i%2==1))\n'
        '  auto_column_width(ws)\n'
        '  freeze_header(ws)\n'
        '  wb.save("__OUTPUT__")\n\n'
        '多 Sheet：wb.create_sheet("明细") 后对新 ws 同样调用样式函数；\n'
        '  两个 sheet 分别用 query_results[-2]["rows"] 和 query_results[-1]["rows"]。\n'
        '返回结果包含 download_url，直接回复"Excel已生成"即可，无需重复链接。'
    )
    input_schema = {
        'type': 'object',
        'properties': {
            'python_code': {
                'type': 'string',
                'description': '完整 openpyxl 代码，使用 PMA 样式库，wb.save("__OUTPUT__") 结尾',
            },
            'title': {
                'type': 'string',
                'description': '文档标题（用于文件命名，如"Q1销售数据"、"项目汇总"）',
            },
        },
        'required': ['python_code'],
    }

    def execute(self, tool_input: dict, context: dict) -> Any:
        tool_input = tool_input or {}
        python_code = tool_input.get('python_code', '').strip()
        title = tool_input.get('title', '').strip()
        user = context.get('user')
        query_results = context.get('query_results') or []

        if not python_code:
            return {'error': 'python_code 参数不能为空'}
        return self._execute_python_code(python_code, title, user, query_results)

    def _execute_python_code(self, code: str, title: str, user, query_results: list) -> dict:
        output_filename = self._make_filename(title)
        os.makedirs(_STORAGE_DIR, exist_ok=True)
        final_path = os.path.join(_STORAGE_DIR, output_filename)

        with tempfile.TemporaryDirectory() as tmpdir:
            # 优先从 DB 加载 pma_xlsx_style skill；否则使用内置版本
            style_code = _FALLBACK_XLSX_STYLE
            try:
                from app.models.cli_skill import CliSkill
                skill = CliSkill.query.filter_by(
                    name='pma_xlsx_style', skill_type='xlsx', is_active=True
                ).first()
                if skill and skill.skill_body:
                    style_code = skill.skill_body
            except Exception as e:
                logger.debug(f'[export_to_excel] 使用内置样式库（DB 加载失败: {e}）')

            Path(os.path.join(tmpdir, 'pma_xlsx_style.py')).write_text(style_code, encoding='utf-8')

            # 把当前 run 的查询结果落到 JSON,供子进程通过 query_results 变量引用
            data_path = os.path.join(tmpdir, '_query_results.json')
            Path(data_path).write_text(
                json.dumps(query_results, ensure_ascii=False, default=str),
                encoding='utf-8',
            )

            tmp_out = os.path.join(tmpdir, 'output.xlsx')
            if '__OUTPUT__' in code:
                code = code.replace('__OUTPUT__', tmp_out)
            else:
                code = re.sub(
                    r'wb\.save\(["\'].*?["\']\)',
                    f'wb.save("{tmp_out}")',
                    code
                )
            prelude = (
                'import json as _json, os as _os\n'
                'with open(_os.path.join(_os.path.dirname(__file__), "_query_results.json"),'
                ' "r", encoding="utf-8") as _f:\n'
                '    query_results = _json.load(_f)\n\n'
            )
            script_path = os.path.join(tmpdir, 'gen.py')
            Path(script_path).write_text(prelude + code, encoding='utf-8')

            try:
                result = subprocess.run(
                    ['python3', script_path],
                    capture_output=True, text=True, timeout=60
                )
            except subprocess.TimeoutExpired:
                return {'error': 'Excel 生成超时（60秒）'}

            if result.returncode != 0:
                logger.error(f'[export_to_excel] openpyxl 执行失败:\n{result.stderr}')
                return {'error': f'Excel 生成失败: {result.stderr[:300]}'}

            if not os.path.exists(tmp_out):
                return {'error': '代码执行成功但未找到输出文件，请确认代码中有 wb.save("__OUTPUT__")'}

            shutil.copy(tmp_out, final_path)

        self._try_upload_nas(final_path, output_filename, user)
        logger.info(f'[export_to_excel] 生成: {output_filename}')
        return {
            'success': True,
            'filename': output_filename,
            'download_url': f'/cli/api/exports/{output_filename}',
            'note': '下载按钮已显示在终端中，只需简短回复"Excel已生成"即可。',
        }

    @staticmethod
    def _make_filename(title: str) -> str:
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        if title:
            safe = re.sub(r'[^\w\u4e00-\u9fff]+', '_', title).strip('_')
            if len(safe) > 60:
                safe = safe[:60].rstrip('_')
            return f'{safe}_{date_str}.xlsx'
        short_id = uuid.uuid4().hex[:8]
        return f'export_{date_str}_{short_id}.xlsx'

    @staticmethod
    def _try_upload_nas(file_path: str, filename: str, user) -> None:
        try:
            from app.utils.synology_webdav_client import get_synology_webdav_client, is_nas_available
            if not is_nas_available():
                return
            client = get_synology_webdav_client()
            remote_dir = f'/exports/cli/{datetime.now().strftime("%Y/%m")}'
            client.ensure_directory(remote_dir)
            with open(file_path, 'rb') as f:
                client.upload(f'{remote_dir}/{filename}', f.read())
        except Exception as e:
            logger.debug(f'[export_to_excel] NAS 上传跳过: {e}')
