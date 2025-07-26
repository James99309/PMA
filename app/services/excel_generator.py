#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel生成服务
用于生成报价单的Excel文档
"""

import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import logging

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Excel生成器"""
    
    def __init__(self):
        # 定义样式
        self.header_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        self.title_font = Font(name='微软雅黑', size=16, bold=True, color='0066CC')
        self.normal_font = Font(name='微软雅黑', size=11)
        self.bold_font = Font(name='微软雅黑', size=11, bold=True)
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    def generate_quotation_excel(self, quotation, current_user=None, export_info=None):
        """生成报价单Excel"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "报价单"
            
            # 设置列宽
            column_widths = {
                'A': 8,   # 序号
                'B': 25,  # 产品名称
                'C': 18,  # 型号
                'D': 30,  # 指标描述
                'E': 12,  # 品牌
                'F': 8,   # 单位
                'G': 8,   # 数量
                'H': 12,  # 单价
                'I': 12,  # 小计
                'J': 15   # MN
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            current_row = 1
            
            # 确定当前语言环境
            from flask import session, g
            try:
                current_language = getattr(g, 'current_language', 'zh') or session.get('language', 'zh')
            except:
                current_language = 'zh'
            
            # 准备国际化文本
            i18n_texts = self._get_i18n_texts(current_language)
            
            # 获取用户所属企业信息
            company_info = self._get_user_company_info(quotation.owner or current_user)
            
            # 标题区域（4行高度）
            logo_row_count = 4
            
            # 左侧Logo区域（4行3栏合并 A1:C4）- 先合并单元格
            ws.merge_cells(f'A{current_row}:C{current_row + logo_row_count - 1}')
            logo_cell = ws[f'A{current_row}']
            logo_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 添加黑色实线边框
            for r in range(current_row, current_row + logo_row_count):
                for c in range(1, 4):  # A到C列
                    cell = ws.cell(row=r, column=c)
                    cell.border = self.thin_border
            
            # 在合并单元格之后插入企业Logo
            logo_height = self._insert_company_logo(ws, quotation.owner or current_user, current_row)
            
            # 如果Logo插入失败，在Logo区域显示企业名称
            if logo_height == 0:
                logger.warning(f"⚠️ Logo插入失败，显示企业名称: {company_info['name']}")
                logo_cell.value = company_info['name']
                logo_cell.font = Font(name='微软雅黑', size=10, bold=True, color='0066CC')
                logo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                logger.info(f"✅ Logo插入成功，高度: {logo_height}px")
                # 清除Logo区域的文本，确保只显示Logo
                logo_cell.value = None
            
            # 设置Logo区域的行高
            for i in range(logo_row_count):
                ws.row_dimensions[current_row + i].height = 25
            
            # 右侧企业名称（2行合并，居中 D1:J2）
            ws.merge_cells(f'D{current_row}:J{current_row + 1}')
            enterprise_name_cell = ws[f'D{current_row}']
            enterprise_name_cell.value = company_info['name']
            enterprise_name_cell.font = Font(name='微软雅黑', size=16, bold=True, color='0066CC')
            enterprise_name_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 添加黑色实线边框
            for r in range(current_row, current_row + 2):
                for c in range(4, 11):  # D到J列
                    cell = ws.cell(row=r, column=c)
                    cell.border = self.thin_border
            
            # 右侧企业详细信息（2行合并，左对齐空一个字符 D3:J4）
            ws.merge_cells(f'D{current_row + 2}:J{current_row + 3}')
            detail_cell = ws[f'D{current_row + 2}']
            company_detail = f" {company_info['address']}\n {company_info['phone']}\n {company_info['website']}"
            detail_cell.value = company_detail
            detail_cell.font = Font(name='微软雅黑', size=10, color='666666')
            detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # 添加黑色实线边框
            for r in range(current_row + 2, current_row + logo_row_count):
                for c in range(4, 11):  # D到J列
                    cell = ws.cell(row=r, column=c)
                    cell.border = self.thin_border
            
            current_row += logo_row_count  # 移动到Logo区域之后，不空行直接对接详情区域
            
            # 详情区域（字段和内容都左对齐，空一个字符位置）
            # 获取客户和联系人信息（优先使用导出信息）
            company_name = self._get_company_name(quotation, export_info)
            customer_address = self._get_customer_address(quotation, export_info)
            contact_person = self._get_contact_person(quotation, export_info)
            contact_phone = self._get_contact_phone(quotation, export_info)
            
            info_data = [
                [f" {i18n_texts['company_label']}", f" {company_name}", 
                 f" {i18n_texts['number_label']}", f" {quotation.quotation_number}"],
                [f" {i18n_texts['address_info_label']}", f" {customer_address}", 
                 f" {i18n_texts['date_label']}", f" {quotation.created_at.strftime('%m.%d.%Y') if quotation.created_at else '-'}"],
                [f" {i18n_texts['contact_label']}", f" {contact_person}", 
                 f" {i18n_texts['responsible_label']}", f" {quotation.owner.real_name or quotation.owner.username if quotation.owner else '-'}"],
                [f" {i18n_texts['contact_info_label']}", f" {contact_phone if contact_phone else ''}", 
                 f" {i18n_texts['currency_label']}", f" {'人民币' if current_language == 'zh' else 'CNY'}"]
            ]
            
            for row_data in info_data:
                # 第一个标签列（A列，左对齐，空一个字符）
                ws[f'A{current_row}'] = row_data[0]
                ws[f'A{current_row}'].font = self.bold_font
                ws[f'A{current_row}'].border = self.thin_border
                ws[f'A{current_row}'].alignment = self.left_alignment
                
                # 第一个值列（B-C列合并，两栏空间）
                ws.merge_cells(f'B{current_row}:C{current_row}')
                ws[f'B{current_row}'] = row_data[1]
                ws[f'B{current_row}'].font = self.normal_font
                ws[f'B{current_row}'].border = self.thin_border
                ws[f'B{current_row}'].alignment = self.left_alignment
                
                # 第二个标签列（D列，左对齐，空一个字符）
                ws[f'D{current_row}'] = row_data[2]
                ws[f'D{current_row}'].font = self.bold_font
                ws[f'D{current_row}'].border = self.thin_border
                ws[f'D{current_row}'].alignment = self.left_alignment
                
                # 第二个值列（E-J列合并）
                ws.merge_cells(f'E{current_row}:J{current_row}')
                ws[f'E{current_row}'] = row_data[3]
                ws[f'E{current_row}'].font = self.normal_font
                ws[f'E{current_row}'].border = self.thin_border
                ws[f'E{current_row}'].alignment = self.left_alignment
                
                # 为未合并的单元格添加边框
                for c in range(1, 11):  # A到J列
                    cell = ws.cell(row=current_row, column=c)
                    if not cell.border.left.style:  # 如果还没有设置边框
                        cell.border = self.thin_border
                
                current_row += 1
            
            # 无空行 - 详情区域后直接对接报价单
            
            # 报价单标题（居中显示）
            ws.merge_cells(f'A{current_row}:J{current_row}')
            ws[f'A{current_row}'] = i18n_texts['quotation_title']
            ws[f'A{current_row}'].font = Font(name='微软雅黑', size=14, bold=True, color='0066CC')
            ws[f'A{current_row}'].alignment = self.center_alignment
            current_row += 1
            
            # 产品明细表头（与PDF保持一致）
            headers = [
                i18n_texts['item_label'], '产品名称', '产品型号', 
                i18n_texts['description_label'], '品牌', '单位',
                i18n_texts['qty_label'], i18n_texts['unit_price_label'], 
                i18n_texts['total_price_label'], 'MN'
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                # 统一字体大小与产品明细内容一致
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_alignment
            
            current_row += 1
            
            # 产品明细数据
            if quotation.details:
                for idx, detail in enumerate(quotation.details, 1):
                    row_data = [
                        idx,
                        detail.product_name or '-',
                        detail.product_model or '-',
                        detail.product_desc or '-',
                        detail.brand or '-',
                        detail.unit or '-',
                        detail.quantity or 0,
                        detail.unit_price or 0,
                        detail.total_price or 0,
                        detail.product_mn or '-'
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.font = self.normal_font
                        cell.border = self.thin_border
                        
                        # 设置对齐方式
                        if col_idx == 1:  # 序号
                            cell.alignment = self.center_alignment
                        elif col_idx in [7, 8, 9]:  # 数量、单价、小计
                            cell.alignment = self.right_alignment
                            if col_idx in [8, 9]:  # 单价、小计格式化为货币
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = self.left_alignment
                    
                    current_row += 1
            else:
                # 无数据行
                ws.merge_cells(f'A{current_row}:J{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = '暂无产品明细'
                cell.font = Font(name='微软雅黑', size=11, color='666666')
                cell.border = self.thin_border
                cell.alignment = self.center_alignment
                current_row += 1
            
            current_row += 1  # 空行
            
            # 明细表最后一行：总价和金额（在右侧）- 总计区域
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = ''  # 左侧空白
            
            ws[f'I{current_row}'] = i18n_texts['total_label']
            # 统一字体大小与产品明细内容一致
            ws[f'I{current_row}'].font = Font(name='微软雅黑', size=11, bold=True)
            ws[f'I{current_row}'].alignment = self.right_alignment
            ws[f'I{current_row}'].border = self.thin_border
            
            ws[f'J{current_row}'] = quotation.amount or 0
            # 统一字体大小与产品明细内容一致
            ws[f'J{current_row}'].font = Font(name='微软雅黑', size=11, bold=True, color='0066CC')
            ws[f'J{current_row}'].alignment = self.right_alignment
            ws[f'J{current_row}'].number_format = '¥#,##0.00'
            ws[f'J{current_row}'].border = self.thin_border
            
            # 为总计区域的所有单元格添加黑色实线边框
            for c in range(1, 11):  # A到J列
                cell = ws.cell(row=current_row, column=c)
                cell.border = self.thin_border
            
            current_row += 3  # 空两行
            
            # 备注部分（与PDF格式保持一致）
            ws.merge_cells(f'A{current_row}:J{current_row}')
            ws[f'A{current_row}'] = i18n_texts['notes_title']
            ws[f'A{current_row}'].font = Font(name='微软雅黑', size=12, bold=True)
            ws[f'A{current_row}'].alignment = self.left_alignment
            current_row += 1
            
            # 备注内容（使用与PDF相同的默认备注）
            notes_content = [
                i18n_texts['delivery_clause'],
                i18n_texts['delivery_time'],
                i18n_texts['warranty'],
                i18n_texts['payment_terms'],
                i18n_texts['others']
            ]
            
            for note in notes_content:
                ws.merge_cells(f'A{current_row}:J{current_row}')
                ws[f'A{current_row}'] = note
                ws[f'A{current_row}'].font = Font(name='微软雅黑', size=10)
                ws[f'A{current_row}'].alignment = self.left_alignment
                current_row += 1
            
            current_row += 4  # 备注后空3行
            
            # 最右侧签字确认
            ws.merge_cells(f'H{current_row}:J{current_row}')
            ws[f'H{current_row}'] = '签字确认'
            ws[f'H{current_row}'].font = Font(name='微软雅黑', size=12, bold=True)
            ws[f'H{current_row}'].alignment = self.right_alignment
            
            current_row += 1  # 换行
            current_row += 5  # 空四行
            
            # 左侧下划线两栏 (A-B列)
            ws.merge_cells(f'A{current_row}:B{current_row}')
            supplier_line_cell = ws[f'A{current_row}']
            supplier_line_cell.value = '________________________'
            supplier_line_cell.font = Font(name='微软雅黑', size=11)
            supplier_line_cell.alignment = self.left_alignment
            
            # 右侧靠右二栏下划线 (I-J列)
            ws.merge_cells(f'I{current_row}:J{current_row}')
            customer_line_cell = ws[f'I{current_row}']
            customer_line_cell.value = '________________________'
            customer_line_cell.font = Font(name='微软雅黑', size=11)
            customer_line_cell.alignment = self.right_alignment
            
            current_row += 1
            
            # 下面写公司名称（左侧 - 账户所在企业名称）
            ws.merge_cells(f'A{current_row}:B{current_row}')
            # 获取账户所在企业名称（当前用户所在企业）
            account_company = company_info['name']  # 用户所在企业名称
            ws[f'A{current_row}'] = account_company
            # 字体大小与产品明细一致
            ws[f'A{current_row}'].font = Font(name='微软雅黑', size=11)
            ws[f'A{current_row}'].alignment = self.left_alignment
            
            # 右侧客户公司名称（目标企业名称）
            ws.merge_cells(f'I{current_row}:J{current_row}')
            ws[f'I{current_row}'] = company_name  # 客户企业名称
            # 字体大小与产品明细一致
            ws[f'I{current_row}'].font = Font(name='微软雅黑', size=11)
            ws[f'I{current_row}'].alignment = self.right_alignment
            
            current_row += 2  # 空行后放置页脚
            
            # 页脚信息
            ws.merge_cells(f'A{current_row}:J{current_row}')
            ws[f'A{current_row}'] = f'生成时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")} | 此报价单由系统自动生成'
            ws[f'A{current_row}'].font = Font(name='微软雅黑', size=9, color='666666')
            ws[f'A{current_row}'].alignment = self.center_alignment
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"生成报价单Excel失败: {str(e)}")
            raise
    
    def _get_user_company_info(self, user):
        """获取用户所属企业的信息"""
        # 默认企业信息（EVERTAC）
        default_info = {
            'name': 'EVERTAC SOLUTIONS',
            'address': '武威路88号19楼6楼, 普陀区 上海市, 中国 (200335)',
            'phone': '021-62596028',
            'email': 'info@evertac.net',
            'website': 'www.evertac.net'
        }
        
        if not user or not user.company_name:
            return default_info
        
        try:
            from app.models.dictionary import Dictionary
            
            # 查找用户所属企业的字典记录
            company_dict = Dictionary.query.filter_by(
                type='company',
                value=user.company_name,
                is_active=True
            ).first()
            
            if company_dict:
                # 如果找到企业字典记录，使用字典中的企业信息
                company_info = {
                    'name': company_dict.value,  # 企业名称
                    'address': company_dict.address or default_info['address'],
                    'phone': company_dict.phone or default_info['phone'],
                    'email': company_dict.email or default_info['email'],
                    'website': company_dict.website or default_info['website']
                }
                
                logger.info(f"✅ Excel生成器使用企业字典中的企业信息: {user.company_name}")
                return company_info
            else:
                # 如果没有找到企业字典记录，使用默认信息但更新企业名称
                logger.info(f"⚠️ Excel生成器：用户 {user.username} 的企业 {user.company_name} 在字典中未找到，使用默认信息")
                default_info['name'] = user.company_name
                return default_info
                
        except Exception as e:
            logger.error(f"❌ Excel生成器获取用户企业信息失败: {e}")
            # 出错时使用默认信息
            return default_info
    
    def _insert_company_logo(self, ws, user, row):
        """在Excel中插入企业Logo"""
        try:
            logger.info(f"🔍 开始插入Logo - 用户: {user.username if user else 'None'}, 企业: {user.company_name if user and hasattr(user, 'company_name') else 'None'}, 行位置: {row}")
            
            if not user or not hasattr(user, 'company_name') or not user.company_name:
                logger.info(f"⚠️ Excel生成器：用户信息不完整，无法插入Logo")
                return 0  # 没有Logo，返回高度0
            
            from app.models.dictionary import Dictionary
            
            # 查找用户所属企业的字典记录
            company_dict = Dictionary.query.filter_by(
                type='company',
                value=user.company_name,
                is_active=True
            ).first()
            
            logger.info(f"🔍 企业字典查询结果: {company_dict.value if company_dict else 'None'}, Logo内容: {'有' if company_dict and company_dict.logo_content else '无'}")
            
            if company_dict and company_dict.logo_content:
                try:
                    # 解码Base64图片数据
                    logo_data = base64.b64decode(company_dict.logo_content)
                    
                    # 检查是否为SVG格式
                    if logo_data.startswith(b'<?xml') or logo_data.startswith(b'<svg'):
                        logger.info(f"⚠️ 检测到SVG格式Logo，Excel不支持SVG，跳过Logo插入: {user.company_name}")
                        return 0
                    
                    logo_io = io.BytesIO(logo_data)
                    
                    # 创建openpyxl图片对象
                    img = OpenpyxlImage(logo_io)
                    
                    # 调整Logo大小（缩小30%）
                    img.width = 168  # 240 * 0.7 = 168像素（缩小30%）
                    img.height = 56  # 80 * 0.7 = 56像素，保持比例
                    
                    # 将Logo插入到指定位置（A1:C4区域，4行3栏合并），保持原始比例
                    # 使用简单的字符串锚点 + 手动偏移实现居中
                    
                    logger.info(f"🎯 Logo尺寸: {img.width}x{img.height}px, 插入位置: A{row}")
                    
                    # 设置锚点到A列起始位置
                    img.anchor = f'A{row}'
                    
                    # 尝试通过属性设置偏移量实现居中（如果支持的话）
                    try:
                        from openpyxl.utils.units import pixels_to_EMU
                        
                        # 计算居中偏移量
                        merge_width = 192   # 合并区域总宽度
                        merge_height = 100  # 合并区域总高度
                        center_offset_x = (merge_width - img.width) // 2   # 水平居中偏移
                        center_offset_y = (merge_height - img.height) // 2  # 垂直居中偏移
                        
                        logger.info(f"📏 居中计算: 区域{merge_width}x{merge_height}px, 偏移({center_offset_x}, {center_offset_y})px")
                        
                        # 如果锚点对象支持偏移属性，设置偏移量
                        if hasattr(img, 'anchor') and hasattr(img.anchor, '__dict__'):
                            # 尝试创建锚点对象并设置偏移
                            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
                            anchor_obj = AnchorMarker(
                                col=0, row=row-1,
                                colOff=pixels_to_EMU(center_offset_x),
                                rowOff=pixels_to_EMU(center_offset_y)
                            )
                            img.anchor = anchor_obj
                            logger.info("✅ 使用AnchorMarker设置居中偏移")
                        else:
                            logger.info("⚠️ 使用简单锚点，无法设置偏移")
                            
                    except Exception as offset_error:
                        logger.warning(f"⚠️ 设置居中偏移失败，使用简单锚点: {offset_error}")
                        img.anchor = f'A{row}'
                    
                    ws.add_image(img)
                    
                    logger.info(f"✅ Excel中成功插入企业Logo: {user.company_name}, 大小: {img.width}x{img.height}, 位置: A{row}")
                    # 返回非零高度表示插入成功
                    return img.height if img.height > 0 else 56  # 确保返回正确的高度值
                    
                except Exception as img_error:
                    logger.error(f"❌ Excel插入Logo图片处理失败: {img_error}")
                    return 0
                
            else:
                logger.info(f"⚠️ Excel生成器：用户 {user.username} 的企业 {user.company_name} 没有Logo")
                return 0
                
        except Exception as e:
            logger.error(f"❌ Excel插入企业Logo失败: {e}")
            return 0
    
    def _get_i18n_texts(self, language='zh'):
        """获取国际化文本（与PDF生成器保持一致）"""
        if language == 'en':
            return {
                # 地址标签
                'address_label': 'Address',
                'phone_label': 'Phone',
                'website_label': 'Website',
                
                # 报价单标题
                'quotation_title': 'QUOTATION',
                
                # 信息表格标签
                'project_label': 'Project',
                'company_label': 'Company',
                'address_info_label': 'Address',
                'date_label': 'Date',
                'contact_label': 'Contact',
                'responsible_label': 'Responsible',
                'contact_info_label': 'Contact Info',
                'currency_label': 'Currency',
                'number_label': 'Number',
                'version_label': 'Version',
                
                # 产品表格标签
                'item_label': 'Item',
                'description_label': 'Description',
                'qty_label': 'Qty',
                'unit_price_label': 'Unit Price',
                'total_price_label': 'Total Price',
                'subtotal_label': 'Subtotal',
                'total_label': 'Total',
                
                # 备注标签
                'notes_title': 'Terms and Conditions',
                'delivery_clause': 'Delivery Terms: On-site delivery',
                'delivery_time': 'Delivery Time: 4-6 weeks after contract signing',
                'warranty': 'Warranty: 24 months after goods handover',
                'payment_terms': 'Payment Terms: Payment within 30 days after contract signing',
                'others': 'Others: Price valid for 30 days'
            }
        else:
            return {
                # 地址标签
                'address_label': '地址',
                'phone_label': '电话',
                'website_label': '网址',
                
                # 报价单标题
                'quotation_title': '报价单',
                
                # 信息表格标签
                'project_label': '项目',
                'company_label': '公司',
                'address_info_label': '公司地址',
                'date_label': '日期',
                'contact_label': '联系人',
                'responsible_label': '报价负责人',
                'contact_info_label': '联系方式',
                'currency_label': '货币',
                'number_label': '编号',
                'version_label': '版本',
                
                # 产品表格标签
                'item_label': '项目',
                'description_label': '描述',
                'qty_label': '数量',
                'unit_price_label': '单价',
                'total_price_label': '总价',
                'subtotal_label': '小计',
                'total_label': '总计',
                
                # 备注标签
                'notes_title': '条款和条件',
                'delivery_clause': '交付条款：项目现场交付',
                'delivery_time': '交付时间：合同签订后 4-6 周',
                'warranty': '质保：货物移交后24个月',
                'payment_terms': '付款条款：合同签订后30天内付款',
                'others': '其他：价格有效期30天'
            }
    
    def _get_company_name(self, quotation, export_info=None):
        """获取公司名称（与PDF生成器保持一致）"""
        if not quotation or not quotation.project:
            return ""
        
        # 优先使用导出信息中的客户
        if export_info and export_info.get('customer'):
            customer_info = export_info['customer']
            if customer_info and customer_info.get('company_name'):
                return str(customer_info['company_name']).strip()
        
        # 然后尝试从项目获取
        if quotation.project and hasattr(quotation.project, 'company') and quotation.project.company:
            if hasattr(quotation.project.company, 'company_name') and quotation.project.company.company_name:
                return str(quotation.project.company.company_name).strip()
        
        return ""
    
    def _get_customer_address(self, quotation, export_info=None):
        """获取客户地址（与PDF生成器保持一致）"""
        if not quotation:
            return ""
        
        # 优先使用导出信息中的客户地址
        if export_info and export_info.get('customer'):
            customer_info = export_info['customer']
            if customer_info and customer_info.get('address'):
                return str(customer_info['address']).strip()
        
        # 尝试从项目location获取
        if quotation.project and hasattr(quotation.project, 'location') and quotation.project.location:
            return str(quotation.project.location).strip()
        
        return ""
    
    def _get_contact_person(self, quotation, export_info=None):
        """获取联系人（与PDF生成器保持一致）"""
        if not quotation:
            return ""
        
        # 优先使用导出信息中的联系人
        if export_info and export_info.get('contact'):
            contact_info = export_info['contact']
            if contact_info and contact_info.get('name'):
                return str(contact_info['name']).strip()
        
        # 然后尝试从项目获取
        if quotation.project and hasattr(quotation.project, 'company') and quotation.project.company:
            if hasattr(quotation.project.company, 'contact_person') and quotation.project.company.contact_person:
                return str(quotation.project.company.contact_person).strip()
        
        return ""
    
    def _get_contact_phone(self, quotation, export_info=None):
        """获取联系人电话（与PDF生成器保持一致）"""
        if not quotation:
            return ""
        
        # 优先使用导出信息中的联系人电话
        if export_info and export_info.get('contact'):
            contact_info = export_info['contact']
            if contact_info and contact_info.get('phone'):
                return str(contact_info['phone']).strip()
        
        # 然后尝试从项目获取
        if quotation.project and hasattr(quotation.project, 'company') and quotation.project.company:
            if hasattr(quotation.project.company, 'contact_phone') and quotation.project.company.contact_phone:
                return str(quotation.project.company.contact_phone).strip()
        
        return ""

# 创建全局实例
excel_generator = ExcelGenerator()

# 便捷函数
def generate_quotation_excel(quotation, current_user=None):
    """生成报价单Excel的便捷函数"""
    return excel_generator.generate_quotation_excel(quotation, current_user) 
    return excel_generator.generate_quotation_excel(quotation, current_user) 