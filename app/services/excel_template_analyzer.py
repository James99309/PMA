#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel模板精确分析器
用于分析Excel电子表格并生成精确的PDF模板配置
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Tuple
import re

logger = logging.getLogger(__name__)

class ExcelTemplateAnalyzer:
    """Excel模板精确分析器"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.template_config = {}
        
    def analyze_excel_file(self, excel_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        分析Excel文件，提取完整的结构和格式信息
        
        Args:
            excel_path: Excel文件路径
            sheet_name: 工作表名称（如果不指定则使用第一个工作表）
            
        Returns:
            Dict: 完整的模板配置信息
        """
        try:
            print(f"🔍 开始分析Excel文件: {excel_path}")
            
            # 加载Excel文件
            self.workbook = openpyxl.load_workbook(excel_path, data_only=False)
            
            if sheet_name:
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
            
            print(f"📊 工作表: {self.worksheet.title}")
            print(f"📐 使用范围: {self.worksheet.calculate_dimension()}")
            
            # 执行全面分析
            config = {
                'file_info': self._extract_file_info(excel_path),
                'sheet_info': self._extract_sheet_info(),
                'layout_structure': self._analyze_layout_structure(),
                'cell_formats': self._extract_cell_formats(),
                'merged_cells': self._extract_merged_cells(),
                'table_regions': self._identify_table_regions(),
                'data_fields': self._identify_data_fields(),
                'styling_rules': self._extract_styling_rules(),
                'print_settings': self._extract_print_settings()
            }
            
            self.template_config = config
            print("✅ Excel分析完成")
            return config
            
        except Exception as e:
            logger.error(f"❌ Excel分析失败: {e}")
            raise
    
    def _extract_file_info(self, excel_path: str) -> Dict[str, Any]:
        """提取文件基本信息"""
        file_path = Path(excel_path)
        return {
            'filename': file_path.name,
            'path': str(file_path.absolute()),
            'size': file_path.stat().st_size,
            'sheets': [sheet.title for sheet in self.workbook.worksheets]
        }
    
    def _extract_sheet_info(self) -> Dict[str, Any]:
        """提取工作表信息"""
        return {
            'title': self.worksheet.title,
            'dimensions': self.worksheet.calculate_dimension(),
            'max_row': self.worksheet.max_row,
            'max_column': self.worksheet.max_column,
            'column_dimensions': {
                col: {
                    'width': self.worksheet.column_dimensions[col].width,
                    'hidden': self.worksheet.column_dimensions[col].hidden
                }
                for col in self.worksheet.column_dimensions
            },
            'row_dimensions': {
                row: {
                    'height': self.worksheet.row_dimensions[row].height,
                    'hidden': self.worksheet.row_dimensions[row].hidden
                }
                for row in self.worksheet.row_dimensions
            }
        }
    
    def _analyze_layout_structure(self) -> Dict[str, Any]:
        """分析布局结构"""
        structure = {
            'header_region': self._find_header_region(),
            'content_regions': self._find_content_regions(),
            'table_regions': self._find_table_regions(),
            'footer_region': self._find_footer_region(),
            'logo_position': self._find_logo_position(),
            'title_position': self._find_title_position()
        }
        return structure
    
    def _extract_cell_formats(self) -> Dict[str, Dict[str, Any]]:
        """提取所有单元格的格式信息"""
        cell_formats = {}
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"
                
                cell_formats[cell_ref] = {
                    'value': cell.value,
                    'data_type': cell.data_type,
                    'font': self._extract_font_info(cell.font),
                    'fill': self._extract_fill_info(cell.fill),
                    'border': self._extract_border_info(cell.border),
                    'alignment': self._extract_alignment_info(cell.alignment),
                    'number_format': cell.number_format,
                    'coordinate': {'row': row, 'column': col}
                }
        
        return cell_formats
    
    def _extract_font_info(self, font: Font) -> Dict[str, Any]:
        """提取字体信息"""
        return {
            'name': font.name,
            'size': font.size,
            'bold': font.bold,
            'italic': font.italic,
            'underline': font.underline,
            'strike': font.strike,
            'color': font.color.rgb if font.color and font.color.rgb else None
        }
    
    def _extract_fill_info(self, fill: Fill) -> Dict[str, Any]:
        """提取填充信息"""
        return {
            'fill_type': fill.fill_type,
            'start_color': fill.start_color.rgb if fill.start_color and fill.start_color.rgb else None,
            'end_color': fill.end_color.rgb if fill.end_color and fill.end_color.rgb else None
        }
    
    def _extract_border_info(self, border: Border) -> Dict[str, Any]:
        """提取边框信息"""
        return {
            'left': {'style': border.left.style, 'color': border.left.color.rgb if border.left.color and border.left.color.rgb else None},
            'right': {'style': border.right.style, 'color': border.right.color.rgb if border.right.color and border.right.color.rgb else None},
            'top': {'style': border.top.style, 'color': border.top.color.rgb if border.top.color and border.top.color.rgb else None},
            'bottom': {'style': border.bottom.style, 'color': border.bottom.color.rgb if border.bottom.color and border.bottom.color.rgb else None}
        }
    
    def _extract_alignment_info(self, alignment: Alignment) -> Dict[str, Any]:
        """提取对齐信息"""
        return {
            'horizontal': alignment.horizontal,
            'vertical': alignment.vertical,
            'text_rotation': alignment.textRotation,
            'wrap_text': alignment.wrapText,
            'shrink_to_fit': alignment.shrinkToFit,
            'indent': alignment.indent
        }
    
    def _extract_merged_cells(self) -> List[Dict[str, Any]]:
        """提取合并单元格信息"""
        merged_cells = []
        for merged_range in self.worksheet.merged_cells.ranges:
            merged_cells.append({
                'range': str(merged_range),
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'width': merged_range.max_col - merged_range.min_col + 1,
                'height': merged_range.max_row - merged_range.min_row + 1
            })
        return merged_cells
    
    def _identify_table_regions(self) -> List[Dict[str, Any]]:
        """识别表格区域"""
        table_regions = []
        
        # 查找具有边框的连续区域作为表格
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if self._has_full_border(cell):
                    # 找到表格起始点，扩展寻找整个表格区域
                    table_region = self._expand_table_region(row, col)
                    if table_region and table_region not in table_regions:
                        table_regions.append(table_region)
        
        return table_regions
    
    def _identify_data_fields(self) -> Dict[str, Any]:
        """识别数据字段位置"""
        data_fields = {}
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"
                
                if cell.value:
                    value_str = str(cell.value).strip()
                    
                    # 识别可能的数据字段模式
                    field_type = self._identify_field_type(value_str)
                    if field_type:
                        data_fields[cell_ref] = {
                            'value': value_str,
                            'field_type': field_type,
                            'coordinate': {'row': row, 'column': col},
                            'suggested_mapping': self._suggest_field_mapping(value_str, field_type)
                        }
        
        return data_fields
    
    def _extract_styling_rules(self) -> Dict[str, Any]:
        """提取样式规则"""
        return {
            'default_font': self._get_most_common_font(),
            'header_style': self._get_header_style(),
            'table_style': self._get_table_style(),
            'border_styles': self._get_border_styles(),
            'color_scheme': self._extract_color_scheme()
        }
    
    def _extract_print_settings(self) -> Dict[str, Any]:
        """提取打印设置"""
        page_setup = self.worksheet.page_setup
        return {
            'orientation': page_setup.orientation,
            'paper_size': page_setup.paperSize,
            'margins': {
                'left': self.worksheet.page_margins.left,
                'right': self.worksheet.page_margins.right,
                'top': self.worksheet.page_margins.top,
                'bottom': self.worksheet.page_margins.bottom
            },
            'scale': page_setup.scale,
            'fit_to_width': page_setup.fitToWidth,
            'fit_to_height': page_setup.fitToHeight
        }
    
    def _find_header_region(self) -> Dict[str, Any]:
        """查找表头区域"""
        # 简单实现：前3行作为表头
        return {
            'start_row': 1,
            'end_row': 3,
            'start_col': 1,
            'end_col': self.worksheet.max_column
        }
    
    def _find_content_regions(self) -> List[Dict[str, Any]]:
        """查找内容区域"""
        return []  # 待实现具体逻辑
    
    def _find_table_regions(self) -> List[Dict[str, Any]]:
        """查找表格区域"""
        return []  # 待实现具体逻辑
    
    def _find_footer_region(self) -> Dict[str, Any]:
        """查找页脚区域"""
        return {
            'start_row': max(1, self.worksheet.max_row - 2),
            'end_row': self.worksheet.max_row,
            'start_col': 1,
            'end_col': self.worksheet.max_column
        }
    
    def _find_logo_position(self) -> Dict[str, Any]:
        """查找Logo位置"""
        # 查找可能的Logo位置（通常在左上角或包含"LOGO"文本的区域）
        for row in range(1, min(6, self.worksheet.max_row + 1)):
            for col in range(1, min(6, self.worksheet.max_column + 1)):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value and 'logo' in str(cell.value).lower():
                    return {'row': row, 'column': col}
        
        # 默认左上角位置
        return {'row': 1, 'column': 1}
    
    def _find_title_position(self) -> Dict[str, Any]:
        """查找标题位置"""
        # 查找字体最大或包含"报价单"等关键词的单元格
        max_font_size = 0
        title_position = {'row': 1, 'column': 1}
        
        for row in range(1, min(10, self.worksheet.max_row + 1)):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value and cell.font.size:
                    if cell.font.size > max_font_size:
                        max_font_size = cell.font.size
                        title_position = {'row': row, 'column': col}
                    
                    # 检查是否包含标题关键词
                    value_str = str(cell.value).lower()
                    if any(keyword in value_str for keyword in ['报价单', '报价', 'quotation', '发票', 'invoice']):
                        return {'row': row, 'column': col}
        
        return title_position
    
    def _has_full_border(self, cell) -> bool:
        """检查单元格是否有完整边框"""
        border = cell.border
        return (border.left.style and border.right.style and 
                border.top.style and border.bottom.style)
    
    def _expand_table_region(self, start_row: int, start_col: int) -> Dict[str, Any]:
        """扩展表格区域"""
        # 简单实现，待完善
        return {
            'start_row': start_row,
            'end_row': start_row + 5,  # 假设表格高度
            'start_col': start_col,
            'end_col': start_col + 8   # 假设表格宽度
        }
    
    def _identify_field_type(self, value: str) -> str:
        """识别字段类型"""
        value_lower = value.lower()
        
        if any(keyword in value_lower for keyword in ['编号', 'number', 'no']):
            return 'number_field'
        elif any(keyword in value_lower for keyword in ['日期', 'date', '时间', 'time']):
            return 'date_field'
        elif any(keyword in value_lower for keyword in ['金额', 'amount', '价格', 'price', '总计', 'total']):
            return 'currency_field'
        elif any(keyword in value_lower for keyword in ['公司', 'company', '客户', 'customer']):
            return 'company_field'
        elif any(keyword in value_lower for keyword in ['项目', 'project']):
            return 'project_field'
        elif any(keyword in value_lower for keyword in ['产品', 'product', '名称', 'name']):
            return 'product_field'
        else:
            return 'text_field'
    
    def _suggest_field_mapping(self, value: str, field_type: str) -> str:
        """建议字段映射"""
        mapping_suggestions = {
            'number_field': 'quotation.quotation_number',
            'date_field': 'quotation.created_at',
            'currency_field': 'quotation.amount',
            'company_field': 'quotation.project.company.company_name',
            'project_field': 'quotation.project.project_name',
            'product_field': 'quotation.details[].product_name'
        }
        return mapping_suggestions.get(field_type, 'custom_field')
    
    def _get_most_common_font(self) -> Dict[str, Any]:
        """获取最常用的字体"""
        font_usage = {}
        
        for row in range(1, self.worksheet.max_row + 1):
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row, column=col)
                font_key = f"{cell.font.name}_{cell.font.size}"
                font_usage[font_key] = font_usage.get(font_key, 0) + 1
        
        if font_usage:
            most_common = max(font_usage, key=font_usage.get)
            name, size = most_common.split('_')
            return {'name': name, 'size': float(size)}
        
        return {'name': '宋体', 'size': 12}
    
    def _get_header_style(self) -> Dict[str, Any]:
        """获取表头样式"""
        # 分析前几行的样式作为表头样式
        return {}
    
    def _get_table_style(self) -> Dict[str, Any]:
        """获取表格样式"""
        return {}
    
    def _get_border_styles(self) -> Dict[str, Any]:
        """获取边框样式"""
        return {}
    
    def _extract_color_scheme(self) -> Dict[str, Any]:
        """提取色彩方案"""
        return {}
    
    def save_config(self, output_path: str):
        """保存配置到JSON文件"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(self.template_config, f, ensure_ascii=False, indent=2, default=str)
            print(f"✅ 配置已保存到: {output_path}")
        except Exception as e:
            logger.error(f"❌ 保存配置失败: {e}")
            raise
    
    def generate_mapping_template(self) -> Dict[str, str]:
        """生成字段映射模板"""
        mapping_template = {}
        
        if 'data_fields' in self.template_config:
            for cell_ref, field_info in self.template_config['data_fields'].items():
                mapping_template[cell_ref] = field_info['suggested_mapping']
        
        return mapping_template