from flask import Blueprint, render_template, render_template_string, request, jsonify, url_for, send_file
from types import SimpleNamespace
from flask_login import login_required, current_user
from flask_babel import gettext as _
from app.decorators import permission_required
from app.permissions import has_permission
from app.models.quotation import Quotation, QuotationDetail
from app.models.project import Project
from app.models.product import Product
from app.models.user import User, Affiliation
from app.utils.access_control import get_viewable_data
from app.components.stage_analytics import StageAnalyticsComponent
from app.utils.chinese_mapping_manager import mapping_manager
from sqlalchemy import func, and_, or_, extract, case
from datetime import datetime, timedelta
from app import db
import logging
from app.utils.dictionary_helpers import project_stage_label
from config import Config

logger = logging.getLogger(__name__)

product_analysis = Blueprint('product_analysis', __name__)

def _get_latest_quotation_subquery():
    """
    创建最新报价单子查询

    返回每个项目的最新报价单ID（基于报价单创建时间）
    这确保了当一个项目有多个报价单时，只统计最新的那一个

    返回: 包含 project_id 和 latest_quotation_id 的子查询
    """
    latest_quotation_subq = db.session.query(
        Quotation.project_id,
        func.max(Quotation.id).label('latest_quotation_id')
    ).group_by(Quotation.project_id).subquery()

    return latest_quotation_subq

def _calculate_product_analysis_stats(base_query, current_month=None):
    """
    模组化的产品分析统计计算函数
    
    参数:
    - base_query: 基础查询对象（已应用权限和筛选条件）
    - current_month: 本月开始时间（可选，用于计算本月新增）
    
    返回: 包含各项统计数据的字典
    """
    try:
        # 计算总体统计
        total_stats = base_query.with_entities(
            func.sum(QuotationDetail.total_price).label('total_amount'),
            func.sum(QuotationDetail.quantity).label('total_quantity'),
            func.count(QuotationDetail.id).label('record_count')
        ).first()
        
        total_amount = float(total_stats.total_amount) if total_stats.total_amount else 0
        total_quantity = int(total_stats.total_quantity) if total_stats.total_quantity else 0
        
        # 计算平均单价（每个单位的价格，不需要除以10000）
        avg_unit_price = (total_amount / total_quantity) if total_quantity > 0 else 0
        
        # 计算本月新增数据（如果提供了current_month）
        monthly_quantity = 0
        monthly_amount = 0
        if current_month:
            monthly_stats = base_query.filter(
                QuotationDetail.created_at >= current_month
            ).with_entities(
                func.sum(QuotationDetail.total_price).label('monthly_amount'),
                func.sum(QuotationDetail.quantity).label('monthly_quantity')
            ).first()
            
            monthly_quantity = int(monthly_stats.monthly_quantity) if monthly_stats.monthly_quantity else 0
            monthly_amount = float(monthly_stats.monthly_amount) if monthly_stats.monthly_amount else 0
        
        return {
            'total_amount': total_amount,           # 总金额（原始金额，元）
            'total_quantity': total_quantity,       # 总数量
            'avg_unit_price': avg_unit_price,       # 平均单价（元/个）
            'monthly_quantity': monthly_quantity,   # 本月新增数量
            'monthly_amount': monthly_amount,       # 本月新增金额（原始金额，元）
            'record_count': int(total_stats.record_count) if total_stats.record_count else 0
        }
        
    except Exception as e:
        logger.error(f"统计计算失败: {e}")
        return {
            'total_amount': 0,
            'total_quantity': 0,
            'avg_unit_price': 0,
            'monthly_quantity': 0,
            'monthly_amount': 0,
            'record_count': 0
        }

def apply_permission_based_filters(query, current_user, quotation_alias=Quotation, project_alias=Project):
    """
    基于四级权限系统应用数据过滤
    """
    # 如果没有用户，返回空查询
    if not current_user:
        return query.filter(False)
    
    # 检查是否是管理员或CEO
    from app.permissions import is_admin_or_ceo
    if is_admin_or_ceo():
        return query
    
    # 检查用户是否有植入分析查看权限
    if not (hasattr(current_user, 'has_permission') and current_user.has_permission('product_analysis', 'view')):
        # 如果没有权限，返回空查询
        return query.filter(False)

    # 获取权限级别
    permission_level = current_user.get_permission_level('product_analysis')
    
    if permission_level == 'system':
        # 系统级权限：可以查看所有数据
        return query
    elif permission_level == 'company' and current_user.company_name:
        # 企业级权限：可以查看企业下所有数据
        company_user_ids = [u.id for u in User.query.filter_by(company_name=current_user.company_name).all()]
        return query.filter(project_alias.owner_id.in_(company_user_ids))
    elif permission_level == 'department' and current_user.department and current_user.company_name:
        # 部门级权限：可以查看部门下所有数据
        dept_user_ids = [u.id for u in User.query.filter(
            User.department == current_user.department,
            User.company_name == current_user.company_name
        ).all()]
        return query.filter(project_alias.owner_id.in_(dept_user_ids))
    else:
        # 个人级权限：应用传统的权限过滤逻辑
        permission_filters = []
        
        # 1. 自己创建的报价单
        permission_filters.append(quotation_alias.owner_id == current_user.id)
        
        # 2. 归属关系 - 使用子查询优化
        affiliation_subquery = db.session.query(Affiliation.owner_id).filter(
            Affiliation.viewer_id == current_user.id
        ).scalar_subquery()
        permission_filters.append(quotation_alias.owner_id.in_(affiliation_subquery))
        
        # 3. 销售负责人相关项目
        permission_filters.append(project_alias.vendor_sales_manager_id == current_user.id)
        
        # 4. 基于权限配置的项目类型过滤
        # 注：各角色可见的项目类型通过权限配置系统的 content_filters 控制
        # 例如：渠道经理配置 content_filters = {"project_type": ["channel_follow"]}
        permission = current_user.get_permission_config('product_analysis')
        if permission and hasattr(permission, 'content_filters') and permission.content_filters:
            content_filters = permission.content_filters
            if isinstance(content_filters, dict) and 'project_type' in content_filters:
                allowed_types = content_filters.get('project_type', [])
                if allowed_types:
                    permission_filters.append(project_alias.project_type.in_(allowed_types))
        
        # 应用权限过滤条件
        if permission_filters:
            return query.filter(or_(*permission_filters))
        else:
            # 如果没有任何权限，返回空查询
            return query.filter(False)

# 阶段配置现在统一由 stage_configs.py 管理

def get_stage_label(stage_key):
    """获取阶段中文标签"""
    return project_stage_label(stage_key, 'zh')

# 创建阶段统计组件实例
def create_stage_analytics_component():
    """创建产品分析阶段统计组件"""
    # 数据源配置
    data_source_config = {
        'primary_table': Project,
        'join_tables': [
            {
                'table': Quotation,
                'join_on': 'Quotation.project_id == Project.id'
            },
            {
                'table': QuotationDetail, 
                'join_on': 'QuotationDetail.quotation_id == Quotation.id'
            },
            {
                'table': User,
                'join_on': 'Quotation.owner_id == User.id'
            }
        ],
        'aggregate_fields': {
            'quantity': 'func.sum(QuotationDetail.quantity)',
            'amount': 'func.sum(QuotationDetail.total_price)'
        },
        'filter_fields': {
            'category': {
                'type': 'subquery',
                'main_table': QuotationDetail,
                'main_fields': ['product_name', 'product_model'],
                'subquery': {
                    'table': Product,
                    'fields': ['product_name', 'model'],
                    'condition': {'field': 'category'}
                }
            },
            'product_name': {
                'type': 'direct',
                'table': QuotationDetail,
                'field': 'product_name'
            },
            'product_model': {
                'type': 'direct', 
                'table': QuotationDetail,
                'field': 'product_model'
            }
        }
    }
    
    # 权限配置
    permission_config = {
        'filter_function': apply_permission_based_filters
    }
    
    return StageAnalyticsComponent(
        stage_config_name='project',
        data_source_config=data_source_config,
        permission_config=permission_config
    )

@product_analysis.route('/analysis')
@login_required
@permission_required('product_analysis', 'view')
def analysis():
    """产品分析主页面 - Tailwind 仪表板"""
    from app.services.product_attribution import get_analysis_view_scope
    scope = get_analysis_view_scope(current_user)
    return render_template('product_analysis/tw_analysis.html', view_scope=scope, now_year=datetime.now().year)

@product_analysis.route('/api/filter_options')
@login_required
@permission_required('product_analysis', 'view')
def get_filter_options():
    """获取筛选选项的联动数据"""
    try:
        category = request.args.get('category')
        product_name = request.args.get('product_name')
        
        # 构建基础查询
        query = db.session.query(Product)

        # 获取产品类别选项（不受其他筛选条件影响）
        from app.models.product_code import ProductCategory
        categories = db.session.query(ProductCategory.name).order_by(
            ProductCategory.id
        ).all()

        # 根据已选择的条件进行筛选
        if category:
            # 通过ProductCategory关联过滤
            category_obj = ProductCategory.query.filter_by(name=category).first()
            if category_obj:
                query = query.filter(Product.category_id == category_obj.id)
        if product_name:
            query = query.filter(Product.product_name == product_name)
        
        # 获取产品名称选项
        product_names = query.with_entities(Product.product_name).distinct().filter(
            Product.product_name.isnot(None)
        ).order_by(Product.product_name).all()
        
        # 获取产品型号选项
        product_models = query.with_entities(Product.model).distinct().filter(
            Product.model.isnot(None)
        ).order_by(Product.model).all()
        
        return jsonify({
            'success': True,
            'categories': [c[0] for c in categories],
            'product_names': [p[0] for p in product_names],
            'product_models': [m[0] for m in product_models]
        })
        
    except Exception as e:
        logger.error(f"获取筛选选项失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@product_analysis.route('/api/analysis_data')
@login_required
@permission_required('product_analysis', 'view')
def get_analysis_data():
    """获取产品分析数据 - 性能优化版本，只统计每个项目的最新报价单"""
    try:
        # 获取筛选参数
        category = request.args.get('category')
        product_name = request.args.get('product_name')
        product_model = request.args.get('product_model')

        # 获取分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))  # 默认每页50条

        # 获取最新报价单子查询
        latest_quotation_subq = _get_latest_quotation_subquery()

        # 性能优化：直接在主查询中处理权限逻辑，避免先查询所有可见报价单
        # 构建基础查询 - 避免因Product表重复记录导致的重复统计
        # 重要：只查询最新报价单的产品明细
        query = db.session.query(
            QuotationDetail.id,
            QuotationDetail.product_name,
            QuotationDetail.product_model,
            QuotationDetail.product_desc,
            QuotationDetail.quantity,
            QuotationDetail.discount,
            QuotationDetail.unit_price,
            QuotationDetail.total_price,
            QuotationDetail.product_mn,
            User,  # 获取完整的User对象
            Project.id.label('project_id'),
            Project.project_name,
            Project.current_stage,
            Quotation.id.label('quotation_id'),
            Quotation.quotation_number,
            QuotationDetail.updated_at,
            QuotationDetail.created_at
        ).join(
            Quotation, QuotationDetail.quotation_id == Quotation.id
        ).join(
            latest_quotation_subq,
            and_(
                Quotation.project_id == latest_quotation_subq.c.project_id,
                Quotation.id == latest_quotation_subq.c.latest_quotation_id
            )
        ).join(
            Project, Quotation.project_id == Project.id
        ).join(
            User, Quotation.owner_id == User.id
        )
        
        # 应用基于权限系统的数据过滤
        query = apply_permission_based_filters(query, current_user)

        # 应用筛选条件
        if category:
            # 使用子查询获取指定类别的产品，避免因Product表重复记录导致的重复统计
            from app.models.product_code import ProductCategory

            category_obj = ProductCategory.query.filter_by(name=category).first()
            if category_obj:
                category_products = db.session.query(
                    Product.product_name,
                    Product.model
                ).filter(
                    Product.category_id == category_obj.id
                ).distinct().subquery()

                query = query.filter(
                    and_(
                        QuotationDetail.product_name == category_products.c.product_name,
                        QuotationDetail.product_model == category_products.c.model
                    )
                )
        
        if product_name and product_name.strip():
            query = query.filter(QuotationDetail.product_name == product_name)
        
        if product_model and product_model.strip():
            query = query.filter(QuotationDetail.product_model == product_model)
        
        # 性能优化：合并 count + stats + monthly 为单次聚合查询（原为4次独立查询）
        current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        stats_result = query.with_entities(
            func.count(QuotationDetail.id).label('total_count'),
            func.sum(QuotationDetail.total_price).label('total_amount'),
            func.sum(QuotationDetail.quantity).label('total_quantity'),
            func.sum(
                case(
                    (QuotationDetail.created_at >= current_month, QuotationDetail.quantity),
                    else_=0
                )
            ).label('monthly_quantity')
        ).first()

        total_count = int(stats_result.total_count) if stats_result.total_count else 0
        total_amount = float(stats_result.total_amount) if stats_result.total_amount else 0
        total_quantity = int(stats_result.total_quantity) if stats_result.total_quantity else 0
        monthly_increase = int(stats_result.monthly_quantity) if stats_result.monthly_quantity else 0
        avg_unit_price = (total_amount / 10000 / total_quantity) if total_quantity > 0 else 0

        # 执行分页查询
        try:
            results = query.order_by(QuotationDetail.updated_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
        except Exception as e:
            logger.warning(f"使用updated_at排序失败: {str(e)}, 尝试使用id排序")
            try:
                db.session.rollback()
                results = query.order_by(QuotationDetail.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
            except Exception as e2:
                logger.error(f"产品分析查询失败: {str(e2)}")
                db.session.rollback()
                results = []

        # 格式化数据
        data = []
        for row in results:
            item = {
                'id': row.id,
                'product_name': row.product_name,
                'product_model': row.product_model,
                'product_desc': row.product_desc,
                'quantity': row.quantity,
                'discount': f"{row.discount * 100:.1f}%" if row.discount else "100.0%",
                'unit_price': float(row.unit_price) if row.unit_price else 0,
                'total_price': float(row.total_price) if row.total_price else 0,
                'product_mn': row.product_mn,
                'owner_name': row.owner_name,
                'owner_real_name': row.owner_real_name,
                'company_name': row.company_name,
                'project_id': row.project_id,
                'project_name': row.project_name,
                'current_stage': row.current_stage,
                'quotation_id': row.quotation_id,
                'quotation_number': row.quotation_number,
                'updated_at': row.updated_at.strftime('%Y-%m-%d %H:%M') if row.updated_at else '',
                'created_at': row.created_at.strftime('%Y-%m-%d %H:%M') if row.created_at else ''
            }
            data.append(item)
        
        # 计算分页信息
        total_pages = (total_count + per_page - 1) // per_page
        has_next = page < total_pages
        has_prev = page > 1
        
        return jsonify({
            'success': True,
            'data': data,
            'statistics': {
                'total_amount': total_amount,
                'total_count': total_quantity,
                'monthly_increase': monthly_increase,
                'avg_unit_price': avg_unit_price
            },
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': total_pages,
                'has_next': has_next,
                'has_prev': has_prev
            }
        })
        
    except Exception as e:
        logger.error(f"获取产品分析数据失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# 创建并注册阶段统计组件API端点
stage_analytics_component = create_stage_analytics_component()
stage_analytics_component.create_api_endpoint(
    product_analysis, 
    '/api/stage_statistics',
    permission_decorator=login_required
)


@product_analysis.route('/api/products/filter')
@login_required
@permission_required('product_analysis', 'view')
def products_list_ajax():
    """植入产品分析AJAX筛选端点 - 只统计每个项目的最新报价单"""
    try:
        # 获取筛选参数
        search = request.args.get('search', '').strip()
        category = request.args.get('category', '').strip()
        product_name = request.args.get('product_name', '').strip()
        product_model = request.args.get('product_model', '').strip()

        # 获取分页参数
        offset = request.args.get('offset', 0, type=int)
        limit = request.args.get('limit', 50, type=int)

        # 排序参数
        sort_field = request.args.get('sort_field', '')
        sort_direction = request.args.get('sort_direction', 'asc')

        # 获取最新报价单子查询
        latest_quotation_subq = _get_latest_quotation_subquery()

        # 构建基础查询 - 避免因Product表重复记录导致的重复统计
        # 重要：只查询最新报价单的产品明细
        query = db.session.query(
            QuotationDetail.id,
            QuotationDetail.product_name,
            QuotationDetail.product_model,
            QuotationDetail.product_desc,
            QuotationDetail.quantity,
            QuotationDetail.discount,
            QuotationDetail.unit_price,
            QuotationDetail.total_price,
            QuotationDetail.product_mn,
            User,  # 获取完整的User对象
            Project.id.label('project_id'),
            Project.project_name,
            Project.current_stage,
            Quotation.id.label('quotation_id'),
            Quotation.quotation_number,
            QuotationDetail.updated_at,
            QuotationDetail.created_at
        ).join(
            Quotation, QuotationDetail.quotation_id == Quotation.id
        ).join(
            latest_quotation_subq,
            and_(
                Quotation.project_id == latest_quotation_subq.c.project_id,
                Quotation.id == latest_quotation_subq.c.latest_quotation_id
            )
        ).join(
            Project, Quotation.project_id == Project.id
        ).join(
            User, Quotation.owner_id == User.id
        )
        
        # 应用基于权限系统的数据过滤
        query = apply_permission_based_filters(query, current_user)
        
        # 应用搜索过滤
        if search:
            search_filter = or_(
                QuotationDetail.product_name.ilike(f'%{search}%'),
                QuotationDetail.product_model.ilike(f'%{search}%'),
                QuotationDetail.product_desc.ilike(f'%{search}%'),
                QuotationDetail.product_mn.ilike(f'%{search}%'),
                Project.project_name.ilike(f'%{search}%'),
                Quotation.quotation_number.ilike(f'%{search}%'),
                User.username.ilike(f'%{search}%'),
                User.real_name.ilike(f'%{search}%')
            )
            query = query.filter(search_filter)

        # 应用筛选条件
        if category:
            # 使用子查询获取指定类别的产品，避免因Product表重复记录导致的重复统计
            from app.models.product_code import ProductCategory

            category_obj = ProductCategory.query.filter_by(name=category).first()
            if category_obj:
                category_products = db.session.query(
                    Product.product_name,
                    Product.model
                ).filter(
                    Product.category_id == category_obj.id
                ).distinct().subquery()

                query = query.filter(
                    and_(
                        QuotationDetail.product_name == category_products.c.product_name,
                        QuotationDetail.product_model == category_products.c.model
                    )
                )
        
        if product_name and product_name.strip():
            query = query.filter(QuotationDetail.product_name == product_name)
        
        if product_model and product_model.strip():
            query = query.filter(QuotationDetail.product_model == product_model)
        
        # 获取总数（用于分页）
        total_count = query.count()
        
        # 动态排序（产品分析复杂查询需要特殊处理）
        if sort_field and sort_direction:
            print(f"🔍 产品分析排序调试 - 字段: {sort_field}, 方向: {sort_direction}")
            
            # 字段映射（已经JOIN的表字段）
            field_mapping = {
                # QuotationDetail 表字段
                'product_name': QuotationDetail.product_name,
                'product_model': QuotationDetail.product_model,
                'quantity': QuotationDetail.quantity,
                'unit_price': QuotationDetail.unit_price,
                'total_price': QuotationDetail.total_price,
                'product_mn': QuotationDetail.product_mn,
                'created_at': QuotationDetail.created_at,
                'updated_at': QuotationDetail.updated_at,
                'discount': QuotationDetail.discount,
                # 关联表字段（已JOIN）
                'owner_name': User.real_name,  # 负责人名称
                'project_name': Project.project_name,  # 项目名称
                'current_stage': Project.current_stage,  # 项目阶段
                'quotation_number': Quotation.quotation_number  # 报价单编号
            }
            
            if sort_field in field_mapping:
                sort_column = field_mapping[sort_field]
                if sort_direction.lower() == 'desc':
                    query = query.order_by(sort_column.desc())
                    print(f"✅ 应用降序排序: {sort_field}")
                else:
                    query = query.order_by(sort_column.asc())
                    print(f"✅ 应用升序排序: {sort_field}")
            else:
                print(f"❌ 未知排序字段: {sort_field}")
                # 默认排序
                query = query.order_by(QuotationDetail.updated_at.desc())
        else:
            print("📊 使用默认排序: updated_at desc")
            # 默认排序
            query = query.order_by(QuotationDetail.updated_at.desc())
        
        # 执行分页查询
        try:
            results = query.offset(offset).limit(limit).all()
            logger.info(f"植入产品分析查询成功: 总计 {total_count} 条，offset={offset}, limit={limit}, 返回 {len(results)} 条")
        except Exception as e:
            logger.warning(f"查询失败: {str(e)}, 尝试使用id排序")
            try:
                # 回滚失败的事务
                db.session.rollback()
                results = query.order_by(QuotationDetail.id.desc()).offset(offset).limit(limit).all()
                logger.info(f"植入产品分析查询(id排序)成功: 总计 {total_count} 条，offset={offset}, limit={limit}, 返回 {len(results)} 条")
            except Exception as e2:
                logger.error(f"产品分析查询失败: {str(e2)}")
                # 回滚失败的事务
                db.session.rollback()
                results = []
        
        # 使用模板渲染表格行HTML
        
        if not results:
            from app.utils.mobile_helpers import is_mobile_request
            if is_mobile_request():
                html = '<div class="text-center py-4">暂无符合条件的数据</div>'
            else:
                html = '<tr><td colspan="13" class="text-center py-4">暂无符合条件的数据</td></tr>'
        else:
            # 创建简单的数据结构供模板使用
            formatted_results = []
            for row in results:
                # 使用命名空间对象，更简单可靠
                # 注意：现在row[9]是完整的User对象，row[10]开始是project_id等
                formatted_row = SimpleNamespace(
                    id=row[0],
                    product_name=row[1],
                    product_model=row[2],
                    product_desc=row[3],
                    quantity=row[4],
                    discount=row[5],
                    unit_price=row[6],
                    total_price=row[7],
                    product_mn=row[8],
                    owner=row[9],  # 完整的User对象
                    project_id=row[10],
                    project_name=row[11],
                    current_stage=row[12],
                    current_stage_label=get_stage_label(row[12] or 'unset'),
                    quotation_id=row[13],
                    quotation_number=row[14],
                    updated_at=row[15],
                    created_at=row[16]
                )
                formatted_results.append(formatted_row)
            
            # 检测移动端并使用智能移动卡片
            from app.utils.mobile_helpers import is_mobile_request
            
            if is_mobile_request():
                # 智能移动卡片配置 - 产品植入分析
                smart_mobile_card = {
                    'module': 'product_analysis',
                    'title_field': {'field': 'product_name'},
                    'badges': [
                        {'field': 'current_stage_label', 'renderer': 'project_stage'}
                    ],
                    'details': [
                        {'field': 'product_model', 'label': '产品型号'},
                        {'field': 'product_mn', 'label': '产品料号'},
                        {'field': 'project_name', 'label': '关联项目', 'renderer': 'project_link'},
                        {'field': 'quotation_number', 'label': '报价单号'},
                        {'field': 'quantity', 'label': '数量'},
                        {'field': 'total_price', 'label': '总价', 'format': 'currency'},
                        {'field': 'updated_at', 'label': '更新时间', 'format': 'date'}
                    ]
                }
                
                # 使用智能移动卡片模板渲染
                html = render_template_string('''
                {% from 'macros/ui_helpers.html' import render_smart_mobile_cards %}
                {{ render_smart_mobile_cards(items, card_config) }}
                ''', items=formatted_results, card_config=smart_mobile_card)
            else:
                # 桌面端使用传统表格行渲染
                html = render_template('product_analysis/product_analysis_rows_simple.html', items=formatted_results)
            logger.info(f"✅ 模板渲染成功，生成了 {len(formatted_results)} 行数据")
        
        # 使用模组化函数计算统计信息
        current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ajax_stats_data = _calculate_product_analysis_stats(query, current_month)
        
        # 计算是否还有更多数据
        has_more = (offset + len(results)) < total_count
        
        return jsonify({
            'success': True,
            'html': html,
            'has_more': has_more,
            'total_count': total_count,
            'loaded_count': offset + len(results),
            'statistics': {
                # 使用模组化统计计算结果，转换为万元以兼容通用模组
                # 总项目数
                'total_count': ajax_stats_data['record_count'],
                'total_amount': round(ajax_stats_data['total_amount'] / 10000, 2),  # 转换为万元
                
                # 本月新增（包含数量和金额）
                'monthly_increase_count': ajax_stats_data['monthly_quantity'],
                'monthly_increase_amount': round(ajax_stats_data['monthly_amount'] / 10000, 2),  # 转换为万元
                
                # 平均单价
                'avg_unit_price_count': 1,  # 固定显示1个单位
                'avg_unit_price_amount': round(ajax_stats_data['avg_unit_price'] / 10000, 4)  # 转换为万元，保留4位小数
            }
        })
        
    except Exception as e:
        logger.error(f"植入产品分析AJAX筛选失败: {str(e)}", exc_info=True)
        return jsonify({
            'success': False, 
            'error': f'筛选失败: {str(e)}',
            'html': '<tr><td colspan="13" class="text-center py-4 text-danger">加载失败，请刷新页面重试</td></tr>'
        }), 500

@product_analysis.route('/api/export_analysis')
@login_required
@permission_required('product_analysis', 'view')
def export_analysis():
    """导出产品分析数据"""
    try:
        # 获取分析数据
        category = request.args.get('category')
        product_name = request.args.get('product_name')
        product_model = request.args.get('product_model')
        
        # 这里可以实现导出功能，例如导出为Excel
        # 暂时返回成功消息
        return jsonify({
            'success': True,
            'message': '导出功能正在开发中'
        })
        
    except Exception as e:
        logger.error(f"导出产品分析数据失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# V2 API — Tailwind 仪表板
# ============================================================================

def _build_base_detail_query():
    """构建基础明细查询（含最新报价单 JOIN，不含权限过滤）"""
    latest_subq = _get_latest_quotation_subquery()
    q = db.session.query(
        QuotationDetail.id,
        QuotationDetail.product_name,
        QuotationDetail.product_model,
        QuotationDetail.product_desc,
        QuotationDetail.quantity,
        QuotationDetail.unit_price,
        QuotationDetail.total_price,
        QuotationDetail.product_mn,
        QuotationDetail.created_at,
        Quotation.id.label('quotation_id'),
        Quotation.quotation_number,
        Quotation.owner_id.label('quotation_owner_id'),
        Project.id.label('project_id'),
        Project.project_name,
        Project.current_stage,
        Project.owner_id.label('project_owner_id'),
    ).join(
        Quotation, QuotationDetail.quotation_id == Quotation.id
    ).join(
        latest_subq,
        and_(
            Quotation.project_id == latest_subq.c.project_id,
            Quotation.id == latest_subq.c.latest_quotation_id
        )
    ).join(
        Project, Quotation.project_id == Project.id
    ).join(
        User, Quotation.owner_id == User.id
    )
    return q


def _apply_scope_filter(query, scope):
    """根据角色视角过滤查询

    admin  → 无过滤
    se     → 仅关联项目（ProjectMember role=solution_engineer）
    pm     → 仅管理分类下的产品型号
    sales  → 权限层过滤（自己/下属/vendor_sales_manager）
    """
    level = scope.get('level', 'sales')

    if level == 'admin':
        return query

    elif level == 'se':
        pids = scope.get('associated_project_ids', [])
        if pids:
            return query.filter(Project.id.in_(pids))
        return query.filter(False)

    elif level == 'pm':
        cat_ids = scope.get('managed_category_ids', [])
        if cat_ids:
            cat_products_subq = db.session.query(Product.model).filter(
                Product.category_id.in_(cat_ids)
            ).distinct().subquery()
            return query.filter(QuotationDetail.product_model.in_(
                db.session.query(cat_products_subq.c.model)
            ))
        return query.filter(False)

    else:  # sales
        return apply_permission_based_filters(query, current_user)


@product_analysis.route('/tw')
@login_required
@permission_required('product_analysis', 'view')
def tw_analysis():
    """向后兼容：重定向到主路由"""
    from flask import redirect
    return redirect(url_for('product_analysis.analysis'))


@product_analysis.route('/api/v2/overview')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_overview():
    """概览统计卡片数据"""
    try:
        from app.services.product_attribution import get_analysis_view_scope
        scope = get_analysis_view_scope(current_user)
        year = int(request.args.get('year', datetime.now().year))

        q = _build_base_detail_query()
        q = _apply_scope_filter(q, scope)

        # 年份过滤（year=0 表示全部年份）
        if year > 0:
            year_start = datetime(year, 1, 1)
            year_end = datetime(year + 1, 1, 1)
            q_year = q.filter(QuotationDetail.created_at >= year_start, QuotationDetail.created_at < year_end)
        else:
            q_year = q

        stats = q_year.with_entities(
            func.sum(QuotationDetail.total_price).label('total_amount'),
            func.sum(QuotationDetail.quantity).label('total_qty'),
            func.count(func.distinct(Project.id)).label('project_count')
        ).first()

        total_amount = float(stats.total_amount or 0)
        total_qty = int(stats.total_qty or 0)
        project_count = int(stats.project_count or 0)

        # 上月 vs 本月环比
        now = datetime.now()
        this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if now.month == 1:
            last_month_start = datetime(now.year - 1, 12, 1)
        else:
            last_month_start = datetime(now.year, now.month - 1, 1)

        this_month_amt = q.filter(
            QuotationDetail.created_at >= this_month_start
        ).with_entities(func.sum(QuotationDetail.total_price)).scalar() or 0

        last_month_amt = q.filter(
            QuotationDetail.created_at >= last_month_start,
            QuotationDetail.created_at < this_month_start
        ).with_entities(func.sum(QuotationDetail.total_price)).scalar() or 0

        mom_rate = round((float(this_month_amt) - float(last_month_amt)) / float(last_month_amt) * 100, 1) if last_month_amt else 0

        return jsonify({
            'success': True,
            'data': {
                'total_amount': total_amount,
                'total_quantity': total_qty,
                'project_count': project_count,
                'mom_rate': mom_rate,
                'this_month_amount': float(this_month_amt),
                'last_month_amount': float(last_month_amt),
            }
        })
    except Exception as e:
        logger.error(f"v2 overview 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/trend')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_trend():
    """月度趋势数据（滚动12个月：当月前6 + 当月 + 后5）"""
    try:
        from app.services.product_attribution import get_analysis_view_scope
        scope = get_analysis_view_scope(current_user)
        now = datetime.now()

        # 计算滚动窗口起止
        start_month = now.month - 6
        start_year = now.year
        if start_month <= 0:
            start_month += 12
            start_year -= 1
        start_date = datetime(start_year, start_month, 1)

        end_month = now.month + 6
        end_year = now.year
        if end_month > 12:
            end_month -= 12
            end_year += 1
        end_date = datetime(end_year, end_month, 1)

        q = _build_base_detail_query()
        q = _apply_scope_filter(q, scope)

        monthly = q.filter(
            QuotationDetail.created_at >= start_date,
            QuotationDetail.created_at < end_date
        ).with_entities(
            extract('year', QuotationDetail.created_at).label('yr'),
            extract('month', QuotationDetail.created_at).label('mn'),
            func.sum(QuotationDetail.total_price).label('amount'),
            func.sum(QuotationDetail.quantity).label('qty')
        ).group_by('yr', 'mn').order_by('yr', 'mn').all()

        monthly_map = {(int(r.yr), int(r.mn)): {'amount': float(r.amount or 0), 'qty': int(r.qty or 0)} for r in monthly}

        months = []
        y, m = start_year, start_month
        for _ in range(12):
            d = monthly_map.get((y, m), {'amount': 0, 'qty': 0})
            months.append({
                'year': y,
                'month': m,
                'label': f'{m}月',
                'amount': d['amount'],
                'quantity': d['qty'],
                'is_current': (y == now.year and m == now.month)
            })
            m += 1
            if m > 12:
                m = 1
                y += 1

        return jsonify({'success': True, 'data': months})
    except Exception as e:
        logger.error(f"v2 trend 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/distribution')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_distribution():
    """分布数据（按类别/阶段）"""
    try:
        from app.services.product_attribution import get_analysis_view_scope
        from app.models.product_code import ProductCategory
        scope = get_analysis_view_scope(current_user)
        year = int(request.args.get('year', datetime.now().year))

        q = _build_base_detail_query()
        q = _apply_scope_filter(q, scope)

        if year > 0:
            year_start = datetime(year, 1, 1)
            year_end = datetime(year + 1, 1, 1)
            q_year = q.filter(QuotationDetail.created_at >= year_start, QuotationDetail.created_at < year_end)
        else:
            q_year = q

        # 按阶段分布
        stage_dist = q_year.with_entities(
            Project.current_stage,
            func.sum(QuotationDetail.total_price).label('amount'),
            func.count(QuotationDetail.id).label('count')
        ).group_by(Project.current_stage).all()

        stages = [{
            'stage': r.current_stage or 'unknown',
            'label': get_stage_label(r.current_stage) if r.current_stage else '未知',
            'amount': float(r.amount or 0),
            'count': int(r.count or 0)
        } for r in stage_dist]

        # 按类别+产品分布（单次查询，Python 层聚合为嵌套结构）
        try:
            prod_rows = q_year.join(
                Product, QuotationDetail.product_model == Product.model
            ).join(
                ProductCategory, Product.category_id == ProductCategory.id
            ).with_entities(
                ProductCategory.name.label('cat_name'),
                QuotationDetail.product_model,
                func.max(QuotationDetail.product_name).label('product_name'),
                func.sum(QuotationDetail.total_price).label('amount'),
                func.count(QuotationDetail.id).label('count')
            ).group_by(
                ProductCategory.name,
                QuotationDetail.product_model
            ).order_by(
                ProductCategory.name,
                func.sum(QuotationDetail.total_price).desc()
            ).all()

            cat_map = {}
            for r in prod_rows:
                if r.cat_name not in cat_map:
                    cat_map[r.cat_name] = {'name': r.cat_name, 'amount': 0, 'count': 0, 'products': []}
                cat_map[r.cat_name]['amount'] += float(r.amount or 0)
                cat_map[r.cat_name]['count'] += int(r.count or 0)
                cat_map[r.cat_name]['products'].append({
                    'name': r.product_name,
                    'model': r.product_model,
                    'amount': float(r.amount or 0),
                    'count': int(r.count or 0)
                })
            categories = sorted(cat_map.values(), key=lambda x: x['amount'], reverse=True)
        except Exception:
            db.session.rollback()
            categories = []

        return jsonify({'success': True, 'data': {'stages': stages, 'categories': categories}})
    except Exception as e:
        logger.error(f"v2 distribution 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/ranking')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_ranking():
    """排名数据（含引用次数、系数、积分）— 显示所有活跃产品"""
    try:
        from app.services.product_attribution import get_analysis_view_scope
        scope = get_analysis_view_scope(current_user)
        year = int(request.args.get('year', datetime.now().year))

        # 1. 构建当年报价明细聚合子查询（保留权限过滤）
        q = _build_base_detail_query()
        q = _apply_scope_filter(q, scope)

        if year > 0:
            year_start = datetime(year, 1, 1)
            year_end = datetime(year + 1, 1, 1)
            q_year = q.filter(QuotationDetail.created_at >= year_start, QuotationDetail.created_at < year_end)
        else:
            q_year = q

        detail_stats = q_year.with_entities(
            QuotationDetail.product_mn,
            func.sum(QuotationDetail.total_price).label('amount'),
            func.sum(QuotationDetail.quantity).label('qty'),
            func.count(func.distinct(QuotationDetail.quotation_id)).label('citation_count')
        ).group_by(QuotationDetail.product_mn).subquery()

        # 2. 执行子查询并转为 dict
        stats_rows = db.session.query(
            detail_stats.c.product_mn,
            detail_stats.c.amount,
            detail_stats.c.qty,
            detail_stats.c.citation_count
        ).all()
        stats_map = {r.product_mn: r for r in stats_rows if r.product_mn}

        # 3. 查所有活跃产品
        products = Product.query.filter(Product.status != 'discontinued').all()

        # 4. 遍历所有产品，合并数据
        data = []
        for p in products:
            s = stats_map.get(p.product_mn)
            retail_price = float(p.retail_price) if p.retail_price else 0
            category = p.category_name
            data.append({
                'name': p.name or '',
                'model': p.model or '',
                'mn': p.product_mn or '',
                'amount': float(s.amount or 0) if s else 0,
                'quantity': int(s.qty or 0) if s else 0,
                'citations': int(s.citation_count or 0) if s else 0,
                'retail_price': retail_price,
                'category': category,
            })

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        logger.error(f"v2 ranking 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/coefficient-stats')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_coefficient_stats():
    """产品统计卡片数据（旧积分系数统计已废弃）"""
    try:
        active_products = Product.query.filter(Product.status != 'discontinued').all()
        total_active = len(active_products)

        return jsonify({'success': True, 'data': {
            'total_active': total_active,
            'with_citations': 0,
            'median_citations': 0,
            'avg_coefficient': 0,
            'attention_count': 0
        }})
    except Exception as e:
        logger.error(f"v2 coefficient-stats 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/detail')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_detail():
    """明细列表（支持级联产品筛选、阶段、负责人、订单量、关联客户）"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        q = _build_filtered_detail_query()
        total = q.count()

        # 汇总统计（供概览卡片同步）
        summary_row = q.with_entities(
            func.sum(QuotationDetail.total_price).label('total_amount'),
            func.sum(QuotationDetail.quantity).label('total_quantity'),
            func.count(func.distinct(Project.id)).label('project_count')
        ).first()
        summary = {
            'total_amount': float(summary_row.total_amount or 0),
            'total_quantity': int(summary_row.total_quantity or 0),
            'project_count': int(summary_row.project_count or 0)
        }

        sort_by = request.args.get('sort_by', 'created_at')
        sort_dir = request.args.get('sort_dir', 'desc')
        sort_map = {
            'product_name': QuotationDetail.product_name,
            'product_model': QuotationDetail.product_model,
            'quantity': QuotationDetail.quantity,
            'total_price': QuotationDetail.total_price,
            'project_name': Project.project_name,
            'created_at': QuotationDetail.created_at,
        }
        sort_col = sort_map.get(sort_by, QuotationDetail.created_at)
        order = sort_col.desc() if sort_dir == 'desc' else sort_col.asc()

        results = q.order_by(order).offset(
            (page - 1) * per_page
        ).limit(per_page).all()

        # 预加载用户名
        owner_ids_set = list({r.quotation_owner_id for r in results})
        owners = {u.id: u for u in User.query.filter(User.id.in_(owner_ids_set)).all()} if owner_ids_set else {}

        # 批量查询订单量：按 (project_id, product_mn) 聚合
        project_ids = list({r.project_id for r in results})
        product_mns = list({r.product_mn for r in results if r.product_mn})
        order_qty_map = {}
        if project_ids and product_mns:
            from app.models.pricing_order import PricingOrder, PricingOrderDetail
            order_rows = db.session.query(
                PricingOrder.project_id,
                PricingOrderDetail.product_mn,
                func.sum(PricingOrderDetail.quantity).label('qty')
            ).join(
                PricingOrder, PricingOrderDetail.pricing_order_id == PricingOrder.id
            ).filter(
                PricingOrder.project_id.in_(project_ids),
                PricingOrderDetail.product_mn.in_(product_mns)
            ).group_by(
                PricingOrder.project_id, PricingOrderDetail.product_mn
            ).all()
            for row in order_rows:
                order_qty_map[(row.project_id, row.product_mn)] = int(row.qty or 0)

        # 批量查询关联客户（分销商/代理商）
        from app.models.project_customer_association import ProjectCustomerAssociation
        from app.models.customer import Company
        from collections import defaultdict
        customer_map = {}
        if project_ids:
            assoc_rows = db.session.query(
                ProjectCustomerAssociation.project_id,
                Company.company_name,
                Company.company_type
            ).join(
                Company, ProjectCustomerAssociation.company_id == Company.id
            ).filter(
                ProjectCustomerAssociation.project_id.in_(project_ids),
                Company.company_type.in_(['distributor', 'dealer']),
                Company.is_deleted == False
            ).all()
            proj_customers = defaultdict(lambda: {'distributor': [], 'dealer': []})
            for row in assoc_rows:
                proj_customers[row.project_id][row.company_type].append(row.company_name)
            for pid, types in proj_customers.items():
                names = types['distributor'] + types['dealer']
                customer_map[pid] = ' / '.join(names)

        data = []
        for r in results:
            owner = owners.get(r.quotation_owner_id)
            data.append({
                'product_name': r.product_name or '',
                'product_model': r.product_model or '',
                'product_mn': r.product_mn or '',
                'quantity': int(r.quantity or 0),
                'order_qty': order_qty_map.get((r.project_id, r.product_mn), 0),
                'unit_price': float(r.unit_price or 0),
                'total_price': float(r.total_price or 0),
                'project_name': r.project_name or '',
                'project_id': r.project_id,
                'stage': get_stage_label(r.current_stage) if r.current_stage else '',
                'customer': customer_map.get(r.project_id, ''),
                'quotation_number': r.quotation_number or '',
                'quotation_id': r.quotation_id,
                'owner': owner.real_name if owner else '',
                'created_at': r.created_at.strftime('%Y-%m-%d') if r.created_at else ''
            })

        return jsonify({
            'success': True,
            'data': data,
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'summary': summary
        })
    except Exception as e:
        logger.error(f"v2 detail 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/filter_options')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_filter_options():
    """获取筛选选项数据（分类/子分类/产品型号级联 + 阶段 + 负责人）"""
    try:
        from app.models.product_code import ProductCategory, ProductSubcategory

        category_id = request.args.get('category_id', type=int)
        subcategory_id = request.args.get('subcategory_id', type=int)

        # 分类列表（始终返回）
        categories = ProductCategory.query.order_by(
            ProductCategory.display_order, ProductCategory.id
        ).all()
        cat_list = [{'id': c.id, 'name': c.name} for c in categories]

        # 子分类列表（当指定 category_id 时）
        sub_list = []
        if category_id:
            subs = ProductSubcategory.query.filter_by(
                category_id=category_id
            ).order_by(ProductSubcategory.display_order, ProductSubcategory.id).all()
            sub_list = [{'id': s.id, 'name': s.name} for s in subs]

        # 产品型号列表（当指定 subcategory_id 时，按 model 去重）
        model_list = []
        if subcategory_id:
            products = Product.query.filter(
                Product.subcategory_id == subcategory_id,
                Product.is_deleted == False
            ).order_by(Product.product_name).all()
            seen_model = set()
            for p in products:
                if p.model and p.model not in seen_model:
                    seen_model.add(p.model)
                    model_list.append({'model': p.model, 'name': p.product_name})

        # 阶段列表（静态）
        stages = [
            {'value': 'discover', 'label': _('发现')},
            {'value': 'embed', 'label': _('植入')},
            {'value': 'pre_tender', 'label': _('预招标')},
            {'value': 'tendering', 'label': _('招标中')},
            {'value': 'quoted', 'label': _('已报价')},
            {'value': 'awarded', 'label': _('中标')},
            {'value': 'signed', 'label': _('签约')},
            {'value': 'lost', 'label': _('丢失')},
            {'value': 'paused', 'label': _('暂停')},
        ]

        # 负责人列表（从现有报价单数据中提取活跃用户）
        from app.services.product_attribution import get_analysis_view_scope
        scope = get_analysis_view_scope(current_user)
        q = _build_base_detail_query()
        q = _apply_scope_filter(q, scope)
        owner_ids = [r[0] for r in q.with_entities(
            func.distinct(Quotation.owner_id)
        ).all()]
        owners_list = User.query.filter(User.id.in_(owner_ids)).order_by(User.real_name).all() if owner_ids else []
        owner_list = [{'id': u.id, 'name': u.real_name or u.username} for u in owners_list]

        # 关联客户列表（分销商/代理商）
        from app.models.project_customer_association import ProjectCustomerAssociation
        from app.models.customer import Company
        project_ids = [r[0] for r in q.with_entities(func.distinct(Project.id)).all()]
        customer_list = []
        if project_ids:
            customers = db.session.query(
                Company.id, Company.company_name, Company.company_type
            ).join(
                ProjectCustomerAssociation, ProjectCustomerAssociation.company_id == Company.id
            ).filter(
                ProjectCustomerAssociation.project_id.in_(project_ids),
                Company.company_type.in_(['distributor', 'dealer']),
                Company.is_deleted == False
            ).distinct().order_by(Company.company_name).all()
            customer_list = [{'id': c.id, 'name': c.company_name, 'type': c.company_type} for c in customers]

        return jsonify({
            'success': True,
            'categories': cat_list,
            'subcategories': sub_list,
            'models': model_list,
            'stages': stages,
            'owners': owner_list,
            'customers': customer_list
        })
    except Exception as e:
        logger.error(f"filter_options 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@product_analysis.route('/api/v2/quotation_items/<int:quotation_id>')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_quotation_items(quotation_id):
    """获取报价单明细产品信息（不含价格）"""
    try:
        quotation = Quotation.query.get_or_404(quotation_id)
        project = Project.query.get(quotation.project_id) if quotation.project_id else None

        items = QuotationDetail.query.filter_by(
            quotation_id=quotation_id
        ).order_by(QuotationDetail.id).all()

        data = []
        for item in items:
            data.append({
                'id': item.id,
                'product_name': item.product_name or '',
                'product_model': item.product_model or '',
                'product_mn': item.product_mn or '',
                'brand': item.brand or '',
                'unit': item.unit or '',
                'quantity': int(item.quantity or 0),
                'product_desc': item.product_desc or '',
                'is_accessory': item.is_accessory,
                'parent_item_id': item.parent_item_id,
            })

        return jsonify({
            'success': True,
            'quotation_number': quotation.quotation_number or '',
            'project_name': project.project_name if project else '',
            'items': data
        })
    except Exception as e:
        logger.error(f"获取报价单明细失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def _build_filtered_detail_query():
    """从请求参数构建带筛选的明细查询，供 detail 和 export 共用"""
    from app.services.product_attribution import get_analysis_view_scope
    scope = get_analysis_view_scope(current_user)

    year = int(request.args.get('year', datetime.now().year))
    category = request.args.get('category')
    category_id = request.args.get('category_id', type=int)
    subcategory_id = request.args.get('subcategory_id', type=int)
    product_mn = request.args.get('product_mn', '').strip()
    product_name = request.args.get('product_name')
    product_model = request.args.get('product_model')
    stages_str = request.args.get('stages', '').strip()
    owner_id = request.args.get('owner_id', type=int)
    company_id = request.args.get('company_id', type=int)
    search = request.args.get('search', '').strip()

    q = _build_base_detail_query()
    q = _apply_scope_filter(q, scope)

    if year > 0:
        year_start = datetime(year, 1, 1)
        year_end = datetime(year + 1, 1, 1)
        q = q.filter(QuotationDetail.created_at >= year_start, QuotationDetail.created_at < year_end)

    if product_model:
        q = q.filter(QuotationDetail.product_model == product_model)
    elif product_mn:
        q = q.filter(QuotationDetail.product_mn == product_mn)
    elif subcategory_id:
        sub_mns = db.session.query(Product.product_mn).filter(
            Product.subcategory_id == subcategory_id, Product.is_deleted == False
        ).distinct().subquery()
        q = q.filter(QuotationDetail.product_mn.in_(db.session.query(sub_mns.c.product_mn)))
    elif category_id:
        cat_mns = db.session.query(Product.product_mn).filter(
            Product.category_id == category_id, Product.is_deleted == False
        ).distinct().subquery()
        q = q.filter(QuotationDetail.product_mn.in_(db.session.query(cat_mns.c.product_mn)))
    elif category:
        from app.models.product_code import ProductCategory
        cat = ProductCategory.query.filter_by(name=category).first()
        if cat:
            cat_models = db.session.query(Product.model).filter(
                Product.category_id == cat.id
            ).distinct().subquery()
            q = q.filter(QuotationDetail.product_model.in_(db.session.query(cat_models.c.model)))

    if stages_str:
        stages_list = [s.strip() for s in stages_str.split(',') if s.strip()]
        if len(stages_list) == 1:
            q = q.filter(Project.current_stage == stages_list[0])
        elif stages_list:
            q = q.filter(Project.current_stage.in_(stages_list))
    if owner_id:
        q = q.filter(Quotation.owner_id == owner_id)
    if company_id:
        from app.models.project_customer_association import ProjectCustomerAssociation
        assoc_pids = db.session.query(ProjectCustomerAssociation.project_id).filter(
            ProjectCustomerAssociation.company_id == company_id
        ).subquery()
        q = q.filter(Project.id.in_(db.session.query(assoc_pids.c.project_id)))
    if product_name:
        q = q.filter(QuotationDetail.product_name == product_name)
    if search:
        q = q.filter(or_(
            QuotationDetail.product_name.ilike(f'%{search}%'),
            QuotationDetail.product_model.ilike(f'%{search}%'),
            Project.project_name.ilike(f'%{search}%')
        ))

    return q


@product_analysis.route('/api/v2/export')
@login_required
@permission_required('product_analysis', 'view')
def api_v2_export():
    """导出明细数据为 Excel（仅管理员）"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': _('仅管理员可使用导出功能')}), 403
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from collections import defaultdict

        q = _build_filtered_detail_query()

        sort_by = request.args.get('sort_by', 'created_at')
        sort_dir = request.args.get('sort_dir', 'desc')
        sort_map = {
            'product_name': QuotationDetail.product_name,
            'product_model': QuotationDetail.product_model,
            'quantity': QuotationDetail.quantity,
            'total_price': QuotationDetail.total_price,
            'project_name': Project.project_name,
            'created_at': QuotationDetail.created_at,
        }
        sort_col = sort_map.get(sort_by, QuotationDetail.created_at)
        order = sort_col.desc() if sort_dir == 'desc' else sort_col.asc()
        results = q.order_by(order).all()

        if not results:
            return jsonify({'success': False, 'message': _('暂无数据')}), 404

        # 批量预加载
        owner_ids_set = list({r.quotation_owner_id for r in results})
        owners = {u.id: u for u in User.query.filter(User.id.in_(owner_ids_set)).all()} if owner_ids_set else {}

        project_ids = list({r.project_id for r in results})
        product_mns = list({r.product_mn for r in results if r.product_mn})

        # 批量订单量
        from app.models.pricing_order import PricingOrder, PricingOrderDetail
        order_qty_map = {}
        if project_ids and product_mns:
            order_rows = db.session.query(
                PricingOrder.project_id, PricingOrderDetail.product_mn,
                func.sum(PricingOrderDetail.quantity).label('qty')
            ).join(PricingOrder, PricingOrderDetail.pricing_order_id == PricingOrder.id
            ).filter(
                PricingOrder.project_id.in_(project_ids),
                PricingOrderDetail.product_mn.in_(product_mns)
            ).group_by(PricingOrder.project_id, PricingOrderDetail.product_mn).all()
            for row in order_rows:
                order_qty_map[(row.project_id, row.product_mn)] = int(row.qty or 0)

        # 批量关联客户
        from app.models.project_customer_association import ProjectCustomerAssociation
        from app.models.customer import Company
        customer_map = {}
        if project_ids:
            assoc_rows = db.session.query(
                ProjectCustomerAssociation.project_id, Company.company_name, Company.company_type
            ).join(Company, ProjectCustomerAssociation.company_id == Company.id
            ).filter(
                ProjectCustomerAssociation.project_id.in_(project_ids),
                Company.company_type.in_(['distributor', 'dealer']),
                Company.is_deleted == False
            ).all()
            proj_customers = defaultdict(lambda: {'distributor': [], 'dealer': []})
            for row in assoc_rows:
                proj_customers[row.project_id][row.company_type].append(row.company_name)
            for pid, types in proj_customers.items():
                names = types['distributor'] + types['dealer']
                customer_map[pid] = ' / '.join(names)

        # 生成 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = _('产品明细')

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        normal_font = Font(name='微软雅黑', size=10)
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [_('产品'), _('型号'), _('MN编码'), _('设计数量'), _('订单量'),
                   _('单价'), _('金额'), _('项目'), _('关联客户'), _('阶段'),
                   _('负责人'), _('日期')]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, r in enumerate(results, 2):
            owner = owners.get(r.quotation_owner_id)
            order_qty = order_qty_map.get((r.project_id, r.product_mn), 0)
            row_data = [
                r.product_name or '',
                r.product_model or '',
                r.product_mn or '',
                int(r.quantity or 0),
                order_qty if order_qty else '',
                float(r.unit_price or 0),
                float(r.total_price or 0),
                r.project_name or '',
                customer_map.get(r.project_id, ''),
                get_stage_label(r.current_stage) if r.current_stage else '',
                owner.real_name if owner else '',
                r.created_at.strftime('%Y-%m-%d') if r.created_at else ''
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                if col_idx in (4, 5, 6, 7):
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align

        # 列宽
        col_widths = [18, 14, 14, 10, 10, 12, 14, 24, 24, 10, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        year = int(request.args.get('year', datetime.now().year))
        year_label = _('全部') if year == 0 else str(year)
        filename = f'product_detail_{year_label}_{datetime.now().strftime("%m%d")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"v2 export 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500